import tkinter as tk
import statistics as stat
import numpy as np
from tkinter import filedialog as fd
import openpyxl as op
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
import math 

def my_round(n, ndigits):
    part = n * 10 ** ndigits
    delta = part - int(part)
    # always my_round "away from 0"
    if delta >= 0.5 or -0.5 < delta <= 0:
        part = math.ceil(part)
    else:
        part = math.floor(part)
    return part / (10 ** ndigits)


def open_dialog():
    file_name = fd.askopenfilename()
    f = open(file_name)
    stroka_path = str(f).split('\'')
    global txt
    txt = stroka_path[1]
    txt = txt.replace('/', '\\')
    global wb
    wb = op.load_workbook(filename = txt)
    global ws
    ws = wb.active   

#справка окна root
def spravka_open_root():
	toplevel_spravka_root.deiconify()
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		toplevel_spravka_root.geometry(f"401x331+{(root.winfo_x())+910}+{(root.winfo_y())}")
	else:
		toplevel_spravka_root.geometry(f"401x730+{(root.winfo_x())+1310}+{(root.winfo_y())}")

def create_input():
	open_dialog()

	#QCA
	ws['C1'].value = 'QCA1'
	if check_var_gr_2.get() >= 2:
		ws['D1'].value = 'QCA2'
	if check_var_gr_2.get() >= 3:
		ws['E1'].value = 'QCA3'
	if check_var_gr_2.get() >= 4:
		ws['F1'].value = 'QCA4'
	if check_var_gr_2.get() >= 5:
		ws['G1'].value = 'QCA5'
	if check_var_gr_2.get() >= 6:
		ws['H1'].value = 'QCA6'


	#НУМЕРАЦИЯ
	ws['B2'].value = 1
	ws['B3'].value = 2
	ws['B4'].value = 3
	if check_var_gr_3.get() >= 4:
		ws['B5'].value = 4
	if check_var_gr_3.get() >= 5:
		ws['B6'].value = 5

	

	#QCB
	if check_var_gr_1.get() >= 2:

		ws['C8'].value = 'QCB1'
		if check_var_gr_2.get() >= 2:
			ws['D8'].value = 'QCB2'
		if check_var_gr_2.get() >= 3:
			ws['E8'].value = 'QCB3'
		if check_var_gr_2.get() >= 4:
			ws['F8'].value = 'QCB4'
		if check_var_gr_2.get() >= 5:
			ws['G8'].value = 'QCB5'
		if check_var_gr_2.get() >= 6:
			ws['H8'].value = 'QCB6'

			#НУМЕРАЦИЯ
		ws['B9'].value = 1
		ws['B10'].value = 2
		ws['B11'].value = 3
		if check_var_gr_3.get() >= 4:
			ws['B12'].value = 4
		if check_var_gr_3.get() >= 5:
			ws['B13'].value = 5


	#QCС
	if check_var_gr_1.get() >= 3:

		ws['C15'].value = 'QCС1'
		if check_var_gr_2.get() >= 2:
			ws['D15'].value = 'QCС2'
		if check_var_gr_2.get() >= 3:
			ws['E15'].value = 'QCС3'
		if check_var_gr_2.get() >= 4:
			ws['F15'].value = 'QCС4'
		if check_var_gr_2.get() >= 5:
			ws['G15'].value = 'QCС5'
		if check_var_gr_2.get() >= 6:
			ws['H15'].value = 'QCС6'

			#НУМЕРАЦИЯ
		ws['B16'].value = 1
		ws['B17'].value = 2
		ws['B18'].value = 3
		if check_var_gr_3.get() >= 4:
			ws['B19'].value = 4
		if check_var_gr_3.get() >= 5:
			ws['B20'].value = 5


	#QCD
	if check_var_gr_1.get() >= 4:

		ws['C22'].value = 'QCD1'
		if check_var_gr_2.get() >= 2:
			ws['D22'].value = 'QCD2'
		if check_var_gr_2.get() >= 3:
			ws['E22'].value = 'QCD3'
		if check_var_gr_2.get() >= 4:
			ws['F22'].value = 'QCD4'
		if check_var_gr_2.get() >= 5:
			ws['G22'].value = 'QCD5'
		if check_var_gr_2.get() >= 6:
			ws['H22'].value = 'QCD6'

			#НУМЕРАЦИЯ  
		ws['B23'].value = 1
		ws['B24'].value = 2
		ws['B25'].value = 3
		if check_var_gr_3.get() >= 4:
			ws['B26'].value = 4
		if check_var_gr_3.get() >= 5:
			ws['B27'].value = 5

	#QCE
	if check_var_gr_1.get() >= 5:

		ws['C29'].value = 'QCE1'
		if check_var_gr_2.get() >= 2:
			ws['D29'].value = 'QCE2'
		if check_var_gr_2.get() >= 3:
			ws['E29'].value = 'QCE3'
		if check_var_gr_2.get() >= 4:
			ws['F29'].value = 'QCE4'
		if check_var_gr_2.get() >= 5:
			ws['G29'].value = 'QCE5'
		if check_var_gr_2.get() >= 6:
			ws['H29'].value = 'QCE6'

			#НУМЕРАЦИЯ
		ws['B30'].value = 1
		ws['B31'].value = 2
		ws['B32'].value = 3
		if check_var_gr_3.get() >= 4:
			ws['B33'].value = 4
		if check_var_gr_3.get() >= 5:
			ws['B34'].value = 5


	#QCF
	if check_var_gr_1.get() >= 6:

		ws['C36'].value = 'QCF1'
		if check_var_gr_2.get() >= 2:
			ws['D36'].value = 'QCF2'
		if check_var_gr_2.get() >= 3:
			ws['E36'].value = 'QCF3'
		if check_var_gr_2.get() >= 4:
			ws['F36'].value = 'QCF4'
		if check_var_gr_2.get() >= 5:
			ws['G36'].value = 'QCF5'
		if check_var_gr_2.get() >= 6:
			ws['H36'].value = 'QCF6'

			#НУМЕРАЦИЯ
		ws['B37'].value = 1
		ws['B38'].value = 2
		ws['B39'].value = 3
		if check_var_gr_3.get() >= 4:
			ws['B40'].value = 4
		if check_var_gr_3.get() >= 5:
			ws['B41'].value = 5


	thin_border(ws, 'A1:A1')
	thin_border(ws, 'C2:E6')
	if check_var_gr_1.get() >= 2:
		thin_border(ws, 'A8:A8')
		thin_border(ws, 'C9:E13')
	if check_var_gr_1.get() >= 3:
		thin_border(ws, 'A15:A15')
		thin_border(ws, 'C16:E20')
	if check_var_gr_1.get() >= 4:
		thin_border(ws, 'C23:E27')
		thin_border(ws, 'A22:A22')
	

	cols_c(ws, 'B2:B80')

	wb.save(txt)		


def input_this():
	open_dialog()
	A_column = ws['A']
	all_A_column_values = []

	for i in A_column[0:]:		
		all_A_column_values.append(i.value)


	n_kletki = []
	spisok_conc = []
	schet = 1
	#цикл определяет количество None в списке значений и сразу вносит значения концентраций в отдельный список
	for i in all_A_column_values:
		if i == None:
			schet += 1
		else:
			spisok_conc.append(i)
			n_kletki.append(schet)
			schet += 1


	#QCA
	QCA_val = []
	#QCA1
	try:
		QCA1_val = ws['C{}:C{}'.format(n_kletki[0]+1, n_kletki[1]-2)]
	except:
		QCA1_val = ws['C{}:C{}'.format(n_kletki[0]+1, n_kletki[0]+10)]
	QCA1_input = []
	for i in QCA1_val:
		QCA1_input.append(i[0].value)
	QCA_val.append(QCA1_input)

	if check_var_gr_2.get() >= 2:
	#QCA2
		try:	
			QCA2_val = ws['D{}:D{}'.format(n_kletki[0]+1, n_kletki[1]-2)]
		except:
			QCA2_val = ws['D{}:D{}'.format(n_kletki[0]+1, n_kletki[0]+10)]
		QCA2_input = []
		for i in QCA2_val:
			QCA2_input.append(i[0].value)
		QCA_val.append(QCA2_input)
	if check_var_gr_2.get() >= 3:

	#QCA3
		try:	
			QCA3_val = ws['E{}:E{}'.format(n_kletki[0]+1, n_kletki[1]-2)]
		except:
			QCA3_val = ws['E{}:E{}'.format(n_kletki[0]+1, n_kletki[0]+10)]
		QCA3_input = []
		for i in QCA3_val:
			QCA3_input.append(i[0].value)
		QCA_val.append(QCA3_input)
	if check_var_gr_2.get() >= 4:
	#QCA4
		try:
			QCA4_val = ws['F{}:F{}'.format(n_kletki[0]+1, n_kletki[1]-2)]
		except:
			QCA4_val = ws['F{}:F{}'.format(n_kletki[0]+1, n_kletki[0])+10]
		QCA4_input = []
		for i in QCA4_val:
			QCA4_input.append(i[0].value)	
		QCA_val.append(QCA4_input)
	if check_var_gr_2.get() >= 5:
	#QCA5
		try:
			QCA5_val = ws['G{}:G{}'.format(n_kletki[0]+1, n_kletki[1]-2)]
		except:
			QCA5_val = ws['G{}:G{}'.format(n_kletki[0]+1, n_kletki[0]+10)]
		QCA5_input = []
		for i in QCA5_val:
			QCA5_input.append(i[0].value)	
		QCA_val.append(QCA5_input)
	if check_var_gr_2.get() >= 6:
	#QCA6
		try:
			QCA6_val = ws['H{}:H{}'.format(n_kletki[0]+1, n_kletki[1]-2)]
		except:
			QCA6_val = ws['H{}:H{}'.format(n_kletki[0]+1, n_kletki[0]+10)]
		QCA6_input = []
		for i in QCA6_val:
			QCA6_input.append(i[0].value)	
		QCA_val.append(QCA6_input)

	

	if check_var_gr_1.get() >= 2:
		#QCB
		QCB_val = []
		#QCB1
		try:
			QCB1_val = ws['C{}:C{}'.format(n_kletki[1]+1, n_kletki[2]-2)]
		except:
			QCB1_val = ws['C{}:C{}'.format(n_kletki[1]+1, n_kletki[1]+10)]

		QCB1_input = []
		for i in QCB1_val:
			QCB1_input.append(i[0].value)
		QCB_val.append(QCB1_input)

		if check_var_gr_2.get() >= 2:
		#QCB2
			try:
				QCB2_val = ws['D{}:D{}'.format(n_kletki[1]+1, n_kletki[2]-2)]
			except:
				QCB2_val = ws['D{}:D{}'.format(n_kletki[1]+1, n_kletki[1]+10)]

			QCB2_input = []
			for i in QCB2_val:
				QCB2_input.append(i[0].value)
			QCB_val.append(QCB2_input)

		if check_var_gr_2.get() >= 3:
		#QCB3	
			try:
				QCB3_val = ws['E{}:E{}'.format(n_kletki[1]+1, n_kletki[2]-2)]
			except:
				QCB3_val = ws['E{}:E{}'.format(n_kletki[1]+1, n_kletki[1]+10)]

			QCB3_input = []
			for i in QCB3_val:
				QCB3_input.append(i[0].value)
			QCB_val.append(QCB3_input)

		if check_var_gr_2.get() >= 4:
		#QCB4
			try:
				QCB4_val = ws['F{}:F{}'.format(n_kletki[1]+1, n_kletki[2]-2)]
			except:
				QCB4_val = ws['F{}:F{}'.format(n_kletki[1]+1, n_kletki[1]+10)]

			QCB4_input = []
			for i in QCB4_val:
				QCB4_input.append(i[0].value)	
			QCB_val.append(QCB4_input)

		if check_var_gr_2.get() >= 5:
		#QCB5
			try:
				QCB5_val = ws['G{}:G{}'.format(n_kletki[1]+1, n_kletki[2]-2)]
			except:
				QCB5_val = ws['G{}:G{}'.format(n_kletki[1]+1, n_kletki[1]+10)]

			QCB5_input = []
			for i in QCB5_val:
				QCB5_input.append(i[0].value)	
			QCB_val.append(QCB5_input)
		if check_var_gr_2.get() >= 6:
		#QCB6
			try:
				QCB6_val = ws['H{}:H{}'.format(n_kletki[1]+1, n_kletki[2]-2)]
			except:
				QCB6_val = ws['H{}:H{}'.format(n_kletki[1]+1, n_kletki[1]+10)]
			QCB6_input = []
			for i in QCB6_val:
				QCB6_input.append(i[0].value)	
			QCB_val.append(QCB6_input)


	if check_var_gr_1.get() >= 3:
		#QCC
		QCC_val = []
		#QCC1
		try:
			QCC1_val = ws['C{}:C{}'.format(n_kletki[2]+1, n_kletki[3]-2)]
		except:
			QCC1_val = ws['C{}:C{}'.format(n_kletki[2]+1, n_kletki[2]+10)]
		QCC1_input = []
		for i in QCC1_val:
			QCC1_input.append(i[0].value)
		QCC_val.append(QCC1_input)

		if check_var_gr_2.get() >= 2:
		#QCC2	
			try:
				QCC2_val = ws['D{}:D{}'.format(n_kletki[2]+1, n_kletki[3]-2)]
			except:
				QCC2_val = ws['D{}:D{}'.format(n_kletki[2]+1, n_kletki[2]+10)]
			QCC2_input = []
			for i in QCC2_val:
				QCC2_input.append(i[0].value)
			QCC_val.append(QCC2_input)
		if check_var_gr_2.get() >= 3:

		#QCC3	
			try:
				QCC3_val = ws['E{}:E{}'.format(n_kletki[2]+1, n_kletki[3]-2)]
			except:
				QCC3_val = ws['E{}:E{}'.format(n_kletki[2]+1, n_kletki[2]+10)]
			QCC3_input = []
			for i in QCC3_val:
				QCC3_input.append(i[0].value)
			QCC_val.append(QCC3_input)

		if check_var_gr_2.get() >= 4:
		#QCC4
			try:
				QCC4_val = ws['F{}:F{}'.format(n_kletki[2]+1, n_kletki[3]-2)]
			except:
				QCC4_val = ws['F{}:F{}'.format(n_kletki[2]+1, n_kletki[2]+10)]
			QCC4_input = []
			for i in QCC4_val:
				QCC4_input.append(i[0].value)	
			QCC_val.append(QCC4_input)

		if check_var_gr_2.get() >= 5:
		#QCC5
			try:
				QCC5_val = ws['G{}:G{}'.format(n_kletki[2]+1, n_kletki[3]-2)]
			except:
				QCC5_val = ws['G{}:G{}'.format(n_kletki[2]+1, n_kletki[2]+10)]

			QCC5_input = []
			for i in QCC5_val:
				QCC5_input.append(i[0].value)	
			QCC_val.append(QCC5_input)

		if check_var_gr_2.get() >= 6:
		#QCC6
			try:
				QCC6_val = ws['H{}:H{}'.format(n_kletki[2]+1, n_kletki[3]-2)]
			except:
				QCC6_val = ws['H{}:H{}'.format(n_kletki[2]+1, n_kletki[2]+10)]

			QCC6_input = []
			for i in QCC6_val:
				QCC6_input.append(i[0].value)	
			QCC_val.append(QCC6_input)



	if check_var_gr_1.get() >= 4:
		#QCD
		QCD_val = []
		#QCD1
		try:
			QCD1_val = ws['C{}:C{}'.format(n_kletki[3]+1, n_kletki[4]-2)]
		except:
			QCD1_val = ws['C{}:C{}'.format(n_kletki[3]+1, n_kletki[3]+10)]
		QCD1_input = []
		for i in QCD1_val:
			QCD1_input.append(i[0].value)
		QCD_val.append(QCD1_input)

		if check_var_gr_2.get() >= 2:
		#QCD2
			try:
				QCD2_val = ws['D{}:D{}'.format(n_kletki[3]+1, n_kletki[4]-2)]
			except:
				QCD2_val = ws['D{}:D{}'.format(n_kletki[3]+1, n_kletki[3]+10)]

			QCD2_input = []
			for i in QCD2_val:
				QCD2_input.append(i[0].value)
			QCD_val.append(QCD2_input)

		if check_var_gr_2.get() >= 3:
		#QCD3
			try:
				QCD3_val = ws['E{}:E{}'.format(n_kletki[3]+1, n_kletki[4]-2)]
			except:
				QCD3_val = ws['E{}:E{}'.format(n_kletki[3]+1, n_kletki[3]+10)]
			QCD3_input = []
			for i in QCD3_val:
				QCD3_input.append(i[0].value)
			QCD_val.append(QCD3_input)

		if check_var_gr_2.get() >= 4:
		#QCD4
			try:
				QCD4_val = ws['F{}:F{}'.format(n_kletki[3]+1, n_kletki[4]-2)]
			except:
				QCD4_val = ws['F{}:F{}'.format(n_kletki[3]+1, n_kletki[3]+10)]

			QCD4_input = []
			for i in QCD4_val:
				QCD4_input.append(i[0].value)	
			QCD_val.append(QCD4_input)

		if check_var_gr_2.get() >= 5:
		#QCD5
			try:
				QCD5_val = ws['G{}:G{}'.format(n_kletki[3]+1, n_kletki[4]-2)]
			except:
				QCD5_val = ws['G{}:G{}'.format(n_kletki[3]+1, n_kletki[3]+10)]

			QCD5_input = []
			for i in QCD5_val:
				QCD5_input.append(i[0].value)	
			QCD_val.append(QCD5_input)

		if check_var_gr_2.get() >= 6:
		#QCD6
			try:
				QCD6_val = ws['H{}:H{}'.format(n_kletki[3]+1, n_kletki[4]-2)]
			except:
				QCD6_val = ws['H{}:H{}'.format(n_kletki[3]+1, n_kletki[3]+10)]
			QCD6_input = []
			for i in QCD6_val:
				QCD6_input.append(i[0].value)	
			QCD_val.append(QCD6_input)


	if check_var_gr_1.get() >= 5:
		#QCE
		QCE_val = []
		#QCE1
		try:
			QCE1_val = ws['C{}:C{}'.format(n_kletki[4]+1, n_kletki[5]-2)]
		except:
			QCE1_val = ws['C{}:C{}'.format(n_kletki[4]+1, n_kletki[4]+10)]
		QCE1_input = []
		for i in QCE1_val:
			QCE1_input.append(i[0].value)
		QCE_val.append(QCE1_input)

		if check_var_gr_2.get() >= 2:
		#QCE2	
			try:
				QCE2_val = ws['D{}:D{}'.format(n_kletki[4]+1, n_kletki[5]-2)]
			except:
				QCE2_val = ws['D{}:D{}'.format(n_kletki[4]+1, n_kletki[4]+10)]
			QCE2_input = []
			for i in QCE2_val:
				QCE2_input.append(i[0].value)
			QCE_val.append(QCE2_input)

		if check_var_gr_2.get() >= 3:
		#QCE3
			try:
				QCE3_val = ws['E{}:E{}'.format(n_kletki[4]+1, n_kletki[5]-2)]
			except:
				QCE3_val = ws['E{}:E{}'.format(n_kletki[4]+1, n_kletki[4]+10)]
			QCE3_input = []
			for i in QCE3_val:
				QCE3_input.append(i[0].value)
			QCE_val.append(QCE3_input)

		if check_var_gr_2.get() >= 4:
		#QCE4
			try:
				QCE4_val = ws['F{}:F{}'.format(n_kletki[4]+1, n_kletki[5]-2)]
			except:
				QCE4_val = ws['F{}:F{}'.format(n_kletki[4]+1, n_kletki[4]+10)]
			QCE4_input = []
			for i in QCE4_val:
				QCE4_input.append(i[0].value)	
			QCE_val.append(QCE4_input)

		if check_var_gr_2.get() >= 5:
		#QCE5
			try:
				QCE5_val = ws['G{}:G{}'.format(n_kletki[4]+1, n_kletki[5]-2)]
			except:
				QCE5_val = ws['G{}:G{}'.format(n_kletki[4]+1, n_kletki[4]+10)]
			QCE5_input = []
			for i in QCE5_val:
				QCE5_input.append(i[0].value)	
			QCE_val.append(QCE5_input)

		if check_var_gr_2.get() >= 6:
		#QCE6
			try:
				QCE6_val = ws['H{}:H{}'.format(n_kletki[4]+1, n_kletki[5]-2)]
			except:
				QCE6_val = ws['H{}:H{}'.format(n_kletki[4]+1, n_kletki[4]+10)]
			QCE6_input = []
			for i in QCE6_val:
				QCE6_input.append(i[0].value)	
			QCE_val.append(QCE6_input)


	if check_var_gr_1.get() >= 6:
		#QCF
		QCF_val = []
		#QCF1
		QCF1_val = ws['C{}:C{}'.format(n_kletki[5]+1, n_kletki[5]+11)]
		QCF1_input = []
		for i in QCF1_val:
			QCF1_input.append(i[0].value)
		QCF_val.append(QCF1_input)

		if check_var_gr_2.get() >= 2:
		#QCF2	
			QCF2_val = ws['D{}:D{}'.format(n_kletki[5]+1, n_kletki[5]+11)]
			QCF2_input = []
			for i in QCF2_val:
				QCF2_input.append(i[0].value)
			QCF_val.append(QCF2_input)
		if check_var_gr_2.get() >= 3:

		#QCF3	
			QCF3_val = ws['E{}:E{}'.format(n_kletki[5]+1, n_kletki[5]+11)]
			QCF3_input = []
			for i in QCF3_val:
				QCF3_input.append(i[0].value)
			QCF_val.append(QCF3_input)
		if check_var_gr_2.get() >= 4:
		#QCF4
			QCF4_val = ws['F{}:F{}'.format(n_kletki[5]+1, n_kletki[5]+11)]
			QCF4_input = []
			for i in QCF4_val:
				QCF4_input.append(i[0].value)	
			QCF_val.append(QCF4_input)
		if check_var_gr_2.get() >= 5:
		#QCF5
			QCF5_val = ws['G{}:G{}'.format(n_kletki[5]+1, n_kletki[5]+11)]
			QCF5_input = []
			for i in QCF5_val:
				QCF5_input.append(i[0].value)	
			QCF_val.append(QCF5_input)
		if check_var_gr_2.get() >= 6:
		#QCF6
			QCF6_val = ws['H{}:H{}'.format(n_kletki[5]+1, n_kletki[5]+11)]
			QCF6_input = []
			for i in QCF6_val:
				QCF6_input.append(i[0].value)	
			QCF_val.append(QCF6_input)	




	#QCA

	#QCA1
	entr_MAIN_I.delete(0, "end")
	entr_MAIN_I.insert(tk.END, '{}'.format(spisok_conc[0]))

	entr_A1_I.delete(0, "end")
	entr_A1_I.insert(tk.END, '{}'.format(QCA_val[0][0]))

	entr_A2_I.delete(0, "end")
	entr_A2_I.insert(tk.END, '{}'.format(QCA_val[0][1]))

	entr_A3_I.delete(0, "end")
	entr_A3_I.insert(tk.END, '{}'.format(QCA_val[0][2]))
	if check_var_gr_3.get() >= 4:
		entr_A4_I.delete(0, "end")
		entr_A4_I.insert(tk.END, '{}'.format(QCA_val[0][3]))
	if check_var_gr_3.get() >= 5:	
		entr_A5_I.delete(0, "end")
		entr_A5_I.insert(tk.END, '{}'.format(QCA_val[0][4]))
	if check_var_gr_3.get() >= 6:
		entr_A6_I.delete(0, "end")
		entr_A6_I.insert(tk.END, '{}'.format(QCA_val[0][5]))
	if check_var_gr_3.get() >= 7:
		entr_A7_I.delete(0, "end")
		entr_A7_I.insert(tk.END, '{}'.format(QCA_val[0][6]))
	if check_var_gr_3.get() >= 8:
		entr_A8_I.delete(0, "end")
		entr_A8_I.insert(tk.END, '{}'.format(QCA_val[0][7]))
	if check_var_gr_3.get() >= 9:
		entr_A9_I.delete(0, "end")
		entr_A9_I.insert(tk.END, '{}'.format(QCA_val[0][8]))
	if check_var_gr_3.get() >= 10:	
		entr_A10_I.delete(0, "end")
		entr_A10_I.insert(tk.END, '{}'.format(QCA_val[0][9]))


	#QCA2
	if check_var_gr_2.get() >= 2:
		entr_B1_I.delete(0, "end")
		entr_B1_I.insert(tk.END, '{}'.format(QCA_val[1][0]))
		entr_B2_I.delete(0, "end")
		entr_B2_I.insert(tk.END, '{}'.format(QCA_val[1][1]))
		entr_B3_I.delete(0, "end")
		entr_B3_I.insert(tk.END, '{}'.format(QCA_val[1][2]))
		if check_var_gr_3.get() >= 4:
			entr_B4_I.delete(0, "end")
			entr_B4_I.insert(tk.END, '{}'.format(QCA_val[1][3]))
		if check_var_gr_3.get() >= 5:	
			entr_B5_I.delete(0, "end")
			entr_B5_I.insert(tk.END, '{}'.format(QCA_val[1][4]))
		if check_var_gr_3.get() >= 6:
			entr_B6_I.delete(0, "end")
			entr_B6_I.insert(tk.END, '{}'.format(QCA_val[1][5]))
		if check_var_gr_3.get() >= 7:
			entr_B7_I.delete(0, "end")
			entr_B7_I.insert(tk.END, '{}'.format(QCA_val[1][6]))
		if check_var_gr_3.get() >= 8:
			entr_B8_I.delete(0, "end")
			entr_B8_I.insert(tk.END, '{}'.format(QCA_val[1][7]))
		if check_var_gr_3.get() >= 9:
			entr_B9_I.delete(0, "end")
			entr_B9_I.insert(tk.END, '{}'.format(QCA_val[1][8]))
		if check_var_gr_3.get() >= 10:
			entr_B10_I.delete(0, "end")
			entr_B10_I.insert(tk.END, '{}'.format(QCA_val[1][9]))


	#QCA3
	if check_var_gr_2.get() >= 3:
		entr_C1_I.delete(0, "end")
		entr_C1_I.insert(tk.END, '{}'.format(QCA_val[2][0]))
		entr_C2_I.delete(0, "end")
		entr_C2_I.insert(tk.END, '{}'.format(QCA_val[2][1]))
		entr_C3_I.delete(0, "end")
		entr_C3_I.insert(tk.END, '{}'.format(QCA_val[2][2]))
		if check_var_gr_3.get() >= 4:
			entr_C4_I.delete(0, "end")
			entr_C4_I.insert(tk.END, '{}'.format(QCA_val[2][3]))
		if check_var_gr_3.get() >= 5:
			entr_C5_I.delete(0, "end")	
			entr_C5_I.insert(tk.END, '{}'.format(QCA_val[2][4]))
		if check_var_gr_3.get() >= 6:
			entr_C6_I.delete(0, "end")
			entr_C6_I.insert(tk.END, '{}'.format(QCA_val[2][5]))
		if check_var_gr_3.get() >= 7:
			entr_C7_I.delete(0, "end")
			entr_C7_I.insert(tk.END, '{}'.format(QCA_val[2][6]))
		if check_var_gr_3.get() >= 8:
			entr_C8_I.delete(0, "end")
			entr_C8_I.insert(tk.END, '{}'.format(QCA_val[2][7]))
		if check_var_gr_3.get() >= 9:
			entr_C9_I.delete(0, "end")
			entr_C9_I.insert(tk.END, '{}'.format(QCA_val[2][8]))
		if check_var_gr_3.get() >= 10:
			entr_C10_I.delete(0, "end")
			entr_C10_I.insert(tk.END, '{}'.format(QCA_val[2][9]))



	#QCA4
	if check_var_gr_2.get() >= 4:
		entr_D1_I.delete(0, "end")
		entr_D1_I.insert(tk.END, '{}'.format(QCA_val[3][0]))
		entr_D2_I.delete(0, "end")
		entr_D2_I.insert(tk.END, '{}'.format(QCA_val[3][1]))
		entr_D3_I.delete(0, "end")
		entr_D3_I.insert(tk.END, '{}'.format(QCA_val[3][2]))
		if check_var_gr_3.get() >= 4:
			entr_D4_I.delete(0, "end")
			entr_D4_I.insert(tk.END, '{}'.format(QCA_val[3][3]))
		if check_var_gr_3.get() >= 5:	
			entr_D5_I.delete(0, "end")
			entr_D5_I.insert(tk.END, '{}'.format(QCA_val[3][4]))
		if check_var_gr_3.get() >= 6:
			entr_D6_I.delete(0, "end")
			entr_D6_I.insert(tk.END, '{}'.format(QCA_val[3][5]))
		if check_var_gr_3.get() >= 7:
			entr_D7_I.delete(0, "end")
			entr_D7_I.insert(tk.END, '{}'.format(QCA_val[3][6]))
		if check_var_gr_3.get() >= 8:
			entr_D8_I.delete(0, "end")
			entr_D8_I.insert(tk.END, '{}'.format(QCA_val[3][7]))
		if check_var_gr_3.get() >= 9:
			entr_D9_I.delete(0, "end")
			entr_D9_I.insert(tk.END, '{}'.format(QCA_val[3][8]))
		if check_var_gr_3.get() >= 10:
			entr_D10_I.delete(0, "end")
			entr_D10_I.insert(tk.END, '{}'.format(QCA_val[3][9]))


	#QCA5
	if check_var_gr_2.get() >= 5:
		entr_E1_I.delete(0, "end")
		entr_E1_I.insert(tk.END, '{}'.format(QCA_val[4][0]))
		entr_E2_I.delete(0, "end")
		entr_E2_I.insert(tk.END, '{}'.format(QCA_val[4][1]))
		entr_E3_I.delete(0, "end")
		entr_E3_I.insert(tk.END, '{}'.format(QCA_val[4][2]))
		if check_var_gr_3.get() >= 4:
			entr_E4_I.delete(0, "end")
			entr_E4_I.insert(tk.END, '{}'.format(QCA_val[4][3]))
		if check_var_gr_3.get() >= 5:
			entr_E5_I.delete(0, "end")	
			entr_E5_I.insert(tk.END, '{}'.format(QCA_val[4][4]))
		if check_var_gr_3.get() >= 6:
			entr_E6_I.delete(0, "end")
			entr_E6_I.insert(tk.END, '{}'.format(QCA_val[4][5]))
		if check_var_gr_3.get() >= 7:
			entr_E7_I.delete(0, "end")
			entr_E7_I.insert(tk.END, '{}'.format(QCA_val[4][6]))
		if check_var_gr_3.get() >= 8:
			entr_E8_I.delete(0, "end")
			entr_E8_I.insert(tk.END, '{}'.format(QCA_val[4][7]))
		if check_var_gr_3.get() >= 9:
			entr_E9_I.delete(0, "end")
			entr_E9_I.insert(tk.END, '{}'.format(QCA_val[4][8]))
		if check_var_gr_3.get() >= 10:
			entr_E10_I.delete(0, "end")
			entr_E10_I.insert(tk.END, '{}'.format(QCA_val[4][9]))


	#QCA6
	if check_var_gr_2.get() >= 6:
		entr_F1_I.delete(0, "end")
		entr_F1_I.insert(tk.END, '{}'.format(QCA_val[5][0]))
		entr_F2_I.delete(0, "end")
		entr_F2_I.insert(tk.END, '{}'.format(QCA_val[5][1]))
		entr_F3_I.delete(0, "end")
		entr_F3_I.insert(tk.END, '{}'.format(QCA_val[5][2]))
		if check_var_gr_3.get() >= 4:
			entr_F4_I.delete(0, "end")
			entr_F4_I.insert(tk.END, '{}'.format(QCA_val[5][3]))
		if check_var_gr_3.get() >= 5:
			entr_F5_I.delete(0, "end")	
			entr_F5_I.insert(tk.END, '{}'.format(QCA_val[5][4]))
		if check_var_gr_3.get() >= 6:
			entr_F6_I.delete(0, "end")
			entr_F6_I.insert(tk.END, '{}'.format(QCA_val[5][5]))
		if check_var_gr_3.get() >= 7:
			entr_F7_I.delete(0, "end")
			entr_F7_I.insert(tk.END, '{}'.format(QCA_val[5][6]))
		if check_var_gr_3.get() >= 8:
			entr_F8_I.delete(0, "end")
			entr_F8_I.insert(tk.END, '{}'.format(QCA_val[5][7]))
		if check_var_gr_3.get() >= 9:
			entr_F9_I.delete(0, "end")
			entr_F9_I.insert(tk.END, '{}'.format(QCA_val[5][8]))
		if check_var_gr_3.get() >= 10:
			entr_F10_I.delete(0, "end")
			entr_F10_I.insert(tk.END, '{}'.format(QCA_val[5][9]))


	#QCB

	if check_var_gr_1.get() >= 2:
	#QCB1
		entr_MAIN_II.delete(0, "end")
		entr_MAIN_II.insert(tk.END, '{}'.format(spisok_conc[1]))
		entr_A1_II.delete(0, "end")
		entr_A1_II.insert(tk.END, '{}'.format(QCB_val[0][0]))
		entr_A2_II.delete(0, "end")
		entr_A2_II.insert(tk.END, '{}'.format(QCB_val[0][1]))
		entr_A3_II.delete(0, "end")
		entr_A3_II.insert(tk.END, '{}'.format(QCB_val[0][2]))
		if check_var_gr_3.get() >= 4:
			entr_A4_II.delete(0, "end")
			entr_A4_II.insert(tk.END, '{}'.format(QCB_val[0][3]))
		if check_var_gr_3.get() >= 5:
			entr_A5_II.delete(0, "end")	
			entr_A5_II.insert(tk.END, '{}'.format(QCB_val[0][4]))
		if check_var_gr_3.get() >= 6:
			entr_A6_II.delete(0, "end")
			entr_A6_II.insert(tk.END, '{}'.format(QCB_val[0][5]))
		if check_var_gr_3.get() >= 7:
			entr_A7_II.delete(0, "end")
			entr_A7_II.insert(tk.END, '{}'.format(QCB_val[0][6]))
		if check_var_gr_3.get() >= 8:
			entr_A8_II.delete(0, "end")
			entr_A8_II.insert(tk.END, '{}'.format(QCB_val[0][7]))
		if check_var_gr_3.get() >= 9:
			entr_A9_II.delete(0, "end")
			entr_A9_II.insert(tk.END, '{}'.format(QCB_val[0][8]))
		if check_var_gr_3.get() >= 10:
			entr_A10_II.delete(0, "end")		
			entr_A10_II.insert(tk.END, '{}'.format(QCB_val[0][9]))


		#QCB2
		if check_var_gr_2.get() >= 2:
			entr_B1_II.delete(0, "end")
			entr_B1_II.insert(tk.END, '{}'.format(QCB_val[1][0]))
			entr_B2_II.delete(0, "end")
			entr_B2_II.insert(tk.END, '{}'.format(QCB_val[1][1]))
			entr_B3_II.delete(0, "end")
			entr_B3_II.insert(tk.END, '{}'.format(QCB_val[1][2]))
			if check_var_gr_3.get() >= 4:
				entr_B4_II.delete(0, "end")
				entr_B4_II.insert(tk.END, '{}'.format(QCB_val[1][3]))
			if check_var_gr_3.get() >= 5:	
				entr_B5_II.delete(0, "end")
				entr_B5_II.insert(tk.END, '{}'.format(QCB_val[1][4]))
			if check_var_gr_3.get() >= 6:
				entr_B6_II.delete(0, "end")
				entr_B6_II.insert(tk.END, '{}'.format(QCB_val[1][5]))
			if check_var_gr_3.get() >= 7:
				entr_B7_II.delete(0, "end")
				entr_B7_II.insert(tk.END, '{}'.format(QCB_val[1][6]))
			if check_var_gr_3.get() >= 8:
				entr_B8_II.delete(0, "end")
				entr_B8_II.insert(tk.END, '{}'.format(QCB_val[1][7]))
			if check_var_gr_3.get() >= 9:
				entr_B9_II.delete(0, "end")
				entr_B9_II.insert(tk.END, '{}'.format(QCB_val[1][8]))
			if check_var_gr_3.get() >= 10:
				entr_B10_II.delete(0, "end")
				entr_B10_II.insert(tk.END, '{}'.format(QCB_val[1][9]))

		#QCB3
		if check_var_gr_2.get() >= 3:
			entr_C1_II.delete(0, "end")
			entr_C1_II.insert(tk.END, '{}'.format(QCB_val[2][0]))
			entr_C2_II.delete(0, "end")
			entr_C2_II.insert(tk.END, '{}'.format(QCB_val[2][1]))
			entr_C3_II.delete(0, "end")
			entr_C3_II.insert(tk.END, '{}'.format(QCB_val[2][2]))
			if check_var_gr_3.get() >= 4:
				entr_C4_II.delete(0, "end")
				entr_C4_II.insert(tk.END, '{}'.format(QCB_val[2][3]))
			if check_var_gr_3.get() >= 5:	
				entr_C5_II.delete(0, "end")
				entr_C5_II.insert(tk.END, '{}'.format(QCB_val[2][4]))
			if check_var_gr_3.get() >= 6:
				entr_C6_II.delete(0, "end")
				entr_C6_II.insert(tk.END, '{}'.format(QCB_val[2][5]))
			if check_var_gr_3.get() >= 7:
				entr_C7_II.delete(0, "end")
				entr_C7_II.insert(tk.END, '{}'.format(QCB_val[2][6]))
			if check_var_gr_3.get() >= 8:
				entr_C8_II.delete(0, "end")
				entr_C8_II.insert(tk.END, '{}'.format(QCB_val[2][7]))
			if check_var_gr_3.get() >= 9:
				entr_C9_II.delete(0, "end")
				entr_C9_II.insert(tk.END, '{}'.format(QCB_val[2][8]))
			if check_var_gr_3.get() >= 10:
				entr_C10_II.delete(0, "end")
				entr_C10_II.insert(tk.END, '{}'.format(QCB_val[2][9]))


			#QCB4
		if check_var_gr_2.get() >= 4:
			entr_D1_II.delete(0, "end")
			entr_D1_II.insert(tk.END, '{}'.format(QCB_val[3][0]))
			entr_D2_II.delete(0, "end")
			entr_D2_II.insert(tk.END, '{}'.format(QCB_val[3][1]))
			entr_D3_II.delete(0, "end")
			entr_D3_II.insert(tk.END, '{}'.format(QCB_val[3][2]))
			if check_var_gr_3.get() >= 4:
				entr_D4_II.delete(0, "end")
				entr_D4_II.insert(tk.END, '{}'.format(QCB_val[3][3]))
			if check_var_gr_3.get() >= 5:	
				entr_D5_II.delete(0, "end")
				entr_D5_II.insert(tk.END, '{}'.format(QCB_val[3][4]))
			if check_var_gr_3.get() >= 6:
				entr_D6_II.delete(0, "end")
				entr_D6_II.insert(tk.END, '{}'.format(QCB_val[3][5]))
			if check_var_gr_3.get() >= 7:
				entr_D7_II.delete(0, "end")
				entr_D7_II.insert(tk.END, '{}'.format(QCB_val[3][6]))
			if check_var_gr_3.get() >= 8:
				entr_D8_II.delete(0, "end")
				entr_D8_II.insert(tk.END, '{}'.format(QCB_val[3][7]))
			if check_var_gr_3.get() >= 9:
				entr_D9_II.delete(0, "end")
				entr_D9_II.insert(tk.END, '{}'.format(QCB_val[3][8]))
			if check_var_gr_3.get() >= 10:
				entr_D10_II.delete(0, "end")
				entr_D10_II.insert(tk.END, '{}'.format(QCB_val[3][9]))


			#QCB5
		if check_var_gr_2.get() >= 5:
			entr_E1_II.delete(0, "end")
			entr_E1_II.insert(tk.END, '{}'.format(QCB_val[4][0]))
			entr_E2_II.delete(0, "end")
			entr_E2_II.insert(tk.END, '{}'.format(QCB_val[4][1]))
			entr_E3_II.delete(0, "end")
			entr_E3_II.insert(tk.END, '{}'.format(QCB_val[4][2]))
			if check_var_gr_3.get() >= 4:
				entr_E4_II.delete(0, "end")
				entr_E4_II.insert(tk.END, '{}'.format(QCB_val[4][3]))
			if check_var_gr_3.get() >= 5:	
				entr_E5_II.delete(0, "end")
				entr_E5_II.insert(tk.END, '{}'.format(QCB_val[4][4]))
			if check_var_gr_3.get() >= 6:
				entr_E6_II.delete(0, "end")
				entr_E6_II.insert(tk.END, '{}'.format(QCB_val[4][5]))
			if check_var_gr_3.get() >= 7:
				entr_E7_II.delete(0, "end")
				entr_E7_II.insert(tk.END, '{}'.format(QCB_val[4][6]))
			if check_var_gr_3.get() >= 8:
				entr_E8_II.delete(0, "end")
				entr_E8_II.insert(tk.END, '{}'.format(QCB_val[4][7]))
			if check_var_gr_3.get() >= 9:
				entr_E9_II.delete(0, "end")
				entr_E9_II.insert(tk.END, '{}'.format(QCB_val[4][8]))
			if check_var_gr_3.get() >= 10:
				entr_E10_II.delete(0, "end")
				entr_E10_II.insert(tk.END, '{}'.format(QCB_val[4][9]))


		#QCB6
		if check_var_gr_2.get() >= 6:
			entr_F1_II.delete(0, "end")
			entr_F1_II.insert(tk.END, '{}'.format(QCB_val[5][0]))
			entr_F2_II.delete(0, "end")
			entr_F2_II.insert(tk.END, '{}'.format(QCB_val[5][1]))
			entr_F3_II.delete(0, "end")
			entr_F3_II.insert(tk.END, '{}'.format(QCB_val[5][2]))
			if check_var_gr_3.get() >= 4:
				entr_F4_II.delete(0, "end")
				entr_F4_II.insert(tk.END, '{}'.format(QCB_val[5][3]))
			if check_var_gr_3.get() >= 5:	
				entr_F5_II.delete(0, "end")
				entr_F5_II.insert(tk.END, '{}'.format(QCB_val[5][4]))
			if check_var_gr_3.get() >= 6:
				entr_F6_II.delete(0, "end")
				entr_F6_II.insert(tk.END, '{}'.format(QCB_val[5][5]))
			if check_var_gr_3.get() >= 7:
				entr_F7_II.delete(0, "end")
				entr_F7_II.insert(tk.END, '{}'.format(QCB_val[5][6]))
			if check_var_gr_3.get() >= 8:
				entr_F8_II.delete(0, "end")
				entr_F8_II.insert(tk.END, '{}'.format(QCB_val[5][7]))
			if check_var_gr_3.get() >= 9:
				entr_F9_II.delete(0, "end")
				entr_F9_II.insert(tk.END, '{}'.format(QCB_val[5][8]))
			if check_var_gr_3.get() >= 10:
				entr_F10_II.delete(0, "end")
				entr_F10_II.insert(tk.END, '{}'.format(QCB_val[5][9]))



		#QCC
	if check_var_gr_1.get() >= 3:
		#QCC1
		entr_MAIN_III.delete(0, "end")
		entr_MAIN_III.insert(tk.END, '{}'.format(spisok_conc[2]))
		entr_A1_III.delete(0, "end")
		entr_A1_III.insert(tk.END, '{}'.format(QCC_val[0][0]))
		entr_A2_III.delete(0, "end")
		entr_A2_III.insert(tk.END, '{}'.format(QCC_val[0][1]))
		entr_A3_III.delete(0, "end")
		entr_A3_III.insert(tk.END, '{}'.format(QCC_val[0][2]))
		if check_var_gr_3.get() >= 4:
			entr_A4_III.delete(0, "end")
			entr_A4_III.insert(tk.END, '{}'.format(QCC_val[0][3]))
		if check_var_gr_3.get() >= 5:	
			entr_A5_III.delete(0, "end")
			entr_A5_III.insert(tk.END, '{}'.format(QCC_val[0][4]))
		if check_var_gr_3.get() >= 6:
			entr_A6_III.delete(0, "end")
			entr_A6_III.insert(tk.END, '{}'.format(QCC_val[0][5]))
		if check_var_gr_3.get() >= 7:
			entr_A7_III.delete(0, "end")
			entr_A7_III.insert(tk.END, '{}'.format(QCC_val[0][6]))
		if check_var_gr_3.get() >= 8:
			entr_A8_III.delete(0, "end")
			entr_A8_III.insert(tk.END, '{}'.format(QCC_val[0][7]))
		if check_var_gr_3.get() >= 9:
			entr_A9_III.delete(0, "end")
			entr_A9_III.insert(tk.END, '{}'.format(QCC_val[0][8]))
		if check_var_gr_3.get() >= 10:	
			entr_A10_III.delete(0, "end")
			entr_A10_III.insert(tk.END, '{}'.format(QCC_val[0][9]))


			#QCC2
		if check_var_gr_2.get() >= 2:
			entr_B1_III.delete(0, "end")
			entr_B1_III.insert(tk.END, '{}'.format(QCC_val[1][0]))
			entr_B2_III.delete(0, "end")
			entr_B2_III.insert(tk.END, '{}'.format(QCC_val[1][1]))
			entr_B3_III.delete(0, "end")
			entr_B3_III.insert(tk.END, '{}'.format(QCC_val[1][2]))
			if check_var_gr_3.get() >= 4:
				entr_B4_III.delete(0, "end")
				entr_B4_III.insert(tk.END, '{}'.format(QCC_val[1][3]))
			if check_var_gr_3.get() >= 5:	
				entr_B5_III.delete(0, "end")
				entr_B5_III.insert(tk.END, '{}'.format(QCC_val[1][4]))
			if check_var_gr_3.get() >= 6:
				entr_B6_III.delete(0, "end")
				entr_B6_III.insert(tk.END, '{}'.format(QCC_val[1][5]))
			if check_var_gr_3.get() >= 7:
				entr_B7_III.delete(0, "end")
				entr_B7_III.insert(tk.END, '{}'.format(QCC_val[1][6]))
			if check_var_gr_3.get() >= 8:
				entr_B8_III.delete(0, "end")
				entr_B8_III.insert(tk.END, '{}'.format(QCC_val[1][7]))
			if check_var_gr_3.get() >= 9:
				entr_B9_III.delete(0, "end")
				entr_B9_III.insert(tk.END, '{}'.format(QCC_val[1][8]))
			if check_var_gr_3.get() >= 10:
				entr_B10_III.delete(0, "end")
				entr_B10_III.insert(tk.END, '{}'.format(QCC_val[1][9]))


			#QCC3
		if check_var_gr_2.get() >= 3:
			entr_C1_III.delete(0, "end")
			entr_C1_III.insert(tk.END, '{}'.format(QCC_val[2][0]))
			entr_C2_III.delete(0, "end")
			entr_C2_III.insert(tk.END, '{}'.format(QCC_val[2][1]))
			entr_C3_III.delete(0, "end")
			entr_C3_III.insert(tk.END, '{}'.format(QCC_val[2][2]))
			if check_var_gr_3.get() >= 4:
				entr_C4_III.delete(0, "end")
				entr_C4_III.insert(tk.END, '{}'.format(QCC_val[2][3]))
			if check_var_gr_3.get() >= 5:	
				entr_C5_III.delete(0, "end")
				entr_C5_III.insert(tk.END, '{}'.format(QCC_val[2][4]))
			if check_var_gr_3.get() >= 6:
				entr_C6_III.delete(0, "end")
				entr_C6_III.insert(tk.END, '{}'.format(QCC_val[2][5]))
			if check_var_gr_3.get() >= 7:
				entr_C7_III.delete(0, "end")
				entr_C7_III.insert(tk.END, '{}'.format(QCC_val[2][6]))
			if check_var_gr_3.get() >= 8:
				entr_C8_III.delete(0, "end")
				entr_C8_III.insert(tk.END, '{}'.format(QCC_val[2][7]))
			if check_var_gr_3.get() >= 9:
				entr_C9_III.delete(0, "end")
				entr_C9_III.insert(tk.END, '{}'.format(QCC_val[2][8]))
			if check_var_gr_3.get() >= 10:
				entr_C10_III.delete(0, "end")
				entr_C10_III.insert(tk.END, '{}'.format(QCC_val[2][9]))



			#QCC4
		if check_var_gr_2.get() >= 4:
			entr_D1_III.delete(0, "end")
			entr_D1_III.insert(tk.END, '{}'.format(QCC_val[3][0]))
			entr_D2_III.delete(0, "end")
			entr_D2_III.insert(tk.END, '{}'.format(QCC_val[3][1]))
			entr_D3_III.delete(0, "end")
			entr_D3_III.insert(tk.END, '{}'.format(QCC_val[3][2]))
			if check_var_gr_3.get() >= 4:
				entr_D4_III.delete(0, "end")
				entr_D4_III.insert(tk.END, '{}'.format(QCC_val[3][3]))
			if check_var_gr_3.get() >= 5:	
				entr_D5_III.delete(0, "end")
				entr_D5_III.insert(tk.END, '{}'.format(QCC_val[3][4]))
			if check_var_gr_3.get() >= 6:
				entr_D6_III.delete(0, "end")
				entr_D6_III.insert(tk.END, '{}'.format(QCC_val[3][5]))
			if check_var_gr_3.get() >= 7:
				entr_D7_III.delete(0, "end")
				entr_D7_III.insert(tk.END, '{}'.format(QCC_val[3][6]))
			if check_var_gr_3.get() >= 8:
				entr_D8_III.delete(0, "end")
				entr_D8_III.insert(tk.END, '{}'.format(QCC_val[3][7]))
			if check_var_gr_3.get() >= 9:
				entr_D9_III.delete(0, "end")
				entr_D9_III.insert(tk.END, '{}'.format(QCC_val[3][8]))
			if check_var_gr_3.get() >= 10:
				entr_D10_III.delete(0, "end")
				entr_D10_III.insert(tk.END, '{}'.format(QCC_val[3][9]))


			#QCC5
		if check_var_gr_2.get() >= 5:
			entr_E1_III.delete(0, "end")
			entr_E1_III.insert(tk.END, '{}'.format(QCC_val[4][0]))
			entr_E2_III.delete(0, "end")
			entr_E2_III.insert(tk.END, '{}'.format(QCC_val[4][1]))
			entr_E3_III.delete(0, "end")
			entr_E3_III.insert(tk.END, '{}'.format(QCC_val[4][2]))
			if check_var_gr_3.get() >= 4:
				entr_E4_III.delete(0, "end")
				entr_E4_III.insert(tk.END, '{}'.format(QCC_val[4][3]))
			if check_var_gr_3.get() >= 5:	
				entr_E5_III.delete(0, "end")
				entr_E5_III.insert(tk.END, '{}'.format(QCC_val[4][4]))
			if check_var_gr_3.get() >= 6:
				entr_E6_III.delete(0, "end")
				entr_E6_III.insert(tk.END, '{}'.format(QCC_val[4][5]))
			if check_var_gr_3.get() >= 7:
				entr_E7_III.delete(0, "end")
				entr_E7_III.insert(tk.END, '{}'.format(QCC_val[4][6]))
			if check_var_gr_3.get() >= 8:
				entr_E8_III.delete(0, "end")
				entr_E8_III.insert(tk.END, '{}'.format(QCC_val[4][7]))
			if check_var_gr_3.get() >= 9:
				entr_E9_III.delete(0, "end")
				entr_E9_III.insert(tk.END, '{}'.format(QCC_val[4][8]))
			if check_var_gr_3.get() >= 10:
				entr_E10_III.delete(0, "end")
				entr_E10_III.insert(tk.END, '{}'.format(QCC_val[4][9]))


			#QCC6
		if check_var_gr_2.get() >= 6:
			entr_F1_III.delete(0, "end")
			entr_F1_III.insert(tk.END, '{}'.format(QCC_val[5][0]))
			entr_F2_III.delete(0, "end")
			entr_F2_III.insert(tk.END, '{}'.format(QCC_val[5][1]))
			entr_F3_III.delete(0, "end")
			entr_F3_III.insert(tk.END, '{}'.format(QCC_val[5][2]))
			if check_var_gr_3.get() >= 4:
				entr_F4_III.delete(0, "end")
				entr_F4_III.insert(tk.END, '{}'.format(QCC_val[5][3]))
			if check_var_gr_3.get() >= 5:	
				entr_F5_III.delete(0, "end")
				entr_F5_III.insert(tk.END, '{}'.format(QCC_val[5][4]))
			if check_var_gr_3.get() >= 6:
				entr_F6_III.delete(0, "end")
				entr_F6_III.insert(tk.END, '{}'.format(QCC_val[5][5]))
			if check_var_gr_3.get() >= 7:
				entr_F7_III.delete(0, "end")
				entr_F7_III.insert(tk.END, '{}'.format(QCC_val[5][6]))
			if check_var_gr_3.get() >= 8:
				entr_F8_III.delete(0, "end")
				entr_F8_III.insert(tk.END, '{}'.format(QCC_val[5][7]))
			if check_var_gr_3.get() >= 9:
				entr_F9_III.delete(0, "end")
				entr_F9_III.insert(tk.END, '{}'.format(QCC_val[5][8]))
			if check_var_gr_3.get() >= 10:
				entr_F10_III.delete(0, "end")
				entr_F10_III.insert(tk.END, '{}'.format(QCC_val[5][9]))


	#QCD
	if check_var_gr_1.get() >= 4:
		#QCD1
		entr_MAIN_IV.delete(0, "end")
		entr_MAIN_IV.insert(tk.END, '{}'.format(spisok_conc[3]))
		entr_A1_IV.delete(0, "end")
		entr_A1_IV.insert(tk.END, '{}'.format(QCD_val[0][0]))
		entr_A2_IV.delete(0, "end")
		entr_A2_IV.insert(tk.END, '{}'.format(QCD_val[0][1]))
		entr_A3_IV.delete(0, "end")
		entr_A3_IV.insert(tk.END, '{}'.format(QCD_val[0][2]))
		if check_var_gr_3.get() >= 4:
			entr_A4_IV.delete(0, "end")
			entr_A4_IV.insert(tk.END, '{}'.format(QCD_val[0][3]))
		if check_var_gr_3.get() >= 5:	
			entr_A5_IV.delete(0, "end")
			entr_A5_IV.insert(tk.END, '{}'.format(QCD_val[0][4]))
		if check_var_gr_3.get() >= 6:
			entr_A6_IV.delete(0, "end")
			entr_A6_IV.insert(tk.END, '{}'.format(QCD_val[0][5]))
		if check_var_gr_3.get() >= 7:
			entr_A7_IV.delete(0, "end")
			entr_A7_IV.insert(tk.END, '{}'.format(QCD_val[0][6]))
		if check_var_gr_3.get() >= 8:
			entr_A8_IV.delete(0, "end")
			entr_A8_IV.insert(tk.END, '{}'.format(QCD_val[0][7]))
		if check_var_gr_3.get() >= 9:
			entr_A9_IV.delete(0, "end")
			entr_A9_IV.insert(tk.END, '{}'.format(QCD_val[0][8]))
		if check_var_gr_3.get() >= 10:
			entr_A10_IV.delete(0, "end")	
			entr_A10_IV.insert(tk.END, '{}'.format(QCD_val[0][9]))


		#QCD2
		if check_var_gr_2.get() >= 2:
			entr_B1_IV.delete(0, "end")
			entr_B1_IV.insert(tk.END, '{}'.format(QCD_val[1][0]))
			entr_B2_IV.delete(0, "end")
			entr_B2_IV.insert(tk.END, '{}'.format(QCD_val[1][1]))
			entr_B3_IV.delete(0, "end")
			entr_B3_IV.insert(tk.END, '{}'.format(QCD_val[1][2]))
			if check_var_gr_3.get() >= 4:
				entr_B4_IV.delete(0, "end")
				entr_B4_IV.insert(tk.END, '{}'.format(QCD_val[1][3]))
			if check_var_gr_3.get() >= 5:	
				entr_B5_IV.delete(0, "end")
				entr_B5_IV.insert(tk.END, '{}'.format(QCD_val[1][4]))
			if check_var_gr_3.get() >= 6:
				entr_B6_IV.delete(0, "end")
				entr_B6_IV.insert(tk.END, '{}'.format(QCD_val[1][5]))
			if check_var_gr_3.get() >= 7:
				entr_B7_IV.delete(0, "end")
				entr_B7_IV.insert(tk.END, '{}'.format(QCD_val[1][6]))
			if check_var_gr_3.get() >= 8:
				entr_B8_IV.delete(0, "end")
				entr_B8_IV.insert(tk.END, '{}'.format(QCD_val[1][7]))
			if check_var_gr_3.get() >= 9:
				entr_B9_IV.delete(0, "end")
				entr_B9_IV.insert(tk.END, '{}'.format(QCD_val[1][8]))
			if check_var_gr_3.get() >= 10:
				entr_B10_IV.delete(0, "end")
				entr_B10_IV.insert(tk.END, '{}'.format(QCD_val[1][9]))


		#QCD3
		if check_var_gr_2.get() >= 3:
			entr_C1_IV.delete(0, "end")
			entr_C1_IV.insert(tk.END, '{}'.format(QCD_val[2][0]))
			entr_C2_IV.delete(0, "end")
			entr_C2_IV.insert(tk.END, '{}'.format(QCD_val[2][1]))
			entr_C3_IV.delete(0, "end")
			entr_C3_IV.insert(tk.END, '{}'.format(QCD_val[2][2]))
			if check_var_gr_3.get() >= 4:
				entr_C4_IV.delete(0, "end")
				entr_C4_IV.insert(tk.END, '{}'.format(QCD_val[2][3]))
			if check_var_gr_3.get() >= 5:	
				entr_C5_IV.delete(0, "end")
				entr_C5_IV.insert(tk.END, '{}'.format(QCD_val[2][4]))
			if check_var_gr_3.get() >= 6:
				entr_C6_IV.delete(0, "end")
				entr_C6_IV.insert(tk.END, '{}'.format(QCD_val[2][5]))
			if check_var_gr_3.get() >= 7:
				entr_C7_IV.delete(0, "end")
				entr_C7_IV.insert(tk.END, '{}'.format(QCD_val[2][6]))
			if check_var_gr_3.get() >= 8:
				entr_C8_IV.delete(0, "end")
				entr_C8_IV.insert(tk.END, '{}'.format(QCD_val[2][7]))
			if check_var_gr_3.get() >= 9:
				entr_C9_IV.delete(0, "end")
				entr_C9_IV.insert(tk.END, '{}'.format(QCD_val[2][8]))
			if check_var_gr_3.get() >= 10:
				entr_C10_IV.delete(0, "end")
				entr_C10_IV.insert(tk.END, '{}'.format(QCD_val[2][9]))


		#QCD4
		if check_var_gr_2.get() >= 4:
			entr_D1_IV.delete(0, "end")
			entr_D1_IV.insert(tk.END, '{}'.format(QCD_val[3][0]))
			entr_D2_IV.delete(0, "end")
			entr_D2_IV.insert(tk.END, '{}'.format(QCD_val[3][1]))
			entr_D3_IV.delete(0, "end")
			entr_D3_IV.insert(tk.END, '{}'.format(QCD_val[3][2]))
			if check_var_gr_3.get() >= 4:
				entr_D4_IV.delete(0, "end")
				entr_D4_IV.insert(tk.END, '{}'.format(QCD_val[3][3]))
			if check_var_gr_3.get() >= 5:	
				entr_D5_IV.delete(0, "end")
				entr_D5_IV.insert(tk.END, '{}'.format(QCD_val[3][4]))
			if check_var_gr_3.get() >= 6:
				entr_D6_IV.delete(0, "end")
				entr_D6_IV.insert(tk.END, '{}'.format(QCD_val[3][5]))
			if check_var_gr_3.get() >= 7:
				entr_D7_IV.delete(0, "end")
				entr_D7_IV.insert(tk.END, '{}'.format(QCD_val[3][6]))
			if check_var_gr_3.get() >= 8:
				entr_D8_IV.delete(0, "end")
				entr_D8_IV.insert(tk.END, '{}'.format(QCD_val[3][7]))
			if check_var_gr_3.get() >= 9:
				entr_D9_IV.delete(0, "end")
				entr_D9_IV.insert(tk.END, '{}'.format(QCD_val[3][8]))
			if check_var_gr_3.get() >= 10:
				entr_D10_IV.delete(0, "end")
				entr_D10_IV.insert(tk.END, '{}'.format(QCD_val[3][9]))


		#QCD5
		if check_var_gr_2.get() >= 5:
			entr_E1_IV.delete(0, "end")
			entr_E1_IV.insert(tk.END, '{}'.format(QCD_val[4][0]))
			entr_E2_IV.delete(0, "end")
			entr_E2_IV.insert(tk.END, '{}'.format(QCD_val[4][1]))
			entr_E3_IV.delete(0, "end")
			entr_E3_IV.insert(tk.END, '{}'.format(QCD_val[4][2]))
			if check_var_gr_3.get() >= 4:
				entr_E4_IV.delete(0, "end")
				entr_E4_IV.insert(tk.END, '{}'.format(QCD_val[4][3]))
			if check_var_gr_3.get() >= 5:	
				entr_E5_IV.delete(0, "end")
				entr_E5_IV.insert(tk.END, '{}'.format(QCD_val[4][4]))
			if check_var_gr_3.get() >= 6:
				entr_E6_IV.delete(0, "end")
				entr_E6_IV.insert(tk.END, '{}'.format(QCD_val[4][5]))
			if check_var_gr_3.get() >= 7:
				entr_E7_IV.delete(0, "end")
				entr_E7_IV.insert(tk.END, '{}'.format(QCD_val[4][6]))
			if check_var_gr_3.get() >= 8:
				entr_E8_IV.delete(0, "end")
				entr_E8_IV.insert(tk.END, '{}'.format(QCD_val[4][7]))
			if check_var_gr_3.get() >= 9:
				entr_E9_IV.delete(0, "end")
				entr_E9_IV.insert(tk.END, '{}'.format(QCD_val[4][8]))
			if check_var_gr_3.get() >= 10:
				entr_E10_IV.delete(0, "end")
				entr_E10_IV.insert(tk.END, '{}'.format(QCD_val[4][9]))


		#QCD6
		if check_var_gr_2.get() >= 6:
			entr_F1_IV.delete(0, "end")
			entr_F1_IV.insert(tk.END, '{}'.format(QCD_val[5][0]))
			entr_F2_IV.delete(0, "end")
			entr_F2_IV.insert(tk.END, '{}'.format(QCD_val[5][1]))
			entr_F3_IV.delete(0, "end")
			entr_F3_IV.insert(tk.END, '{}'.format(QCD_val[5][2]))
			if check_var_gr_3.get() >= 4:
				entr_F4_IV.delete(0, "end")
				entr_F4_IV.insert(tk.END, '{}'.format(QCD_val[5][3]))
			if check_var_gr_3.get() >= 5:	
				entr_F5_IV.delete(0, "end")
				entr_F5_IV.insert(tk.END, '{}'.format(QCD_val[5][4]))
			if check_var_gr_3.get() >= 6:
				entr_F6_IV.delete(0, "end")
				entr_F6_IV.insert(tk.END, '{}'.format(QCD_val[5][5]))
			if check_var_gr_3.get() >= 7:
				entr_F7_IV.delete(0, "end")
				entr_F7_IV.insert(tk.END, '{}'.format(QCD_val[5][6]))
			if check_var_gr_3.get() >= 8:
				entr_F8_IV.delete(0, "end")
				entr_F8_IV.insert(tk.END, '{}'.format(QCD_val[5][7]))
			if check_var_gr_3.get() >= 9:
				entr_F9_IV.delete(0, "end")
				entr_F9_IV.insert(tk.END, '{}'.format(QCD_val[5][8]))
			if check_var_gr_3.get() >= 10:
				entr_F10_IV.delete(0, "end")
				entr_F10_IV.insert(tk.END, '{}'.format(QCD_val[5][9]))					


	#QCE
	if check_var_gr_1.get() >= 5:
		#QCE1
		entr_MAIN_V.delete(0, "end")
		entr_MAIN_V.insert(tk.END, '{}'.format(spisok_conc[4]))
		entr_A1_V.delete(0, "end")
		entr_A1_V.insert(tk.END, '{}'.format(QCE_val[0][0]))
		entr_A2_V.delete(0, "end")
		entr_A2_V.insert(tk.END, '{}'.format(QCE_val[0][1]))
		entr_A3_V.delete(0, "end")
		entr_A3_V.insert(tk.END, '{}'.format(QCE_val[0][2]))
		if check_var_gr_3.get() >= 4:
			entr_A4_V.delete(0, "end")
			entr_A4_V.insert(tk.END, '{}'.format(QCE_val[0][3]))
		if check_var_gr_3.get() >= 5:
			entr_A5_V.delete(0, "end")	
			entr_A5_V.insert(tk.END, '{}'.format(QCE_val[0][4]))
		if check_var_gr_3.get() >= 6:
			entr_A6_V.delete(0, "end")
			entr_A6_V.insert(tk.END, '{}'.format(QCE_val[0][5]))
		if check_var_gr_3.get() >= 7:
			entr_A7_V.delete(0, "end")
			entr_A7_V.insert(tk.END, '{}'.format(QCE_val[0][6]))
		if check_var_gr_3.get() >= 8:
			entr_A8_V.delete(0, "end")
			entr_A8_V.insert(tk.END, '{}'.format(QCE_val[0][7]))
		if check_var_gr_3.get() >= 9:
			entr_A9_V.delete(0, "end")
			entr_A9_V.insert(tk.END, '{}'.format(QCE_val[0][8]))
		if check_var_gr_3.get() >= 10:
			entr_A10_V.delete(0, "end")	
			entr_A10_V.insert(tk.END, '{}'.format(QCE_val[0][9]))


		#QCE2
		if check_var_gr_2.get() >= 2:
			entr_B1_V.delete(0, "end")
			entr_B1_V.insert(tk.END, '{}'.format(QCE_val[1][0]))
			entr_B2_V.delete(0, "end")
			entr_B2_V.insert(tk.END, '{}'.format(QCE_val[1][1]))
			entr_B3_V.delete(0, "end")
			entr_B3_V.insert(tk.END, '{}'.format(QCE_val[1][2]))
			if check_var_gr_3.get() >= 4:
				entr_B4_V.delete(0, "end")
				entr_B4_V.insert(tk.END, '{}'.format(QCE_val[1][3]))
			if check_var_gr_3.get() >= 5:
				entr_B5_V.delete(0, "end")	
				entr_B5_V.insert(tk.END, '{}'.format(QCE_val[1][4]))
			if check_var_gr_3.get() >= 6:
				entr_B6_V.delete(0, "end")
				entr_B6_V.insert(tk.END, '{}'.format(QCE_val[1][5]))
			if check_var_gr_3.get() >= 7:
				entr_B7_V.delete(0, "end")
				entr_B7_V.insert(tk.END, '{}'.format(QCE_val[1][6]))
			if check_var_gr_3.get() >= 8:
				entr_B8_V.delete(0, "end")
				entr_B8_V.insert(tk.END, '{}'.format(QCE_val[1][7]))
			if check_var_gr_3.get() >= 9:
				entr_B9_V.delete(0, "end")
				entr_B9_V.insert(tk.END, '{}'.format(QCE_val[1][8]))
			if check_var_gr_3.get() >= 10:
				entr_B10_V.delete(0, "end")
				entr_B10_V.insert(tk.END, '{}'.format(QCE_val[1][9]))

		#QCE3
		if check_var_gr_2.get() >= 3:
			entr_C1_V.delete(0, "end")
			entr_C1_V.insert(tk.END, '{}'.format(QCE_val[2][0]))
			entr_C2_V.delete(0, "end")
			entr_C2_V.insert(tk.END, '{}'.format(QCE_val[2][1]))
			entr_C3_V.delete(0, "end")
			entr_C3_V.insert(tk.END, '{}'.format(QCE_val[2][2]))
			if check_var_gr_3.get() >= 4:
				entr_C4_V.delete(0, "end")
				entr_C4_V.insert(tk.END, '{}'.format(QCE_val[2][3]))
			if check_var_gr_3.get() >= 5:
				entr_C5_V.delete(0, "end")	
				entr_C5_V.insert(tk.END, '{}'.format(QCE_val[2][4]))
			if check_var_gr_3.get() >= 6:
				entr_C6_V.delete(0, "end")
				entr_C6_V.insert(tk.END, '{}'.format(QCE_val[2][5]))
			if check_var_gr_3.get() >= 7:
				entr_C7_V.delete(0, "end")
				entr_C7_V.insert(tk.END, '{}'.format(QCE_val[2][6]))
			if check_var_gr_3.get() >= 8:
				entr_C8_V.delete(0, "end")
				entr_C8_V.insert(tk.END, '{}'.format(QCE_val[2][7]))
			if check_var_gr_3.get() >= 9:
				entr_C9_V.delete(0, "end")
				entr_C9_V.insert(tk.END, '{}'.format(QCE_val[2][8]))
			if check_var_gr_3.get() >= 10:
				entr_C10_V.delete(0, "end")
				entr_C10_V.insert(tk.END, '{}'.format(QCE_val[2][9]))


			#QCE4
		if check_var_gr_2.get() >= 4:
			entr_D1_V.delete(0, "end")
			entr_D1_V.insert(tk.END, '{}'.format(QCE_val[3][0]))
			entr_D2_V.delete(0, "end")
			entr_D2_V.insert(tk.END, '{}'.format(QCE_val[3][1]))
			entr_D3_V.delete(0, "end")
			entr_D3_V.insert(tk.END, '{}'.format(QCE_val[3][2]))
			if check_var_gr_3.get() >= 4:
				entr_D4_V.delete(0, "end")
				entr_D4_V.insert(tk.END, '{}'.format(QCE_val[3][3]))
			if check_var_gr_3.get() >= 5:
				entr_D5_V.delete(0, "end")	
				entr_D5_V.insert(tk.END, '{}'.format(QCE_val[3][4]))
			if check_var_gr_3.get() >= 6:
				entr_D6_V.delete(0, "end")
				entr_D6_V.insert(tk.END, '{}'.format(QCE_val[3][5]))
			if check_var_gr_3.get() >= 7:
				entr_D7_V.delete(0, "end")
				entr_D7_V.insert(tk.END, '{}'.format(QCE_val[3][6]))
			if check_var_gr_3.get() >= 8:
				entr_D8_V.delete(0, "end")
				entr_D8_V.insert(tk.END, '{}'.format(QCE_val[3][7]))
			if check_var_gr_3.get() >= 9:
				entr_D9_V.delete(0, "end")
				entr_D9_V.insert(tk.END, '{}'.format(QCE_val[3][8]))
			if check_var_gr_3.get() >= 10:
				entr_D10_V.delete(0, "end")
				entr_D10_V.insert(tk.END, '{}'.format(QCE_val[3][9]))


		#QCE5
		if check_var_gr_2.get() >= 5:
			entr_E1_V.delete(0, "end")
			entr_E1_V.insert(tk.END, '{}'.format(QCE_val[4][0]))
			entr_E2_V.delete(0, "end")
			entr_E2_V.insert(tk.END, '{}'.format(QCE_val[4][1]))
			entr_E3_V.delete(0, "end")
			entr_E3_V.insert(tk.END, '{}'.format(QCE_val[4][2]))
			if check_var_gr_3.get() >= 4:
				entr_E4_V.delete(0, "end")
				entr_E4_V.insert(tk.END, '{}'.format(QCE_val[4][3]))
			if check_var_gr_3.get() >= 5:
				entr_E5_V.delete(0, "end")	
				entr_E5_V.insert(tk.END, '{}'.format(QCE_val[4][4]))
			if check_var_gr_3.get() >= 6:
				entr_E6_V.delete(0, "end")
				entr_E6_V.insert(tk.END, '{}'.format(QCE_val[4][5]))
			if check_var_gr_3.get() >= 7:
				entr_E7_V.delete(0, "end")
				entr_E7_V.insert(tk.END, '{}'.format(QCE_val[4][6]))
			if check_var_gr_3.get() >= 8:
				entr_E8_V.delete(0, "end")
				entr_E8_V.insert(tk.END, '{}'.format(QCE_val[4][7]))
			if check_var_gr_3.get() >= 9:
				entr_E9_V.delete(0, "end")
				entr_E9_V.insert(tk.END, '{}'.format(QCE_val[4][8]))
			if check_var_gr_3.get() >= 10:
				entr_E10_V.delete(0, "end")
				entr_E10_V.insert(tk.END, '{}'.format(QCE_val[4][9]))


		#QCE6
		if check_var_gr_2.get() >= 6:
			entr_F1_V.delete(0, "end")
			entr_F1_V.insert(tk.END, '{}'.format(QCE_val[5][0]))
			entr_F2_V.delete(0, "end")
			entr_F2_V.insert(tk.END, '{}'.format(QCE_val[5][1]))
			entr_F3_V.delete(0, "end")
			entr_F3_V.insert(tk.END, '{}'.format(QCE_val[5][2]))
			if check_var_gr_3.get() >= 4:
				entr_F4_V.delete(0, "end")
				entr_F4_V.insert(tk.END, '{}'.format(QCE_val[5][3]))
			if check_var_gr_3.get() >= 5:
				entr_F5_V.delete(0, "end")	
				entr_F5_V.insert(tk.END, '{}'.format(QCE_val[5][4]))
			if check_var_gr_3.get() >= 6:
				entr_F6_V.delete(0, "end")
				entr_F6_V.insert(tk.END, '{}'.format(QCE_val[5][5]))
			if check_var_gr_3.get() >= 7:
				entr_F7_V.delete(0, "end")
				entr_F7_V.insert(tk.END, '{}'.format(QCE_val[5][6]))
			if check_var_gr_3.get() >= 8:
				entr_F8_V.delete(0, "end")
				entr_F8_V.insert(tk.END, '{}'.format(QCE_val[5][7]))
			if check_var_gr_3.get() >= 9:
				entr_F9_V.delete(0, "end")
				entr_F9_V.insert(tk.END, '{}'.format(QCE_val[5][8]))
			if check_var_gr_3.get() >= 10:
				entr_F10_V.delete(0, "end")
				entr_F10_V.insert(tk.END, '{}'.format(QCE_val[5][9]))	


	#QCF
	if check_var_gr_1.get() >= 6:
		#QCF1
		entr_MAIN_VI.delete(0, "end")
		entr_MAIN_VI.insert(tk.END, '{}'.format(spisok_conc[5]))
		entr_A1_VI.delete(0, "end")
		entr_A1_VI.insert(tk.END, '{}'.format(QCF_val[0][0]))
		entr_A2_VI.delete(0, "end")
		entr_A2_VI.insert(tk.END, '{}'.format(QCF_val[0][1]))
		entr_A3_VI.delete(0, "end")
		entr_A3_VI.insert(tk.END, '{}'.format(QCF_val[0][2]))		
		if check_var_gr_3.get() >= 4:
			entr_A4_VI.delete(0, "end")
			entr_A4_VI.insert(tk.END, '{}'.format(QCF_val[0][3]))
		if check_var_gr_3.get() >= 5:	
			entr_A5_VI.delete(0, "end")
			entr_A5_VI.insert(tk.END, '{}'.format(QCF_val[0][4]))
		if check_var_gr_3.get() >= 6:
			entr_A6_VI.delete(0, "end")
			entr_A6_VI.insert(tk.END, '{}'.format(QCF_val[0][5]))
		if check_var_gr_3.get() >= 7:
			entr_A7_VI.delete(0, "end")
			entr_A7_VI.insert(tk.END, '{}'.format(QCF_val[0][6]))
		if check_var_gr_3.get() >= 8:
			entr_A8_VI.delete(0, "end")
			entr_A8_VI.insert(tk.END, '{}'.format(QCF_val[0][7]))
		if check_var_gr_3.get() >= 9:
			entr_A9_VI.delete(0, "end")
			entr_A9_VI.insert(tk.END, '{}'.format(QCF_val[0][8]))
		if check_var_gr_3.get() >= 10:
			entr_A10_VI.delete(0, "end")	
			entr_A10_VI.insert(tk.END, '{}'.format(QCF_val[0][9]))


		#QCF2
		if check_var_gr_2.get() >= 2:
			entr_B1_VI.delete(0, "end")
			entr_B1_VI.insert(tk.END, '{}'.format(QCF_val[1][0]))
			entr_B2_VI.delete(0, "end")
			entr_B2_VI.insert(tk.END, '{}'.format(QCF_val[1][1]))
			entr_B3_VI.delete(0, "end")
			entr_B3_VI.insert(tk.END, '{}'.format(QCF_val[1][2]))
			if check_var_gr_3.get() >= 4:
				entr_B4_VI.delete(0, "end")
				entr_B4_VI.insert(tk.END, '{}'.format(QCF_val[1][3]))
			if check_var_gr_3.get() >= 5:	
				entr_B5_VI.delete(0, "end")
				entr_B5_VI.insert(tk.END, '{}'.format(QCF_val[1][4]))
			if check_var_gr_3.get() >= 6:
				entr_B6_VI.delete(0, "end")
				entr_B6_VI.insert(tk.END, '{}'.format(QCF_val[1][5]))
			if check_var_gr_3.get() >= 7:
				entr_B7_VI.delete(0, "end")
				entr_B7_VI.insert(tk.END, '{}'.format(QCF_val[1][6]))
			if check_var_gr_3.get() >= 8:
				entr_B8_VI.delete(0, "end")
				entr_B8_VI.insert(tk.END, '{}'.format(QCF_val[1][7]))
			if check_var_gr_3.get() >= 9:
				entr_B9_VI.delete(0, "end")
				entr_B9_VI.insert(tk.END, '{}'.format(QCF_val[1][8]))
			if check_var_gr_3.get() >= 10:
				entr_B10_VI.delete(0, "end")
				entr_B10_VI.insert(tk.END, '{}'.format(QCF_val[1][9]))


			#QCF3
		if check_var_gr_2.get() >= 3:
			entr_C1_VI.delete(0, "end")
			entr_C1_VI.insert(tk.END, '{}'.format(QCF_val[2][0]))
			entr_C2_VI.delete(0, "end")
			entr_C2_VI.insert(tk.END, '{}'.format(QCF_val[2][1]))
			entr_C3_VI.delete(0, "end")
			entr_C3_VI.insert(tk.END, '{}'.format(QCF_val[2][2]))
			if check_var_gr_3.get() >= 4:
				entr_C4_VI.delete(0, "end")
				entr_C4_VI.insert(tk.END, '{}'.format(QCF_val[2][3]))
			if check_var_gr_3.get() >= 5:	
				entr_C5_VI.delete(0, "end")
				entr_C5_VI.insert(tk.END, '{}'.format(QCF_val[2][4]))
			if check_var_gr_3.get() >= 6:
				entr_C6_VI.delete(0, "end")
				entr_C6_VI.insert(tk.END, '{}'.format(QCF_val[2][5]))
			if check_var_gr_3.get() >= 7:
				entr_C7_VI.delete(0, "end")
				entr_C7_VI.insert(tk.END, '{}'.format(QCF_val[2][6]))
			if check_var_gr_3.get() >= 8:
				entr_C8_VI.delete(0, "end")
				entr_C8_VI.insert(tk.END, '{}'.format(QCF_val[2][7]))
			if check_var_gr_3.get() >= 9:
				entr_C9_VI.delete(0, "end")
				entr_C9_VI.insert(tk.END, '{}'.format(QCF_val[2][8]))
			if check_var_gr_3.get() >= 10:
				entr_C10_VI.delete(0, "end")
				entr_C10_VI.insert(tk.END, '{}'.format(QCF_val[2][9]))


		#QCF4
		if check_var_gr_2.get() >= 4:
			entr_D1_VI.delete(0, "end")
			entr_D1_VI.insert(tk.END, '{}'.format(QCF_val[3][0]))
			entr_D2_VI.delete(0, "end")
			entr_D2_VI.insert(tk.END, '{}'.format(QCF_val[3][1]))
			entr_D3_VI.delete(0, "end")
			entr_D3_VI.insert(tk.END, '{}'.format(QCF_val[3][2]))
			if check_var_gr_3.get() >= 4:
				entr_D4_VI.delete(0, "end")
				entr_D4_VI.insert(tk.END, '{}'.format(QCF_val[3][3]))
			if check_var_gr_3.get() >= 5:	
				entr_D5_VI.delete(0, "end")
				entr_D5_VI.insert(tk.END, '{}'.format(QCF_val[3][4]))
			if check_var_gr_3.get() >= 6:
				entr_D6_VI.delete(0, "end")
				entr_D6_VI.insert(tk.END, '{}'.format(QCF_val[3][5]))
			if check_var_gr_3.get() >= 7:
				entr_D7_VI.delete(0, "end")
				entr_D7_VI.insert(tk.END, '{}'.format(QCF_val[3][6]))
			if check_var_gr_3.get() >= 8:
				entr_D8_VI.delete(0, "end")
				entr_D8_VI.insert(tk.END, '{}'.format(QCF_val[3][7]))
			if check_var_gr_3.get() >= 9:
				entr_D9_VI.delete(0, "end")
				entr_D9_VI.insert(tk.END, '{}'.format(QCF_val[3][8]))
			if check_var_gr_3.get() >= 10:
				entr_D10_VI.delete(0, "end")
				entr_D10_VI.insert(tk.END, '{}'.format(QCF_val[3][9]))


		#QCF5
		if check_var_gr_2.get() >= 5:
			entr_E1_VI.delete(0, "end")
			entr_E1_VI.insert(tk.END, '{}'.format(QCF_val[4][0]))
			entr_E2_VI.delete(0, "end")
			entr_E2_VI.insert(tk.END, '{}'.format(QCF_val[4][1]))
			entr_E3_VI.delete(0, "end")
			entr_E3_VI.insert(tk.END, '{}'.format(QCF_val[4][2]))
			if check_var_gr_3.get() >= 4:
				entr_E4_VI.delete(0, "end")
				entr_E4_VI.insert(tk.END, '{}'.format(QCF_val[4][3]))
			if check_var_gr_3.get() >= 5:	
				entr_E5_VI.delete(0, "end")
				entr_E5_VI.insert(tk.END, '{}'.format(QCF_val[4][4]))
			if check_var_gr_3.get() >= 6:
				entr_E6_VI.delete(0, "end")
				entr_E6_VI.insert(tk.END, '{}'.format(QCF_val[4][5]))
			if check_var_gr_3.get() >= 7:
				entr_E7_VI.delete(0, "end")
				entr_E7_VI.insert(tk.END, '{}'.format(QCF_val[4][6]))
			if check_var_gr_3.get() >= 8:
				entr_E8_VI.delete(0, "end")
				entr_E8_VI.insert(tk.END, '{}'.format(QCF_val[4][7]))
			if check_var_gr_3.get() >= 9:
				entr_E9_VI.delete(0, "end")
				entr_E9_VI.insert(tk.END, '{}'.format(QCF_val[4][8]))
			if check_var_gr_3.get() >= 10:
				entr_E10_VI.delete(0, "end")
				entr_E10_VI.insert(tk.END, '{}'.format(QCF_val[4][9]))


		#QCF6
		if check_var_gr_2.get() >= 6:
			entr_F1_VI.delete(0, "end")
			entr_F1_VI.insert(tk.END, '{}'.format(QCF_val[5][0]))
			entr_F2_VI.delete(0, "end")
			entr_F2_VI.insert(tk.END, '{}'.format(QCF_val[5][1]))
			entr_F3_VI.delete(0, "end")
			entr_F3_VI.insert(tk.END, '{}'.format(QCF_val[5][2]))
			if check_var_gr_3.get() >= 4:
				entr_F4_VI.delete(0, "end")
				entr_F4_VI.insert(tk.END, '{}'.format(QCF_val[5][3]))
			if check_var_gr_3.get() >= 5:	
				entr_F5_VI.delete(0, "end")
				entr_F5_VI.insert(tk.END, '{}'.format(QCF_val[5][4]))
			if check_var_gr_3.get() >= 6:
				entr_F6_VI.delete(0, "end")
				entr_F6_VI.insert(tk.END, '{}'.format(QCF_val[5][5]))
			if check_var_gr_3.get() >= 7:
				entr_F7_VI.delete(0, "end")
				entr_F7_VI.insert(tk.END, '{}'.format(QCF_val[5][6]))
			if check_var_gr_3.get() >= 8:
				entr_F8_VI.delete(0, "end")
				entr_F8_VI.insert(tk.END, '{}'.format(QCF_val[5][7]))
			if check_var_gr_3.get() >= 9:
				entr_F9_VI.delete(0, "end")
				entr_F9_VI.insert(tk.END, '{}'.format(QCF_val[5][8]))
			if check_var_gr_3.get() >= 10:
				entr_F10_VI.delete(0, "end")
				entr_F10_VI.insert(tk.END, '{}'.format(QCF_val[5][9]))	
 
def clear():

	entr_MAIN_I.delete(0, "end")
	entr_A1_I.delete(0, "end")
	entr_A2_I.delete(0, "end")
	entr_A3_I.delete(0, "end")
	
	if check_var_gr_3.get() >= 4:
		entr_A4_I.delete(0, "end")
		
	if check_var_gr_3.get() >= 5:	
		entr_A5_I.delete(0, "end")
		
	if check_var_gr_3.get() >= 6:
		entr_A6_I.delete(0, "end")
		
	if check_var_gr_3.get() >= 7:
		entr_A7_I.delete(0, "end")
		
	if check_var_gr_3.get() >= 8:
		entr_A8_I.delete(0, "end")
		
	if check_var_gr_3.get() >= 9:
		entr_A9_I.delete(0, "end")
		
	if check_var_gr_3.get() >= 10:	
		entr_A10_I.delete(0, "end")
		
	#QCA2
	if check_var_gr_2.get() >= 2:
		entr_B1_I.delete(0, "end")
		
		entr_B2_I.delete(0, "end")
		
		entr_B3_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 4:
			entr_B4_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 5:	
			entr_B5_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 6:
			entr_B6_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 7:
			entr_B7_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 8:
			entr_B8_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 9:
			entr_B9_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 10:
			entr_B10_I.delete(0, "end")
			
	#QCA3
	if check_var_gr_2.get() >= 3:
		entr_C1_I.delete(0, "end")
		
		entr_C2_I.delete(0, "end")
		
		entr_C3_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 4:
			entr_C4_I.delete(0, "end")
			
		if check_var_gr_3.get() >= 5:
			entr_C5_I.delete(0, "end")	
			
		if check_var_gr_3.get() >= 6:
			entr_C6_I.delete(0, "end")
			
		if check_var_gr_3.get() >= 7:
			entr_C7_I.delete(0, "end")
			
		if check_var_gr_3.get() >= 8:
			entr_C8_I.delete(0, "end")
			
		if check_var_gr_3.get() >= 9:
			entr_C9_I.delete(0, "end")
			
		if check_var_gr_3.get() >= 10:
			entr_C10_I.delete(0, "end")
			

	#QCA4
	if check_var_gr_2.get() >= 4:
		entr_D1_I.delete(0, "end")
	
		entr_D2_I.delete(0, "end")
	
		entr_D3_I.delete(0, "end")
	
		if check_var_gr_3.get() >= 4:
			entr_D4_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 5:	
			entr_D5_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 6:
			entr_D6_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 7:
			entr_D7_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 8:
			entr_D8_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 9:
			entr_D9_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 10:
			entr_D10_I.delete(0, "end")
			

	#QCA5
	if check_var_gr_2.get() >= 5:
		entr_E1_I.delete(0, "end")
		entr_E2_I.delete(0, "end")
		entr_E3_I.delete(0, "end")
	
		if check_var_gr_3.get() >= 4:
			entr_E4_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 5:
			entr_E5_I.delete(0, "end")	
		
		if check_var_gr_3.get() >= 6:
			entr_E6_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 7:
			entr_E7_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 8:
			entr_E8_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 9:
			entr_E9_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 10:
			entr_E10_I.delete(0, "end")
			


	#QCA6
	if check_var_gr_2.get() >= 6:
		entr_F1_I.delete(0, "end")
		entr_F2_I.delete(0, "end")
		entr_F3_I.delete(0, "end")
	
		if check_var_gr_3.get() >= 4:
			entr_F4_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 5:
			entr_F5_I.delete(0, "end")	
		
		if check_var_gr_3.get() >= 6:
			entr_F6_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 7:
			entr_F7_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 8:
			entr_F8_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 9:
			entr_F9_I.delete(0, "end")
		
		if check_var_gr_3.get() >= 10:
			entr_F10_I.delete(0, "end")
			


	#QCB

	if check_var_gr_1.get() >= 2:

		entr_MAIN_II.delete(0, "end")
		entr_A1_II.delete(0, "end")
		entr_A2_II.delete(0, "end")
		entr_A3_II.delete(0, "end")
	
		if check_var_gr_3.get() >= 4:
			entr_A4_II.delete(0, "end")
		
		if check_var_gr_3.get() >= 5:
			entr_A5_II.delete(0, "end")	
		
		if check_var_gr_3.get() >= 6:
			entr_A6_II.delete(0, "end")
		
		if check_var_gr_3.get() >= 7:
			entr_A7_II.delete(0, "end")
		
		if check_var_gr_3.get() >= 8:
			entr_A8_II.delete(0, "end")
		
		if check_var_gr_3.get() >= 9:
			entr_A9_II.delete(0, "end")
		
		if check_var_gr_3.get() >= 10:
			entr_A10_II.delete(0, "end")		
			


		#QCB2
		if check_var_gr_2.get() >= 2:

			entr_B1_II.delete(0, "end")
			entr_B2_II.delete(0, "end")
			entr_B3_II.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_B4_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_B5_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_B6_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_B7_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_B8_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_B9_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_B10_II.delete(0, "end")
				

		#QCB3
		if check_var_gr_2.get() >= 3:

			entr_C1_II.delete(0, "end")
			entr_C2_II.delete(0, "end")
			entr_C3_II.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_C4_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_C5_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_C6_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_C7_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_C8_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_C9_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_C10_II.delete(0, "end")
				
			#QCB4
		if check_var_gr_2.get() >= 4:

			entr_D1_II.delete(0, "end")
			entr_D2_II.delete(0, "end")
			entr_D3_II.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_D4_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_D5_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_D6_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_D7_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_D8_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_D9_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_D10_II.delete(0, "end")
				

			#QCB5
		if check_var_gr_2.get() >= 5:

			entr_E1_II.delete(0, "end")
			entr_E2_II.delete(0, "end")
			entr_E3_II.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_E4_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_E5_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_E6_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_E7_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_E8_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_E9_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_E10_II.delete(0, "end")
				

		#QCB6
		if check_var_gr_2.get() >= 6:

			entr_F1_II.delete(0, "end")
			entr_F2_II.delete(0, "end")
			entr_F3_II.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_F4_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_F5_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_F6_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_F7_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_F8_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_F9_II.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_F10_II.delete(0, "end")
				

	if check_var_gr_1.get() >= 3:
		entr_MAIN_III.delete(0, "end")
		entr_A1_III.delete(0, "end")
		entr_A2_III.delete(0, "end")
		entr_A3_III.delete(0, "end")
	
		if check_var_gr_3.get() >= 4:
			entr_A4_III.delete(0, "end")
		
		if check_var_gr_3.get() >= 5:	
			entr_A5_III.delete(0, "end")
		
		if check_var_gr_3.get() >= 6:
			entr_A6_III.delete(0, "end")
		
		if check_var_gr_3.get() >= 7:
			entr_A7_III.delete(0, "end")
		
		if check_var_gr_3.get() >= 8:
			entr_A8_III.delete(0, "end")
		
		if check_var_gr_3.get() >= 9:
			entr_A9_III.delete(0, "end")
		
		if check_var_gr_3.get() >= 10:	
			entr_A10_III.delete(0, "end")
			


			#QCC2
		if check_var_gr_2.get() >= 2:
			entr_B1_III.delete(0, "end")
			entr_B2_III.delete(0, "end")
			entr_B3_III.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_B4_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_B5_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_B6_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_B7_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_B8_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_B9_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_B10_III.delete(0, "end")
				


			#QCC3
		if check_var_gr_2.get() >= 3:
			entr_C1_III.delete(0, "end")
			entr_C2_III.delete(0, "end")
			entr_C3_III.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_C4_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_C5_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_C6_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_C7_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_C8_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_C9_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_C10_III.delete(0, "end")
				
			#QCC4
		if check_var_gr_2.get() >= 4:
			entr_D1_III.delete(0, "end")
			entr_D2_III.delete(0, "end")
			entr_D3_III.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_D4_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_D5_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_D6_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_D7_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_D8_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_D9_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_D10_III.delete(0, "end")
				
			#QCC5
		if check_var_gr_2.get() >= 5:
			entr_E1_III.delete(0, "end")
			entr_E2_III.delete(0, "end")
			entr_E3_III.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_E4_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_E5_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_E6_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_E7_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_E8_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_E9_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_E10_III.delete(0, "end")
				

			#QCC6
		if check_var_gr_2.get() >= 6:
			entr_F1_III.delete(0, "end")
			entr_F2_III.delete(0, "end")
			entr_F3_III.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_F4_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_F5_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_F6_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_F7_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_F8_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_F9_III.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_F10_III.delete(0, "end")
			

	#QCD
	if check_var_gr_1.get() >= 4:
		entr_MAIN_IV.delete(0, "end")
		entr_A1_IV.delete(0, "end")
		entr_A2_IV.delete(0, "end")
		entr_A3_IV.delete(0, "end")
	
		if check_var_gr_3.get() >= 4:
			entr_A4_IV.delete(0, "end")
		
		if check_var_gr_3.get() >= 5:	
			entr_A5_IV.delete(0, "end")
		
		if check_var_gr_3.get() >= 6:
			entr_A6_IV.delete(0, "end")
		
		if check_var_gr_3.get() >= 7:
			entr_A7_IV.delete(0, "end")
		
		if check_var_gr_3.get() >= 8:
			entr_A8_IV.delete(0, "end")
		
		if check_var_gr_3.get() >= 9:
			entr_A9_IV.delete(0, "end")
		
		if check_var_gr_3.get() >= 10:
			entr_A10_IV.delete(0, "end")	
			
		#QCD2
		if check_var_gr_2.get() >= 2:
			entr_B1_IV.delete(0, "end")
			entr_B2_IV.delete(0, "end")
			entr_B3_IV.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_B4_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_B5_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_B6_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_B7_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_B8_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_B9_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_B10_IV.delete(0, "end")
				


		#QCD3
		if check_var_gr_2.get() >= 3:
			entr_C1_IV.delete(0, "end")
			entr_C2_IV.delete(0, "end")
			entr_C3_IV.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_C4_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_C5_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_C6_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_C7_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_C8_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_C9_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_C10_IV.delete(0, "end")
				

		if check_var_gr_2.get() >= 4:
			entr_D1_IV.delete(0, "end")
			entr_D2_IV.delete(0, "end")
			entr_D3_IV.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_D4_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_D5_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_D6_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_D7_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_D8_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_D9_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_D10_IV.delete(0, "end")
				

		#QCD5
		if check_var_gr_2.get() >= 5:
			entr_E1_IV.delete(0, "end")
			entr_E2_IV.delete(0, "end")
			entr_E3_IV.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_E4_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_E5_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_E6_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_E7_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_E8_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_E9_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_E10_IV.delete(0, "end")
				

		#QCD6
		if check_var_gr_2.get() >= 6:
			entr_F1_IV.delete(0, "end")
			entr_F2_IV.delete(0, "end")
			entr_F3_IV.delete(0, "end")
			
			if check_var_gr_3.get() >= 4:
				entr_F4_IV.delete(0, "end")
				
			if check_var_gr_3.get() >= 5:	
				entr_F5_IV.delete(0, "end")
				
			if check_var_gr_3.get() >= 6:
				entr_F6_IV.delete(0, "end")
				
			if check_var_gr_3.get() >= 7:
				entr_F7_IV.delete(0, "end")
				
			if check_var_gr_3.get() >= 8:
				entr_F8_IV.delete(0, "end")
				
			if check_var_gr_3.get() >= 9:
				entr_F9_IV.delete(0, "end")
				
			if check_var_gr_3.get() >= 10:
				entr_F10_IV.delete(0, "end")
					

	#QCE
	if check_var_gr_1.get() >= 5:
		
		entr_MAIN_V.delete(0, "end")
		entr_A1_V.delete(0, "end")
		entr_A2_V.delete(0, "end")
		entr_A3_V.delete(0, "end")
		
		if check_var_gr_3.get() >= 4:
			entr_A4_V.delete(0, "end")
			
		if check_var_gr_3.get() >= 5:
			entr_A5_V.delete(0, "end")	
			
		if check_var_gr_3.get() >= 6:
			entr_A6_V.delete(0, "end")
			
		if check_var_gr_3.get() >= 7:
			entr_A7_V.delete(0, "end")
			
		if check_var_gr_3.get() >= 8:
			entr_A8_V.delete(0, "end")
			
		if check_var_gr_3.get() >= 9:
			entr_A9_V.delete(0, "end")
			
		if check_var_gr_3.get() >= 10:
			entr_A10_V.delete(0, "end")	
			

		#QCE2
		if check_var_gr_2.get() >= 2:
			entr_B1_V.delete(0, "end")
			entr_B2_V.delete(0, "end")
			entr_B3_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 4:
				entr_B4_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 5:
				entr_B5_V.delete(0, "end")	
				
			if check_var_gr_3.get() >= 6:
				entr_B6_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 7:
				entr_B7_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 8:
				entr_B8_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 9:
				entr_B9_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 10:
				entr_B10_V.delete(0, "end")
				

		#QCE3
		if check_var_gr_2.get() >= 3:
			entr_C1_V.delete(0, "end")
		
			entr_C2_V.delete(0, "end")
		
			entr_C3_V.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_C4_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:
				entr_C5_V.delete(0, "end")	
			
			if check_var_gr_3.get() >= 6:
				entr_C6_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_C7_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_C8_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_C9_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_C10_V.delete(0, "end")
				


			#QCE4
		if check_var_gr_2.get() >= 4:
			entr_D1_V.delete(0, "end")
			
			entr_D2_V.delete(0, "end")
			
			entr_D3_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 4:
				entr_D4_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 5:
				entr_D5_V.delete(0, "end")	
				
			if check_var_gr_3.get() >= 6:
				entr_D6_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 7:
				entr_D7_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 8:
				entr_D8_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 9:
				entr_D9_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 10:
				entr_D10_V.delete(0, "end")
				


		#QCE5
		if check_var_gr_2.get() >= 5:
			entr_E1_V.delete(0, "end")
			
			entr_E2_V.delete(0, "end")
			
			entr_E3_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 4:
				entr_E4_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 5:
				entr_E5_V.delete(0, "end")	
				
			if check_var_gr_3.get() >= 6:
				entr_E6_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 7:
				entr_E7_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 8:
				entr_E8_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 9:
				entr_E9_V.delete(0, "end")
				
			if check_var_gr_3.get() >= 10:
				entr_E10_V.delete(0, "end")
				


		#QCE6
		if check_var_gr_2.get() >= 6:
			entr_F1_V.delete(0, "end")
		
			entr_F2_V.delete(0, "end")
		
			entr_F3_V.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_F4_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:
				entr_F5_V.delete(0, "end")	
			
			if check_var_gr_3.get() >= 6:
				entr_F6_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_F7_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_F8_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_F9_V.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_F10_V.delete(0, "end")
				


	#QCF
	if check_var_gr_1.get() >= 6:
		#QCF1
		entr_MAIN_VI.delete(0, "end")
		
		entr_A1_VI.delete(0, "end")
		
		entr_A2_VI.delete(0, "end")
		
		entr_A3_VI.delete(0, "end")
				
		if check_var_gr_3.get() >= 4:
			entr_A4_VI.delete(0, "end")
			
		if check_var_gr_3.get() >= 5:	
			entr_A5_VI.delete(0, "end")
			
		if check_var_gr_3.get() >= 6:
			entr_A6_VI.delete(0, "end")
			
		if check_var_gr_3.get() >= 7:
			entr_A7_VI.delete(0, "end")
			
		if check_var_gr_3.get() >= 8:
			entr_A8_VI.delete(0, "end")
			
		if check_var_gr_3.get() >= 9:
			entr_A9_VI.delete(0, "end")
			
		if check_var_gr_3.get() >= 10:
			entr_A10_VI.delete(0, "end")	
			


		#QCF2
		if check_var_gr_2.get() >= 2:
			entr_B1_VI.delete(0, "end")
			
			entr_B2_VI.delete(0, "end")
			
			entr_B3_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 4:
				entr_B4_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 5:	
				entr_B5_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 6:
				entr_B6_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 7:
				entr_B7_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 8:
				entr_B8_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 9:
				entr_B9_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 10:
				entr_B10_VI.delete(0, "end")
				


			#QCF3
		if check_var_gr_2.get() >= 3:
			entr_C1_VI.delete(0, "end")
			
			entr_C2_VI.delete(0, "end")
			
			entr_C3_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 4:
				entr_C4_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 5:	
				entr_C5_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 6:
				entr_C6_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 7:
				entr_C7_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 8:
				entr_C8_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 9:
				entr_C9_VI.delete(0, "end")
				
			if check_var_gr_3.get() >= 10:
				entr_C10_VI.delete(0, "end")
				


		#QCF4
		if check_var_gr_2.get() >= 4:
			entr_D1_VI.delete(0, "end")
		
			entr_D2_VI.delete(0, "end")
		
			entr_D3_VI.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_D4_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_D5_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_D6_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_D7_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_D8_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_D9_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_D10_VI.delete(0, "end")
				


		#QCF5
		if check_var_gr_2.get() >= 5:
			entr_E1_VI.delete(0, "end")
		
			entr_E2_VI.delete(0, "end")
		
			entr_E3_VI.delete(0, "end")
		
			if check_var_gr_3.get() >= 4:
				entr_E4_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 5:	
				entr_E5_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 6:
				entr_E6_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 7:
				entr_E7_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 8:
				entr_E8_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 9:
				entr_E9_VI.delete(0, "end")
			
			if check_var_gr_3.get() >= 10:
				entr_E10_VI.delete(0, "end")
				


		#QCF6
		if check_var_gr_2.get() >= 6:
			entr_F1_VI.delete(0, "end")
	
			entr_F2_VI.delete(0, "end")
	
			entr_F3_VI.delete(0, "end")
	
			if check_var_gr_3.get() >= 4:
				entr_F4_VI.delete(0, "end")
		
			if check_var_gr_3.get() >= 5:	
				entr_F5_VI.delete(0, "end")
		
			if check_var_gr_3.get() >= 6:
				entr_F6_VI.delete(0, "end")
		
			if check_var_gr_3.get() >= 7:
				entr_F7_VI.delete(0, "end")
		
			if check_var_gr_3.get() >= 8:
				entr_F8_VI.delete(0, "end")
		
			if check_var_gr_3.get() >= 9:
				entr_F9_VI.delete(0, "end")
		
			if check_var_gr_3.get() >= 10:
				entr_F10_VI.delete(0, "end")
				

    

#функция РАВНЕНИЕ ПО ЦЕНТРУ EXCEL 
def cols_c(ws, cell_range):
	alignment = Alignment(horizontal='center', vertical='center')
	cols_center = ws[cell_range]
	for col in cols_center:
		for cell in col:
			cell.alignment = alignment


#функция thin бордюра
def thin_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows_thin = ws[cell_range]
    for row in rows_thin:
        for cell in row:
            cell.border = border


def start_std():
	global check_var_gr_1
	check_var_gr_1.set(4)
	check_var_gr_2.set(3)
	check_var_gr_3.set(5)
	start()


def start():
	root_open.destroy()
	global root
	root = tk.Tk()
	root.title('AnovaD')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		root.geometry('900x300+300+180')
	else:
		root.geometry('1300x700+200+180')
	root.resizable(False,False)


	if check_var_gr_1.get() >= 1:

		
		labl_MAIN_I = tk.Label(root, text = 'QCA')
		labl_MAIN_I.place(x = 27, y = 5)
		global entr_MAIN_I
		entr_MAIN_I = tk.Entry(root, width = 6, relief= 'solid')
		entr_MAIN_I.place(x = 25, y = 25)
		
		labl_A_1_I  = tk.Label(root, text = "Обр. 1")
		labl_A_1_I.place(x = 25, y = 50)

		global entr_A1_I
		entr_A1_I = tk.Entry(root, width = 6, relief= 'solid')
		entr_A1_I.place(x = 25, y = 70)
		global entr_A2_I
		entr_A2_I = tk.Entry(root, width = 6, relief='solid')
		entr_A2_I.place(x = 25, y = 90)
		global entr_A3_I
		entr_A3_I = tk.Entry(root, width = 6, relief='solid')
		entr_A3_I.place(x = 25, y = 110)


		if check_var_gr_3.get() >= 4:

			global entr_A4_I
			entr_A4_I = tk.Entry(root, width = 6, relief='solid')
			entr_A4_I.place(x = 25, y = 130)

		if check_var_gr_3.get() >= 5:	

			global entr_A5_I
			entr_A5_I = tk.Entry(root, width = 6, relief='solid')
			entr_A5_I.place(x = 25, y = 150)


		if check_var_gr_3.get() >= 6:
			
			global entr_A6_I
			entr_A6_I = tk.Entry(root, width = 6, relief='solid')
			entr_A6_I.place(x = 25, y = 170)

		if check_var_gr_3.get() >= 7:

			global entr_A7_I
			entr_A7_I = tk.Entry(root, width = 6, relief='solid')
			entr_A7_I.place(x = 25, y = 190)

		if check_var_gr_3.get() >= 8:

			global entr_A8_I
			entr_A8_I = tk.Entry(root, width = 6, relief='solid')
			entr_A8_I.place(x = 25, y = 210)

		if check_var_gr_3.get() >= 9:	

			global entr_A9_I
			entr_A9_I = tk.Entry(root, width = 6, relief='solid')
			entr_A9_I.place(x = 25, y = 230)

		if check_var_gr_3.get() >= 10:	
			global entr_A10_I
			entr_A10_I = tk.Entry(root, width = 6, relief='solid')
			entr_A10_I.place(x = 25, y = 250)
			
		
		if check_var_gr_2.get() >= 2:

			labl_B_1_I  = tk.Label(root, text = "Обр. 2")
			labl_B_1_I.place(x = 75, y = 50)	

			global entr_B1_I
			entr_B1_I = tk.Entry(root, width = 6, relief='solid')
			entr_B1_I.place(x = 75, y = 70)
			global entr_B2_I
			entr_B2_I = tk.Entry(root, width = 6, relief='solid')
			entr_B2_I.place(x = 75, y = 90)
			global entr_B3_I
			entr_B3_I = tk.Entry(root, width = 6, relief='solid')
			entr_B3_I.place(x = 75, y = 110)


			if check_var_gr_3.get() >= 4:

				global entr_B4_I
				entr_B4_I = tk.Entry(root, width = 6, relief='solid')
				entr_B4_I.place(x = 75, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_B5_I
				entr_B5_I = tk.Entry(root, width = 6, relief='solid')
				entr_B5_I.place(x = 75, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_B6_I
				entr_B6_I = tk.Entry(root, width = 6, relief='solid')
				entr_B6_I.place(x = 75, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_B7_I
				entr_B7_I = tk.Entry(root, width = 6, relief='solid')
				entr_B7_I.place(x = 75, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_B8_I
				entr_B8_I = tk.Entry(root, width = 6, relief='solid')
				entr_B8_I.place(x = 75, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_B9_I
				entr_B9_I = tk.Entry(root, width = 6, relief='solid')
				entr_B9_I.place(x = 75, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_B10_I
				entr_B10_I = tk.Entry(root, width = 6, relief='solid')
				entr_B10_I.place(x = 75, y = 250)
			

		
		if check_var_gr_2.get() >= 3:

			labl_C_1_I  = tk.Label(root, text = "Обр. 3")
			labl_C_1_I.place(x = 125, y = 50)

			global entr_C1_I 
			entr_C1_I = tk.Entry(root, width = 6, relief='solid')
			entr_C1_I.place(x = 125, y = 70)
			global entr_C2_I
			entr_C2_I = tk.Entry(root, width = 6, relief='solid')
			entr_C2_I.place(x = 125, y = 90)
			global entr_C3_I
			entr_C3_I = tk.Entry(root, width = 6, relief='solid')
			entr_C3_I.place(x = 125, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_C4_I
				entr_C4_I = tk.Entry(root, width = 6, relief='solid')
				entr_C4_I.place(x = 125, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_C5_I
				entr_C5_I = tk.Entry(root, width = 6, relief='solid')
				entr_C5_I.place(x = 125, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_C6_I
				entr_C6_I = tk.Entry(root, width = 6, relief='solid')
				entr_C6_I.place(x = 125, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_C7_I
				entr_C7_I = tk.Entry(root, width = 6, relief='solid')
				entr_C7_I.place(x = 125, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_C8_I
				entr_C8_I = tk.Entry(root, width = 6, relief='solid')
				entr_C8_I.place(x = 125, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_C9_I
				entr_C9_I = tk.Entry(root, width = 6, relief='solid')
				entr_C9_I.place(x = 125, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_C10_I
				entr_C10_I = tk.Entry(root, width = 6, relief='solid')
				entr_C10_I.place(x = 125, y = 250)


		if check_var_gr_2.get() >= 4:

			labl_D_1_I  = tk.Label(root, text = "Обр. 4")
			labl_D_1_I.place(x = 175, y = 50)

			global entr_D1_I 
			entr_D1_I = tk.Entry(root, width = 6, relief='solid')
			entr_D1_I.place(x = 175, y = 70)
			global entr_D2_I
			entr_D2_I = tk.Entry(root, width = 6, relief='solid')
			entr_D2_I.place(x = 175, y = 90)
			global entr_D3_I
			entr_D3_I = tk.Entry(root, width = 6, relief='solid')
			entr_D3_I.place(x = 175, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_D4_I
				entr_D4_I = tk.Entry(root, width = 6, relief='solid')
				entr_D4_I.place(x = 175, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_D5_I
				entr_D5_I = tk.Entry(root, width = 6, relief='solid')
				entr_D5_I.place(x = 175, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_D6_I
				entr_D6_I = tk.Entry(root, width = 6, relief='solid')
				entr_D6_I.place(x = 175, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_D7_I
				entr_D7_I = tk.Entry(root, width = 6, relief='solid')
				entr_D7_I.place(x = 175, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_D8_I
				entr_D8_I = tk.Entry(root, width = 6, relief='solid')
				entr_D8_I.place(x = 175, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_D9_I
				entr_D9_I = tk.Entry(root, width = 6, relief='solid')
				entr_D9_I.place(x = 175, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_D10_I
				entr_D10_I = tk.Entry(root, width = 6, relief='solid')
				entr_D10_I.place(x = 175, y = 250)



		if check_var_gr_2.get() >= 5:

			labl_E_1_I  = tk.Label(root, text = "Обр. 5")
			labl_E_1_I.place(x = 225, y = 50)

			global entr_E1_I 
			entr_E1_I = tk.Entry(root, width = 6, relief='solid')
			entr_E1_I.place(x = 225, y = 70)
			global entr_E2_I
			entr_E2_I = tk.Entry(root, width = 6, relief='solid')
			entr_E2_I.place(x = 225, y = 90)
			global entr_E3_I
			entr_E3_I = tk.Entry(root, width = 6, relief='solid')
			entr_E3_I.place(x = 225, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_E4_I
				entr_E4_I = tk.Entry(root, width = 6, relief='solid')
				entr_E4_I.place(x = 225, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_E5_I
				entr_E5_I = tk.Entry(root, width = 6, relief='solid')
				entr_E5_I.place(x = 225, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_E6_I
				entr_E6_I = tk.Entry(root, width = 6, relief='solid')
				entr_E6_I.place(x = 225, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_E7_I
				entr_E7_I = tk.Entry(root, width = 6, relief='solid')
				entr_E7_I.place(x = 225, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_E8_I
				entr_E8_I = tk.Entry(root, width = 6, relief='solid')
				entr_E8_I.place(x = 225, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_E9_I
				entr_E9_I = tk.Entry(root, width = 6, relief='solid')
				entr_E9_I.place(x = 225, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_E10_I
				entr_E10_I = tk.Entry(root, width = 6, relief='solid')
				entr_E10_I.place(x = 225, y = 250)



		if check_var_gr_2.get() >= 6:

			labl_F_1_I  = tk.Label(root, text = "Обр. 6")
			labl_F_1_I.place(x = 275, y = 50)

			global entr_F1_I 
			entr_F1_I = tk.Entry(root, width = 6, relief='solid')
			entr_F1_I.place(x = 275, y = 70)
			global entr_F2_I
			entr_F2_I = tk.Entry(root, width = 6, relief='solid')
			entr_F2_I.place(x = 275, y = 90)
			global entr_F3_I
			entr_F3_I = tk.Entry(root, width = 6, relief='solid')
			entr_F3_I.place(x = 275, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_F4_I
				entr_F4_I = tk.Entry(root, width = 6, relief='solid')
				entr_F4_I.place(x = 275, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_F5_I
				entr_F5_I = tk.Entry(root, width = 6, relief='solid')
				entr_F5_I.place(x = 275, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_F6_I
				entr_F6_I = tk.Entry(root, width = 6, relief='solid')
				entr_F6_I.place(x = 275, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_F7_I
				entr_F7_I = tk.Entry(root, width = 6, relief='solid')
				entr_F7_I.place(x = 275, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_F8_I
				entr_F8_I = tk.Entry(root, width = 6, relief='solid')
				entr_F8_I.place(x = 275, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_F9_I
				entr_F9_I = tk.Entry(root, width = 6, relief='solid')
				entr_F9_I.place(x = 275, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_F10_I
				entr_F10_I = tk.Entry(root, width = 6, relief='solid')
				entr_F10_I.place(x = 275, y = 250)



	######################################################################################################
	if check_var_gr_1.get() >= 2:

		
		labl_MAIN_II = tk.Label(root, text = 'QCB')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_MAIN_II.place(x = 222, y = 5)
		else:
			labl_MAIN_II.place(x = 362, y = 5)

		global entr_MAIN_II
		entr_MAIN_II = tk.Entry(root, width = 6, relief= 'solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_MAIN_II.place(x = 220, y = 25)
		else:
			entr_MAIN_II.place(x = 360, y = 25)
		
		labl_A_1_II  = tk.Label(root, text = "Обр. 1")
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_A_1_II.place(x = 220, y = 50)
		else:	
			labl_A_1_II.place(x = 360, y = 50)

		global entr_A1_II
		entr_A1_II = tk.Entry(root, width = 6, relief= 'solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A1_II.place(x = 220, y = 70)
		else:
			entr_A1_II.place(x = 360, y = 70)

		global entr_A2_II
		entr_A2_II = tk.Entry(root, width = 6, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A2_II.place(x = 220, y = 90)
		else:
			entr_A2_II.place(x = 360, y = 90)

		global entr_A3_II
		entr_A3_II = tk.Entry(root, width = 6, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A3_II.place(x = 220, y = 110)
		else:
			entr_A3_II.place(x = 360, y = 110)


		if check_var_gr_3.get() >= 4:

			global entr_A4_II
			entr_A4_II = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_A4_II.place(x = 220, y = 130)
			else:
				entr_A4_II.place(x = 360, y = 130)

		if check_var_gr_3.get() >= 5:	

			global entr_A5_II
			entr_A5_II = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_A5_II.place(x = 220, y = 150)
			else:
				entr_A5_II.place(x = 360, y = 150)


		if check_var_gr_3.get() >= 6:
			
			global entr_A6_II
			entr_A6_II = tk.Entry(root, width = 6, relief='solid')
			entr_A6_II.place(x = 360, y = 170)

		if check_var_gr_3.get() >= 7:

			global entr_A7_II
			entr_A7_II = tk.Entry(root, width = 6, relief='solid')
			entr_A7_II.place(x = 360, y = 190)

		if check_var_gr_3.get() >= 8:

			global entr_A8_II
			entr_A8_II = tk.Entry(root, width = 6, relief='solid')
			entr_A8_II.place(x = 360, y = 210)

		if check_var_gr_3.get() >= 9:	

			global entr_A9_II
			entr_A9_II = tk.Entry(root, width = 6, relief='solid')
			entr_A9_II.place(x = 360, y = 230)

		if check_var_gr_3.get() >= 10:	
			global entr_A10_II
			entr_A10_II = tk.Entry(root, width = 6, relief='solid')
			entr_A10_II.place(x = 360, y = 250)
			
		
		if check_var_gr_2.get() >= 2:

			labl_B_1_II  = tk.Label(root, text = "Обр. 2")
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				labl_B_1_II.place(x = 270, y = 50)	
			else:
				labl_B_1_II.place(x = 410, y = 50)	

			global entr_B1_II
			entr_B1_II = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B1_II.place(x = 270, y = 70)
			else:
				entr_B1_II.place(x = 410, y = 70)

			global entr_B2_II
			entr_B2_II = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B2_II.place(x = 270, y = 90)
			else:
				entr_B2_II.place(x = 410, y = 90)

			global entr_B3_II
			entr_B3_II = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B3_II.place(x = 270, y = 110)
			else:
				entr_B3_II.place(x = 410, y = 110)


			if check_var_gr_3.get() >= 4:

				global entr_B4_II
				entr_B4_II = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_B4_II.place(x = 270, y = 130)
				else:
					entr_B4_II.place(x = 410, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_B5_II
				entr_B5_II = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_B5_II.place(x = 270, y = 150)
				else:
					entr_B5_II.place(x = 410, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_B6_II
				entr_B6_II = tk.Entry(root, width = 6, relief='solid')
				entr_B6_II.place(x = 410, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_B7_II
				entr_B7_II = tk.Entry(root, width = 6, relief='solid')
				entr_B7_II.place(x = 410, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_B8_II
				entr_B8_II = tk.Entry(root, width = 6, relief='solid')
				entr_B8_II.place(x = 410, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_B9_II
				entr_B9_II = tk.Entry(root, width = 6, relief='solid')
				entr_B9_II.place(x = 410, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_B10_II
				entr_B10_II = tk.Entry(root, width = 6, relief='solid')
				entr_B10_II.place(x = 410, y = 250)
			

		
		if check_var_gr_2.get() >= 3:

			labl_C_1_II  = tk.Label(root, text = "Обр. 3")
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				labl_C_1_II.place(x = 320, y = 50)
			else:
				labl_C_1_II.place(x = 460, y = 50)

			global entr_C1_II 
			entr_C1_II = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C1_II.place(x = 320, y = 70)
			else:
				entr_C1_II.place(x = 460, y = 70)

			global entr_C2_II
			entr_C2_II = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C2_II.place(x = 320, y = 90)
			else:
				entr_C2_II.place(x = 460, y = 90)

			global entr_C3_II
			entr_C3_II = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C3_II.place(x = 320, y = 110)
			else:
				entr_C3_II.place(x = 460, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_C4_II
				entr_C4_II = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_C4_II.place(x = 320, y = 130)
				else:
					entr_C4_II.place(x = 460, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_C5_II
				entr_C5_II = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_C5_II.place(x = 320, y = 150)
				else:
					entr_C5_II.place(x = 460, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_C6_II
				entr_C6_II = tk.Entry(root, width = 6, relief='solid')
				entr_C6_II.place(x = 460, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_C7_II
				entr_C7_II = tk.Entry(root, width = 6, relief='solid')
				entr_C7_II.place(x = 460, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_C8_II
				entr_C8_II = tk.Entry(root, width = 6, relief='solid')
				entr_C8_II.place(x = 460, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_C9_II
				entr_C9_II = tk.Entry(root, width = 6, relief='solid')
				entr_C9_II.place(x = 460, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_C10_II
				entr_C10_II = tk.Entry(root, width = 6, relief='solid')
				entr_C10_II.place(x = 460, y = 250)


		if check_var_gr_2.get() >= 4:

			labl_D_1_II  = tk.Label(root, text = "Обр. 4")
			labl_D_1_II.place(x = 510, y = 50)

			global entr_D1_II 
			entr_D1_II = tk.Entry(root, width = 6, relief='solid')
			entr_D1_II.place(x = 510, y = 70)
			global entr_D2_II
			entr_D2_II = tk.Entry(root, width = 6, relief='solid')
			entr_D2_II.place(x = 510, y = 90)
			global entr_D3_II
			entr_D3_II = tk.Entry(root, width = 6, relief='solid')
			entr_D3_II.place(x = 510, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_D4_II
				entr_D4_II = tk.Entry(root, width = 6, relief='solid')
				entr_D4_II.place(x = 510, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_D5_II
				entr_D5_II = tk.Entry(root, width = 6, relief='solid')
				entr_D5_II.place(x = 510, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_D6_II
				entr_D6_II = tk.Entry(root, width = 6, relief='solid')
				entr_D6_II.place(x = 510, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_D7_II
				entr_D7_II = tk.Entry(root, width = 6, relief='solid')
				entr_D7_II.place(x = 510, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_D8_II
				entr_D8_II = tk.Entry(root, width = 6, relief='solid')
				entr_D8_II.place(x = 510, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_D9_II
				entr_D9_II = tk.Entry(root, width = 6, relief='solid')
				entr_D9_II.place(x = 510, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_D10_II
				entr_D10_II = tk.Entry(root, width = 6, relief='solid')
				entr_D10_II.place(x = 510, y = 250)



		if check_var_gr_2.get() >= 5:

			labl_E_1_II  = tk.Label(root, text = "Обр. 5")
			labl_E_1_II.place(x = 560, y = 50)

			global entr_E1_II 
			entr_E1_II = tk.Entry(root, width = 6, relief='solid')
			entr_E1_II.place(x = 560, y = 70)
			global entr_E2_II
			entr_E2_II = tk.Entry(root, width = 6, relief='solid')
			entr_E2_II.place(x = 560, y = 90)
			global entr_E3_II
			entr_E3_II = tk.Entry(root, width = 6, relief='solid')
			entr_E3_II.place(x = 560, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_E4_II
				entr_E4_II = tk.Entry(root, width = 6, relief='solid')
				entr_E4_II.place(x = 560, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_E5_II
				entr_E5_II = tk.Entry(root, width = 6, relief='solid')
				entr_E5_II.place(x = 560, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_E6_II
				entr_E6_II = tk.Entry(root, width = 6, relief='solid')
				entr_E6_II.place(x = 560, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_E7_II
				entr_E7_II = tk.Entry(root, width = 6, relief='solid')
				entr_E7_II.place(x = 560, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_E8_II
				entr_E8_II = tk.Entry(root, width = 6, relief='solid')
				entr_E8_II.place(x = 560, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_E9_II
				entr_E9_II = tk.Entry(root, width = 6, relief='solid')
				entr_E9_II.place(x = 560, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_E10_II
				entr_E10_II = tk.Entry(root, width = 6, relief='solid')
				entr_E10_II.place(x = 560, y = 250)



		if check_var_gr_2.get() >= 6:

			labl_F_1_II  = tk.Label(root, text = "Обр. 6")
			labl_F_1_II.place(x = 610, y = 50)

			global entr_F1_II 
			entr_F1_II = tk.Entry(root, width = 6, relief='solid')
			entr_F1_II.place(x = 610, y = 70)
			global entr_F2_II
			entr_F2_II = tk.Entry(root, width = 6, relief='solid')
			entr_F2_II.place(x = 610, y = 90)
			global entr_F3_II
			entr_F3_II = tk.Entry(root, width = 6, relief='solid')
			entr_F3_II.place(x = 610, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_F4_II
				entr_F4_II = tk.Entry(root, width = 6, relief='solid')
				entr_F4_II.place(x = 610, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_F5_II
				entr_F5_II = tk.Entry(root, width = 6, relief='solid')
				entr_F5_II.place(x = 610, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_F6_II
				entr_F6_II = tk.Entry(root, width = 6, relief='solid')
				entr_F6_II.place(x = 610, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_F7_II
				entr_F7_II = tk.Entry(root, width = 6, relief='solid')
				entr_F7_II.place(x = 610, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_F8_II
				entr_F8_II = tk.Entry(root, width = 6, relief='solid')
				entr_F8_II.place(x = 610, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_F9_II
				entr_F9_II = tk.Entry(root, width = 6, relief='solid')
				entr_F9_II.place(x = 610, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_F10_II
				entr_F10_II = tk.Entry(root, width = 6, relief='solid')
				entr_F10_II.place(x = 610, y = 250)

	##########################################################################################

	if check_var_gr_1.get() >= 3:

		labl_MAIN_III = tk.Label(root, text = 'QCC')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_MAIN_III.place(x = 417, y = 5)
		else:
			labl_MAIN_III.place(x = 697, y = 5)

		global entr_MAIN_III
		entr_MAIN_III = tk.Entry(root, width = 6, relief= 'solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_MAIN_III.place(x = 415, y = 25)
		else:
			entr_MAIN_III.place(x = 695, y = 25)
		
		labl_A_1_III  = tk.Label(root, text = "Обр. 1")
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_A_1_III.place(x = 415, y = 50)
		else:
			labl_A_1_III.place(x = 695, y = 50)

		global entr_A1_III
		entr_A1_III = tk.Entry(root, width = 6, relief= 'solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A1_III.place(x = 415, y = 70)
		else:
			entr_A1_III.place(x = 695, y = 70)

		global entr_A2_III
		entr_A2_III = tk.Entry(root, width = 6, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A2_III.place(x = 415, y = 90)
		else:
			entr_A2_III.place(x = 695, y = 90)

		global entr_A3_III
		entr_A3_III = tk.Entry(root, width = 6, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A3_III.place(x = 415, y = 110)
		else:
			entr_A3_III.place(x = 695, y = 110)


		if check_var_gr_3.get() >= 4:

			global entr_A4_III
			entr_A4_III = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_A4_III.place(x = 415, y = 130)
			else:
				entr_A4_III.place(x = 695, y = 130)

		if check_var_gr_3.get() >= 5:	

			global entr_A5_III
			entr_A5_III = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_A5_III.place(x = 415, y = 150)
			else:
				entr_A5_III.place(x = 695, y = 150)


		if check_var_gr_3.get() >= 6:
			
			global entr_A6_III
			entr_A6_III = tk.Entry(root, width = 6, relief='solid')
			entr_A6_III.place(x = 695, y = 170)

		if check_var_gr_3.get() >= 7:

			global entr_A7_III
			entr_A7_III = tk.Entry(root, width = 6, relief='solid')
			entr_A7_III.place(x = 695, y = 190)

		if check_var_gr_3.get() >= 8:

			global entr_A8_III
			entr_A8_III = tk.Entry(root, width = 6, relief='solid')
			entr_A8_III.place(x = 695, y = 210)

		if check_var_gr_3.get() >= 9:	

			global entr_A9_III
			entr_A9_III = tk.Entry(root, width = 6, relief='solid')
			entr_A9_III.place(x = 695, y = 230)

		if check_var_gr_3.get() >= 10:	
			global entr_A10_III
			entr_A10_III = tk.Entry(root, width = 6, relief='solid')
			entr_A10_III.place(x = 695, y = 250)
			
		
		if check_var_gr_2.get() >= 2:

			labl_B_1_III  = tk.Label(root, text = "Обр. 2")
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				labl_B_1_III.place(x = 465, y = 50)	
			else:
				labl_B_1_III.place(x = 745, y = 50)	

			global entr_B1_III
			entr_B1_III = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B1_III.place(x = 465, y = 70)
			else:
				entr_B1_III.place(x = 745, y = 70)

			global entr_B2_III
			entr_B2_III = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B2_III.place(x = 465, y = 90)
			else:
				entr_B2_III.place(x = 745, y = 90)

			global entr_B3_III
			entr_B3_III = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B3_III.place(x = 465, y = 110)
			else:
				entr_B3_III.place(x = 745, y = 110)


			if check_var_gr_3.get() >= 4:

				global entr_B4_III
				entr_B4_III = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_B4_III.place(x = 465, y = 130)
				else:
					entr_B4_III.place(x = 745, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_B5_III
				entr_B5_III = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_B5_III.place(x = 465, y = 150)
				else:
					entr_B5_III.place(x = 745, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_B6_III
				entr_B6_III = tk.Entry(root, width = 6, relief='solid')
				entr_B6_III.place(x = 745, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_B7_III
				entr_B7_III = tk.Entry(root, width = 6, relief='solid')
				entr_B7_III.place(x = 745, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_B8_III
				entr_B8_III = tk.Entry(root, width = 6, relief='solid')
				entr_B8_III.place(x = 745, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_B9_III
				entr_B9_III = tk.Entry(root, width = 6, relief='solid')
				entr_B9_III.place(x = 745, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_B10_III
				entr_B10_III = tk.Entry(root, width = 6, relief='solid')
				entr_B10_III.place(x = 745, y = 250)
			

		
		if check_var_gr_2.get() >= 3:

			labl_C_1_III  = tk.Label(root, text = "Обр. 3")
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				labl_C_1_III.place(x = 515, y = 50)
			else:
				labl_C_1_III.place(x = 795, y = 50)

			global entr_C1_III 
			entr_C1_III = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C1_III.place(x = 515, y = 70)
			else:
				entr_C1_III.place(x = 795, y = 70)

			global entr_C2_III
			entr_C2_III = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C2_III.place(x = 515, y = 90)
			else:
				entr_C2_III.place(x = 795, y = 90)

			global entr_C3_III
			entr_C3_III = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C3_III.place(x = 515, y = 110)
			else:
				entr_C3_III.place(x = 795, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_C4_III
				entr_C4_III = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_C4_III.place(x = 515, y = 130)
				else:
					entr_C4_III.place(x = 795, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_C5_III
				entr_C5_III = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_C5_III.place(x = 515, y = 150)
				else:
					entr_C5_III.place(x = 795, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_C6_III
				entr_C6_III = tk.Entry(root, width = 6, relief='solid')
				entr_C6_III.place(x = 795, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_C7_III
				entr_C7_III = tk.Entry(root, width = 6, relief='solid')
				entr_C7_III.place(x = 795, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_C8_III
				entr_C8_III = tk.Entry(root, width = 6, relief='solid')
				entr_C8_III.place(x = 795, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_C9_III
				entr_C9_III = tk.Entry(root, width = 6, relief='solid')
				entr_C9_III.place(x = 795, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_C10_III
				entr_C10_III = tk.Entry(root, width = 6, relief='solid')
				entr_C10_III.place(x = 795, y = 250)


		if check_var_gr_2.get() >= 4:

			labl_D_1_III  = tk.Label(root, text = "Обр. 4")
			labl_D_1_III.place(x = 845, y = 50)

			global entr_D1_III 
			entr_D1_III = tk.Entry(root, width = 6, relief='solid')
			entr_D1_III.place(x = 845, y = 70)
			global entr_D2_III
			entr_D2_III = tk.Entry(root, width = 6, relief='solid')
			entr_D2_III.place(x = 845, y = 90)
			global entr_D3_III
			entr_D3_III = tk.Entry(root, width = 6, relief='solid')
			entr_D3_III.place(x = 845, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_D4_III
				entr_D4_III = tk.Entry(root, width = 6, relief='solid')
				entr_D4_III.place(x = 845, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_D5_III
				entr_D5_III = tk.Entry(root, width = 6, relief='solid')
				entr_D5_III.place(x = 845, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_D6_III
				entr_D6_III = tk.Entry(root, width = 6, relief='solid')
				entr_D6_III.place(x = 845, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_D7_III
				entr_D7_III = tk.Entry(root, width = 6, relief='solid')
				entr_D7_III.place(x = 845, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_D8_III
				entr_D8_III = tk.Entry(root, width = 6, relief='solid')
				entr_D8_III.place(x = 845, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_D9_III
				entr_D9_III = tk.Entry(root, width = 6, relief='solid')
				entr_D9_III.place(x = 845, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_D10_III
				entr_D10_III = tk.Entry(root, width = 6, relief='solid')
				entr_D10_III.place(x = 845, y = 250)



		if check_var_gr_2.get() >= 5:

			labl_E_1_III  = tk.Label(root, text = "Обр. 5")
			labl_E_1_III.place(x = 895, y = 50)

			global entr_E1_III 
			entr_E1_III = tk.Entry(root, width = 6, relief='solid')
			entr_E1_III.place(x = 895, y = 70)
			global entr_E2_III
			entr_E2_III = tk.Entry(root, width = 6, relief='solid')
			entr_E2_III.place(x = 895, y = 90)
			global entr_E3_III
			entr_E3_III = tk.Entry(root, width = 6, relief='solid')
			entr_E3_III.place(x = 895, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_E4_III
				entr_E4_III = tk.Entry(root, width = 6, relief='solid')
				entr_E4_III.place(x = 895, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_E5_III
				entr_E5_III = tk.Entry(root, width = 6, relief='solid')
				entr_E5_III.place(x = 895, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_E6_III
				entr_E6_III = tk.Entry(root, width = 6, relief='solid')
				entr_E6_III.place(x = 895, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_E7_III
				entr_E7_III = tk.Entry(root, width = 6, relief='solid')
				entr_E7_III.place(x = 895, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_E8_III
				entr_E8_III = tk.Entry(root, width = 6, relief='solid')
				entr_E8_III.place(x = 895, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_E9_III
				entr_E9_III = tk.Entry(root, width = 6, relief='solid')
				entr_E9_III.place(x = 895, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_E10_III
				entr_E10_III = tk.Entry(root, width = 6, relief='solid')
				entr_E10_III.place(x = 895, y = 250)



		if check_var_gr_2.get() >= 6:

			labl_F_1_III  = tk.Label(root, text = "Обр. 6")
			labl_F_1_III.place(x = 945, y = 50)

			global entr_F1_III 
			entr_F1_III = tk.Entry(root, width = 6, relief='solid')
			entr_F1_III.place(x = 945, y = 70)
			global entr_F2_III
			entr_F2_III = tk.Entry(root, width = 6, relief='solid')
			entr_F2_III.place(x = 945, y = 90)
			global entr_F3_III
			entr_F3_III = tk.Entry(root, width = 6, relief='solid')
			entr_F3_III.place(x = 945, y = 110)

			if check_var_gr_3.get() >= 4:

				global entr_F4_III
				entr_F4_III = tk.Entry(root, width = 6, relief='solid')
				entr_F4_III.place(x = 945, y = 130)

			if check_var_gr_3.get() >= 5:

				global entr_F5_III
				entr_F5_III = tk.Entry(root, width = 6, relief='solid')
				entr_F5_III.place(x = 945, y = 150)

			if check_var_gr_3.get() >= 6:

				global entr_F6_III
				entr_F6_III = tk.Entry(root, width = 6, relief='solid')
				entr_F6_III.place(x = 945, y = 170)

			if check_var_gr_3.get() >= 7:

				global entr_F7_III
				entr_F7_III = tk.Entry(root, width = 6, relief='solid')
				entr_F7_III.place(x = 945, y = 190)

			if check_var_gr_3.get() >= 8:

				global entr_F8_III
				entr_F8_III = tk.Entry(root, width = 6, relief='solid')
				entr_F8_III.place(x = 945, y = 210)

			if check_var_gr_3.get() >= 9:

				global entr_F9_III
				entr_F9_III = tk.Entry(root, width = 6, relief='solid')
				entr_F9_III.place(x = 945, y = 230)

			if check_var_gr_3.get() >= 10:

				global entr_F10_III
				entr_F10_III = tk.Entry(root, width = 6, relief='solid')
				entr_F10_III.place(x = 945, y = 250)

	#####################################################################################################################################
	if check_var_gr_1.get() >= 4:

		
		labl_MAIN_IV = tk.Label(root, text = 'QCD')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_MAIN_IV.place(x = 612, y = 5)
		else:
			labl_MAIN_IV.place(x = 27, y = 300)

		global entr_MAIN_IV
		entr_MAIN_IV = tk.Entry(root, width = 6, relief= 'solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_MAIN_IV.place(x = 610, y = 25)
		else:
			entr_MAIN_IV.place(x = 25, y = 320)
		
		labl_A_1_IV  = tk.Label(root, text = "Обр. 1")
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_A_1_IV.place(x = 610, y = 50)
		else:
			labl_A_1_IV.place(x = 25, y = 345)

		global entr_A1_IV
		entr_A1_IV = tk.Entry(root, width = 6, relief= 'solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A1_IV.place(x = 610, y = 70)
		else:
			entr_A1_IV.place(x = 25, y = 365)

		global entr_A2_IV
		entr_A2_IV = tk.Entry(root, width = 6, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A2_IV.place(x = 610, y = 90)
		else:
			entr_A2_IV.place(x = 25, y = 385)

		global entr_A3_IV
		entr_A3_IV = tk.Entry(root, width = 6, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_A3_IV.place(x = 610, y = 110)
		else:
			entr_A3_IV.place(x = 25, y = 405)


		if check_var_gr_3.get() >= 4:

			global entr_A4_IV
			entr_A4_IV = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_A4_IV.place(x = 610, y = 130)
			else:
				entr_A4_IV.place(x = 25, y = 425)

		if check_var_gr_3.get() >= 5:	

			global entr_A5_IV
			entr_A5_IV = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_A5_IV.place(x = 610, y = 150)
			else:
				entr_A5_IV.place(x = 25, y = 445)


		if check_var_gr_3.get() >= 6:
			
			global entr_A6_IV
			entr_A6_IV = tk.Entry(root, width = 6, relief='solid')
			entr_A6_IV.place(x = 25, y = 465)

		if check_var_gr_3.get() >= 7:

			global entr_A7_IV
			entr_A7_IV = tk.Entry(root, width = 6, relief='solid')
			entr_A7_IV.place(x = 25, y = 485)

		if check_var_gr_3.get() >= 8:

			global entr_A8_IV
			entr_A8_IV = tk.Entry(root, width = 6, relief='solid')
			entr_A8_IV.place(x = 25, y = 505)

		if check_var_gr_3.get() >= 9:	

			global entr_A9_IV
			entr_A9_IV = tk.Entry(root, width = 6, relief='solid')
			entr_A9_IV.place(x = 25, y = 525)

		if check_var_gr_3.get() >= 10:	
			global entr_A10_IV
			entr_A10_IV = tk.Entry(root, width = 6, relief='solid')
			entr_A10_IV.place(x = 25, y = 545)
			
		


		if check_var_gr_2.get() >= 2:

			labl_B_1_IV  = tk.Label(root, text = "Обр. 2")
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				labl_B_1_IV.place(x = 660, y = 50)	
			else:
				labl_B_1_IV.place(x = 75, y = 345)	

			global entr_B1_IV
			entr_B1_IV = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B1_IV.place(x = 660, y = 70)
			else:
				entr_B1_IV.place(x = 75, y = 365)

			global entr_B2_IV
			entr_B2_IV = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B2_IV.place(x = 660, y = 90)
			else:
				entr_B2_IV.place(x = 75, y = 385)

			global entr_B3_IV
			entr_B3_IV = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_B3_IV.place(x = 660, y = 110)
			else:
				entr_B3_IV.place(x = 75, y = 405)


			if check_var_gr_3.get() >= 4:

				global entr_B4_IV
				entr_B4_IV = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_B4_IV.place(x = 660, y = 130)
				else:
					entr_B4_IV.place(x = 75, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_B5_IV
				entr_B5_IV = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_B5_IV.place(x = 660, y = 150)
				else:
					entr_B5_IV.place(x = 75, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_B6_IV
				entr_B6_IV = tk.Entry(root, width = 6, relief='solid')
				entr_B6_IV.place(x = 75, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_B7_IV
				entr_B7_IV = tk.Entry(root, width = 6, relief='solid')
				entr_B7_IV.place(x = 75, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_B8_IV
				entr_B8_IV = tk.Entry(root, width = 6, relief='solid')
				entr_B8_IV.place(x = 75, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_B9_IV
				entr_B9_IV = tk.Entry(root, width = 6, relief='solid')
				entr_B9_IV.place(x = 75, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_B10_IV
				entr_B10_IV = tk.Entry(root, width = 6, relief='solid')
				entr_B10_IV.place(x = 75, y = 545)
			

		
		if check_var_gr_2.get() >= 3:

			labl_C_1_IV  = tk.Label(root, text = "Обр. 3")
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				labl_C_1_IV.place(x = 710, y = 50)
			else:
				labl_C_1_IV.place(x = 125, y = 345)

			global entr_C1_IV
			entr_C1_IV = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C1_IV.place(x = 710, y = 70)
			else:
				entr_C1_IV.place(x = 125, y = 365)

			global entr_C2_IV
			entr_C2_IV = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C2_IV.place(x = 710, y = 90)
			else:
				entr_C2_IV.place(x = 125, y = 385)

			global entr_C3_IV
			entr_C3_IV = tk.Entry(root, width = 6, relief='solid')
			if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
				entr_C3_IV.place(x = 710, y = 110)
			else:
				entr_C3_IV.place(x = 125, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_C4_IV
				entr_C4_IV = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_C4_IV.place(x = 710, y = 130)
				else:
					entr_C4_IV.place(x = 125, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_C5_IV
				entr_C5_IV = tk.Entry(root, width = 6, relief='solid')
				if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
					entr_C5_IV.place(x = 710, y = 150)
				else:
					entr_C5_IV.place(x = 125, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_C6_IV
				entr_C6_IV = tk.Entry(root, width = 6, relief='solid')
				entr_C6_IV.place(x = 125, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_C7_IV
				entr_C7_IV = tk.Entry(root, width = 6, relief='solid')
				entr_C7_IV.place(x = 125, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_C8_IV
				entr_C8_IV = tk.Entry(root, width = 6, relief='solid')
				entr_C8_IV.place(x = 125, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_C9_IV
				entr_C9_IV = tk.Entry(root, width = 6, relief='solid')
				entr_C9_IV.place(x = 125, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_C10_IV
				entr_C10_IV = tk.Entry(root, width = 6, relief='solid')
				entr_C10_IV.place(x = 125, y = 545)


		if check_var_gr_2.get() >= 4:

			labl_D_1_IV  = tk.Label(root, text = "Обр. 4")
			labl_D_1_IV.place(x = 175, y = 345)

			global entr_D1_IV
			entr_D1_IV = tk.Entry(root, width = 6, relief='solid')
			entr_D1_IV.place(x = 175, y = 365)
			global entr_D2_IV
			entr_D2_IV = tk.Entry(root, width = 6, relief='solid')
			entr_D2_IV.place(x = 175, y = 385)
			global entr_D3_IV
			entr_D3_IV = tk.Entry(root, width = 6, relief='solid')
			entr_D3_IV.place(x = 175, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_D4_IV
				entr_D4_IV = tk.Entry(root, width = 6, relief='solid')
				entr_D4_IV.place(x = 175, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_D5_IV
				entr_D5_IV = tk.Entry(root, width = 6, relief='solid')
				entr_D5_IV.place(x = 175, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_D6_IV
				entr_D6_IV = tk.Entry(root, width = 6, relief='solid')
				entr_D6_IV.place(x = 175, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_D7_IV
				entr_D7_IV = tk.Entry(root, width = 6, relief='solid')
				entr_D7_IV.place(x = 175, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_D8_IV
				entr_D8_IV = tk.Entry(root, width = 6, relief='solid')
				entr_D8_IV.place(x = 175, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_D9_IV
				entr_D9_IV = tk.Entry(root, width = 6, relief='solid')
				entr_D9_IV.place(x = 175, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_D10_IV
				entr_D10_IV = tk.Entry(root, width = 6, relief='solid')
				entr_D10_IV.place(x = 175, y = 545)



		if check_var_gr_2.get() >= 5:

			labl_E_1_IV  = tk.Label(root, text = "Обр. 5")
			labl_E_1_IV.place(x = 225, y = 345)

			global entr_E1_IV 
			entr_E1_IV = tk.Entry(root, width = 6, relief='solid')
			entr_E1_IV.place(x = 225, y = 365)
			global entr_E2_IV
			entr_E2_IV = tk.Entry(root, width = 6, relief='solid')
			entr_E2_IV.place(x = 225, y = 385)
			global entr_E3_IV
			entr_E3_IV = tk.Entry(root, width = 6, relief='solid')
			entr_E3_IV.place(x = 225, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_E4_IV
				entr_E4_IV = tk.Entry(root, width = 6, relief='solid')
				entr_E4_IV.place(x = 225, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_E5_IV
				entr_E5_IV = tk.Entry(root, width = 6, relief='solid')
				entr_E5_IV.place(x = 225, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_E6_IV
				entr_E6_IV = tk.Entry(root, width = 6, relief='solid')
				entr_E6_IV.place(x = 225, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_E7_IV
				entr_E7_IV = tk.Entry(root, width = 6, relief='solid')
				entr_E7_IV.place(x = 225, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_E8_IV
				entr_E8_IV = tk.Entry(root, width = 6, relief='solid')
				entr_E8_IV.place(x = 225, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_E9_IV
				entr_E9_IV = tk.Entry(root, width = 6, relief='solid')
				entr_E9_IV.place(x = 225, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_E10_IV
				entr_E10_IV = tk.Entry(root, width = 6, relief='solid')
				entr_E10_IV.place(x = 225, y = 545)



		if check_var_gr_2.get() >= 6:

			labl_F_1_IV  = tk.Label(root, text = "Обр. 6")
			labl_F_1_IV.place(x = 275, y = 345)

			global entr_F1_IV
			entr_F1_IV = tk.Entry(root, width = 6, relief='solid')
			entr_F1_IV.place(x = 275, y = 365)
			global entr_F2_IV
			entr_F2_IV = tk.Entry(root, width = 6, relief='solid')
			entr_F2_IV.place(x = 275, y = 385)
			global entr_F3_IV
			entr_F3_IV = tk.Entry(root, width = 6, relief='solid')
			entr_F3_IV.place(x = 275, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_F4_IV
				entr_F4_IV = tk.Entry(root, width = 6, relief='solid')
				entr_F4_IV.place(x = 275, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_F5_IV
				entr_F5_IV = tk.Entry(root, width = 6, relief='solid')
				entr_F5_IV.place(x = 275, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_F6_IV
				entr_F6_IV = tk.Entry(root, width = 6, relief='solid')
				entr_F6_IV.place(x = 275, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_F7_IV
				entr_F7_IV = tk.Entry(root, width = 6, relief='solid')
				entr_F7_IV.place(x = 275, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_F8_IV
				entr_F8_IV = tk.Entry(root, width = 6, relief='solid')
				entr_F8_IV.place(x = 275, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_F9_IV
				entr_F9_IV = tk.Entry(root, width = 6, relief='solid')
				entr_F9_IV.place(x = 275, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_F10_IV
				entr_F10_IV = tk.Entry(root, width = 6, relief='solid')
				entr_F10_IV.place(x = 275, y = 545)

	######################################################################################################
	if check_var_gr_1.get() >= 5:

		
		labl_MAIN_V = tk.Label(root, text = 'QCE')
		labl_MAIN_V.place(x = 362, y = 300)
		global entr_MAIN_V
		entr_MAIN_V = tk.Entry(root, width = 6, relief= 'solid')
		entr_MAIN_V.place(x = 360, y = 320)
		
		labl_A_1_V  = tk.Label(root, text = "Обр. 1")
		labl_A_1_V.place(x = 360, y = 345)

		global entr_A1_V
		entr_A1_V = tk.Entry(root, width = 6, relief= 'solid')
		entr_A1_V.place(x = 360, y = 365)
		global entr_A2_V
		entr_A2_V = tk.Entry(root, width = 6, relief='solid')
		entr_A2_V.place(x = 360, y = 385)
		global entr_A3_V
		entr_A3_V = tk.Entry(root, width = 6, relief='solid')
		entr_A3_V.place(x = 360, y = 405)


		if check_var_gr_3.get() >= 4:

			global entr_A4_V
			entr_A4_V = tk.Entry(root, width = 6, relief='solid')
			entr_A4_V.place(x = 360, y = 425)

		if check_var_gr_3.get() >= 5:	

			global entr_A5_V
			entr_A5_V = tk.Entry(root, width = 6, relief='solid')
			entr_A5_V.place(x = 360, y = 445)


		if check_var_gr_3.get() >= 6:
			
			global entr_A6_V
			entr_A6_V = tk.Entry(root, width = 6, relief='solid')
			entr_A6_V.place(x = 360, y = 465)

		if check_var_gr_3.get() >= 7:

			global entr_A7_V
			entr_A7_V = tk.Entry(root, width = 6, relief='solid')
			entr_A7_V.place(x = 360, y = 485)

		if check_var_gr_3.get() >= 8:

			global entr_A8_V
			entr_A8_V = tk.Entry(root, width = 6, relief='solid')
			entr_A8_V.place(x = 360, y = 505)

		if check_var_gr_3.get() >= 9:	

			global entr_A9_V
			entr_A9_V = tk.Entry(root, width = 6, relief='solid')
			entr_A9_V.place(x = 360, y = 525)

		if check_var_gr_3.get() >= 10:	
			global entr_A10_V
			entr_A10_V = tk.Entry(root, width = 6, relief='solid')
			entr_A10_V.place(x = 360, y = 545)
			
		
		if check_var_gr_2.get() >= 2:

			labl_B_1_V  = tk.Label(root, text = "Обр. 2")
			labl_B_1_V.place(x = 410, y = 345)	

			global entr_B1_V
			entr_B1_V = tk.Entry(root, width = 6, relief='solid')
			entr_B1_V.place(x = 410, y = 365)
			global entr_B2_V
			entr_B2_V = tk.Entry(root, width = 6, relief='solid')
			entr_B2_V.place(x = 410, y = 385)
			global entr_B3_V
			entr_B3_V= tk.Entry(root, width = 6, relief='solid')
			entr_B3_V.place(x = 410, y = 405)


			if check_var_gr_3.get() >= 4:

				global entr_B4_V
				entr_B4_V = tk.Entry(root, width = 6, relief='solid')
				entr_B4_V.place(x = 410, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_B5_V
				entr_B5_V = tk.Entry(root, width = 6, relief='solid')
				entr_B5_V.place(x = 410, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_B6_V
				entr_B6_V = tk.Entry(root, width = 6, relief='solid')
				entr_B6_V.place(x = 410, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_B7_V
				entr_B7_V = tk.Entry(root, width = 6, relief='solid')
				entr_B7_V.place(x = 410, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_B8_V
				entr_B8_V = tk.Entry(root, width = 6, relief='solid')
				entr_B8_V.place(x = 410, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_B9_V
				entr_B9_V = tk.Entry(root, width = 6, relief='solid')
				entr_B9_V.place(x = 410, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_B10_V
				entr_B10_V = tk.Entry(root, width = 6, relief='solid')
				entr_B10_V.place(x = 410, y = 545)
			

		
		if check_var_gr_2.get() >= 3:

			labl_C_1_V  = tk.Label(root, text = "Обр. 3")
			labl_C_1_V.place(x = 460, y = 345)

			global entr_C1_V 
			entr_C1_V = tk.Entry(root, width = 6, relief='solid')
			entr_C1_V.place(x = 460, y = 365)
			global entr_C2_V
			entr_C2_V = tk.Entry(root, width = 6, relief='solid')
			entr_C2_V.place(x = 460, y = 385)
			global entr_C3_V
			entr_C3_V = tk.Entry(root, width = 6, relief='solid')
			entr_C3_V.place(x = 460, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_C4_V
				entr_C4_V = tk.Entry(root, width = 6, relief='solid')
				entr_C4_V.place(x = 460, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_C5_V
				entr_C5_V = tk.Entry(root, width = 6, relief='solid')
				entr_C5_V.place(x = 460, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_C6_V
				entr_C6_V = tk.Entry(root, width = 6, relief='solid')
				entr_C6_V.place(x = 460, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_C7_V
				entr_C7_V = tk.Entry(root, width = 6, relief='solid')
				entr_C7_V.place(x = 460, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_C8_V
				entr_C8_V = tk.Entry(root, width = 6, relief='solid')
				entr_C8_V.place(x = 460, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_C9_V
				entr_C9_V = tk.Entry(root, width = 6, relief='solid')
				entr_C9_V.place(x = 460, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_C10_V
				entr_C10_V = tk.Entry(root, width = 6, relief='solid')
				entr_C10_V.place(x = 460, y = 545)


		if check_var_gr_2.get() >= 4:

			labl_D_1_V  = tk.Label(root, text = "Обр. 4")
			labl_D_1_V.place(x = 510, y = 345)

			global entr_D1_V 
			entr_D1_V = tk.Entry(root, width = 6, relief='solid')
			entr_D1_V.place(x = 510, y = 365)
			global entr_D2_V
			entr_D2_V = tk.Entry(root, width = 6, relief='solid')
			entr_D2_V.place(x = 510, y = 385)
			global entr_D3_V
			entr_D3_V = tk.Entry(root, width = 6, relief='solid')
			entr_D3_V.place(x = 510, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_D4_V
				entr_D4_V = tk.Entry(root, width = 6, relief='solid')
				entr_D4_V.place(x = 510, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_D5_V
				entr_D5_V = tk.Entry(root, width = 6, relief='solid')
				entr_D5_V.place(x = 510, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_D6_V
				entr_D6_V = tk.Entry(root, width = 6, relief='solid')
				entr_D6_V.place(x = 510, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_D7_V
				entr_D7_V = tk.Entry(root, width = 6, relief='solid')
				entr_D7_V.place(x = 510, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_D8_V
				entr_D8_V = tk.Entry(root, width = 6, relief='solid')
				entr_D8_V.place(x = 510, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_D9_V
				entr_D9_V = tk.Entry(root, width = 6, relief='solid')
				entr_D9_V.place(x = 510, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_D10_V
				entr_D10_V = tk.Entry(root, width = 6, relief='solid')
				entr_D10_V.place(x = 510, y = 545)



		if check_var_gr_2.get() >= 5:

			labl_E_1_V  = tk.Label(root, text = "Обр. 5")
			labl_E_1_V.place(x = 560, y = 345)

			global entr_E1_V 
			entr_E1_V = tk.Entry(root, width = 6, relief='solid')
			entr_E1_V.place(x = 560, y = 365)
			global entr_E2_V
			entr_E2_V = tk.Entry(root, width = 6, relief='solid')
			entr_E2_V.place(x = 560, y = 385)
			global entr_E3_V
			entr_E3_V = tk.Entry(root, width = 6, relief='solid')
			entr_E3_V.place(x = 560, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_E4_V
				entr_E4_V = tk.Entry(root, width = 6, relief='solid')
				entr_E4_V.place(x = 560, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_E5_V
				entr_E5_V = tk.Entry(root, width = 6, relief='solid')
				entr_E5_V.place(x = 560, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_E6_V
				entr_E6_V = tk.Entry(root, width = 6, relief='solid')
				entr_E6_V.place(x = 560, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_E7_V
				entr_E7_V = tk.Entry(root, width = 6, relief='solid')
				entr_E7_V.place(x = 560, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_E8_V
				entr_E8_V = tk.Entry(root, width = 6, relief='solid')
				entr_E8_V.place(x = 560, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_E9_V
				entr_E9_V = tk.Entry(root, width = 6, relief='solid')
				entr_E9_V.place(x = 560, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_E10_V
				entr_E10_V = tk.Entry(root, width = 6, relief='solid')
				entr_E10_V.place(x = 560, y = 545)



		if check_var_gr_2.get() >= 6:

			labl_F_1_V  = tk.Label(root, text = "Обр. 6")
			labl_F_1_V.place(x = 610, y = 345)

			global entr_F1_V 
			entr_F1_V = tk.Entry(root, width = 6, relief='solid')
			entr_F1_V.place(x = 610, y = 365)
			global entr_F2_V
			entr_F2_V = tk.Entry(root, width = 6, relief='solid')
			entr_F2_V.place(x = 610, y = 385)
			global entr_F3_V
			entr_F3_V = tk.Entry(root, width = 6, relief='solid')
			entr_F3_V.place(x = 610, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_F4_V
				entr_F4_V = tk.Entry(root, width = 6, relief='solid')
				entr_F4_V.place(x = 610, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_F5_V
				entr_F5_V = tk.Entry(root, width = 6, relief='solid')
				entr_F5_V.place(x = 610, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_F6_V
				entr_F6_V = tk.Entry(root, width = 6, relief='solid')
				entr_F6_V.place(x = 610, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_F7_V
				entr_F7_V = tk.Entry(root, width = 6, relief='solid')
				entr_F7_V.place(x = 610, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_F8_V
				entr_F8_V = tk.Entry(root, width = 6, relief='solid')
				entr_F8_V.place(x = 610, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_F9_V
				entr_F9_V = tk.Entry(root, width = 6, relief='solid')
				entr_F9_V.place(x = 610, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_F10_V
				entr_F10_V = tk.Entry(root, width = 6, relief='solid')
				entr_F10_V.place(x = 610, y = 545)

	##########################################################################################

	if check_var_gr_1.get() >= 6:

		labl_MAIN_VI = tk.Label(root, text = 'QCF')
		labl_MAIN_VI.place(x = 697, y = 300)
		global entr_MAIN_VI
		entr_MAIN_VI = tk.Entry(root, width = 6, relief= 'solid')
		entr_MAIN_VI.place(x = 695, y = 320)
		
		labl_A_1_VI  = tk.Label(root, text = "Обр. 1")
		labl_A_1_VI.place(x = 695, y = 345)

		global entr_A1_VI
		entr_A1_VI = tk.Entry(root, width = 6, relief= 'solid')
		entr_A1_VI.place(x = 695, y = 365)
		global entr_A2_VI
		entr_A2_VI = tk.Entry(root, width = 6, relief='solid')
		entr_A2_VI.place(x = 695, y = 385)
		global entr_A3_VI
		entr_A3_VI = tk.Entry(root, width = 6, relief='solid')
		entr_A3_VI.place(x = 695, y = 405)


		if check_var_gr_3.get() >= 4:

			global entr_A4_VI
			entr_A4_VI = tk.Entry(root, width = 6, relief='solid')
			entr_A4_VI.place(x = 695, y = 425)

		if check_var_gr_3.get() >= 5:	

			global entr_A5_VI
			entr_A5_VI = tk.Entry(root, width = 6, relief='solid')
			entr_A5_VI.place(x = 695, y = 445)


		if check_var_gr_3.get() >= 6:
			
			global entr_A6_VI
			entr_A6_VI = tk.Entry(root, width = 6, relief='solid')
			entr_A6_VI.place(x = 695, y = 465)

		if check_var_gr_3.get() >= 7:

			global entr_A7_VI
			entr_A7_VI = tk.Entry(root, width = 6, relief='solid')
			entr_A7_VI.place(x = 695, y = 485)

		if check_var_gr_3.get() >= 8:

			global entr_A8_VI
			entr_A8_VI = tk.Entry(root, width = 6, relief='solid')
			entr_A8_VI.place(x = 695, y = 505)

		if check_var_gr_3.get() >= 9:	

			global entr_A9_VI
			entr_A9_VI = tk.Entry(root, width = 6, relief='solid')
			entr_A9_VI.place(x = 695, y = 525)

		if check_var_gr_3.get() >= 10:	
			global entr_A10_VI
			entr_A10_VI = tk.Entry(root, width = 6, relief='solid')
			entr_A10_VI.place(x = 695, y = 545)
			
		
		if check_var_gr_2.get() >= 2:

			labl_B_1_VI  = tk.Label(root, text = "Обр. 2")
			labl_B_1_VI.place(x = 745, y = 345)	

			global entr_B1_VI
			entr_B1_VI = tk.Entry(root, width = 6, relief='solid')
			entr_B1_VI.place(x = 745, y = 365)
			global entr_B2_VI
			entr_B2_VI = tk.Entry(root, width = 6, relief='solid')
			entr_B2_VI.place(x = 745, y = 385)
			global entr_B3_VI
			entr_B3_VI = tk.Entry(root, width = 6, relief='solid')
			entr_B3_VI.place(x = 745, y = 405)


			if check_var_gr_3.get() >= 4:

				global entr_B4_VI
				entr_B4_VI = tk.Entry(root, width = 6, relief='solid')
				entr_B4_VI.place(x = 745, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_B5_VI
				entr_B5_VI = tk.Entry(root, width = 6, relief='solid')
				entr_B5_VI.place(x = 745, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_B6_VI
				entr_B6_VI = tk.Entry(root, width = 6, relief='solid')
				entr_B6_VI.place(x = 745, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_B7_VI
				entr_B7_VI = tk.Entry(root, width = 6, relief='solid')
				entr_B7_VI.place(x = 745, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_B8_VI
				entr_B8_VI = tk.Entry(root, width = 6, relief='solid')
				entr_B8_VI.place(x = 745, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_B9_VI
				entr_B9_VI = tk.Entry(root, width = 6, relief='solid')
				entr_B9_VI.place(x = 745, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_B10_VI
				entr_B10_VI = tk.Entry(root, width = 6, relief='solid')
				entr_B10_VI.place(x = 745, y = 545)
			

		
		if check_var_gr_2.get() >= 3:

			labl_C_1_VI  = tk.Label(root, text = "Обр. 3")
			labl_C_1_VI.place(x = 795, y = 345)

			global entr_C1_VI 
			entr_C1_VI = tk.Entry(root, width = 6, relief='solid')
			entr_C1_VI.place(x = 795, y = 365)
			global entr_C2_VI
			entr_C2_VI = tk.Entry(root, width = 6, relief='solid')
			entr_C2_VI.place(x = 795, y = 385)
			global entr_C3_VI
			entr_C3_VI = tk.Entry(root, width = 6, relief='solid')
			entr_C3_VI.place(x = 795, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_C4_VI
				entr_C4_VI = tk.Entry(root, width = 6, relief='solid')
				entr_C4_VI.place(x = 795, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_C5_VI
				entr_C5_VI = tk.Entry(root, width = 6, relief='solid')
				entr_C5_VI.place(x = 795, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_C6_VI
				entr_C6_VI = tk.Entry(root, width = 6, relief='solid')
				entr_C6_VI.place(x = 795, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_C7_VI
				entr_C7_VI = tk.Entry(root, width = 6, relief='solid')
				entr_C7_VI.place(x = 795, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_C8_VI
				entr_C8_VI = tk.Entry(root, width = 6, relief='solid')
				entr_C8_VI.place(x = 795, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_C9_VI
				entr_C9_VI = tk.Entry(root, width = 6, relief='solid')
				entr_C9_VI.place(x = 795, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_C10_VI
				entr_C10_VI = tk.Entry(root, width = 6, relief='solid')
				entr_C10_VI.place(x = 795, y = 545)


		if check_var_gr_2.get() >= 4:

			labl_D_1_VI  = tk.Label(root, text = "Обр. 4")
			labl_D_1_VI.place(x = 845, y = 345)

			global entr_D1_VI 
			entr_D1_VI = tk.Entry(root, width = 6, relief='solid')
			entr_D1_VI.place(x = 845, y = 365)
			global entr_D2_VI
			entr_D2_VI = tk.Entry(root, width = 6, relief='solid')
			entr_D2_VI.place(x = 845, y = 385)
			global entr_D3_VI
			entr_D3_VI = tk.Entry(root, width = 6, relief='solid')
			entr_D3_VI.place(x = 845, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_D4_VI
				entr_D4_VI = tk.Entry(root, width = 6, relief='solid')
				entr_D4_VI.place(x = 845, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_D5_VI
				entr_D5_VI = tk.Entry(root, width = 6, relief='solid')
				entr_D5_VI.place(x = 845, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_D6_VI
				entr_D6_VI = tk.Entry(root, width = 6, relief='solid')
				entr_D6_VI.place(x = 845, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_D7_VI
				entr_D7_VI = tk.Entry(root, width = 6, relief='solid')
				entr_D7_VI.place(x = 845, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_D8_VI
				entr_D8_VI = tk.Entry(root, width = 6, relief='solid')
				entr_D8_VI.place(x = 845, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_D9_VI
				entr_D9_VI = tk.Entry(root, width = 6, relief='solid')
				entr_D9_VI.place(x = 845, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_D10_VI
				entr_D10_VI = tk.Entry(root, width = 6, relief='solid')
				entr_D10_VI.place(x = 845, y = 545)



		if check_var_gr_2.get() >= 5:

			labl_E_1_VI  = tk.Label(root, text = "Обр. 5")
			labl_E_1_VI.place(x = 895, y = 345)

			global entr_E1_VI 
			entr_E1_VI = tk.Entry(root, width = 6, relief='solid')
			entr_E1_VI.place(x = 895, y = 365)
			global entr_E2_VI
			entr_E2_VI = tk.Entry(root, width = 6, relief='solid')
			entr_E2_VI.place(x = 895, y = 385)
			global entr_E3_VI
			entr_E3_VI = tk.Entry(root, width = 6, relief='solid')
			entr_E3_VI.place(x = 895, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_E4_VI
				entr_E4_VI = tk.Entry(root, width = 6, relief='solid')
				entr_E4_VI.place(x = 895, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_E5_VI
				entr_E5_VI = tk.Entry(root, width = 6, relief='solid')
				entr_E5_VI.place(x = 895, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_E6_VI
				entr_E6_VI = tk.Entry(root, width = 6, relief='solid')
				entr_E6_VI.place(x = 895, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_E7_VI
				entr_E7_VI = tk.Entry(root, width = 6, relief='solid')
				entr_E7_VI.place(x = 895, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_E8_VI
				entr_E8_VI = tk.Entry(root, width = 6, relief='solid')
				entr_E8_VI.place(x = 895, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_E9_VI
				entr_E9_VI = tk.Entry(root, width = 6, relief='solid')
				entr_E9_VI.place(x = 895, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_E10_VI
				entr_E10_VI = tk.Entry(root, width = 6, relief='solid')
				entr_E10_VI.place(x = 895, y = 545)



		if check_var_gr_2.get() >= 6:

			labl_F_1_VI  = tk.Label(root, text = "Обр. 6")
			labl_F_1_VI.place(x = 945, y = 345)

			global entr_F1_VI 
			entr_F1_VI = tk.Entry(root, width = 6, relief='solid')
			entr_F1_VI.place(x = 945, y = 365)
			global entr_F2_VI
			entr_F2_VI = tk.Entry(root, width = 6, relief='solid')
			entr_F2_VI.place(x = 945, y = 385)
			global entr_F3_VI
			entr_F3_VI = tk.Entry(root, width = 6, relief='solid')
			entr_F3_VI.place(x = 945, y = 405)

			if check_var_gr_3.get() >= 4:

				global entr_F4_VI
				entr_F4_VI = tk.Entry(root, width = 6, relief='solid')
				entr_F4_VI.place(x = 945, y = 425)

			if check_var_gr_3.get() >= 5:

				global entr_F5_VI
				entr_F5_VI = tk.Entry(root, width = 6, relief='solid')
				entr_F5_VI.place(x = 945, y = 445)

			if check_var_gr_3.get() >= 6:

				global entr_F6_VI
				entr_F6_VI = tk.Entry(root, width = 6, relief='solid')
				entr_F6_VI.place(x = 945, y = 465)

			if check_var_gr_3.get() >= 7:

				global entr_F7_VI
				entr_F7_VI = tk.Entry(root, width = 6, relief='solid')
				entr_F7_VI.place(x = 945, y = 485)

			if check_var_gr_3.get() >= 8:

				global entr_F8_VI
				entr_F8_VI = tk.Entry(root, width = 6, relief='solid')
				entr_F8_VI.place(x = 945, y = 505)

			if check_var_gr_3.get() >= 9:

				global entr_F9_VI
				entr_F9_VI = tk.Entry(root, width = 6, relief='solid')
				entr_F9_VI.place(x = 945, y = 525)

			if check_var_gr_3.get() >= 10:

				global entr_F10_VI
				entr_F10_VI = tk.Entry(root, width = 6, relief='solid')
				entr_F10_VI.place(x = 945, y = 545)




	
	butt_1 = tk.Button(root,text = 'save', width = 8, command = calc)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		butt_1.place(x = 800, y = 95)
	else:
		butt_1.place(x = 1220, y = 95)

	butt_2 = tk.Button(root,text = 'open file\nto save', width = 8, command = open_dialog)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		butt_2.place(x = 800, y = 50)
	else:
		butt_2.place(x = 1220, y = 50)


	butt_3 = tk.Button(root,text = 'input', width = 8, command = input_this)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		butt_3.place(x = 800, y = 20)
	else:
		butt_3.place(x = 1220, y = 20)

	butt_4 = tk.Button(root,text = '+', command = create_input)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		butt_4.place(x = 782, y = 20)

	butt_5 = tk.Button(root, text = '?', command = spravka_open_root, width = 2, bg = 'pink')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		butt_5.place(x = 870, y = 265)
	else:
		butt_5.place(x = 1270, y = 665)

	butt_6 = tk.Button(root, text = 'back', command = back_to_first_widow, bg = 'yellow')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		butt_6.place(x = 825, y = 265)
	else:
		butt_6.place(x = 1225, y = 665)
	
	butt_7 = tk.Button(root, text = 'clear', command = clear, bg = '#9cffd0')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		butt_7.place(x = 825, y = 235)
	else:
		butt_7.place(x = 1225, y = 635)


	labl_concent = tk.Label(root,text = 'Введите концентрацию\n(пример: ng/mL)')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_concent.place(x = 222, y = 185)
	else:
		labl_concent.place(x = 1028, y = 150)

	global entr_concent
	entr_concent = tk.Entry(root, width = 10, relief='solid')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		entr_concent.place(x = 256, y = 228)
	else:
		entr_concent.place(x = 1063, y = 187)

	global check_razryad
	check_razryad = tk.IntVar()
	check_razryad.set(1)

	labl_2 = tk.Label(root,text = 'Выберите кол-во знаков\nпосле запятой')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_2.place(x = 22, y = 185)
	else:
		labl_2.place(x = 1025, y = 50)

	check_razryad_0 = tk.Radiobutton(root,variable = check_razryad, value = 0)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		check_razryad_0.place(x = 39, y = 220)
	else:
		check_razryad_0.place(x = 1044, y = 85)


	labl_check_0 = tk.Label(root,text = '0')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_check_0.place(x = 43, y = 239)
	else:
		labl_check_0.place(x = 1048, y = 104)

	check_razryad_1 = tk.Radiobutton(root,variable = check_razryad, value = 1)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		check_razryad_1.place(x = 59, y = 220)
	else:
		check_razryad_1.place(x = 1064, y = 85)

	labl_check_1 = tk.Label(root,text = '1')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_check_1.place(x = 63, y = 239)
	else:
		labl_check_1.place(x = 1068, y = 104)

	check_razryad_2 = tk.Radiobutton(root,variable = check_razryad, value = 2)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		check_razryad_2.place(x = 79, y = 220)
	else:
		check_razryad_2.place(x = 1084, y = 85)

	labl_check_2 = tk.Label(root,text = '2')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_check_2.place(x = 83, y = 239)
	else:
		labl_check_2.place(x = 1088, y = 104)


	check_razryad_3 = tk.Radiobutton(root,variable = check_razryad, value = 3)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		check_razryad_3.place(x = 99, y = 220)
	else:
		check_razryad_3.place(x = 1104, y = 85)

	labl_check_3 = tk.Label(root,text = '3')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_check_3.place(x = 103, y = 239)
	else:
		labl_check_3.place(x = 1108, y = 104)

	check_razryad_4 = tk.Radiobutton(root, variable = check_razryad, value = 4)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		check_razryad_4.place(x = 119, y = 220)
	else:
		check_razryad_4.place(x = 1124, y = 85)


	labl_check_4 = tk.Label(root,text = '4')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_check_4.place(x = 123, y = 239)
	else:
		labl_check_4.place(x = 1128, y = 104)

	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_norm = tk.Label(root,text = 'Введите значение критериев\n(если необходимо изменить значение по умолчанию)')
	else:
		labl_norm = tk.Label(root,text = 'Введите значение\nкритериев\n(если необходимо\nизменить значение\nпо умолчанию)')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_norm.place(x = 430, y = 185)
	else:
		labl_norm.place(x = 1040, y = 235)



	labl_norm_QCA = tk.Label(root, text = 'QCA')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		labl_norm_QCA.place(x = 430, y = 228)
	else:
		labl_norm_QCA.place(x = 1078, y = 320)



	global entr_NORM_QCA1
	entr_NORM_QCA1 = tk.Entry(root,width = 4, relief='solid')
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		entr_NORM_QCA1.place(x = 465, y = 228)
	else:
		entr_NORM_QCA1.place(x = 1080, y = 340)
	


	if check_var_gr_1.get() >= 2:
		labl_norm_QCB = tk.Label(root, text = 'QCB')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_norm_QCB.place(x = 512, y = 228)
		else:
			labl_norm_QCB.place(x = 1078, y = 360)

		global entr_NORM_QCB1
		entr_NORM_QCB1 = tk.Entry(root,width = 4, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_NORM_QCB1.place(x = 547, y = 228)
		else:
			entr_NORM_QCB1.place(x = 1080, y = 380)
		

	if check_var_gr_1.get() >= 3:
		labl_norm_QCC = tk.Label(root, text = 'QCC')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_norm_QCC.place(x = 592, y = 228)
		else:
			labl_norm_QCC.place(x = 1078, y = 400)

		global entr_NORM_QCC1
		entr_NORM_QCC1 = tk.Entry(root,width = 4, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_NORM_QCC1.place(x = 627, y = 228)
		else:
			entr_NORM_QCC1.place(x = 1080, y = 420)
		

	if check_var_gr_1.get() >= 4:
		labl_norm_QCD = tk.Label(root, text = 'QCD')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			labl_norm_QCD.place(x = 672, y = 228)
		else:
			labl_norm_QCD.place(x = 1078, y = 440)
		global entr_NORM_QCD1

		entr_NORM_QCD1 = tk.Entry(root,width = 4, relief='solid')
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			entr_NORM_QCD1.place(x = 707, y = 228)
		else:
			entr_NORM_QCD1.place(x = 1080, y = 460)
		

	if check_var_gr_1.get() >= 5:
		labl_norm_QCE = tk.Label(root, text = 'QCE')
		labl_norm_QCE.place(x = 1078, y = 480)
		global entr_NORM_QCE1
		entr_NORM_QCE1 = tk.Entry(root,width = 4, relief='solid')
		entr_NORM_QCE1.place(x = 1080, y = 500)
		

	if check_var_gr_1.get() >= 6:
		labl_norm_QCF = tk.Label(root, text = 'QCF')
		labl_norm_QCF.place(x = 1078, y = 520)
		global entr_NORM_QCF1
		entr_NORM_QCF1 = tk.Entry(root,width = 4, relief='solid')
		entr_NORM_QCF1.place(x = 1080, y = 540)



#КУСОК ОТВЕАЮЩИЙ ЗА ВВОД СОХРАНЕННОГО ТЕКСТА
#QCA
#QCA 1 
	try:
		entr_MAIN_I.insert(tk.END, save_MAIN_I)
	except:
		pass
	try:		
		entr_A1_I.insert(tk.END, save_A1_I)
	except:
		pass
	try:
		entr_A2_I.insert(tk.END, save_A2_I)
	except:
		pass
	try:
		entr_A3_I.insert(tk.END, save_A3_I)
	except:
		pass
	try:
		entr_A4_I.insert(tk.END, save_A4_I)
	except:
		pass
	try:
		entr_A5_I.insert(tk.END, save_A5_I)
	except:
		pass
	try:
		entr_A6_I.insert(tk.END, save_A6_I)
	except:
		pass
	try:
		entr_A7_I.insert(tk.END, save_A7_I)
	except:
		pass
	try:
		entr_A8_I.insert(tk.END, save_A8_I)
	except:
		pass
	try:
		entr_A9_I.insert(tk.END, save_A9_I)
	except:
		pass
	try:
		entr_A10_I.insert(tk.END, save_A10_I)
	except:
		pass
#QCA 2
	try:
		entr_B1_I.insert(tk.END, save_B1_I)
	except:
		pass
	try:
		entr_B2_I.insert(tk.END, save_B2_I)
	except:
		pass
	try:
		entr_B3_I.insert(tk.END, save_B3_I)
	except:
		pass
	try:
		entr_B4_I.insert(tk.END, save_B4_I)
	except:
		pass
	try:
		entr_B5_I.insert(tk.END, save_B5_I)
	except:
		pass
	try:
		entr_B6_I.insert(tk.END, save_B6_I)
	except:
		pass
	try:
		entr_B7_I.insert(tk.END, save_B7_I)
	except:
		pass
	try:
		entr_B8_I.insert(tk.END, save_B8_I)
	except:
		pass
	try:
		entr_B9_I.insert(tk.END, save_B9_I)
	except:
		pass
	try:
		entr_B10_I.insert(tk.END, save_B10_I)
	except:
		pass
#QCA 3
	try:
		entr_C1_I.insert(tk.END, save_C1_I)
	except:
		pass
	try:
		entr_C2_I.insert(tk.END, save_C2_I)
	except:
		pass
	try:
		entr_C3_I.insert(tk.END, save_C3_I)
	except:
		pass
	try:
		entr_C4_I.insert(tk.END, save_C4_I)
	except:
		pass
	try:
		entr_C5_I.insert(tk.END, save_C5_I)
	except:
		pass
	try:
		entr_C6_I.insert(tk.END, save_C6_I)
	except:
		pass
	try:
		entr_C7_I.insert(tk.END, save_C7_I)
	except:
		pass
	try:
		entr_C8_I.insert(tk.END, save_C8_I)
	except:
		pass
	try:
		entr_C9_I.insert(tk.END, save_C9_I)
	except:
		pass
	try:
		entr_C10_I.insert(tk.END, save_C10_I)
	except:
		pass

#QCA 4
	try:
		entr_D1_I.insert(tk.END, save_D1_I)
	except:
		pass
	try:
		entr_D2_I.insert(tk.END, save_D2_I)
	except:
		pass
	try:
		entr_D3_I.insert(tk.END, save_D3_I)
	except:
		pass
	try:
		entr_D4_I.insert(tk.END, save_D4_I)
	except:
		pass
	try:
		entr_D5_I.insert(tk.END, save_D5_I)
	except:
		pass
	try:
		entr_D6_I.insert(tk.END, save_D6_I)
	except:
		pass
	try:
		entr_D7_I.insert(tk.END, save_D7_I)
	except:
		pass
	try:
		entr_D8_I.insert(tk.END, save_D8_I)
	except:
		pass
	try:
		entr_D9_I.insert(tk.END, save_D9_I)
	except:
		pass
	try:
		entr_D10_I.insert(tk.END, save_D10_I)
	except:
		pass
#QCA 5
	try:
		entr_E1_I.insert(tk.END, save_E1_I)
	except:
		pass
	try:
		entr_E2_I.insert(tk.END, save_E2_I)
	except:
		pass
	try:
		entr_E3_I.insert(tk.END, save_E3_I)
	except:
		pass
	try:
		entr_E4_I.insert(tk.END, save_E4_I)
	except:
		pass
	try:
		entr_E5_I.insert(tk.END, save_E5_I)
	except:
		pass
	try:
		entr_E6_I.insert(tk.END, save_E6_I)
	except:
		pass
	try:
		entr_E7_I.insert(tk.END, save_E7_I)
	except:
		pass
	try:
		entr_E8_I.insert(tk.END, save_E8_I)
	except:
		pass
	try:
		entr_E9_I.insert(tk.END, save_E9_I)
	except:
		pass
	try:
		entr_E10_I.insert(tk.END, save_E10_I)
	except:
		pass

#QCA 6
	try:
		entr_F1_I.insert(tk.END, save_F1_I)
	except:
		pass
	try:
		entr_F2_I.insert(tk.END, save_F2_I)
	except:
		pass
	try:
		entr_F3_I.insert(tk.END, save_F3_I)
	except:
		pass
	try:
		entr_F4_I.insert(tk.END, save_F4_I)
	except:
		pass
	try:
		entr_F5_I.insert(tk.END, save_F5_I)
	except:
		pass
	try:
		entr_F6_I.insert(tk.END, save_F6_I)
	except:
		pass
	try:
		entr_F7_I.insert(tk.END, save_F7_I)
	except:
		pass
	try:
		entr_F8_I.insert(tk.END, save_F8_I)
	except:
		pass
	try:
		entr_F9_I.insert(tk.END, save_F9_I)
	except:
		pass
	try:
		entr_F10_I.insert(tk.END, save_F10_I)
	except:
		pass






#QCB
#QCB 1 
	try:
		entr_MAIN_II.insert(tk.END, save_MAIN_II)
	except:
		pass
	try:
		#entr_A1_I.delete(0, "end")
		entr_A1_II.insert(tk.END, save_A1_II)
	except:
		pass
	try:
		entr_A2_II.insert(tk.END, save_A2_II)
	except:
		pass
	try:
		entr_A3_II.insert(tk.END, save_A3_II)
	except:
		pass
	try:
		entr_A4_II.insert(tk.END, save_A4_II)
	except:
		pass
	try:
		entr_A5_II.insert(tk.END, save_A5_II)
	except:
		pass
	try:
		entr_A6_II.insert(tk.END, save_A6_II)
	except:
		pass
	try:
		entr_A7_II.insert(tk.END, save_A7_II)
	except:
		pass
	try:
		entr_A8_II.insert(tk.END, save_A8_II)
	except:
		pass
	try:
		entr_A9_II.insert(tk.END, save_A9_II)
	except:
		pass
	try:
		entr_A10_II.insert(tk.END, save_A10_II)
	except:
		pass
#QCB 2
	try:
		entr_B1_II.insert(tk.END, save_B1_II)
	except:
		pass
	try:
		entr_B2_II.insert(tk.END, save_B2_II)
	except:
		pass
	try:
		entr_B3_II.insert(tk.END, save_B3_II)
	except:
		pass
	try:
		entr_B4_II.insert(tk.END, save_B4_II)
	except:
		pass
	try:
		entr_B5_II.insert(tk.END, save_B5_II)
	except:
		pass
	try:
		entr_B6_II.insert(tk.END, save_B6_II)
	except:
		pass
	try:
		entr_B7_II.insert(tk.END, save_B7_II)
	except:
		pass
	try:
		entr_B8_II.insert(tk.END, save_B8_II)
	except:
		pass
	try:
		entr_B9_II.insert(tk.END, save_B9_II)
	except:
		pass
	try:
		entr_B10_II.insert(tk.END, save_B10_II)
	except:
		pass
#QCB 3
	try:
		entr_C1_II.insert(tk.END, save_C1_II)
	except:
		pass
	try:
		entr_C2_II.insert(tk.END, save_C2_II)
	except:
		pass
	try:
		entr_C3_II.insert(tk.END, save_C3_II)
	except:
		pass
	try:
		entr_C4_II.insert(tk.END, save_C4_II)
	except:
		pass
	try:
		entr_C5_II.insert(tk.END, save_C5_II)
	except:
		pass
	try:
		entr_C6_II.insert(tk.END, save_C6_II)
	except:
		pass
	try:
		entr_C7_II.insert(tk.END, save_C7_II)
	except:
		pass
	try:
		entr_C8_II.insert(tk.END, save_C8_II)
	except:
		pass
	try:
		entr_C9_II.insert(tk.END, save_C9_II)
	except:
		pass
	try:
		entr_C10_II.insert(tk.END, save_C10_II)
	except:
		pass

#QCB 4
	try:
		entr_D1_II.insert(tk.END, save_D1_II)
	except:
		pass
	try:
		entr_D2_II.insert(tk.END, save_D2_II)
	except:
		pass
	try:
		entr_D3_II.insert(tk.END, save_D3_II)
	except:
		pass
	try:
		entr_D4_II.insert(tk.END, save_D4_II)
	except:
		pass
	try:
		entr_D5_II.insert(tk.END, save_D5_II)
	except:
		pass
	try:
		entr_D6_II.insert(tk.END, save_D6_II)
	except:
		pass
	try:
		entr_D7_II.insert(tk.END, save_D7_II)
	except:
		pass
	try:
		entr_D8_II.insert(tk.END, save_D8_II)
	except:
		pass
	try:
		entr_D9_II.insert(tk.END, save_D9_II)
	except:
		pass
	try:
		entr_D10_II.insert(tk.END, save_D10_II)
	except:
		pass
#QCB 5
	try:
		entr_E1_II.insert(tk.END, save_E1_II)
	except:
		pass
	try:
		entr_E2_II.insert(tk.END, save_E2_II)
	except:
		pass
	try:
		entr_E3_II.insert(tk.END, save_E3_II)
	except:
		pass
	try:
		entr_E4_II.insert(tk.END, save_E4_II)
	except:
		pass
	try:
		entr_E5_II.insert(tk.END, save_E5_II)
	except:
		pass
	try:
		entr_E6_II.insert(tk.END, save_E6_II)
	except:
		pass
	try:
		entr_E7_II.insert(tk.END, save_E7_II)
	except:
		pass
	try:
		entr_E8_II.insert(tk.END, save_E8_II)
	except:
		pass
	try:
		entr_E9_II.insert(tk.END, save_E9_II)
	except:
		pass
	try:
		entr_E10_II.insert(tk.END, save_E10_II)
	except:
		pass

#QCB 6
	try:
		entr_F1_II.insert(tk.END, save_F1_II)
	except:
		pass
	try:
		entr_F2_II.insert(tk.END, save_F2_II)
	except:
		pass
	try:
		entr_F3_II.insert(tk.END, save_F3_II)
	except:
		pass
	try:
		entr_F4_II.insert(tk.END, save_F4_II)
	except:
		pass
	try:
		entr_F5_II.insert(tk.END, save_F5_II)
	except:
		pass
	try:
		entr_F6_II.insert(tk.END, save_F6_II)
	except:
		pass
	try:
		entr_F7_II.insert(tk.END, save_F7_II)
	except:
		pass
	try:
		entr_F8_II.insert(tk.END, save_F8_II)
	except:
		pass
	try:
		entr_F9_II.insert(tk.END, save_F9_II)
	except:
		pass
	try:
		entr_F10_II.insert(tk.END, save_F10_II)
	except:
		pass



#QCC
#QCC 1 
	try:
		entr_MAIN_III.insert(tk.END, save_MAIN_III)
	except:
		pass

	try:
		#entr_A1_I.delete(0, "end")
		entr_A1_III.insert(tk.END, save_A1_III)
	except:
		pass
	try:
		entr_A2_III.insert(tk.END, save_A2_III)
	except:
		pass
	try:
		entr_A3_III.insert(tk.END, save_A3_III)
	except:
		pass
	try:
		entr_A4_III.insert(tk.END, save_A4_III)
	except:
		pass
	try:
		entr_A5_III.insert(tk.END, save_A5_III)
	except:
		pass
	try:
		entr_A6_III.insert(tk.END, save_A6_III)
	except:
		pass
	try:
		entr_A7_III.insert(tk.END, save_A7_III)
	except:
		pass
	try:
		entr_A8_III.insert(tk.END, save_A8_III)
	except:
		pass
	try:
		entr_A9_III.insert(tk.END, save_A9_III)
	except:
		pass
	try:
		entr_A10_III.insert(tk.END, save_A10_III)
	except:
		pass
#QCC 2
	try:
		entr_B1_III.insert(tk.END, save_B1_III)
	except:
		pass
	try:
		entr_B2_III.insert(tk.END, save_B2_III)
	except:
		pass
	try:
		entr_B3_III.insert(tk.END, save_B3_III)
	except:
		pass
	try:
		entr_B4_III.insert(tk.END, save_B4_III)
	except:
		pass
	try:
		entr_B5_III.insert(tk.END, save_B5_III)
	except:
		pass
	try:
		entr_B6_III.insert(tk.END, save_B6_III)
	except:
		pass
	try:
		entr_B7_III.insert(tk.END, save_B7_III)
	except:
		pass
	try:
		entr_B8_III.insert(tk.END, save_B8_III)
	except:
		pass
	try:
		entr_B9_III.insert(tk.END, save_B9_III)
	except:
		pass
	try:
		entr_B10_III.insert(tk.END, save_B10_III)
	except:
		pass
#QCC 3
	try:
		entr_C1_III.insert(tk.END, save_C1_III)
	except:
		pass
	try:
		entr_C2_III.insert(tk.END, save_C2_III)
	except:
		pass
	try:
		entr_C3_III.insert(tk.END, save_C3_III)
	except:
		pass
	try:
		entr_C4_III.insert(tk.END, save_C4_III)
	except:
		pass
	try:
		entr_C5_III.insert(tk.END, save_C5_III)
	except:
		pass
	try:
		entr_C6_III.insert(tk.END, save_C6_III)
	except:
		pass
	try:
		entr_C7_III.insert(tk.END, save_C7_III)
	except:
		pass
	try:
		entr_C8_III.insert(tk.END, save_C8_III)
	except:
		pass
	try:
		entr_C9_III.insert(tk.END, save_C9_III)
	except:
		pass
	try:
		entr_C10_III.insert(tk.END, save_C10_III)
	except:
		pass

#QCC 4
	try:
		entr_D1_III.insert(tk.END, save_D1_III)
	except:
		pass
	try:
		entr_D2_III.insert(tk.END, save_D2_III)
	except:
		pass
	try:
		entr_D3_III.insert(tk.END, save_D3_III)
	except:
		pass
	try:
		entr_D4_III.insert(tk.END, save_D4_III)
	except:
		pass
	try:
		entr_D5_III.insert(tk.END, save_D5_III)
	except:
		pass
	try:
		entr_D6_III.insert(tk.END, save_D6_III)
	except:
		pass
	try:
		entr_D7_III.insert(tk.END, save_D7_III)
	except:
		pass
	try:
		entr_D8_III.insert(tk.END, save_D8_III)
	except:
		pass
	try:
		entr_D9_III.insert(tk.END, save_D9_III)
	except:
		pass
	try:
		entr_D10_III.insert(tk.END, save_D10_III)
	except:
		pass
#QCC 5
	try:
		entr_E1_III.insert(tk.END, save_E1_III)
	except:
		pass
	try:
		entr_E2_III.insert(tk.END, save_E2_III)
	except:
		pass
	try:
		entr_E3_III.insert(tk.END, save_E3_III)
	except:
		pass
	try:
		entr_E4_III.insert(tk.END, save_E4_III)
	except:
		pass
	try:
		entr_E5_III.insert(tk.END, save_E5_III)
	except:
		pass
	try:
		entr_E6_III.insert(tk.END, save_E6_III)
	except:
		pass
	try:
		entr_E7_III.insert(tk.END, save_E7_III)
	except:
		pass
	try:
		entr_E8_III.insert(tk.END, save_E8_III)
	except:
		pass
	try:
		entr_E9_III.insert(tk.END, save_E9_III)
	except:
		pass
	try:
		entr_E10_III.insert(tk.END, save_E10_III)
	except:
		pass

#QCC 6
	try:
		entr_F1_III.insert(tk.END, save_F1_III)
	except:
		pass
	try:
		entr_F2_III.insert(tk.END, save_F2_III)
	except:
		pass
	try:
		entr_F3_III.insert(tk.END, save_F3_III)
	except:
		pass
	try:
		entr_F4_III.insert(tk.END, save_F4_III)
	except:
		pass
	try:
		entr_F5_III.insert(tk.END, save_F5_III)
	except:
		pass
	try:
		entr_F6_III.insert(tk.END, save_F6_III)
	except:
		pass
	try:
		entr_F7_III.insert(tk.END, save_F7_III)
	except:
		pass
	try:
		entr_F8_III.insert(tk.END, save_F8_III)
	except:
		pass
	try:
		entr_F9_III.insert(tk.END, save_F9_III)
	except:
		pass
	try:
		entr_F10_III.insert(tk.END, save_F10_III)
	except:
		pass


#QCD
#QCD 1 
	try:
		entr_MAIN_IV.insert(tk.END, save_MAIN_IV)
	except:
		pass
	try:
		#entr_A1_I.delete(0, "end")
		entr_A1_IV.insert(tk.END, save_A1_IV)
	except:
		pass
	try:
		entr_A2_IV.insert(tk.END, save_A2_IV)
	except:
		pass
	try:
		entr_A3_IV.insert(tk.END, save_A3_IV)
	except:
		pass
	try:
		entr_A4_IV.insert(tk.END, save_A4_IV)
	except:
		pass
	try:
		entr_A5_IV.insert(tk.END, save_A5_IV)
	except:
		pass
	try:
		entr_A6_IV.insert(tk.END, save_A6_IV)
	except:
		pass
	try:
		entr_A7_IV.insert(tk.END, save_A7_IV)
	except:
		pass
	try:
		entr_A8_IV.insert(tk.END, save_A8_IV)
	except:
		pass
	try:
		entr_A9_IV.insert(tk.END, save_A9_IV)
	except:
		pass
	try:
		entr_A10_IV.insert(tk.END, save_A10_IV)
	except:
		pass
#QCD 2
	try:
		entr_B1_IV.insert(tk.END, save_B1_IV)
	except:
		pass
	try:
		entr_B2_IV.insert(tk.END, save_B2_IV)
	except:
		pass
	try:
		entr_B3_IV.insert(tk.END, save_B3_IV)
	except:
		pass
	try:
		entr_B4_IV.insert(tk.END, save_B4_IV)
	except:
		pass
	try:
		entr_B5_IV.insert(tk.END, save_B5_IV)
	except:
		pass
	try:
		entr_B6_IV.insert(tk.END, save_B6_IV)
	except:
		pass
	try:
		entr_B7_IV.insert(tk.END, save_B7_IV)
	except:
		pass
	try:
		entr_B8_IV.insert(tk.END, save_B8_IV)
	except:
		pass
	try:
		entr_B9_IV.insert(tk.END, save_B9_IV)
	except:
		pass
	try:
		entr_B10_IV.insert(tk.END, save_B10_IV)
	except:
		pass
#QCD 3
	try:
		entr_C1_IV.insert(tk.END, save_C1_IV)
	except:
		pass
	try:
		entr_C2_IV.insert(tk.END, save_C2_IV)
	except:
		pass
	try:
		entr_C3_IV.insert(tk.END, save_C3_IV)
	except:
		pass
	try:
		entr_C4_IV.insert(tk.END, save_C4_IV)
	except:
		pass
	try:
		entr_C5_IV.insert(tk.END, save_C5_IV)
	except:
		pass
	try:
		entr_C6_IV.insert(tk.END, save_C6_IV)
	except:
		pass
	try:
		entr_C7_IV.insert(tk.END, save_C7_IV)
	except:
		pass
	try:
		entr_C8_IV.insert(tk.END, save_C8_IV)
	except:
		pass
	try:
		entr_C9_IV.insert(tk.END, save_C9_IV)
	except:
		pass
	try:
		entr_C10_IV.insert(tk.END, save_C10_IV)
	except:
		pass

#QCD 4
	try:
		entr_D1_IV.insert(tk.END, save_D1_IV)
	except:
		pass
	try:
		entr_D2_IV.insert(tk.END, save_D2_IV)
	except:
		pass
	try:
		entr_D3_IV.insert(tk.END, save_D3_IV)
	except:
		pass
	try:
		entr_D4_IV.insert(tk.END, save_D4_IV)
	except:
		pass
	try:
		entr_D5_IV.insert(tk.END, save_D5_IV)
	except:
		pass
	try:
		entr_D6_IV.insert(tk.END, save_D6_IV)
	except:
		pass
	try:
		entr_D7_IV.insert(tk.END, save_D7_IV)
	except:
		pass
	try:
		entr_D8_IV.insert(tk.END, save_D8_IV)
	except:
		pass
	try:
		entr_D9_IV.insert(tk.END, save_D9_IV)
	except:
		pass
	try:
		entr_D10_IV.insert(tk.END, save_D10_IV)
	except:
		pass
#QCD 5
	try:
		entr_E1_IV.insert(tk.END, save_E1_IV)
	except:
		pass
	try:
		entr_E2_IV.insert(tk.END, save_E2_IV)
	except:
		pass
	try:
		entr_E3_IV.insert(tk.END, save_E3_IV)
	except:
		pass
	try:
		entr_E4_IV.insert(tk.END, save_E4_IV)
	except:
		pass
	try:
		entr_E5_IV.insert(tk.END, save_E5_IV)
	except:
		pass
	try:
		entr_E6_IV.insert(tk.END, save_E6_IV)
	except:
		pass
	try:
		entr_E7_IV.insert(tk.END, save_E7_IV)
	except:
		pass
	try:
		entr_E8_IV.insert(tk.END, save_E8_IV)
	except:
		pass
	try:
		entr_E9_IV.insert(tk.END, save_E9_IV)
	except:
		pass
	try:
		entr_E10_IV.insert(tk.END, save_E10_IV)
	except:
		pass

#QCD 6
	try:
		entr_F1_IV.insert(tk.END, save_F1_IV)
	except:
		pass
	try:
		entr_F2_IV.insert(tk.END, save_F2_IV)
	except:
		pass
	try:
		entr_F3_IV.insert(tk.END, save_F3_IV)
	except:
		pass
	try:
		entr_F4_IV.insert(tk.END, save_F4_IV)
	except:
		pass
	try:
		entr_F5_IV.insert(tk.END, save_F5_IV)
	except:
		pass
	try:
		entr_F6_IV.insert(tk.END, save_F6_IV)
	except:
		pass
	try:
		entr_F7_IV.insert(tk.END, save_F7_IV)
	except:
		pass
	try:
		entr_F8_IV.insert(tk.END, save_F8_IV)
	except:
		pass
	try:
		entr_F9_IV.insert(tk.END, save_F9_IV)
	except:
		pass
	try:
		entr_F10_IV.insert(tk.END, save_F10_IV)
	except:
		pass



#QCE
#QCE 1 
	try:
		entr_MAIN_V.insert(tk.END, save_MAIN_V)
	except:
		pass
	try:
		#entr_A1_I.delete(0, "end")
		entr_A1_V.insert(tk.END, save_A1_V)
	except:
		pass
	try:
		entr_A2_V.insert(tk.END, save_A2_V)
	except:
		pass
	try:
		entr_A3_V.insert(tk.END, save_A3_V)
	except:
		pass
	try:
		entr_A4_V.insert(tk.END, save_A4_V)
	except:
		pass
	try:
		entr_A5_V.insert(tk.END, save_A5_V)
	except:
		pass
	try:
		entr_A6_V.insert(tk.END, save_A6_V)
	except:
		pass
	try:
		entr_A7_V.insert(tk.END, save_A7_V)
	except:
		pass
	try:
		entr_A8_V.insert(tk.END, save_A8_V)
	except:
		pass
	try:
		entr_A9_V.insert(tk.END, save_A9_V)
	except:
		pass
	try:
		entr_A10_V.insert(tk.END, save_A10_V)
	except:
		pass
#QCE 2
	try:
		entr_B1_V.insert(tk.END, save_B1_V)
	except:
		pass
	try:
		entr_B2_V.insert(tk.END, save_B2_V)
	except:
		pass
	try:
		entr_B3_V.insert(tk.END, save_B3_V)
	except:
		pass
	try:
		entr_B4_V.insert(tk.END, save_B4_V)
	except:
		pass
	try:
		entr_B5_V.insert(tk.END, save_B5_V)
	except:
		pass
	try:
		entr_B6_V.insert(tk.END, save_B6_V)
	except:
		pass
	try:
		entr_B7_V.insert(tk.END, save_B7_V)
	except:
		pass
	try:
		entr_B8_V.insert(tk.END, save_B8_V)
	except:
		pass
	try:
		entr_B9_V.insert(tk.END, save_B9_V)
	except:
		pass
	try:
		entr_B10_V.insert(tk.END, save_B10_V)
	except:
		pass
#QCE 3
	try:
		entr_C1_V.insert(tk.END, save_C1_V)
	except:
		pass
	try:
		entr_C2_V.insert(tk.END, save_C2_V)
	except:
		pass
	try:
		entr_C3_V.insert(tk.END, save_C3_V)
	except:
		pass
	try:
		entr_C4_V.insert(tk.END, save_C4_V)
	except:
		pass
	try:
		entr_C5_V.insert(tk.END, save_C5_V)
	except:
		pass
	try:
		entr_C6_V.insert(tk.END, save_C6_V)
	except:
		pass
	try:
		entr_C7_V.insert(tk.END, save_C7_V)
	except:
		pass
	try:
		entr_C8_V.insert(tk.END, save_C8_V)
	except:
		pass
	try:
		entr_C9_V.insert(tk.END, save_C9_V)
	except:
		pass
	try:
		entr_C10_V.insert(tk.END, save_C10_V)
	except:
		pass

#QCE 4
	try:
		entr_D1_V.insert(tk.END, save_D1_V)
	except:
		pass
	try:
		entr_D2_V.insert(tk.END, save_D2_V)
	except:
		pass
	try:
		entr_D3_V.insert(tk.END, save_D3_V)
	except:
		pass
	try:
		entr_D4_V.insert(tk.END, save_D4_V)
	except:
		pass
	try:
		entr_D5_V.insert(tk.END, save_D5_V)
	except:
		pass
	try:
		entr_D6_V.insert(tk.END, save_D6_V)
	except:
		pass
	try:
		entr_D7_V.insert(tk.END, save_D7_V)
	except:
		pass
	try:
		entr_D8_V.insert(tk.END, save_D8_V)
	except:
		pass
	try:
		entr_D9_V.insert(tk.END, save_D9_V)
	except:
		pass
	try:
		entr_D10_V.insert(tk.END, save_D10_V)
	except:
		pass
#QCE 5
	try:
		entr_E1_V.insert(tk.END, save_E1_V)
	except:
		pass
	try:
		entr_E2_V.insert(tk.END, save_E2_V)
	except:
		pass
	try:
		entr_E3_V.insert(tk.END, save_E3_V)
	except:
		pass
	try:
		entr_E4_V.insert(tk.END, save_E4_V)
	except:
		pass
	try:
		entr_E5_V.insert(tk.END, save_E5_V)
	except:
		pass
	try:
		entr_E6_V.insert(tk.END, save_E6_V)
	except:
		pass
	try:
		entr_E7_V.insert(tk.END, save_E7_V)
	except:
		pass
	try:
		entr_E8_V.insert(tk.END, save_E8_V)
	except:
		pass
	try:
		entr_E9_V.insert(tk.END, save_E9_V)
	except:
		pass
	try:
		entr_E10_V.insert(tk.END, save_E10_V)
	except:
		pass

#QCE 6
	try:
		entr_F1_V.insert(tk.END, save_F1_V)
	except:
		pass
	try:
		entr_F2_V.insert(tk.END, save_F2_V)
	except:
		pass
	try:
		entr_F3_V.insert(tk.END, save_F3_V)
	except:
		pass
	try:
		entr_F4_V.insert(tk.END, save_F4_V)
	except:
		pass
	try:
		entr_F5_V.insert(tk.END, save_F5_V)
	except:
		pass
	try:
		entr_F6_V.insert(tk.END, save_F6_V)
	except:
		pass
	try:
		entr_F7_V.insert(tk.END, save_F7_V)
	except:
		pass
	try:
		entr_F8_V.insert(tk.END, save_F8_V)
	except:
		pass
	try:
		entr_F9_V.insert(tk.END, save_F9_V)
	except:
		pass
	try:
		entr_F10_V.insert(tk.END, save_F10_V)
	except:
		pass



#QCF
#QCF 1 
	try:
		entr_MAIN_VI.insert(tk.END, save_MAIN_VI)
	except:
		pass
	try:
		#entr_A1_I.delete(0, "end")
		entr_A1_VI.insert(tk.END, save_A1_VI)
	except:
		pass
	try:
		entr_A2_VI.insert(tk.END, save_A2_VI)
	except:
		pass
	try:
		entr_A3_VI.insert(tk.END, save_A3_VI)
	except:
		pass
	try:
		entr_A4_VI.insert(tk.END, save_A4_VI)
	except:
		pass
	try:
		entr_A5_VI.insert(tk.END, save_A5_VI)
	except:
		pass
	try:
		entr_A6_VI.insert(tk.END, save_A6_VI)
	except:
		pass
	try:
		entr_A7_VI.insert(tk.END, save_A7_VI)
	except:
		pass
	try:
		entr_A8_VI.insert(tk.END, save_A8_VI)
	except:
		pass
	try:
		entr_A9_VI.insert(tk.END, save_A9_VI)
	except:
		pass
	try:
		entr_A10_VI.insert(tk.END, save_A10_VI)
	except:
		pass
#QCF 2
	try:
		entr_B1_VI.insert(tk.END, save_B1_VI)
	except:
		pass
	try:
		entr_B2_VI.insert(tk.END, save_B2_VI)
	except:
		pass
	try:
		entr_B3_VI.insert(tk.END, save_B3_VI)
	except:
		pass
	try:
		entr_B4_VI.insert(tk.END, save_B4_VI)
	except:
		pass
	try:
		entr_B5_VI.insert(tk.END, save_B5_VI)
	except:
		pass
	try:
		entr_B6_VI.insert(tk.END, save_B6_VI)
	except:
		pass
	try:
		entr_B7_VI.insert(tk.END, save_B7_VI)
	except:
		pass
	try:
		entr_B8_VI.insert(tk.END, save_B8_VI)
	except:
		pass
	try:
		entr_B9_VI.insert(tk.END, save_B9_VI)
	except:
		pass
	try:
		entr_B10_VI.insert(tk.END, save_B10_VI)
	except:
		pass
#QCF 3
	try:
		entr_C1_VI.insert(tk.END, save_C1_VI)
	except:
		pass
	try:
		entr_C2_VI.insert(tk.END, save_C2_VI)
	except:
		pass
	try:
		entr_C3_VI.insert(tk.END, save_C3_VI)
	except:
		pass
	try:
		entr_C4_VI.insert(tk.END, save_C4_VI)
	except:
		pass
	try:
		entr_C5_VI.insert(tk.END, save_C5_VI)
	except:
		pass
	try:
		entr_C6_VI.insert(tk.END, save_C6_VI)
	except:
		pass
	try:
		entr_C7_VI.insert(tk.END, save_C7_VI)
	except:
		pass
	try:
		entr_C8_VI.insert(tk.END, save_C8_VI)
	except:
		pass
	try:
		entr_C9_VI.insert(tk.END, save_C9_VI)
	except:
		pass
	try:
		entr_C10_VI.insert(tk.END, save_C10_VI)
	except:
		pass

#QCF 4
	try:
		entr_D1_VI.insert(tk.END, save_D1_VI)
	except:
		pass
	try:
		entr_D2_VI.insert(tk.END, save_D2_VI)
	except:
		pass
	try:
		entr_D3_VI.insert(tk.END, save_D3_VI)
	except:
		pass
	try:
		entr_D4_VI.insert(tk.END, save_D4_VI)
	except:
		pass
	try:
		entr_D5_VI.insert(tk.END, save_D5_VI)
	except:
		pass
	try:
		entr_D6_VI.insert(tk.END, save_D6_VI)
	except:
		pass
	try:
		entr_D7_VI.insert(tk.END, save_D7_VI)
	except:
		pass
	try:
		entr_D8_VI.insert(tk.END, save_D8_VI)
	except:
		pass
	try:
		entr_D9_VI.insert(tk.END, save_D9_VI)
	except:
		pass
	try:
		entr_D10_VI.insert(tk.END, save_D10_VI)
	except:
		pass
#QCF 5
	try:
		entr_E1_VI.insert(tk.END, save_E1_VI)
	except:
		pass
	try:
		entr_E2_VI.insert(tk.END, save_E2_VI)
	except:
		pass
	try:
		entr_E3_VI.insert(tk.END, save_E3_VI)
	except:
		pass
	try:
		entr_E4_VI.insert(tk.END, save_E4_VI)
	except:
		pass
	try:
		entr_E5_VI.insert(tk.END, save_E5_VI)
	except:
		pass
	try:
		entr_E6_VI.insert(tk.END, save_E6_VI)
	except:
		pass
	try:
		entr_E7_VI.insert(tk.END, save_E7_VI)
	except:
		pass
	try:
		entr_E8_VI.insert(tk.END, save_E8_VI)
	except:
		pass
	try:
		entr_E9_VI.insert(tk.END, save_E9_VI)
	except:
		pass
	try:
		entr_E10_VI.insert(tk.END, save_E10_VI)
	except:
		pass

#QCF 6
	try:
		entr_F1_VI.insert(tk.END, save_F1_VI)
	except:
		pass
	try:
		entr_F2_VI.insert(tk.END, save_F2_VI)
	except:
		pass
	try:
		entr_F3_VI.insert(tk.END, save_F3_VI)
	except:
		pass
	try:
		entr_F4_VI.insert(tk.END, save_F4_VI)
	except:
		pass
	try:
		entr_F5_VI.insert(tk.END, save_F5_VI)
	except:
		pass
	try:
		entr_F6_VI.insert(tk.END, save_F6_VI)
	except:
		pass
	try:
		entr_F7_VI.insert(tk.END, save_F7_VI)
	except:
		pass
	try:
		entr_F8_VI.insert(tk.END, save_F8_VI)
	except:
		pass
	try:
		entr_F9_VI.insert(tk.END, save_F9_VI)
	except:
		pass
	try:
		entr_F10_VI.insert(tk.END, save_F10_VI)
	except:
		pass




	def spravka_close_root():
		toplevel_spravka_root.withdraw()

	def peremeshenie_toplevel_root(event):#перемещение окна toplevel за root
		if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
			x = root.winfo_x() + 910
			y = root.winfo_y() 
			toplevel_spravka_root.geometry("+%d+%d" % (x,y))
		else:			
			x = root.winfo_x() + 1310
			y = root.winfo_y() 
			toplevel_spravka_root.geometry("+%d+%d" % (x,y))
		

		#toplevel_spravka.geometry(f"401x300+{(root_open.winfo_x())+7}+{(root_open.winfo_y()+283)}")

	root.bind("<Configure>", peremeshenie_toplevel_root)

	global toplevel_spravka_root
	toplevel_spravka_root = tk.Toplevel()
	toplevel_spravka_root.title('Справка')
	toplevel_spravka_root.resizable(False, False)
	toplevel_spravka_root.withdraw()
	toplevel_spravka_root.overrideredirect(True)
	if check_var_gr_1.get() == 4 and check_var_gr_2.get() == 3 and check_var_gr_3.get() == 5:
		spravka_text_root = tk.Text(toplevel_spravka_root, width = 43, height = 20)
	else:
		spravka_text_root = tk.Text(toplevel_spravka_root, width = 43, height = 45)
	spravka_text_root.place(x = 1, y = 1)
	btn_spravka_close_root = tk.Button(toplevel_spravka_root, text = 'close', command = spravka_close_root, bg = 'pink')
	btn_spravka_close_root.place(x = 355, y = 15)

	#текстовая часть справки 
	listbox_insert_text_root = '''
1. Зачем нужна программа?
  Программа "AnovaD" автоматизирует  
получение данных о внутригрупповой и
межгрупповойпрецизионности методом
однофакторного дисперсионного анализа,
а также сопутствующих показателей. 
  Программа работает с Excel-файлами
расширения ".xlsx". 

2. Как работать с программой?
Для начала работы выберите параметры на
начальном окне. 
В зависимости от вашего выбора программ
создаст окна ввода для ваших  данных для
расчета. Можно начать работу как по 
выбранным параметрам, так и по "стандарту"
принятом в лаборатории. 
Будьте внимательны, если после выбора
параметров, при вводе данных вы обнаружили,
что неправильно сделали выбор, возврат
на предыдущий экран возможен и вам
не придется делать перезапуск ПО и 
заполнять все заново. 
После выбора параметров программа
предложит вам в зависимости от выбора
форму для заполнения. Заполнять нужно
СТРОГО все ячейки, которые будут
участвовать в расчетах: ячейки QCA, QCB и
тд, ячейки под наименованиями Обр.1, Обр.2
и тд.
Также укажите вашу концентрацию, например
"ng/mL", если не указать программа все
равно расчитает, но поле "Concentation"
будет  заполнено без концентрации,
придется   дописывать вручную. пропуски
недопустимы,  программа не сохранит
результат. 
Ячейки норм QCA, QCB, QCC, QCD и тд,
заполняются по необходимости, если оставить
эти ячейки пустыми, то они примут значения
по умолчанию, QCA - 20 и все остальные по
15 для всех таблиц.

После ввода данных результат расчета 
необходимо сохранить в файл. Для этого
нажмите кнопку, которая находится в 
правом верхнем углу программы
"open file to save".
После нажатия клавиши - будет открыто
меню проводника. 
С помощью данного меню выберите уже
существующий файл, в который будет
произведено сохранение, либо создайте
новый, путем нажатия ПКМ(правой клавиши
мыши), на свободном месте внутри
дирректории. После нажатия на ПКМ, 
появится окно выбора.
Выберите пункт --> создать -->
--> выбираем "Лист Microsoft Excel"
(Убедитесь, что создается .xlsx файл),
после выбора данного пункта, проводник
предложит вам ввести имя файла -->
--> вводим имя файла --> после ввода
имени файла --> нажмите на
клавишу клавиатуры "Enter" --> затем 
выберите ЛКМ(левой кнопкой мыши)
только что созданный файл и нажмите
на кнопку окна проводника "открыть".
Теперь файл открыт внутри программы
(т.е. визуально он не отобразится
для пользователя, но программа будет
понимать, в какой файл идет сохранение)
Теперь нажмите кнопку "save", для
сохранения расчетов в файл.

Также программа поддерживает автозаполнение
ячеек Кнопка "input" - свойственна только 
для окна "стандарт", при нажатии на нее
откроется окно проводника в котором нужно
выбрать excel файл в котором будут
содержаться входные данные для
автозаполнения.
ПРОГРАММА ПОДДЕРЖИВАЕТ ТОЛЬКО xlsx файлы на
вход и выход, будьте внимательны! 
Кнопка "+" рядом с кнопкой "input" также
открывает меню проводника в котором нужно
выбрать пустой ".xlsx" файл, данная кнопка в
этом файле выделит те ячейки, которые
необходимо заполнить для "стандарта", чтобы
произвести автозаполнение из Excel-файла.
По точно такому же алгоритму заполняется
Excel файл на вход для любого другого
размера программы, соответственно:
QC - заполняется в левом верхнем углу в
столбце A, далее идет отступ на 1 строку
вниз и на один столбец вправо. 
Соответственно заполнение значений QCA1,
QCA2 и тд начинается с ячейки "C2" и 
смещается влево(в зависимости от количества
QC) и вниз(в зависисмости от n), после
заполнения необходимого количества
показателей QCA, идет смещение на одну
строку вниз и начинается заполнения
показателя QCB. 
Программа для заполнения ориентируется по
концентрации которая указана в столбце "A",
поэтому ее правильное заполнение
обязательно.
Кнопка "open file to save" позволяет
выбрать необходимый файл .xlsx для 
проведения в него расчетов. 
Кнопка "save" - проводит расчеты и
сохраняет данные в выбранном файле.

Связь с разработчиком
email - daniil.popkov@gmail.com
WhatsApp - +7 985 187 81-24
	'''

	for i in listbox_insert_text_root:
		spravka_text_root.insert(tk.END, str(i))

	root.mainloop()




def calc():

	if check_razryad.get() == 0:
		razryad = 0 #'%.0f'
	elif check_razryad.get() == 1:
		razryad = 1 #'%.1f'
	elif check_razryad.get() == 2:
		razryad = 2 #'%.2f'
	elif check_razryad.get() == 3:
		razryad = 3 #'%.3f'
	elif check_razryad.get() == 4:
		razryad = 4 #'%.4f'

	#get


	#функция выдает диалоговое окно с ошибкой, если в колонке A в файле есть данные 
	def error_1():		
		error_1_top = tk.Tk()
		error_1_top.title('Ошибка Сохранения Результатов')
		error_1_top.geometry(f"300x100+{(root.winfo_x())+300}+{(root.winfo_y()+150)}")
		error_1_top.configure(bg = '#ffe5ea')
		error_1_top.after(3000, lambda: error_1_top.destroy())
		label_err_1 = tk.Label(error_1_top, text = 'Выберите\nдругой файл!', bg = '#ffe5ea', font = 11)
		label_err_1.pack(pady = 20)
		error_1_top.mainloop()

	A_column = ws['A']
	all_A_column_values = []
	for i in A_column:
		all_A_column_values.append(i)
	for i in all_A_column_values:	
		if i.value != None:
			error_1()
			return 




	QC_I = float(entr_MAIN_I.get())
	spisok_A_I = []
	spisok_A_I.append(my_round((float(entr_A1_I.get())), check_razryad.get()))
	spisok_A_I.append(my_round((float(entr_A2_I.get())), check_razryad.get()))
	spisok_A_I.append(my_round((float(entr_A3_I.get())), check_razryad.get()))

	if check_var_gr_3.get() >= 4:
		spisok_A_I.append(my_round((float(entr_A4_I.get())), check_razryad.get()))
	if check_var_gr_3.get() >= 5:
		spisok_A_I.append(my_round((float(entr_A5_I.get())), check_razryad.get()))
	if check_var_gr_3.get() >= 6:
		spisok_A_I.append(my_round((float(entr_A6_I.get())), check_razryad.get()))
	if check_var_gr_3.get() >= 7:
		spisok_A_I.append(my_round((float(entr_A7_I.get())), check_razryad.get()))
	if check_var_gr_3.get() >= 8:
		spisok_A_I.append(my_round((float(entr_A8_I.get())), check_razryad.get()))
	if check_var_gr_3.get() >= 9:
		spisok_A_I.append(my_round((float(entr_A9_I.get())), check_razryad.get()))
	if check_var_gr_3.get() >= 10:
		spisok_A_I.append(my_round((float(entr_A10_I.get())), check_razryad.get()))			


	mean_A_I = my_round(stat.mean(spisok_A_I), check_razryad.get())
	Er_A_I = ((mean_A_I - QC_I)/QC_I)*100
	sigma_A_I = (stat.stdev(spisok_A_I)/mean_A_I)*100


	if check_var_gr_2.get() >= 2:

		spisok_B_I = []
		spisok_B_I.append(my_round((float(entr_B1_I.get())), check_razryad.get()))
		spisok_B_I.append(my_round((float(entr_B2_I.get())), check_razryad.get()))
		spisok_B_I.append(my_round((float(entr_B3_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 4:
			spisok_B_I.append(my_round((float(entr_B4_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_B_I.append(my_round((float(entr_B5_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_B_I.append(my_round((float(entr_B6_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_B_I.append(my_round((float(entr_B7_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_B_I.append(my_round((float(entr_B8_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_B_I.append(my_round((float(entr_B9_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:	
			spisok_B_I.append(my_round((float(entr_B10_I.get())), check_razryad.get()))

		mean_B_I = my_round(stat.mean(spisok_B_I), check_razryad.get())

		
		Er_B_I = ((mean_B_I - QC_I)/QC_I)*100

		sigma_B_I = (stat.stdev(spisok_B_I)/mean_B_I)*100



	if check_var_gr_2.get() >= 3:
		
		spisok_C_I = []
		spisok_C_I.append(my_round((float(entr_C1_I.get())), check_razryad.get()))
		spisok_C_I.append(my_round((float(entr_C2_I.get())), check_razryad.get()))
		spisok_C_I.append(my_round((float(entr_C3_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 4:
			spisok_C_I.append(my_round((float(entr_C4_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_C_I.append(my_round((float(entr_C5_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_C_I.append(my_round((float(entr_C6_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_C_I.append(my_round((float(entr_C7_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_C_I.append(my_round((float(entr_C8_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_C_I.append(my_round((float(entr_C9_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_C_I.append(my_round((float(entr_C10_I.get())), check_razryad.get()))
		
		mean_C_I = my_round(stat.mean(spisok_C_I), check_razryad.get())
		Er_C_I = ((mean_C_I - QC_I)/QC_I)*100
		sigma_C_I = (stat.stdev(spisok_C_I)/mean_C_I)*100



	if check_var_gr_2.get() >= 4:
		spisok_D_I = []
		spisok_D_I.append(my_round((float(entr_D1_I.get())), check_razryad.get()))
		spisok_D_I.append(my_round((float(entr_D2_I.get())), check_razryad.get()))
		spisok_D_I.append(my_round((float(entr_D3_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 4:
			spisok_D_I.append(my_round((float(entr_D4_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_D_I.append(my_round((float(entr_D5_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_D_I.append(my_round((float(entr_D6_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_D_I.append(my_round((float(entr_D7_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_D_I.append(my_round((float(entr_D8_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_D_I.append(my_round((float(entr_D9_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_D_I.append(my_round((float(entr_D10_I.get())), check_razryad.get()))
		
		mean_D_I = my_round(stat.mean(spisok_D_I), check_razryad.get())
		Er_D_I = ((mean_D_I - QC_I)/QC_I)*100
		sigma_D_I = (stat.stdev(spisok_D_I)/mean_D_I)*100


	if check_var_gr_2.get() >= 5:
		spisok_E_I = []
		spisok_E_I.append(my_round((float(entr_E1_I.get())), check_razryad.get()))
		spisok_E_I.append(my_round((float(entr_E2_I.get())), check_razryad.get()))
		spisok_E_I.append(my_round((float(entr_E3_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 4:
			spisok_E_I.append(my_round((float(entr_E4_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_E_I.append(my_round((float(entr_E5_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_E_I.append(my_round((float(entr_E6_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_E_I.append(my_round((float(entr_E7_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_E_I.append(my_round((float(entr_E8_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_E_I.append(my_round((float(entr_E9_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_E_I.append(my_round((float(entr_E10_I.get())), check_razryad.get()))

		mean_E_I = my_round(stat.mean(spisok_E_I), check_razryad.get())
		Er_E_I = ((mean_E_I - QC_I)/QC_I)*100
		sigma_E_I = (stat.stdev(spisok_E_I)/mean_E_I)*100

	if check_var_gr_2.get() >= 6:
		
		spisok_F_I = []
		spisok_F_I.append(my_round((float(entr_F1_I.get())), check_razryad.get()))
		spisok_F_I.append(my_round((float(entr_F2_I.get())), check_razryad.get()))
		spisok_F_I.append(my_round((float(entr_F3_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 4:
			spisok_F_I.append(my_round((float(entr_F4_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_F_I.append(my_round((float(entr_F5_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_F_I.append(my_round((float(entr_F6_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_F_I.append(my_round((float(entr_F7_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_F_I.append(my_round((float(entr_F8_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_F_I.append(my_round((float(entr_F9_I.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_F_I.append(my_round((float(entr_F10_I.get())), check_razryad.get()))

		mean_F_I = my_round(stat.mean(spisok_F_I), check_razryad.get())
		Er_F_I = (((mean_F_I - QC_I)/QC_I)*100)
		sigma_F_I = ((stat.stdev(spisok_F_I)/mean_F_I)*100)

	if check_var_gr_1.get() >= 2:

		QC_II = float(entr_MAIN_II.get())
		spisok_A_II = []
		spisok_A_II.append(my_round((float(entr_A1_II.get())), check_razryad.get()))
		spisok_A_II.append(my_round((float(entr_A2_II.get())), check_razryad.get()))
		spisok_A_II.append(my_round((float(entr_A3_II.get())), check_razryad.get()))

		if check_var_gr_3.get() >= 4:
			spisok_A_II.append(my_round((float(entr_A4_II.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_A_II.append(my_round((float(entr_A5_II.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_A_II.append(my_round((float(entr_A6_II.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_A_II.append(my_round((float(entr_A7_II.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_A_II.append(my_round((float(entr_A8_II.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_A_II.append(my_round((float(entr_A9_II.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_A_II.append(my_round((float(entr_A10_II.get()))	, check_razryad.get()))		

		mean_A_II = my_round(stat.mean(spisok_A_II), check_razryad.get())
		Er_A_II = ((mean_A_II - QC_II)/QC_II)*100
		sigma_A_II = ((stat.stdev(spisok_A_II)/mean_A_II)*100)



		if check_var_gr_2.get() >= 2:
			
			spisok_B_II = []
			spisok_B_II.append(my_round((float(entr_B1_II.get())), check_razryad.get()))
			spisok_B_II.append(my_round((float(entr_B2_II.get())), check_razryad.get()))
			spisok_B_II.append(my_round((float(entr_B3_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_B_II.append(my_round((float(entr_B4_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_B_II.append(my_round((float(entr_B5_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_B_II.append(my_round((float(entr_B6_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_B_II.append(my_round((float(entr_B7_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_B_II.append(my_round((float(entr_B8_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_B_II.append(my_round((float(entr_B9_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:	
				spisok_B_II.append(my_round((float(entr_B10_II.get())), check_razryad.get()))

			mean_B_II = my_round(stat.mean(spisok_B_II), check_razryad.get())
			Er_B_II = (((mean_B_II - QC_II)/QC_II)*100)
			sigma_B_II = (stat.stdev(spisok_B_II)/mean_B_II)*100

		if check_var_gr_2.get() >= 3:

			spisok_C_II = []
			spisok_C_II.append(my_round((float(entr_C1_II.get())), check_razryad.get()))
			spisok_C_II.append(my_round((float(entr_C2_II.get())), check_razryad.get()))
			spisok_C_II.append(my_round((float(entr_C3_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_C_II.append(my_round((float(entr_C4_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_C_II.append(my_round((float(entr_C5_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_C_II.append(my_round((float(entr_C6_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_C_II.append(my_round((float(entr_C7_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_C_II.append(my_round((float(entr_C8_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_C_II.append(my_round((float(entr_C9_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_C_II.append(my_round((float(entr_C10_II.get())), check_razryad.get()))
			
			mean_C_II = my_round(stat.mean(spisok_C_II), check_razryad.get())
			Er_C_II = ((mean_C_II - QC_II)/QC_II)*100
			sigma_C_II = (stat.stdev(spisok_C_II)/mean_C_II)*100

		if check_var_gr_2.get() >= 4:
			spisok_D_II = []
			spisok_D_II.append(my_round((float(entr_D1_II.get())), check_razryad.get()))
			spisok_D_II.append(my_round((float(entr_D2_II.get())), check_razryad.get()))
			spisok_D_II.append(my_round((float(entr_D3_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_D_II.append(my_round((float(entr_D4_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_D_II.append(my_round((float(entr_D5_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_D_II.append(my_round((float(entr_D6_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_D_II.append(my_round((float(entr_D7_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_D_II.append(my_round((float(entr_D8_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_D_II.append(my_round((float(entr_D9_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_D_II.append(my_round((float(entr_D10_II.get())), check_razryad.get()))
			

			mean_D_II = my_round(stat.mean(spisok_D_II), check_razryad.get())
			Er_D_II = ((mean_D_II - QC_II)/QC_II)*100
			sigma_D_II = (stat.stdev(spisok_D_II)/mean_D_II)*100


		if check_var_gr_2.get() >= 5:
			spisok_E_II = []
			spisok_E_II.append(my_round((float(entr_E1_II.get())), check_razryad.get()))
			spisok_E_II.append(my_round((float(entr_E2_II.get())), check_razryad.get()))
			spisok_E_II.append(my_round((float(entr_E3_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_E_II.append(my_round((float(entr_E4_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_E_II.append(my_round((float(entr_E5_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_E_II.append(my_round((float(entr_E6_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_E_II.append(my_round((float(entr_E7_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_E_II.append(my_round((float(entr_E8_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_E_II.append(my_round((float(entr_E9_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_E_II.append(my_round((float(entr_E10_II.get())), check_razryad.get()))
			
			mean_E_II = my_round(stat.mean(spisok_E_II), check_razryad.get())
			Er_E_II = ((mean_E_II - QC_II)/QC_II)*100
			sigma_E_II = (stat.stdev(spisok_E_II)/mean_E_II)*100


		if check_var_gr_2.get() >= 6:
		
			spisok_F_II = []
			spisok_F_II.append(my_round((float(entr_F1_II.get())), check_razryad.get()))
			spisok_F_II.append(my_round((float(entr_F2_II.get())), check_razryad.get()))
			spisok_F_II.append(my_round((float(entr_F3_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_F_II.append(my_round((float(entr_F4_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_F_II.append(my_round((float(entr_F5_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_F_II.append(my_round((float(entr_F6_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_F_II.append(my_round((float(entr_F7_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_F_II.append(my_round((float(entr_F8_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_F_II.append(my_round((float(entr_F9_II.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_F_II.append(my_round((float(entr_F10_II.get())), check_razryad.get()))

			mean_F_II = my_round(stat.mean(spisok_F_II), check_razryad.get())
			Er_F_II = ((mean_F_II - QC_II)/QC_II)*100
			sigma_F_II = (stat.stdev(spisok_F_II)/mean_F_II)*100

	if check_var_gr_1.get() >= 3:

		QC_III = float(entr_MAIN_III.get())
		spisok_A_III = []
		spisok_A_III.append(my_round((float(entr_A1_III.get())), check_razryad.get()))
		spisok_A_III.append(my_round((float(entr_A2_III.get())), check_razryad.get()))
		spisok_A_III.append(my_round((float(entr_A3_III.get())), check_razryad.get()))

		if check_var_gr_3.get() >= 4:
			spisok_A_III.append(my_round((float(entr_A4_III.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_A_III.append(my_round((float(entr_A5_III.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_A_III.append(my_round((float(entr_A6_III.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_A_III.append(my_round((float(entr_A7_III.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_A_III.append(my_round((float(entr_A8_III.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_A_III.append(my_round((float(entr_A9_III.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_A_III.append(my_round((float(entr_A10_III.get())), check_razryad.get()))	

		mean_A_III = my_round(stat.mean(spisok_A_III), check_razryad.get())
		Er_A_III = ((mean_A_III - QC_III)/QC_III)*100
		sigma_A_III = (stat.stdev(spisok_A_III)/mean_A_III)*100

		if check_var_gr_2.get() >= 2:
			spisok_B_III = []
			spisok_B_III.append(my_round((float(entr_B1_III.get())), check_razryad.get()))
			spisok_B_III.append(my_round((float(entr_B2_III.get())), check_razryad.get()))
			spisok_B_III.append(my_round((float(entr_B3_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_B_III.append(my_round((float(entr_B4_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_B_III.append(my_round((float(entr_B5_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_B_III.append(my_round((float(entr_B6_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_B_III.append(my_round((float(entr_B7_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_B_III.append(my_round((float(entr_B8_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_B_III.append(my_round((float(entr_B9_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:	
				spisok_B_III.append(my_round((float(entr_B10_III.get())), check_razryad.get()))

			mean_B_III = my_round(stat.mean(spisok_B_III), check_razryad.get())
			Er_B_III = ((mean_B_III - QC_III)/QC_III)*100
			sigma_B_III = (stat.stdev(spisok_B_III)/mean_B_III)*100

		if check_var_gr_2.get() >= 3:
			
			spisok_C_III = []
			spisok_C_III.append(my_round((float(entr_C1_III.get())), check_razryad.get()))
			spisok_C_III.append(my_round((float(entr_C2_III.get())), check_razryad.get()))
			spisok_C_III.append(my_round((float(entr_C3_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_C_III.append(my_round((float(entr_C4_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_C_III.append(my_round((float(entr_C5_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_C_III.append(my_round((float(entr_C6_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_C_III.append(my_round((float(entr_C7_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_C_III.append(my_round((float(entr_C8_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_C_III.append(my_round((float(entr_C9_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_C_III.append(my_round((float(entr_C10_III.get())), check_razryad.get()))
			
			mean_C_III = my_round(stat.mean(spisok_C_III), check_razryad.get())
			Er_C_III = ((mean_C_III - QC_III)/QC_III)*100
			sigma_C_III = (stat.stdev(spisok_C_III)/mean_C_III)*100

		if check_var_gr_2.get() >= 4:
			spisok_D_III = []
			spisok_D_III.append(my_round((float(entr_D1_III.get())), check_razryad.get()))
			spisok_D_III.append(my_round((float(entr_D2_III.get())), check_razryad.get()))
			spisok_D_III.append(my_round((float(entr_D3_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_D_III.append(my_round((float(entr_D4_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_D_III.append(my_round((float(entr_D5_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_D_III.append(my_round((float(entr_D6_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_D_III.append(my_round((float(entr_D7_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_D_III.append(my_round((float(entr_D8_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_D_III.append(my_round((float(entr_D9_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_D_III.append(my_round((float(entr_D10_III.get())), check_razryad.get()))
			

			mean_D_III = my_round(stat.mean(spisok_D_III), check_razryad.get())
			

			Er_D_III = ((mean_D_III - QC_III)/QC_III)*100

			sigma_D_III = (stat.stdev(spisok_D_III)/mean_D_III)*100


		if check_var_gr_2.get() >= 5:
			
			spisok_E_III = []
			spisok_E_III.append(my_round((float(entr_E1_III.get())), check_razryad.get()))
			spisok_E_III.append(my_round((float(entr_E2_III.get())), check_razryad.get()))
			spisok_E_III.append(my_round((float(entr_E3_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_E_III.append(my_round((float(entr_E4_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_E_III.append(my_round((float(entr_E5_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_E_III.append(my_round((float(entr_E6_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_E_III.append(my_round((float(entr_E7_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_E_III.append(my_round((float(entr_E8_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_E_III.append(my_round((float(entr_E9_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_E_III.append(my_round((float(entr_E10_III.get())), check_razryad.get()))
			
			mean_E_III = my_round(stat.mean(spisok_E_III), check_razryad.get())
			Er_E_III = ((mean_E_III - QC_III)/QC_III)*100
			sigma_E_III = (stat.stdev(spisok_E_III)/mean_E_III)*100


		if check_var_gr_2.get() >= 6:
	
			spisok_F_III = []
			spisok_F_III.append(my_round((float(entr_F1_III.get())), check_razryad.get()))
			spisok_F_III.append(my_round((float(entr_F2_III.get())), check_razryad.get()))
			spisok_F_III.append(my_round((float(entr_F3_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_F_III.append(my_round((float(entr_F4_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_F_III.append(my_round((float(entr_F5_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_F_III.append(my_round((float(entr_F6_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_F_III.append(my_round((float(entr_F7_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_F_III.append(my_round((float(entr_F8_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_F_III.append(my_round((float(entr_F9_III.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_F_III.append(my_round((float(entr_F10_III.get())), check_razryad.get()))


			
			mean_F_III = my_round(stat.mean(spisok_F_III), check_razryad.get())
	
			Er_F_III = ((mean_F_III - QC_III)/QC_III)*100
		
			sigma_F_III = (stat.stdev(spisok_F_III)/mean_F_III)*100



# проверка условия на количество групп для 1 группы не требуется т.к. значение по  умолчанию 1

	if check_var_gr_1.get() >= 4:


		QC_IV = float(entr_MAIN_IV.get())
		spisok_A_IV = []
		spisok_A_IV.append(my_round((float(entr_A1_IV.get())), check_razryad.get()))
		spisok_A_IV.append(my_round((float(entr_A2_IV.get())), check_razryad.get()))
		spisok_A_IV.append(my_round((float(entr_A3_IV.get())), check_razryad.get()))

		if check_var_gr_3.get() >= 4:
			spisok_A_IV.append(my_round((float(entr_A4_IV.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_A_IV.append(my_round((float(entr_A5_IV.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_A_IV.append(my_round((float(entr_A6_IV.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_A_IV.append(my_round((float(entr_A7_IV.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_A_IV.append(my_round((float(entr_A8_IV.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_A_IV.append(my_round((float(entr_A9_IV.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_A_IV.append(my_round((float(entr_A10_IV.get())), check_razryad.get()))			

		mean_A_IV = my_round(stat.mean(spisok_A_IV), check_razryad.get())
	
		Er_A_IV = ((mean_A_IV - QC_IV)/QC_IV)*100

		sigma_A_IV = (stat.stdev(spisok_A_IV)/mean_A_IV)*100


		if check_var_gr_2.get() >= 2:
			
			spisok_B_IV = []
			spisok_B_IV.append(my_round((float(entr_B1_IV.get())), check_razryad.get()))
			spisok_B_IV.append(my_round((float(entr_B2_IV.get())), check_razryad.get()))
			spisok_B_IV.append(my_round((float(entr_B3_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_B_IV.append(my_round((float(entr_B4_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_B_IV.append(my_round((float(entr_B5_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_B_IV.append(my_round((float(entr_B6_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_B_IV.append(my_round((float(entr_B7_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_B_IV.append(my_round((float(entr_B8_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_B_IV.append(my_round((float(entr_B9_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:	
				spisok_B_IV.append(my_round((float(entr_B10_IV.get())), check_razryad.get()))

			mean_B_IV = my_round(stat.mean(spisok_B_IV), check_razryad.get())
			Er_B_IV = ((mean_B_IV - QC_IV)/QC_IV)*100
			sigma_B_IV = (stat.stdev(spisok_B_IV)/mean_B_IV)*100



		if check_var_gr_2.get() >= 3:

			spisok_C_IV = []
			spisok_C_IV.append(my_round((float(entr_C1_IV.get())), check_razryad.get()))
			spisok_C_IV.append(my_round((float(entr_C2_IV.get())), check_razryad.get()))
			spisok_C_IV.append(my_round((float(entr_C3_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_C_IV.append(my_round((float(entr_C4_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_C_IV.append(my_round((float(entr_C5_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_C_IV.append(my_round((float(entr_C6_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_C_IV.append(my_round((float(entr_C7_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_C_IV.append(my_round((float(entr_C8_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_C_IV.append(my_round((float(entr_C9_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_C_IV.append(my_round((float(entr_C10_IV.get())), check_razryad.get()))
			
			mean_C_IV = my_round(stat.mean(spisok_C_IV), check_razryad.get())
			Er_C_IV = ((mean_C_IV - QC_IV)/QC_IV)*100
			sigma_C_IV = (stat.stdev(spisok_C_IV)/mean_C_IV)*100


		if check_var_gr_2.get() >= 4:
		
			spisok_D_IV = []
			spisok_D_IV.append(my_round((float(entr_D1_IV.get())), check_razryad.get()))
			spisok_D_IV.append(my_round((float(entr_D2_IV.get())), check_razryad.get()))
			spisok_D_IV.append(my_round((float(entr_D3_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_D_IV.append(my_round((float(entr_D4_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_D_IV.append(my_round((float(entr_D5_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_D_IV.append(my_round((float(entr_D6_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_D_IV.append(my_round((float(entr_D7_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_D_IV.append(my_round((float(entr_D8_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_D_IV.append(my_round((float(entr_D9_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_D_IV.append(my_round((float(entr_D10_IV.get())), check_razryad.get()))
			
			mean_D_IV = my_round(stat.mean(spisok_D_IV), check_razryad.get())	
			Er_D_IV = ((mean_D_IV - QC_IV)/QC_IV)*100
			sigma_D_IV = (stat.stdev(spisok_D_IV)/mean_D_IV)*100


	#проверка условия по количеству групп если больше либо равно 5 этот код выполняется

		if check_var_gr_2.get() >= 5:
			spisok_E_IV = []
			spisok_E_IV.append(my_round((float(entr_E1_IV.get())), check_razryad.get()))
			spisok_E_IV.append(my_round((float(entr_E2_IV.get())), check_razryad.get()))
			spisok_E_IV.append(my_round((float(entr_E3_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_E_IV.append(my_round((float(entr_E4_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_E_IV.append(my_round((float(entr_E5_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_E_IV.append(my_round((float(entr_E6_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_E_IV.append(my_round((float(entr_E7_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_E_IV.append(my_round((float(entr_E8_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_E_IV.append(my_round((float(entr_E9_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_E_IV.append(my_round((float(entr_E10_IV.get())), check_razryad.get()))
			

			mean_E_IV = my_round(stat.mean(spisok_E_IV), check_razryad.get())
			Er_E_IV = ((mean_E_IV - QC_IV)/QC_IV)*100
			sigma_E_IV = (stat.stdev(spisok_E_IV)/mean_E_IV)*100



		if check_var_gr_2.get() >= 6:
		
			spisok_F_IV = []
			spisok_F_IV.append(my_round((float(entr_F1_IV.get())), check_razryad.get()))
			spisok_F_IV.append(my_round((float(entr_F2_IV.get())), check_razryad.get()))
			spisok_F_IV.append(my_round((float(entr_F3_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_F_IV.append(my_round((float(entr_F4_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_F_IV.append(my_round((float(entr_F5_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_F_IV.append(my_round((float(entr_F6_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_F_IV.append(my_round((float(entr_F7_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_F_IV.append(my_round((float(entr_F8_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_F_IV.append(my_round((float(entr_F9_IV.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_F_IV.append(my_round((float(entr_F10_IV.get())), check_razryad.get()))

			mean_F_IV = my_round(stat.mean(spisok_F_IV), check_razryad.get())
			
			Er_F_IV = ((mean_F_IV - QC_IV)/QC_IV)*100

			sigma_F_IV = (stat.stdev(spisok_F_IV)/mean_F_IV)*100



	if check_var_gr_1.get() >= 5:

		QC_V = float(entr_MAIN_V.get())
		spisok_A_V = []
		spisok_A_V.append(my_round((float(entr_A1_V.get())), check_razryad.get()))
		spisok_A_V.append(my_round((float(entr_A2_V.get())), check_razryad.get()))
		spisok_A_V.append(my_round((float(entr_A3_V.get())), check_razryad.get()))

		if check_var_gr_3.get() >= 4:
			spisok_A_V.append(my_round((float(entr_A4_V.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_A_V.append(my_round((float(entr_A5_V.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_A_V.append(my_round((float(entr_A6_V.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_A_V.append(my_round((float(entr_A7_V.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_A_V.append(my_round((float(entr_A8_V.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_A_V.append(my_round((float(entr_A9_V.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_A_V.append(my_round((float(entr_A10_V.get())), check_razryad.get()))	

		#calc

		#среднее по QCA1
		mean_A_V = my_round(stat.mean(spisok_A_V), check_razryad.get())

	
		Er_A_V = ((mean_A_V - QC_V)/QC_V)*100

		sigma_A_V = (stat.stdev(spisok_A_V)/mean_A_V)*100

	# проверка условия по количеству групп если больше либо равно 2 этот код выполняется

		##########################

		if check_var_gr_2.get() >= 2:
		

		##########################

			spisok_B_V = []
			spisok_B_V.append(my_round((float(entr_B1_V.get())), check_razryad.get()))
			spisok_B_V.append(my_round((float(entr_B2_V.get())), check_razryad.get()))
			spisok_B_V.append(my_round((float(entr_B3_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_B_V.append(my_round((float(entr_B4_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_B_V.append(my_round((float(entr_B5_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_B_V.append(my_round((float(entr_B6_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_B_V.append(my_round((float(entr_B7_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_B_V.append(my_round((float(entr_B8_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_B_V.append(my_round((float(entr_B9_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:	
				spisok_B_V.append(my_round((float(entr_B10_V.get())), check_razryad.get()))

			mean_B_V = my_round(stat.mean(spisok_B_V), check_razryad.get())
			
			Er_B_V = ((mean_B_V - QC_V)/QC_V)*100

			sigma_B_V = (stat.stdev(spisok_B_V)/mean_B_V)*100

		if check_var_gr_2.get() >= 3:
			
			spisok_C_V = []
			spisok_C_V.append(my_round((float(entr_C1_V.get())), check_razryad.get()))
			spisok_C_V.append(my_round((float(entr_C2_V.get())), check_razryad.get()))
			spisok_C_V.append(my_round((float(entr_C3_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_C_V.append(my_round((float(entr_C4_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_C_V.append(my_round((float(entr_C5_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_C_V.append(my_round((float(entr_C6_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_C_V.append(my_round((float(entr_C7_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_C_V.append(my_round((float(entr_C8_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_C_V.append(my_round((float(entr_C9_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_C_V.append(my_round((float(entr_C10_V.get())), check_razryad.get()))
			

			mean_C_V = my_round(stat.mean(spisok_C_V), check_razryad.get())
			
			Er_C_V = ((mean_C_V - QC_V)/QC_V)*100
			sigma_C_V = (stat.stdev(spisok_C_V)/mean_C_V)*100

		if check_var_gr_2.get() >= 4:
	
			spisok_D_V = []
			spisok_D_V.append(my_round((float(entr_D1_V.get())), check_razryad.get()))
			spisok_D_V.append(my_round((float(entr_D2_V.get())), check_razryad.get()))
			spisok_D_V.append(my_round((float(entr_D3_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_D_V.append(my_round((float(entr_D4_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_D_V.append(my_round((float(entr_D5_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_D_V.append(my_round((float(entr_D6_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_D_V.append(my_round((float(entr_D7_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_D_V.append(my_round((float(entr_D8_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_D_V.append(my_round((float(entr_D9_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_D_V.append(my_round((float(entr_D10_V.get())), check_razryad.get()))
			

			mean_D_V = my_round(stat.mean(spisok_D_V), check_razryad.get())
			Er_D_V = ((mean_D_V - QC_V)/QC_V)*100
			sigma_D_V = (stat.stdev(spisok_D_V)/mean_D_V)*100
		if check_var_gr_2.get() >= 5:
			
			spisok_E_V = []
			spisok_E_V.append(my_round((float(entr_E1_V.get())), check_razryad.get()))
			spisok_E_V.append(my_round((float(entr_E2_V.get())), check_razryad.get()))
			spisok_E_V.append(my_round((float(entr_E3_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_E_V.append(my_round((float(entr_E4_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_E_V.append(my_round((float(entr_E5_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_E_V.append(my_round((float(entr_E6_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_E_V.append(my_round((float(entr_E7_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_E_V.append(my_round((float(entr_E8_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_E_V.append(my_round((float(entr_E9_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_E_V.append(my_round((float(entr_E10_V.get())), check_razryad.get()))
			

			mean_E_V = my_round(stat.mean(spisok_E_V), check_razryad.get())
			
			Er_E_V = ((mean_E_V - QC_V)/QC_V)*100
			sigma_E_V = (stat.stdev(spisok_E_V)/mean_E_V)*100
		if check_var_gr_2.get() >= 6:
			spisok_F_V = []
			spisok_F_V.append(my_round((float(entr_F1_V.get())), check_razryad.get()))
			spisok_F_V.append(my_round((float(entr_F2_V.get())), check_razryad.get()))
			spisok_F_V.append(my_round((float(entr_F3_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_F_V.append(my_round((float(entr_F4_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_F_V.append(my_round((float(entr_F5_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_F_V.append(my_round((float(entr_F6_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_F_V.append(my_round((float(entr_F7_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_F_V.append(my_round((float(entr_F8_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_F_V.append(my_round((float(entr_F9_V.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_F_V.append(my_round((float(entr_F10_V.get())), check_razryad.get()))


			mean_F_V = my_round(stat.mean(spisok_F_V), check_razryad.get())
			Er_F_V = ((mean_F_V - QC_V)/QC_V)*100
			sigma_F_V = (stat.stdev(spisok_F_V)/mean_F_V)*100

	if check_var_gr_1.get() >= 6:

		QC_VI = float(entr_MAIN_VI.get())
		spisok_A_VI = []
		spisok_A_VI.append(my_round((float(entr_A1_VI.get())), check_razryad.get()))
		spisok_A_VI.append(my_round((float(entr_A2_VI.get())), check_razryad.get()))
		spisok_A_VI.append(my_round((float(entr_A3_VI.get())), check_razryad.get()))

		if check_var_gr_3.get() >= 4:
			spisok_A_VI.append(my_round((float(entr_A4_VI.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 5:
			spisok_A_VI.append(my_round((float(entr_A5_VI.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 6:
			spisok_A_VI.append(my_round((float(entr_A6_VI.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 7:
			spisok_A_VI.append(my_round((float(entr_A7_VI.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 8:
			spisok_A_VI.append(my_round((float(entr_A8_VI.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 9:
			spisok_A_VI.append(my_round((float(entr_A9_VI.get())), check_razryad.get()))
		if check_var_gr_3.get() >= 10:
			spisok_A_VI.append(my_round((float(entr_A10_VI.get())), check_razryad.get()))		

		mean_A_VI = my_round(stat.mean(spisok_A_VI), check_razryad.get())
		Er_A_VI = ((mean_A_VI - QC_VI)/QC_VI)*100
		sigma_A_VI = (stat.stdev(spisok_A_VI)/mean_A_VI)*100

		if check_var_gr_2.get() >= 2:
			
			spisok_B_VI = []
			spisok_B_VI.append(my_round((float(entr_B1_VI.get())), check_razryad.get()))
			spisok_B_VI.append(my_round((float(entr_B2_VI.get())), check_razryad.get()))
			spisok_B_VI.append(my_round((float(entr_B3_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_B_VI.append(my_round((float(entr_B4_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_B_VI.append(my_round((float(entr_B5_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_B_VI.append(my_round((float(entr_B6_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_B_VI.append(my_round((float(entr_B7_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_B_VI.append(my_round((float(entr_B8_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_B_VI.append(my_round((float(entr_B9_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:	
				spisok_B_VI.append(my_round((float(entr_B10_VI.get())), check_razryad.get()))

			mean_B_VI = my_round(stat.mean(spisok_B_VI), check_razryad.get())
			Er_B_VI = ((mean_B_VI - QC_VI)/QC_VI)*100
			sigma_B_VI = (stat.stdev(spisok_B_VI)/mean_B_VI)*100

		if check_var_gr_2.get() >= 3:
			
			spisok_C_VI = []
			spisok_C_VI.append(my_round((float(entr_C1_VI.get())), check_razryad.get()))
			spisok_C_VI.append(my_round((float(entr_C2_VI.get())), check_razryad.get()))
			spisok_C_VI.append(my_round((float(entr_C3_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_C_VI.append(my_round((float(entr_C4_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_C_VI.append(my_round((float(entr_C5_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_C_VI.append(my_round((float(entr_C6_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_C_VI.append(my_round((float(entr_C7_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_C_VI.append(my_round((float(entr_C8_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_C_VI.append(my_round((float(entr_C9_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_C_VI.append(my_round((float(entr_C10_VI.get())), check_razryad.get()))
			

			mean_C_VI = my_round(stat.mean(spisok_C_VI), check_razryad.get())
			Er_C_VI = ((mean_C_VI - QC_VI)/QC_VI)*100
			sigma_C_VI = (stat.stdev(spisok_C_VI)/mean_C_VI)*100

		if check_var_gr_2.get() >= 4:

			spisok_D_VI = []
			spisok_D_VI.append(my_round((float(entr_D1_VI.get())), check_razryad.get()))
			spisok_D_VI.append(my_round((float(entr_D2_VI.get())), check_razryad.get()))
			spisok_D_VI.append(my_round((float(entr_D3_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_D_VI.append(my_round((float(entr_D4_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_D_VI.append(my_round((float(entr_D5_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_D_VI.append(my_round((float(entr_D6_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_D_VI.append(my_round((float(entr_D7_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_D_VI.append(my_round((float(entr_D8_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_D_VI.append(my_round((float(entr_D9_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_D_VI.append(my_round((float(entr_D10_VI.get())), check_razryad.get()))
			
			mean_D_VI = my_round(stat.mean(spisok_D_VI), check_razryad.get())
			Er_D_VI = ((mean_D_VI - QC_VI)/QC_VI)*100

			sigma_D_VI = (stat.stdev(spisok_D_VI)/mean_D_VI)*100

	#проверка условия по количеству групп если больше либо равно 5 этот код выполняется

		##########################
		if check_var_gr_2.get() >= 5:
			
		##########################
			spisok_E_VI = []
			spisok_E_VI.append(my_round((float(entr_E1_VI.get())), check_razryad.get()))
			spisok_E_VI.append(my_round((float(entr_E2_VI.get())), check_razryad.get()))
			spisok_E_VI.append(my_round((float(entr_E3_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_E_VI.append(my_round((float(entr_E4_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_E_VI.append(my_round((float(entr_E5_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_E_VI.append(my_round((float(entr_E6_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_E_VI.append(my_round((float(entr_E7_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_E_VI.append(my_round((float(entr_E8_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_E_VI.append(my_round((float(entr_E9_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_E_VI.append(my_round((float(entr_E10_VI.get())), check_razryad.get()))
			
			mean_E_VI = my_round(stat.mean(spisok_E_VI), check_razryad.get())
			
			Er_E_VI = ((mean_E_VI - QC_VI)/QC_VI)*100

			sigma_E_VI = (stat.stdev(spisok_E_VI)/mean_E_VI)*100
	#проверка условия по количеству групп если больше либо равно 6 этот код выполняется

		##########################
		if check_var_gr_2.get() >= 6:
		

		##########################
			spisok_F_VI = []
			spisok_F_VI.append(my_round((float(entr_F1_VI.get())), check_razryad.get()))
			spisok_F_VI.append(my_round((float(entr_F2_VI.get())), check_razryad.get()))
			spisok_F_VI.append(my_round((float(entr_F3_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 4:
				spisok_F_VI.append(my_round((float(entr_F4_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 5:
				spisok_F_VI.append(my_round((float(entr_F5_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 6:
				spisok_F_VI.append(my_round((float(entr_F6_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 7:
				spisok_F_VI.append(my_round((float(entr_F7_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 8:
				spisok_F_VI.append(my_round((float(entr_F8_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 9:
				spisok_F_VI.append(my_round((float(entr_F9_VI.get())), check_razryad.get()))
			if check_var_gr_3.get() >= 10:
				spisok_F_VI.append(my_round((float(entr_F10_VI.get())), check_razryad.get()))


			#calc
			#среднее по
			mean_F_VI = my_round(stat.mean(spisok_F_VI), check_razryad.get())
			#CKO_F_VI = (stat.stdev(spisok_F_VI)/mean_F_VI)*100
			
			#vntr_gr_smech_F_VI = (mean_F_VI - QC_VI)/QC_VI*100
			
			Er_F_VI = ((mean_F_VI - QC_VI)/QC_VI)*100

			sigma_F_VI = (stat.stdev(spisok_F_VI)/mean_F_VI)*100



	if check_var_gr_1.get() >= 1:
		stat_mean_spisok_QCA = []
		for i in range(0, len(spisok_A_I)):
			stat_mean_spisok_QCA.append(spisok_A_I[i])
			if check_var_gr_2.get() >= 2:
				stat_mean_spisok_QCA.append(spisok_B_I[i])
			if check_var_gr_2.get() >= 3:
				stat_mean_spisok_QCA.append(spisok_C_I[i])
			if check_var_gr_2.get() >= 4:
				stat_mean_spisok_QCA.append(spisok_D_I[i])
			if check_var_gr_2.get() >= 5:	
				stat_mean_spisok_QCA.append(spisok_E_I[i])
			if check_var_gr_2.get() >= 6:	
				stat_mean_spisok_QCA.append(spisok_F_I[i])

		mean_QCA = my_round(stat.mean(stat_mean_spisok_QCA), check_razryad.get())
		#среднее смещение
		mean_smech_QCA = (mean_QCA - QC_I)/QC_I*100

#					QCB
	if check_var_gr_1.get() >= 2:
		stat_mean_spisok_QCB = []
		for i in range(0, len(spisok_A_II)):
			stat_mean_spisok_QCB.append(spisok_A_II[i])
			if check_var_gr_2.get() >= 2:
				stat_mean_spisok_QCB.append(spisok_B_II[i])
			if check_var_gr_2.get() >= 3:
				stat_mean_spisok_QCB.append(spisok_C_II[i])
			if check_var_gr_2.get() >= 4:
				stat_mean_spisok_QCB.append(spisok_D_II[i])
			if check_var_gr_2.get() >= 5:	
				stat_mean_spisok_QCB.append(spisok_E_II[i])
			if check_var_gr_2.get() >= 6:	
				stat_mean_spisok_QCB.append(spisok_F_II[i])

		mean_QCB = my_round(stat.mean(stat_mean_spisok_QCB), check_razryad.get())
		#среднее смещение
		mean_smech_QCB = (mean_QCB - QC_II)/QC_II*100

#					QCC
	if check_var_gr_1.get() >= 3:
		stat_mean_spisok_QCC = []
		for i in range(0, len(spisok_A_III)):
			stat_mean_spisok_QCC.append(spisok_A_III[i])
			if check_var_gr_2.get() >= 2:
				stat_mean_spisok_QCC.append(spisok_B_III[i])
			if check_var_gr_2.get() >= 3:
				stat_mean_spisok_QCC.append(spisok_C_III[i])
			if check_var_gr_2.get() >= 4:
				stat_mean_spisok_QCC.append(spisok_D_III[i])
			if check_var_gr_2.get() >= 5:	
				stat_mean_spisok_QCC.append(spisok_E_III[i])
			if check_var_gr_2.get() >= 6:	
				stat_mean_spisok_QCC.append(spisok_F_III[i])

		mean_QCC = stat.mean(stat_mean_spisok_QCC)
		#среднее смещение
		mean_smech_QCC = (mean_QCC - QC_III)/QC_III*100

#					QCD
	if check_var_gr_1.get() >= 4:
		stat_mean_spisok_QCD = []
		for i in range(0, len(spisok_A_IV)):
			stat_mean_spisok_QCD.append(spisok_A_IV[i])
			if check_var_gr_2.get() >= 2:
				stat_mean_spisok_QCD.append(spisok_B_IV[i])
			if check_var_gr_2.get() >= 3:
				stat_mean_spisok_QCD.append(spisok_C_IV[i])
			if check_var_gr_2.get() >= 4:
				stat_mean_spisok_QCD.append(spisok_D_IV[i])
			if check_var_gr_2.get() >= 5:	
				stat_mean_spisok_QCD.append(spisok_E_IV[i])
			if check_var_gr_2.get() >= 6:	
				stat_mean_spisok_QCD.append(spisok_F_IV[i])

		mean_QCD = my_round(stat.mean(stat_mean_spisok_QCD), check_razryad.get())
		#среднее смещение
		mean_smech_QCD = (mean_QCD - QC_IV)/QC_IV*100

#					QCE
	if check_var_gr_1.get() >= 5:
		stat_mean_spisok_QCE = []
		for i in range(0, len(spisok_A_V)):
			stat_mean_spisok_QCE.append(spisok_A_V[i])
			if check_var_gr_2.get() >= 2:
				stat_mean_spisok_QCE.append(spisok_B_V[i])
			if check_var_gr_2.get() >= 3:
				stat_mean_spisok_QCE.append(spisok_C_V[i])
			if check_var_gr_2.get() >= 4:
				stat_mean_spisok_QCE.append(spisok_D_V[i])
			if check_var_gr_2.get() >= 5:	
				stat_mean_spisok_QCE.append(spisok_E_V[i])
			if check_var_gr_2.get() >= 6:	
				stat_mean_spisok_QCE.append(spisok_F_V[i])

		mean_QCE = my_round(stat.mean(stat_mean_spisok_QCE), check_razryad.get())
		#среднее смещение
		mean_smech_QCE = (mean_QCE - QC_V)/QC_V*100

#					QCF
	if check_var_gr_1.get() >= 6:
		stat_mean_spisok_QCF = []
		for i in range(0, len(spisok_A_VI)):
			stat_mean_spisok_QCE.append(spisok_A_VI[i])
			if check_var_gr_2.get() >= 2:
				stat_mean_spisok_QCF.append(spisok_B_VI[i])
			if check_var_gr_2.get() >= 3:
				stat_mean_spisok_QCF.append(spisok_C_VI[i])
			if check_var_gr_2.get() >= 4:
				stat_mean_spisok_QCF.append(spisok_D_VI[i])
			if check_var_gr_2.get() >= 5:	
				stat_mean_spisok_QCF.append(spisok_E_VI[i])
			if check_var_gr_2.get() >= 6:	
				stat_mean_spisok_QCF.append(spisok_F_VI[i])

		mean_QCF = my_round(stat.mean(stat_mean_spisok_QCF), check_razryad.get())
		#среднее смещение
		mean_smech_QCF = (mean_QCF - QC_VI)/QC_VI*100



#Надписи на весь документ - независимо от таблиц
#Первая колонка

	ws['A1'].value = 'Concentration, {}'.format(entr_concent.get())
	
	#ws.merge_cells('A1:C1')
	if check_var_gr_1.get() >= 2:
		ws['J1'].value = 'Concentration, {}'.format(entr_concent.get())
		#ws.merge_cells('J1:L1')
	if check_var_gr_1.get() >= 3:
		ws['S1'].value = 'Concentration, {}'.format(entr_concent.get())
		#ws.merge_cells('S1:U1')
	if check_var_gr_1.get() >= 4:
		ws['AB1'].value = 'Concentration, {}'.format(entr_concent.get())
		#ws.merge_cells('AB1:AD1')
	if check_var_gr_1.get() >= 5:
		ws['AK1'].value = 'Concentration, {}'.format(entr_concent.get())
		#ws.merge_cells('AK1:AM1')
	if check_var_gr_1.get() >= 6:
		ws['AT1'].value = 'Concentration, {}'.format(entr_concent.get())
		#ws.merge_cells('AT1:AV1')

	
	#ЦИФРЫ QCA
	ws['A2'].value = QC_I
	ws['B3'].value = '1'
	ws['B4'].value = '2'
	ws['B5'].value = '3'
	if check_var_gr_3.get() >= 4:
		ws['B6'].value = '4'
	if check_var_gr_3.get() >= 5:
		ws['B7'].value = '5'
	if check_var_gr_3.get() >= 6:
		ws['B8'].value = '6'
	if check_var_gr_3.get() >= 7:
		ws['B9'].value = '7'
	if check_var_gr_3.get() >= 8:
		ws['B10'].value = '8'
	if check_var_gr_3.get() >= 9:
		ws['B11'].value = '9'
	if check_var_gr_3.get() >= 10:
		ws['B12'].value = '10'

	#ЦИФРЫ QCB
	if check_var_gr_1.get() >= 2:
		ws['J2'].value = QC_II
		ws['K3'].value = '1'
		ws['K4'].value = '2'
		ws['K5'].value = '3'
		if check_var_gr_3.get() >= 4:
			ws['K6'].value = '4'
		if check_var_gr_3.get() >= 5:
			ws['K7'].value = '5'
		if check_var_gr_3.get() >= 6:
			ws['K8'].value = '6'
		if check_var_gr_3.get() >= 7:
			ws['K9'].value = '7'
		if check_var_gr_3.get() >= 8:
			ws['K10'].value = '8'
		if check_var_gr_3.get() >= 9:
			ws['K11'].value = '9'
		if check_var_gr_3.get() >= 10:
			ws['K12'].value = '10'
	
	#ЦИФРЫ QCC
	if check_var_gr_1.get() >= 3:
		ws['S2'].value = QC_III
		ws['T3'].value = '1'
		ws['T4'].value = '2'
		ws['T5'].value = '3'
		if check_var_gr_3.get() >= 4:
			ws['T6'].value = '4'
		if check_var_gr_3.get() >= 5:
			ws['T7'].value = '5'
		if check_var_gr_3.get() >= 6:
			ws['T8'].value = '6'
		if check_var_gr_3.get() >= 7:
			ws['T9'].value = '7'
		if check_var_gr_3.get() >= 8:
			ws['T10'].value = '8'
		if check_var_gr_3.get() >= 9:
			ws['T11'].value = '9'
		if check_var_gr_3.get() >= 10:
			ws['T12'].value = '10'

	#ЦИФРЫ QCD
	if check_var_gr_1.get() >= 4:
		ws['AB2'].value = QC_IV
		ws['AC3'].value = '1'
		ws['AC4'].value = '2'
		ws['AC5'].value = '3'
		if check_var_gr_3.get() >= 4:
			ws['AC6'].value = '4'
		if check_var_gr_3.get() >= 5:
			ws['AC7'].value = '5'
		if check_var_gr_3.get() >= 6:
			ws['AC8'].value = '6'
		if check_var_gr_3.get() >= 7:
			ws['AC9'].value = '7'
		if check_var_gr_3.get() >= 8:
			ws['AC10'].value = '8'
		if check_var_gr_3.get() >= 9:
			ws['AC11'].value = '9'
		if check_var_gr_3.get() >= 10:
			ws['AC12'].value = '10'

	#ЦИФРЫ QCE
	if check_var_gr_1.get() >= 5:
		ws['AK2'].value = QC_V
		ws['AL3'].value = '1'
		ws['AL4'].value = '2'
		ws['AL5'].value = '3'
		if check_var_gr_3.get() >= 4:
			ws['AL6'].value = '4'
		if check_var_gr_3.get() >= 5:
			ws['AL7'].value = '5'
		if check_var_gr_3.get() >= 6:
			ws['AL8'].value = '6'
		if check_var_gr_3.get() >= 7:
			ws['AL9'].value = '7'
		if check_var_gr_3.get() >= 8:
			ws['AL10'].value = '8'
		if check_var_gr_3.get() >= 9:
			ws['AL11'].value = '9'
		if check_var_gr_3.get() >= 10:
			ws['AL12'].value = '10'


	#ЦИФРЫ QCF
	if check_var_gr_1.get() >= 6:
		ws['AT2'].value = QC_VI
		ws['AU3'].value = '1'
		ws['AU4'].value = '2'
		ws['AU5'].value = '3'
		if check_var_gr_3.get() >= 4:
			ws['AU6'].value = '4'
		if check_var_gr_3.get() >= 5:
			ws['AU7'].value = '5'
		if check_var_gr_3.get() >= 6:
			ws['AU8'].value = '6'
		if check_var_gr_3.get() >= 7:
			ws['AU9'].value = '7'
		if check_var_gr_3.get() >= 8:
			ws['AU10'].value = '8'
		if check_var_gr_3.get() >= 9:
			ws['AU11'].value = '9'
		if check_var_gr_3.get() >= 10:
			ws['AU12'].value = '10'




#QC ТАБЛИЦА № 1
	#ОФОРМЛЕНИЕ EXCEL НА ВСЮ ТАБЛИЦУ

	#выравнивание по центру]
	cols_c(ws, 'M15:S32')

	#границы ячеек
	if check_var_gr_1.get() == 1:
		if check_var_gr_3.get() == 3:
			thin_border(ws, 'M15:N24')
		if check_var_gr_3.get() == 4:
			thin_border(ws, 'M15:N25')
		if check_var_gr_3.get() == 5:
			thin_border(ws, 'M15:N26')	
		if check_var_gr_3.get() == 6:
			thin_border(ws, 'M15:N27')
		if check_var_gr_3.get() == 7:
			thin_border(ws, 'M15:N28')	
		if check_var_gr_3.get() == 8:
			thin_border(ws, 'M15:N29')
		if check_var_gr_3.get() == 9:
			thin_border(ws, 'M15:N30')
		if check_var_gr_3.get() == 10:
			thin_border(ws, 'M15:N31')

	if check_var_gr_1.get() == 2:
		if check_var_gr_3.get() == 3:
			thin_border(ws, 'M15:O24')
		if check_var_gr_3.get() == 4:
			thin_border(ws, 'M15:O25')
		if check_var_gr_3.get() == 5:
			thin_border(ws, 'M15:O26')	
		if check_var_gr_3.get() == 6:
			thin_border(ws, 'M15:O27')
		if check_var_gr_3.get() == 7:
			thin_border(ws, 'M15:O28')	
		if check_var_gr_3.get() == 8:
			thin_border(ws, 'M15:O29')
		if check_var_gr_3.get() == 9:
			thin_border(ws, 'M15:O30')
		if check_var_gr_3.get() == 10:
			thin_border(ws, 'M15:O31')

	if check_var_gr_1.get() == 3:
		if check_var_gr_3.get() == 3:
			thin_border(ws, 'M15:P24')
		if check_var_gr_3.get() == 4:
			thin_border(ws, 'M15:P25')
		if check_var_gr_3.get() == 5:
			thin_border(ws, 'M15:P26')	
		if check_var_gr_3.get() == 6:
			thin_border(ws, 'M15:P27')
		if check_var_gr_3.get() == 7:
			thin_border(ws, 'M15:P28')	
		if check_var_gr_3.get() == 8:
			thin_border(ws, 'M15:P29')
		if check_var_gr_3.get() == 9:
			thin_border(ws, 'M15:P30')
		if check_var_gr_3.get() == 10:
			thin_border(ws, 'M15:P31')				

	if check_var_gr_1.get() == 4:
		if check_var_gr_3.get() == 3:
			thin_border(ws, 'M15:Q24')
		if check_var_gr_3.get() == 4:
			thin_border(ws, 'M15:Q25')
		if check_var_gr_3.get() == 5:
			thin_border(ws, 'M15:Q26')	
		if check_var_gr_3.get() == 6:
			thin_border(ws, 'M15:Q27')
		if check_var_gr_3.get() == 7:
			thin_border(ws, 'M15:Q28')	
		if check_var_gr_3.get() == 8:
			thin_border(ws, 'M15:Q29')
		if check_var_gr_3.get() == 9:
			thin_border(ws, 'M15:Q30')
		if check_var_gr_3.get() == 10:
			thin_border(ws, 'M15:Q31')			

	if check_var_gr_1.get() == 5:
		if check_var_gr_3.get() == 3:
			thin_border(ws, 'M15:R24')
		if check_var_gr_3.get() == 4:
			thin_border(ws, 'M15:R25')
		if check_var_gr_3.get() == 5:
			thin_border(ws, 'M15:R26')	
		if check_var_gr_3.get() == 6:
			thin_border(ws, 'M15:R27')
		if check_var_gr_3.get() == 7:
			thin_border(ws, 'M15:R28')	
		if check_var_gr_3.get() == 8:
			thin_border(ws, 'M15:R29')
		if check_var_gr_3.get() == 9:
			thin_border(ws, 'M15:R30')
		if check_var_gr_3.get() == 10:
			thin_border(ws, 'M15:R31')		

	if check_var_gr_1.get() == 6:
		if check_var_gr_3.get() == 3:
			thin_border(ws, 'M15:S24')
		if check_var_gr_3.get() == 4:
			thin_border(ws, 'M15:S25')
		if check_var_gr_3.get() == 5:
			thin_border(ws, 'M15:S26')	
		if check_var_gr_3.get() == 6:
			thin_border(ws, 'M15:S27')
		if check_var_gr_3.get() == 7:
			thin_border(ws, 'M15:S28')	
		if check_var_gr_3.get() == 8:
			thin_border(ws, 'M15:S29')
		if check_var_gr_3.get() == 9:
			thin_border(ws, 'M15:S30')
		if check_var_gr_3.get() == 10:
			thin_border(ws, 'M15:S31')	

	

	#надпись и объединение ячеек под надпись
	ws['M15'].value = 'QC'
	ws['M16'].value = f'Введено, {entr_concent.get()}'
	ws['M17'].value = '№ измерения'
	ws['N17'].value = 'Найдено, последовательность 1'


	#QCA1
	ws['N15'].value = 'QCA1'
	ws['N16'].value = QC_I
	ws['C2'].value = 'QCA1'
	ws['M18'].value = '1'
	ws['N18'].value = my_round(float(entr_A1_I.get()), razryad)
	ws['C3'].value = my_round(float(entr_A1_I.get()), razryad)
	ws['M19'].value = '2'
	ws['N19'].value = my_round(float(entr_A2_I.get()), razryad)
	ws['C4'].value = my_round(float(entr_A2_I.get()), razryad)
	ws['M20'].value = '3'
	ws['N20'].value = my_round(float(entr_A3_I.get()), razryad)
	ws['C5'].value = my_round(float(entr_A3_I.get()), razryad)
	if check_var_gr_3.get() >= 4:
		ws['M21'].value = '4'
		ws['N21'].value = my_round(float(entr_A4_I.get()), razryad)
		ws['C6'].value = my_round(float(entr_A4_I.get()), razryad)
	if check_var_gr_3.get() >= 5:
		ws['M22'].value = '5'
		ws['N22'].value = my_round(float(entr_A5_I.get()), razryad)
		ws['C7'].value = my_round(float(entr_A5_I.get()), razryad)
	if check_var_gr_3.get() >= 6:	
		ws['M23'].value = '6'
		ws['N23'].value = my_round(float(entr_A6_I.get()), razryad)
		ws['C8'].value = my_round(float(entr_A6_I.get()), razryad)
	if check_var_gr_3.get() >= 7:	
		ws['M24'].value = '7'
		ws['N24'].value = my_round(float(entr_A7_I.get()), razryad)
		ws['C9'].value = my_round(float(entr_A7_I.get()), razryad)
	if check_var_gr_3.get() >= 8:	
		ws['M25'].value = '8'
		ws['N25'].value = my_round(float(entr_A8_I.get()), razryad)
		ws['C10'].value = my_round(float(entr_A8_I.get()), razryad)
	if check_var_gr_3.get() >= 9:	
		ws['M26'].value = '9'
		ws['N26'].value = my_round(float(entr_A9_I.get()), razryad)
		ws['C11'].value = my_round(float(entr_A9_I.get()), razryad)
	if check_var_gr_3.get() >= 10:	
		ws['M27'].value = '10'
		ws['N27'].value = my_round(float(entr_A10_I.get()), razryad)
		ws['C12'].value = my_round(float(entr_A10_I.get()), razryad)

	#вывод среднего и остальных показателей в таблицу 
	#НАДПИСЬ ЗНАЧЕНИЙ КОЛОНКИ G БОЛЕЕ В ЭТОЙ ТАБЛИЦЕ НЕ ДУБЛИРОВАТЬ!!!!!!!!!!!
	if check_var_gr_3.get() == 3:
		ws['M21'].value = 'найдено, сред.зн.,\nнг/мл (n=3)'
		ws['M21'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

		ws['M22'].value = 'εR, % (n=3)'
		ws['M23'].value = 'σR, % (n=3)'
		ws['M24'].value = 'Норма |ε| и |σ|, %'

		ws['N21'].value = my_round(mean_A_I, razryad)
		ws['N22'].value = my_round(Er_A_I, 1)
		ws['N23'].value = my_round(sigma_A_I, 1)
		if entr_NORM_QCA1.get() != '':
			ws['N24'].value = '≤{}'.format(entr_NORM_QCA1.get())
		else:
			ws['N24'].value = '≤20'	
		
		if check_var_gr_1.get() >= 2:
			ws['O21'].value = my_round(mean_A_II, razryad)
			ws['O22'].value = my_round(Er_A_II, 1)
			ws['O23'].value = my_round(sigma_A_II, 1)
			if entr_NORM_QCB1.get() != '':
				ws['O24'].value = '≤{}'.format(entr_NORM_QCB1.get())
			else:
				ws['O24'].value = '≤15'

		if check_var_gr_1.get() >= 3:	
			ws['P21'].value = my_round(mean_A_III, razryad)
			ws['P22'].value = my_round(Er_A_III, 1)
			ws['P23'].value = my_round(sigma_A_III, 1)
			if entr_NORM_QCC1.get() != '':
				ws['P24'].value = '≤{}'.format(entr_NORM_QCC1.get())
			else:
				ws['P24'].value = '≤15'

		if check_var_gr_1.get() >= 4:	
			ws['Q21'].value = my_round(mean_A_IV, razryad)
			ws['Q22'].value = my_round(Er_A_IV, 1)
			ws['Q23'].value = my_round(sigma_A_IV, 1)
			if entr_NORM_QCD1.get() != '':
				ws['Q24'].value = '≤{}'.format(entr_NORM_QCD1.get())
			else:
				ws['Q24'].value = '≤15'

		if check_var_gr_1.get() >= 5:	
			ws['R21'].value = my_round(mean_A_V, razryad)
			ws['R22'].value = my_round(Er_A_V, 1)
			ws['R23'].value = my_round(sigma_A_V, 1)
			if entr_NORM_QCE1.get() != '':
				ws['R24'].value = '≤{}'.format(entr_NORM_QCE1.get())
			else:
				ws['R24'].value = '≤15'

		if check_var_gr_1.get() >= 6:	
			ws['S21'].value = my_round(mean_A_VI, razryad)
			ws['S22'].value = my_round(Er_A_VI, 1)
			ws['S23'].value = my_round(sigma_A_VI, 1)
			if entr_NORM_QCF1.get() != '':
				ws['S24'].value = '≤{}'.format(entr_NORM_QCF1.get())
			else:
				ws['S24'].value = '≤15'


	if check_var_gr_3.get() == 4:

		ws['M22'].value = 'найдено, сред.зн.,\nнг/мл (n=4)'
		ws['M22'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
		ws['M23'].value = 'εR, % (n=4)'
		ws['M24'].value = 'σR, % (n=4)'
		ws['M25'].value = 'Норма |ε| и |σ|, %'

		ws['N22'].value = my_round(mean_A_I, razryad)
		ws['N23'].value = my_round(Er_A_I, 1)
		ws['N24'].value = my_round(sigma_A_I	, 1)
		if entr_NORM_QCA1.get() != '':
			ws['N25'].value = '≤{}'.format(entr_NORM_QCA1.get())
		else:
			ws['N25'].value = '≤20'		

		if check_var_gr_1.get() >= 2:
			ws['O22'].value = my_round(mean_A_II, razryad)
			ws['O23'].value = my_round(Er_A_II, 1)
			ws['O24'].value = my_round(sigma_A_II, 1)
			if entr_NORM_QCB1.get() != '':
				ws['O25'].value = '≤{}'.format(entr_NORM_QCB1.get())
			else:
				ws['O25'].value = '≤15'

		if check_var_gr_1.get() >= 3:	
			ws['P22'].value = my_round(mean_A_III, razryad)
			ws['P23'].value = my_round(Er_A_III, 1)
			ws['P24'].value = my_round(sigma_A_III, 1)
			if entr_NORM_QCC1.get() != '':
				ws['P25'].value = '≤{}'.format(entr_NORM_QCC1.get())
			else:
				ws['P25'].value = '≤15'

		if check_var_gr_1.get() >= 4:	
			ws['Q22'].value = my_round(mean_A_IV, razryad)
			ws['Q23'].value = my_round(Er_A_IV, 1)
			ws['Q24'].value = my_round(sigma_A_IV, 1)
			if entr_NORM_QCD1.get() != '':
				ws['Q25'].value = '≤{}'.format(entr_NORM_QCD1.get())
			else:
				ws['Q25'].value = '≤15'

		if check_var_gr_1.get() >= 5:	
			ws['R22'].value = my_round(mean_A_V, razryad)
			ws['R23'].value = my_round(Er_A_V, 1)
			ws['R24'].value = my_round(sigma_A_V, 1)
			if entr_NORM_QCE1.get() != '':
				ws['R25'].value = '≤{}'.format(entr_NORM_QCE1.get())
			else:
				ws['R25'].value = '≤15'

		if check_var_gr_1.get() >= 6:	
			ws['S22'].value = my_round(mean_A_VI, razryad)
			ws['S23'].value = my_round(Er_A_VI, 1)
			ws['S24'].value = my_round(sigma_A_VI, 1)
			if entr_NORM_QCF1.get() != '':
				ws['S25'].value = '≤{}'.format(entr_NORM_QCF1.get())
			else:
				ws['S25'].value = '≤15'

	if check_var_gr_3.get() == 5:
		ws['M23'].value = 'найдено, сред.зн., \nнг/мл (n=5)'
		ws['M23'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
		ws['M24'].value = 'εR, % (n=5)'
		ws['M25'].value = 'σR, % (n=5)'
		ws['M26'].value = 'Норма |ε| и |σ|, %'

		ws['N23'].value = my_round(mean_A_I, razryad)
		ws['N24'].value = my_round(Er_A_I, 1)
		ws['N25'].value = my_round(sigma_A_I, 1)
		if entr_NORM_QCA1.get() != '':
			ws['N26'].value = '≤{}'.format(entr_NORM_QCA1.get())
		else:
			ws['N26'].value = '≤20'	

		if check_var_gr_1.get() >= 2:
			ws['O23'].value = my_round(mean_A_II, razryad)
			ws['O24'].value = my_round(Er_A_II, 1)
			ws['O25'].value = my_round(sigma_A_II, 1)
			if entr_NORM_QCB1.get() != '':
				ws['O26'].value = '≤{}'.format(entr_NORM_QCB1.get())
			else:
				ws['O26'].value = '≤15'	

		if check_var_gr_1.get() >= 3:	
			ws['P23'].value = my_round(mean_A_III, razryad)
			ws['P24'].value = my_round(Er_A_III, 1)
			ws['P25'].value = my_round(sigma_A_III, 1)
			if entr_NORM_QCC1.get() != '':
				ws['P26'].value = '≤{}'.format(entr_NORM_QCC1.get())
			else:
				ws['P26'].value = '≤15'	

		if check_var_gr_1.get() >= 4:	
			ws['Q23'].value = my_round(mean_A_IV, razryad)
			ws['Q24'].value = my_round(Er_A_IV, 1)
			ws['Q25'].value = my_round(sigma_A_IV, 1)
			if entr_NORM_QCD1.get() != '':
				ws['Q26'].value = '≤{}'.format(entr_NORM_QCD1.get())
			else:
				ws['Q26'].value = '≤15'

		if check_var_gr_1.get() >= 5:	
			ws['R23'].value = my_round(mean_A_V, razryad)
			ws['R24'].value = my_round(Er_A_V, 1)
			ws['R25'].value = my_round(sigma_A_V, 1)
			if entr_NORM_QCE1.get() != '':
				ws['R26'].value = '≤{}'.format(entr_NORM_QCE1.get())
			else:
				ws['R26'].value = '≤15'

		if check_var_gr_1.get() >= 6:	
			ws['S23'].value = my_round(mean_A_VI, razryad)
			ws['S24'].value = my_round(Er_A_VI, 1)
			ws['S25'].value = my_round(sigma_A_VI, 1)
			if entr_NORM_QCF1.get() != '':
				ws['S26'].value = '≤{}'.format(entr_NORM_QCF1.get())
			else:
				ws['S26'].value = '≤15'


	if check_var_gr_3.get() == 6:
		ws['M24'].value = 'найдено, сред.зн., \nнг/мл (n=6)'
		ws['M24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
		ws['M25'].value = 'εR, % (n=6)'
		ws['M26'].value = 'σR, % (n=6)'
		ws['M27'].value = 'Норма |ε| и |σ|, %'

		ws['N24'].value = my_round(mean_A_I, razryad)
		ws['N25'].value = my_round(Er_A_I, 1)
		ws['N26'].value = my_round(sigma_A_I, 1)
		if entr_NORM_QCA1.get() != '':
			ws['N27'].value = '≤{}'.format(entr_NORM_QCA1.get())
		else:
			ws['N27'].value = '≤20'		

		if check_var_gr_1.get() >= 2:
			ws['O24'].value = my_round(mean_A_II, razryad)
			ws['O25'].value = my_round(Er_A_II, 1)
			ws['O26'].value = my_round(sigma_A_II, 1)
			if entr_NORM_QCB1.get() != '':
				ws['O27'].value = '≤{}'.format(entr_NORM_QCB1.get())
			else:
				ws['O27'].value = '≤15'

		if check_var_gr_1.get() >= 3:	
			ws['P24'].value = my_round(mean_A_III, razryad)
			ws['P25'].value = my_round(Er_A_III, 1)
			ws['P26'].value = my_round(sigma_A_III, 1)
			if entr_NORM_QCC1.get() != '':
				ws['P27'].value = '≤{}'.format(entr_NORM_QCC1.get())
			else:
				ws['P27'].value = '≤15'

		if check_var_gr_1.get() >= 4:	
			ws['Q24'].value = my_round(mean_A_IV, razryad)
			ws['Q25'].value = my_round(Er_A_IV, 1)
			ws['Q26'].value = my_round(sigma_A_IV, 1)
			if entr_NORM_QCD1.get() != '':
				ws['Q27'].value = '≤{}'.format(entr_NORM_QCD1.get())
			else:
				ws['Q27'].value = '≤15'

		if check_var_gr_1.get() >= 5:	
			ws['R24'].value = my_round(mean_A_V, razryad)
			ws['R25'].value = my_round(Er_A_V, 1)
			ws['R26'].value = my_round(sigma_A_V, 1)
			if entr_NORM_QCE1.get() != '':
				ws['R27'].value = '≤{}'.format(entr_NORM_QCE1.get())
			else:
				ws['R27'].value = '≤15'

		if check_var_gr_1.get() >= 6:	
			ws['S24'].value = my_round(mean_A_VI, razryad)
			ws['S25'].value = my_round(Er_A_VI, 1)
			ws['S26'].value = my_round(sigma_A_VI, 1)
			if entr_NORM_QCF1.get() != '':
				ws['S27'].value = '≤{}'.format(entr_NORM_QCF1.get())
			else:
				ws['S27'].value = '≤15'

	if check_var_gr_3.get() == 7:
		ws['M25'].value = 'найдено, сред.зн., \nнг/мл (n=7)'
		ws['M25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
		ws['M26'].value = 'εR, % (n=7)'
		ws['M27'].value = 'σR, % (n=7)'
		ws['M28'].value = 'Норма |ε| и |σ|, %'

		ws['N25'].value = my_round(mean_A_I, razryad)
		ws['N26'].value = my_round(Er_A_I, 1)
		ws['N27'].value = my_round(sigma_A_I, 1)
		if entr_NORM_QCA1.get() != '':
			ws['N28'].value = '≤{}'.format(entr_NORM_QCA1.get())
		else:
			ws['N28'].value = '≤20'	

		if check_var_gr_1.get() >= 2:
			ws['O25'].value = my_round(mean_A_II, razryad)
			ws['O26'].value = my_round(Er_A_II, 1)
			ws['O27'].value = my_round(sigma_A_II, 1)
			if entr_NORM_QCB1.get() != '':
				ws['O28'].value = '≤{}'.format(entr_NORM_QCB1.get())
			else:
				ws['O28'].value = '≤15'	

		if check_var_gr_1.get() >= 3:	
			ws['P25'].value = my_round(mean_A_III, razryad)
			ws['P26'].value = my_round(Er_A_III, 1)
			ws['P27'].value = my_round(sigma_A_III, 1)
			if entr_NORM_QCC1.get() != '':
				ws['P28'].value = '≤{}'.format(entr_NORM_QCC1.get())
			else:
				ws['P28'].value = '≤15'	

		if check_var_gr_1.get() >= 4:	
			ws['Q25'].value = my_round(mean_A_IV, razryad)
			ws['Q26'].value = my_round(Er_A_IV, 1)
			ws['Q27'].value = my_round(sigma_A_IV, 1)
			if entr_NORM_QCD1.get() != '':
				ws['Q28'].value = '≤{}'.format(entr_NORM_QCD1.get())
			else:
				ws['Q28'].value = '≤15'	

		if check_var_gr_1.get() >= 5:	
			ws['R25'].value = my_round(mean_A_V, razryad)
			ws['R26'].value = my_round(Er_A_V, 1)
			ws['R27'].value = my_round(sigma_A_V, 1)
			if entr_NORM_QCE1.get() != '':
				ws['R28'].value = '≤{}'.format(entr_NORM_QCE1.get())
			else:
				ws['R28'].value = '≤15'	

		if check_var_gr_1.get() >= 6:	
			ws['S25'].value = my_round(mean_A_VI, razryad)
			ws['S26'].value = my_round(Er_A_VI, 1)
			ws['S27'].value = my_round(sigma_A_VI, 1)
			if entr_NORM_QCE1.get() != '':
				ws['S28'].value = '≤{}'.format(entr_NORM_QCF1.get())
			else:
				ws['S28'].value = '≤15'	

	if check_var_gr_3.get() == 8:
		ws['M26'].value = 'найдено, сред.зн., \nнг/мл (n=8)'
		ws['M26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
		ws['M27'].value = 'εR, % (n=8)'
		ws['M28'].value = 'σR, % (n=8)'
		ws['M29'].value = 'Норма |ε| и |σ|, %'

		ws['N26'].value = my_round(mean_A_I, razryad)
		ws['N27'].value = my_round(Er_A_I, 1)
		ws['N28'].value = my_round(sigma_A_I, 1)
		if entr_NORM_QCA1.get() != '':
			ws['N29'].value = '≤{}'.format(entr_NORM_QCA1.get())
		else:
			ws['N29'].value = '≤20'	

		if check_var_gr_1.get() >= 2:
			ws['O26'].value = my_round(mean_A_II, razryad)
			ws['O27'].value = my_round(Er_A_II, 1)
			ws['O28'].value = my_round(sigma_A_II, 1)
			if entr_NORM_QCB1.get() != '':
				ws['O29'].value = '≤{}'.format(entr_NORM_QCB1.get())
			else:
				ws['O29'].value = '≤15'	

		if check_var_gr_1.get() >= 3:	
			ws['P26'].value = my_round(mean_A_III, razryad)
			ws['P27'].value = my_round(Er_A_III, 1)
			ws['P28'].value = my_round(sigma_A_III, 1)
			if entr_NORM_QCC1.get() != '':
				ws['P29'].value = '≤{}'.format(entr_NORM_QCC1.get())
			else:
				ws['P29'].value = '≤15'	

		if check_var_gr_1.get() >= 4:	
			ws['Q26'].value = my_round(mean_A_IV, razryad)
			ws['Q27'].value = my_round(Er_A_IV, 1)
			ws['Q28'].value = my_round(sigma_A_IV, 1)
			if entr_NORM_QCD1.get() != '':
				ws['Q29'].value = '≤{}'.format(entr_NORM_QCD1.get())
			else:
				ws['Q29'].value = '≤15'

		if check_var_gr_1.get() >= 5:	
			ws['R26'].value = my_round(mean_A_V, razryad)
			ws['R27'].value = my_round(Er_A_V, 1)
			ws['R28'].value = my_round(sigma_A_V, 1)
			if entr_NORM_QCE1.get() != '':
				ws['R29'].value = '≤{}'.format(entr_NORM_QCE1.get())
			else:
				ws['R29'].value = '≤15'

		if check_var_gr_1.get() >= 6:	
			ws['S26'].value = my_round(mean_A_VI, razryad)
			ws['S27'].value = my_round(Er_A_VI, 1)
			ws['S28'].value = my_round(sigma_A_VI, 1)
			if entr_NORM_QCF1.get() != '':
				ws['S29'].value = '≤{}'.format(entr_NORM_QCF1.get())
			else:
				ws['S29'].value = '≤15'

	if check_var_gr_3.get() == 9:
		ws['M27'].value = 'найдено, сред.зн., \nнг/мл (n=9)'
		ws['M27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
		ws['M28'].value = 'εR, % (n=9)'
		ws['M29'].value = 'σR, % (n=9)'
		ws['M30'].value = 'Норма |ε| и |σ|, %'

		ws['N27'].value = my_round(mean_A_I, razryad)
		ws['N28'].value = my_round(Er_A_I, 1)
		ws['N29'].value = my_round(sigma_A_I, 1)
		if entr_NORM_QCA1.get() != '':
			ws['N30'].value = '≤{}'.format(entr_NORM_QCA1.get())
		else:
			ws['N30'].value = '≤20'

		if check_var_gr_1.get() >= 2:
			ws['O27'].value = my_round(mean_A_II, razryad)
			ws['O28'].value = my_round(Er_A_II, 1)
			ws['O29'].value = my_round(sigma_A_II, 1)
			if entr_NORM_QCB1.get() != '':
				ws['O30'].value = '≤{}'.format(entr_NORM_QCB1.get())
			else:
				ws['O30'].value = '≤15'

		if check_var_gr_1.get() >= 3:	
			ws['P27'].value = my_round(mean_A_III, razryad)
			ws['P28'].value = my_round(Er_A_III, 1)
			ws['P29'].value = my_round(sigma_A_III, 1)
			if entr_NORM_QCC1.get() != '':
				ws['P30'].value = '≤{}'.format(entr_NORM_QCC1.get())
			else:
				ws['P30'].value = '≤15'

		if check_var_gr_1.get() >= 4:	
			ws['Q27'].value = my_round(mean_A_IV, razryad)
			ws['Q28'].value = my_round(Er_A_IV, 1)
			ws['Q29'].value = my_round(sigma_A_IV, 1)
			if entr_NORM_QCD1.get() != '':
				ws['Q30'].value = '≤{}'.format(entr_NORM_QCD1.get())
			else:
				ws['Q30'].value = '≤15'

		if check_var_gr_1.get() >= 5:	
			ws['R27'].value = my_round(mean_A_V, razryad)
			ws['R28'].value = my_round(Er_A_V, 1)
			ws['R29'].value = my_round(sigma_A_V, 1)
			if entr_NORM_QCE1.get() != '':
				ws['R30'].value = '≤{}'.format(entr_NORM_QCE1.get())
			else:
				ws['R30'].value = '≤15'

		if check_var_gr_1.get() >= 6:	
			ws['S27'].value = my_round(mean_A_VI, razryad)
			ws['S28'].value = my_round(Er_A_VI, 1)
			ws['S29'].value = my_round(sigma_A_VI, 1)
			if entr_NORM_QCF1.get() != '':
				ws['S30'].value = '≤{}'.format(entr_NORM_QCF1.get())
			else:
				ws['S30'].value = '≤15'

	if check_var_gr_3.get() == 10:
		ws['M28'].value = 'найдено, сред.зн., \nнг/мл (n=10)'
		ws['M28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
		ws['M29'].value = 'εR, % (n=10)'
		ws['M30'].value = 'σR, % (n=10)'
		ws['M31'].value = 'Норма |ε| и |σ|, %'

		ws['N28'].value = my_round(mean_A_I, razryad)
		ws['N29'].value = my_round(Er_A_I, 1)
		ws['N30'].value = my_round(sigma_A_I, 1)
		if entr_NORM_QCA1.get() != '':
			ws['N31'].value = '≤{}'.format(entr_NORM_QCA1.get())
		else:
			ws['N31'].value = '≤20'
		

		if check_var_gr_1.get() >= 2:
			ws['O28'].value = my_round(mean_A_II, razryad)
			ws['O29'].value = my_round(Er_A_II, 1)
			ws['O30'].value = my_round(sigma_A_II, 1)
			if entr_NORM_QCB1.get() != '':
				ws['O31'].value = '≤{}'.format(entr_NORM_QCB1.get())
			else:
				ws['O31'].value = '≤15'

		if check_var_gr_1.get() >= 3:	
			ws['P28'].value = my_round(mean_A_III, razryad)
			ws['P29'].value = my_round(Er_A_III, 1)
			ws['P30'].value = my_round(sigma_A_III, 1)
			if entr_NORM_QCC1.get() != '':
				ws['P31'].value = '≤{}'.format(entr_NORM_QCC1.get())
			else:
				ws['P31'].value = '≤15'

		if check_var_gr_1.get() >= 4:	
			ws['Q28'].value = my_round(mean_A_IV, razryad)
			ws['Q29'].value = my_round(Er_A_IV, 1)
			ws['Q30'].value = my_round(sigma_A_IV, 1)
			if entr_NORM_QCD1.get() != '':
				ws['Q31'].value = '≤{}'.format(entr_NORM_QCD1.get())
			else:
				ws['Q31'].value = '≤15'

		if check_var_gr_1.get() >= 5:	
			ws['R28'].value = my_round(mean_A_V, razryad)
			ws['R29'].value = my_round(Er_A_V, 1)
			ws['R30'].value = my_round(sigma_A_V, 1)
			if entr_NORM_QCE1.get() != '':
				ws['R31'].value = '≤{}'.format(entr_NORM_QCE1.get())
			else:
				ws['R31'].value = '≤15'

		if check_var_gr_1.get() >= 6:	
			ws['S28'].value = my_round(mean_A_VI, razryad)
			ws['S29'].value = my_round(Er_A_VI, 1)
			ws['S30'].value = my_round(sigma_A_VI, 1)
			if entr_NORM_QCF1.get() != '':
				ws['S31'].value = '≤{}'.format(entr_NORM_QCF1.get())
			else:
				ws['S31'].value = '≤15'


	if check_var_gr_1.get() >= 2:
		#QCB1
		ws['O15'].value = 'QCB1'
		ws['O16'].value = QC_II
		ws['L2'].value = 'QCB1'
		ws['O18'].value = my_round(float(entr_A1_II.get()), razryad)
		ws['L3'].value = my_round(float(entr_A1_II.get()), razryad)
		ws['O19'].value = my_round(float(entr_A2_II.get()), razryad)
		ws['L4'].value = my_round(float(entr_A2_II.get()), razryad)
		ws['O20'].value = my_round(float(entr_A3_II.get()), razryad)
		ws['L5'].value = my_round(float(entr_A3_II.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['O21'].value = my_round(float(entr_A4_II.get()), razryad)
			ws['L6'].value = my_round(float(entr_A4_II.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['O22'].value = my_round(float(entr_A5_II.get()), razryad)
			ws['L7'].value = my_round(float(entr_A5_II.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['O23'].value = my_round(float(entr_A6_II.get()), razryad)
			ws['L8'].value = my_round(float(entr_A6_II.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['O24'].value = my_round(float(entr_A7_II.get()), razryad)
			ws['L9'].value = my_round(float(entr_A7_II.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['O25'].value = my_round(float(entr_A8_II.get()), razryad)
			ws['L10'].value = my_round(float(entr_A8_II.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['O26'].value = my_round(float(entr_A9_II.get()), razryad)
			ws['L11'].value = my_round(float(entr_A9_II.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['O27'].value = my_round(float(entr_A10_II.get()), razryad)
			ws['L12'].value = my_round(float(entr_A10_II.get()), razryad)

	
	if check_var_gr_1.get() >= 3:
		#QCC1
		ws['P15'].value = 'QCC1'
		ws['P16'].value = QC_III
		ws['U2'].value = 'QCC1'
		ws['P18'].value = my_round(float(entr_A1_III.get()), razryad)
		ws['U3'].value = my_round(float(entr_A1_III.get()), razryad)
		ws['P19'].value = my_round(float(entr_A2_III.get()), razryad)
		ws['U4'].value = my_round(float(entr_A2_III.get()), razryad)
		ws['P20'].value = my_round(float(entr_A3_III.get()), razryad)
		ws['U5'].value = my_round(float(entr_A3_III.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['P21'].value = my_round(float(entr_A4_III.get()), razryad)
			ws['U6'].value = my_round(float(entr_A4_III.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['P22'].value = my_round(float(entr_A5_III.get()), razryad)
			ws['U7'].value = my_round(float(entr_A5_III.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['P23'].value = my_round(float(entr_A6_III.get()), razryad)
			ws['U8'].value = my_round(float(entr_A6_III.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['P24'].value = my_round(float(entr_A7_III.get()), razryad)
			ws['U9'].value = my_round(float(entr_A7_III.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['P25'].value = my_round(float(entr_A8_III.get()), razryad)
			ws['U10'].value = my_round(float(entr_A8_III.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['P26'].value = my_round(float(entr_A9_III.get()), razryad)
			ws['U11'].value = my_round(float(entr_A9_III.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['P27'].value = my_round(float(entr_A10_III.get()), razryad)
			ws['U12'].value = my_round(float(entr_A10_III.get()), razryad)


	
	if check_var_gr_1.get() >= 4:
		#QCD1
		ws['Q15'].value = 'QCD1'
		ws['Q16'].value = QC_IV
		ws['AD2'].value = 'QCD1'
		ws['Q18'].value = my_round(float(entr_A1_IV.get()), razryad)
		ws['AD3'].value = my_round(float(entr_A1_IV.get()), razryad)
		ws['Q19'].value = my_round(float(entr_A2_IV.get()), razryad)
		ws['AD4'].value = my_round(float(entr_A2_IV.get()), razryad)
		ws['Q20'].value = my_round(float(entr_A3_IV.get()), razryad)
		ws['AD5'].value = my_round(float(entr_A3_IV.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['Q21'].value = my_round(float(entr_A4_IV.get()), razryad)
			ws['AD6'].value = my_round(float(entr_A4_IV.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['Q22'].value = my_round(float(entr_A5_IV.get()), razryad)
			ws['AD7'].value = my_round(float(entr_A5_IV.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['Q23'].value = my_round(float(entr_A6_IV.get()), razryad)
			ws['AD8'].value = my_round(float(entr_A6_IV.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['Q24'].value = my_round(float(entr_A7_IV.get()), razryad)
			ws['AD9'].value = my_round(float(entr_A7_IV.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['Q25'].value = my_round(float(entr_A8_IV.get()), razryad)
			ws['AD10'].value = my_round(float(entr_A8_IV.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['Q26'].value = my_round(float(entr_A9_IV.get()), razryad)
			ws['AD11'].value = my_round(float(entr_A9_IV.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['Q27'].value = my_round(float(entr_A10_IV.get()), razryad)
			ws['AD12'].value = my_round(float(entr_A10_IV.get()), razryad)

	if check_var_gr_1.get() >= 5:
		#QCE1
		ws['R15'].value = 'QCE1'
		ws['R16'].value = QC_V
		ws['AM2'].value = 'QCE1'
		ws['R18'].value = my_round(float(entr_A1_V.get()), razryad)
		ws['AM3'].value = my_round(float(entr_A1_V.get()), razryad)
		ws['R19'].value = my_round(float(entr_A2_V.get()), razryad)
		ws['AM4'].value = my_round(float(entr_A2_V.get()), razryad)
		ws['R20'].value = my_round(float(entr_A3_V.get()), razryad)
		ws['AM5'].value = my_round(float(entr_A3_V.get()), razryad)

		if check_var_gr_3.get() >= 4:
			ws['R21'].value = my_round(float(entr_A4_V.get()), razryad)
			ws['AM6'].value = my_round(float(entr_A4_V.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['R22'].value = my_round(float(entr_A5_V.get()), razryad)
			ws['AM7'].value = my_round(float(entr_A5_V.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['R23'].value = my_round(float(entr_A6_V.get()), razryad)
			ws['AM8'].value = my_round(float(entr_A6_V.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['R24'].value = my_round(float(entr_A7_V.get()), razryad)
			ws['AM9'].value = my_round(float(entr_A7_V.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['R25'].value = my_round(float(entr_A8_V.get()), razryad)
			ws['AM10'].value = my_round(float(entr_A8_V.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['R26'].value = my_round(float(entr_A9_V.get()), razryad)
			ws['AM11'].value = my_round(float(entr_A9_V.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['R27'].value = my_round(float(entr_A10_V.get()), razryad)
			ws['AM12'].value = my_round(float(entr_A10_V.get()), razryad)

	

	if check_var_gr_1.get() >= 6:
		#QCF1
		ws['S15'].value = 'QCF1'
		ws['S16'].value = QC_VI
		ws['AV2'].value = 'QCF1'
		ws['S18'].value = my_round(float(entr_A1_VI.get()), razryad)
		ws['AV3'].value = my_round(float(entr_A1_VI.get()), razryad)
		ws['S19'].value = my_round(float(entr_A2_VI.get()), razryad)
		ws['AV4'].value = my_round(float(entr_A2_VI.get()), razryad)
		ws['S20'].value = my_round(float(entr_A3_VI.get()), razryad)
		ws['AV5'].value = my_round(float(entr_A3_VI.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['S21'].value = my_round(float(entr_A4_VI.get()), razryad)
			ws['AV6'].value = my_round(float(entr_A4_VI.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['S22'].value = my_round(float(entr_A5_VI.get()), razryad)
			ws['AV7'].value = my_round(float(entr_A5_VI.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['S23'].value = my_round(float(entr_A6_VI.get()), razryad)
			ws['AV8'].value = my_round(float(entr_A6_VI.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['S24'].value = my_round(float(entr_A7_VI.get()), razryad)
			ws['AV9'].value = my_round(float(entr_A7_VI.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['S25'].value = my_round(float(entr_A8_VI.get()), razryad)
			ws['AV10'].value = my_round(float(entr_A8_VI.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['S26'].value = my_round(float(entr_A9_VI.get()), razryad)
			ws['AV11'].value = my_round(float(entr_A9_VI.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['S27'].value = my_round(float(entr_A10_VI.get()), razryad)
			ws['AV12'].value = my_round(float(entr_A10_VI.get()), razryad)
		



#QC ТАБЛИЦА № 2
	if check_var_gr_2.get() >= 2:
		#выравнивание по центру]
		cols_c(ws, 'U15:AA34')

		#границы ячеек
		if check_var_gr_1.get() == 1:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'U15:V27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'U15:V28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'U15:V29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'U15:V30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'U15:V31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'U15:V32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'U15:V33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'U15:V34')

		if check_var_gr_1.get() == 2:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'U15:W27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'U15:W28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'U15:W29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'U15:W30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'U15:W31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'U15:W32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'U15:W33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'U15:W34')

		if check_var_gr_1.get() == 3:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'U15:X27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'U15:X28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'U15:X29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'U15:X30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'U15:X31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'U15:X32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'U15:X33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'U15:X34')				

		if check_var_gr_1.get() == 4:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'U15:Y27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'U15:Y28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'U15:Y29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'U15:Y30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'U15:Y31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'U15:Y32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'U15:Y33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'U15:Y34')			

		if check_var_gr_1.get() == 5:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'U15:Z27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'U15:Z28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'U15:Z29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'U15:Z30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'U15:Z31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'U15:Z32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'U15:Z33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'U15:Z34')		

		if check_var_gr_1.get() == 6:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'U15:AA27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'U15:AA28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'U15:AA29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'U15:AA30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'U15:AA31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'U15:AA32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'U15:AA33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'U15:AA34')	

		
		#надпись и объединение ячеек под надпись
		ws['U15'].value = 'QC'
		ws['U16'].value = f'Введено, {entr_concent.get()}'
		ws['U17'].value = '№ измерения'
		ws['V17'].value = 'Найдено, последовательность 2'

		#QCA2
		
		ws['V15'].value = 'QCA2'
		ws['V16'].value = QC_I
		ws['D2'].value = 'QCA2'
		ws['U18'].value = '1'
		ws['V18'].value = my_round(float(entr_B1_I.get()), razryad)
		ws['D3'].value = my_round(float(entr_B1_I.get()), razryad)
		ws['U19'].value = '2'
		ws['V19'].value = my_round(float(entr_B2_I.get()), razryad)
		ws['D4'].value = my_round(float(entr_B2_I.get()), razryad)
		ws['U20'].value = '3'
		ws['V20'].value = my_round(float(entr_B3_I.get()), razryad)
		ws['D5'].value = my_round(float(entr_B3_I.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['U21'].value = '4'
			ws['V21'].value = my_round(float(entr_B4_I.get()), razryad)
			ws['D6'].value = my_round(float(entr_B4_I.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['U22'].value = '5'
			ws['V22'].value = my_round(float(entr_B5_I.get()), razryad)
			ws['D7'].value = my_round(float(entr_B5_I.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['U23'].value = '6'
			ws['V23'].value = my_round(float(entr_B6_I.get()), razryad)
			ws['D8'].value = my_round(float(entr_B6_I.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['U24'].value = '7'
			ws['V24'].value = my_round(float(entr_B7_I.get()), razryad)
			ws['D9'].value = my_round(float(entr_B7_I.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['U25'].value = '8'
			ws['V25'].value = my_round(float(entr_B8_I.get()), razryad)
			ws['D10'].value = my_round(float(entr_B8_I.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['U26'].value = '9'
			ws['V26'].value = my_round(float(entr_B9_I.get()), razryad)
			ws['D11'].value = my_round(float(entr_B9_I.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['U27'].value = '10'
			ws['V27'].value = my_round(float(entr_B10_I.get()), razryad)
			ws['D12'].value = my_round(float(entr_B10_I.get()), razryad)


	#РАСЧЕТЫ ANOVA
	#количество таблиц зависит от показателя check_var_gr_2
		if check_var_gr_2.get() >= 2:
			mean_QCA1_QCA2 = my_round(stat.mean([mean_A_I, mean_B_I]), check_razryad.get())
			
			if check_var_gr_1.get() >= 2:
				#QCB2
				mean_QCB1_QCB2 = my_round(stat.mean([mean_A_II, mean_B_II]), check_razryad.get())
				
			if check_var_gr_1.get() >= 3:
				#QCC2			
				mean_QCC1_QCC2 = my_round(stat.mean([mean_A_III, mean_B_III]), check_razryad.get())			

			if check_var_gr_1.get() >= 4:
				#QCD2			
				mean_QCD1_QCD2 = my_round(stat.mean([mean_A_IV, mean_B_IV]), check_razryad.get())		

			if check_var_gr_1.get() >= 5:
				#QCE2			
				mean_QCE1_QCE2 = my_round(stat.mean([mean_A_V, mean_B_V]), check_razryad.get())	

			if check_var_gr_1.get() >= 6:
				#QCF2			
				mean_QCF1_QCF2 = my_round(stat.mean([mean_A_VI, mean_B_VI]), check_razryad.get())
				

	#ЗАПОЛНЕНИЕ ANOVA		
		if check_var_gr_2.get() >= 2:
		 	#междугруппами
			sum_kv_mezhdy_gr_QCA1_QCA2 = len(spisok_A_I)*((mean_A_I - mean_QCA1_QCA2)**2) + len(spisok_B_I)*((mean_B_I - mean_QCA1_QCA2)**2)
			CV_mezhdy_gr_QCA1_QCA2 = (sum_kv_mezhdy_gr_QCA1_QCA2**0.5)/mean_QCA1_QCA2*100
			CV_vnytr_gr_QCA1_QCA2 = (mean_QCA1_QCA2 - QC_I)/QC_I * 100
			#проверка на количество QC
			if check_var_gr_1.get() >= 2:
				#проверка на количество образцов
				if check_var_gr_2.get() >= 2:
				 	#междугруппами
					sum_kv_mezhdy_gr_QCB1_QCB2 = len(spisok_A_II)*((mean_A_II - mean_QCB1_QCB2)**2) + len(spisok_B_II)*((mean_B_II - mean_QCB1_QCB2)**2)
					CV_mezhdy_gr_QCB1_QCB2 = (sum_kv_mezhdy_gr_QCB1_QCB2**0.5)/mean_QCB1_QCB2*100
					CV_vnytr_gr_QCB1_QCB2 = (mean_QCB1_QCB2 - QC_II)/QC_II * 100

			#проверка на количество QC
			if check_var_gr_1.get() >= 3:
				#проверка на количество образцов
				if check_var_gr_2.get() >= 2:
				 	#междугруппами
					sum_kv_mezhdy_gr_QCC1_QCC2 = len(spisok_A_III)*((mean_A_III - mean_QCC1_QCC2)**2) + len(spisok_B_III)*((mean_B_III - mean_QCC1_QCC2)**2)
					CV_mezhdy_gr_QCC1_QCC2 = (sum_kv_mezhdy_gr_QCC1_QCC2**0.5)/mean_QCC1_QCC2*100
					CV_vnytr_gr_QCC1_QCC2 = (mean_QCC1_QCC2 - QC_III)/QC_III * 100

			if check_var_gr_1.get() >= 4:
				#проверка на количество образцов
				if check_var_gr_2.get() >= 2:
					sum_kv_mezhdy_gr_QCD1_QCD2 = len(spisok_A_IV)*((mean_A_IV - mean_QCD1_QCD2)**2) + len(spisok_B_IV)*((mean_B_IV - mean_QCD1_QCD2)**2)
					CV_mezhdy_gr_QCD1_QCD2 = (sum_kv_mezhdy_gr_QCD1_QCD2**0.5)/mean_QCD1_QCD2*100
					#внутригрупп
					CV_vnytr_gr_QCD1_QCD2 = (mean_QCD1_QCD2 - QC_IV)/QC_IV * 100

			if check_var_gr_1.get() >= 5:
				#проверка на количество образцов
				if check_var_gr_2.get() >= 2:
					sum_kv_mezhdy_gr_QCE1_QCE2 = len(spisok_A_V)*((mean_A_V - mean_QCE1_QCE2)**2) + len(spisok_B_V)*((mean_B_V - mean_QCE1_QCE2)**2)
					CV_mezhdy_gr_QCE1_QCE2 = (sum_kv_mezhdy_gr_QCE1_QCE2**0.5)/mean_QCE1_QCE2*100
					CV_vnytr_gr_QCE1_QCE2 = (mean_QCE1_QCE2 - QC_V)/QC_V * 100

			if check_var_gr_1.get() >= 6:
				#проверка на количество образцов
				if check_var_gr_2.get() >= 2:
				 	#междугруппами
					sum_kv_mezhdy_gr_QCF1_QCF2 = len(spisok_A_VI)*((mean_A_VI - mean_QCF1_QCF2)**2) + len(spisok_B_VI)*((mean_B_VI - mean_QCF1_QCF2)**2)
					CV_mezhdy_gr_QCF1_QCF2 = (sum_kv_mezhdy_gr_QCF1_QCF2**0.5)/mean_QCF1_QCF2*100

					#внутригрупп
					CV_vnytr_gr_QCF1_QCF2 = (mean_QCF1_QCF2 - QC_VI)/QC_VI * 100




	#вывод среднего и остальных показателей в таблицу 
		if check_var_gr_3.get() == 3:

			ws['U21'].value = 'найдено, сред.зн.,\nнг/мл (n=3)'
			ws['U21'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

			ws['U22'].value = 'εR, % (n=3)'
			ws['U23'].value = 'σR, % (n=3)'

			ws['U24'].value = 'найдено, сред.зн.,\nнг/мл (n=6)'
			ws['U24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U25'].value = 'εR, % (n=6)'
			ws['U26'].value = 'σR, % (n=6)'
			ws['U27'].value = 'Норма |ε| и |σ|, %'


			ws['V21'].value = my_round(mean_B_I, razryad)
			ws['V22'].value = my_round(Er_B_I, 1)
			ws['V23'].value = my_round(sigma_B_I, 1)

			ws['V24'].value = my_round(mean_QCA1_QCA2, razryad)
			ws['V25'].value = my_round(CV_vnytr_gr_QCA1_QCA2, 1)
			ws['V26'].value = my_round(CV_mezhdy_gr_QCA1_QCA2, 1)
			if entr_NORM_QCA1.get() != '':
				ws['V27'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['V27'].value = '≤20'


			
			if check_var_gr_1.get() >= 2:
				ws['W21'].value = my_round(mean_B_II, razryad)
				ws['W22'].value = my_round(Er_B_II, 1)
				ws['W23'].value = my_round(sigma_B_II, 1)

				ws['W24'].value = my_round(mean_QCB1_QCB2, razryad)
				ws['W25'].value = my_round(CV_vnytr_gr_QCB1_QCB2, 1)
				ws['W26'].value = my_round(CV_mezhdy_gr_QCB1_QCB2, 1)
				if entr_NORM_QCB1.get() != '':
					ws['W27'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['W27'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['X21'].value = my_round(mean_B_III, razryad)
				ws['X22'].value = my_round(Er_B_III, 1)
				ws['X23'].value = my_round(sigma_B_III, 1)

				ws['X24'].value = my_round(mean_QCC1_QCC2, razryad)
				ws['X25'].value = my_round(CV_vnytr_gr_QCC1_QCC2, 1)
				ws['X26'].value = my_round(CV_mezhdy_gr_QCC1_QCC2, 1)
				if entr_NORM_QCC1.get() != '':
					ws['X27'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['X27'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['Y21'].value = my_round(mean_B_IV, razryad)
				ws['Y22'].value = my_round(Er_B_IV, 1)
				ws['Y23'].value = my_round(sigma_B_IV, 1)

				ws['Y24'].value = my_round(mean_QCD1_QCD2, razryad)
				ws['Y25'].value = my_round(CV_vnytr_gr_QCD1_QCD2, 1)
				ws['Y26'].value = my_round(CV_mezhdy_gr_QCD1_QCD2, 1)
				if entr_NORM_QCD1.get() != '':
					ws['Y27'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['Y27'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['Z21'].value = my_round(mean_B_V, razryad)
				ws['Z22'].value = my_round(Er_B_V, 1)
				ws['Z23'].value = my_round(sigma_B_V, 1)

				ws['Z24'].value = my_round(mean_QCE1_QCE2, razryad)
				ws['Z25'].value = my_round(CV_vnytr_gr_QCE1_QCE2, 1)
				ws['Z26'].value = my_round(CV_mezhdy_gr_QCE1_QCE2, 1)
				if entr_NORM_QCE1.get() != '':
					ws['Z27'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['Z27'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AA21'].value = my_round(mean_B_VI, razryad)
				ws['AA22'].value = my_round(Er_B_VI, 1)
				ws['AA23'].value = my_round(sigma_B_VI, 1)

				ws['AA24'].value = my_round(mean_QCF1_QCF2, razryad)
				ws['AA25'].value = my_round(CV_vnytr_gr_QCF1_QCF2, 1)
				ws['AA26'].value = my_round(CV_mezhdy_gr_QCF1_QCF2, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AA27'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AA27'].value = '≤15'
		if check_var_gr_3.get() == 4:

			ws['U22'].value = 'найдено, сред.зн.,\nнг/мл (n=4)'
			ws['U22'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U23'].value = 'εR, % (n=4)'
			ws['U24'].value = 'σR, % (n=4)'
			
			ws['U25'].value = 'найдено, сред.зн.,\nнг/мл (n=8)'
			ws['U25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U26'].value = 'εR, % (n=8)'
			ws['U27'].value = 'σR, % (n=8)'
			ws['U28'].value = 'Норма |ε| и |σ|, %'


			ws['V22'].value = my_round(mean_B_I, razryad)
			ws['V23'].value = my_round(Er_B_I, 1)
			ws['V24'].value = my_round(sigma_B_I, 1)

			ws['V25'].value = my_round(mean_QCA1_QCA2, razryad)
			ws['V26'].value = my_round(CV_vnytr_gr_QCA1_QCA2, 1)
			ws['V27'].value = my_round(CV_mezhdy_gr_QCA1_QCA2, 1)
			if entr_NORM_QCA1.get() != '':
				ws['V28'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['V28'].value = '≤20'	

			if check_var_gr_1.get() >= 2:
				ws['W22'].value = my_round(mean_B_II, razryad)
				ws['W23'].value = my_round(Er_B_II, 1)
				ws['W24'].value = my_round(sigma_B_II, 1)

				ws['W25'].value = my_round(mean_QCB1_QCB2, razryad)
				ws['W26'].value = my_round(CV_vnytr_gr_QCB1_QCB2, 1)
				ws['W27'].value = my_round(CV_mezhdy_gr_QCB1_QCB2, 1)
				if entr_NORM_QCB1.get() != '':
					ws['W28'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['W28'].value = '≤15'	

			if check_var_gr_1.get() >= 3:	
				ws['X22'].value = my_round(mean_B_III, razryad)
				ws['X23'].value = my_round(Er_B_III, 1)
				ws['X24'].value = my_round(sigma_B_III, 1)

				ws['X25'].value = my_round(mean_QCC1_QCC2, razryad)
				ws['X26'].value = my_round(CV_vnytr_gr_QCC1_QCC2, 1)
				ws['X27'].value = my_round(CV_mezhdy_gr_QCC1_QCC2, 1)
				if entr_NORM_QCC1.get() != '':
					ws['X28'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['X28'].value = '≤15'	

			if check_var_gr_1.get() >= 4:	
				ws['Y22'].value = my_round(mean_B_IV, razryad)
				ws['Y23'].value = my_round(Er_B_IV, 1)
				ws['Y24'].value = my_round(sigma_B_IV, 1)

				ws['Y25'].value = my_round(mean_QCD1_QCD2, razryad)
				ws['Y26'].value = my_round(CV_vnytr_gr_QCD1_QCD2, 1)
				ws['Y27'].value = my_round(CV_mezhdy_gr_QCD1_QCD2, 1)
				if entr_NORM_QCD1.get() != '':
					ws['Y28'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['Y28'].value = '≤15'	

			if check_var_gr_1.get() >= 5:	
				ws['Z22'].value = my_round(mean_B_V, razryad)
				ws['Z23'].value = my_round(Er_B_V, 1)
				ws['Z24'].value = my_round(sigma_B_V, 1)

				ws['Z25'].value = my_round(mean_QCE1_QCE2, razryad)
				ws['Z26'].value = my_round(CV_vnytr_gr_QCE1_QCE2, 1)
				ws['Z27'].value = my_round(CV_mezhdy_gr_QCE1_QCE2, 1)
				if entr_NORM_QCE1.get() != '':
					ws['Z28'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['Z28'].value = '≤15'	

			if check_var_gr_1.get() >= 6:	
				ws['AA22'].value = my_round(mean_B_VI, razryad)
				ws['AA23'].value = my_round(Er_B_VI, 1)
				ws['AA24'].value = my_round(sigma_B_VI, 1)

				ws['AA25'].value = my_round(mean_QCF1_QCF2, razryad)
				ws['AA26'].value = my_round(CV_vnytr_gr_QCF1_QCF2, 1)
				ws['AA27'].value = my_round(CV_mezhdy_gr_QCF1_QCF2, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AA28'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AA28'].value = '≤15'	

		if check_var_gr_3.get() == 5:
			ws['U23'].value = 'найдено, сред.зн., \nнг/мл (n=5)'
			ws['U23'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U24'].value = 'εR, % (n=5)'
			ws['U25'].value = 'σR, % (n=5)'
			
			ws['U26'].value = 'найдено, сред.зн.,\nнг/мл (n=10)'
			ws['U26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U27'].value = 'εR, % (n=10)'
			ws['U28'].value = 'σR, % (n=10)'
			ws['U29'].value = 'Норма |ε| и |σ|, %'

			ws['V23'].value = my_round(mean_B_I, razryad)
			ws['V24'].value = my_round(Er_B_I, 1)
			ws['V25'].value = my_round(sigma_B_I, 1)

			ws['V26'].value = my_round(mean_QCA1_QCA2, razryad)
			ws['V27'].value = my_round(CV_vnytr_gr_QCA1_QCA2, 1)
			ws['V28'].value = my_round(CV_mezhdy_gr_QCA1_QCA2, 1)

			if entr_NORM_QCA1.get() != '':
				ws['V29'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['V29'].value = '≤20'			

			if check_var_gr_1.get() >= 2:
				ws['W23'].value = my_round(mean_B_II, razryad)
				ws['W24'].value = my_round(Er_B_II, 1)
				ws['W25'].value = my_round(sigma_B_II, 1)

				ws['W26'].value = my_round(mean_QCB1_QCB2, razryad)
				ws['W27'].value = my_round(CV_vnytr_gr_QCB1_QCB2, 1)
				ws['W28'].value = my_round(CV_mezhdy_gr_QCB1_QCB2, 1)
				if entr_NORM_QCB1.get() != '':
					ws['W29'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['W29'].value = '≤15'		

			if check_var_gr_1.get() >= 3:	
				ws['X23'].value = my_round(mean_B_III, razryad)
				ws['X24'].value = my_round(Er_B_III, 1)
				ws['X25'].value = my_round(sigma_B_III, 1)

				ws['X26'].value = my_round(mean_QCC1_QCC2, razryad)
				ws['X27'].value = my_round(CV_vnytr_gr_QCC1_QCC2, 1)
				ws['X28'].value = my_round(CV_mezhdy_gr_QCC1_QCC2, 1)

				if entr_NORM_QCC1.get() != '':
					ws['X29'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['X29'].value = '≤15'	

			if check_var_gr_1.get() >= 4:	
				ws['Y23'].value = my_round(mean_B_IV, razryad)
				ws['Y24'].value = my_round(Er_B_IV, 1)
				ws['Y25'].value = my_round(sigma_B_IV, 1)

				ws['Y26'].value = my_round(mean_QCD1_QCD2, razryad)
				ws['Y27'].value = my_round(CV_vnytr_gr_QCD1_QCD2, 1)
				ws['Y28'].value = my_round(CV_mezhdy_gr_QCD1_QCD2, 1)
				if entr_NORM_QCD1.get() != '':
					ws['Y29'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['Y29'].value = '≤15'	

			if check_var_gr_1.get() >= 5:	
				ws['Z23'].value = my_round(mean_B_V, razryad)
				ws['Z24'].value = my_round(Er_B_V, 1)
				ws['Z25'].value = my_round(sigma_B_V, 1)

				ws['Z26'].value = my_round(mean_QCE1_QCE2, razryad)
				ws['Z27'].value = my_round(CV_vnytr_gr_QCE1_QCE2, 1)
				ws['Z28'].value = my_round(CV_mezhdy_gr_QCE1_QCE2, 1)
				if entr_NORM_QCE1.get() != '':
					ws['Z29'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['Z29'].value = '≤15'	

			if check_var_gr_1.get() >= 6:	
				ws['AA23'].value = my_round(mean_B_VI, razryad)
				ws['AA24'].value = my_round(Er_B_VI, 1)
				ws['AA25'].value = my_round(sigma_B_VI, 1)

				ws['AA26'].value = my_round(mean_QCF1_QCF2, razryad)
				ws['AA27'].value = my_round(CV_vnytr_gr_QCF1_QCF2, 1)
				ws['AA28'].value = my_round(CV_mezhdy_gr_QCF1_QCF2, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AA29'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AA29'].value = '≤15'	

		if check_var_gr_3.get() == 6:

			ws['U24'].value = 'найдено, сред.зн., \nнг/мл (n=6)'
			ws['U24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U25'].value = 'εR, % (n=6)'
			ws['U26'].value = 'σR, % (n=6)'
			
			ws['U27'].value = 'найдено, сред.зн.,\nнг/мл (n=12)'
			ws['U27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U28'].value = 'εR, % (n=12)'
			ws['U29'].value = 'σR, % (n=12)'
			ws['U30'].value = 'Норма |ε| и |σ|, %'

			ws['V24'].value = my_round(mean_B_I, razryad)
			ws['V25'].value = my_round(Er_B_I, 1)
			ws['V26'].value = my_round(sigma_B_I, 1)

			ws['V27'].value = my_round(mean_QCA1_QCA2, razryad)
			ws['V28'].value = my_round(CV_vnytr_gr_QCA1_QCA2, 1)
			ws['V29'].value = my_round(CV_mezhdy_gr_QCA1_QCA2, 1)
			if entr_NORM_QCA1.get() != '':
				ws['V30'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['V30'].value = '≤20'	

			if check_var_gr_1.get() >= 2:
				ws['W24'].value = my_round(mean_B_II, razryad)
				ws['W25'].value = my_round(Er_B_II, 1)
				ws['W26'].value = my_round(sigma_B_II, 1)

				ws['W27'].value = my_round(mean_QCB1_QCB2, razryad)
				ws['W28'].value = my_round(CV_vnytr_gr_QCB1_QCB2, 1)
				ws['W29'].value = my_round(CV_mezhdy_gr_QCB1_QCB2, 1)
				if entr_NORM_QCB1.get() != '':
					ws['W30'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['W30'].value = '≤15'	

			if check_var_gr_1.get() >= 3:	
				ws['X24'].value = my_round(mean_B_III, razryad)
				ws['X25'].value = my_round(Er_B_III, 1)
				ws['X26'].value = my_round(sigma_B_III, 1)

				ws['X27'].value = my_round(mean_QCC1_QCC2, razryad)
				ws['X28'].value = my_round(CV_vnytr_gr_QCC1_QCC2, 1)
				ws['X29'].value = my_round(CV_mezhdy_gr_QCC1_QCC2, 1)
				if entr_NORM_QCC1.get() != '':
					ws['X30'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['X30'].value = '≤15'	

			if check_var_gr_1.get() >= 4:	
				ws['Y24'].value = my_round(mean_B_IV, razryad)
				ws['Y25'].value = my_round(Er_B_IV, 1)
				ws['Y26'].value = my_round(sigma_B_IV, 1)

				ws['Y27'].value = my_round(mean_QCD1_QCD2, razryad)
				ws['Y28'].value = my_round(CV_vnytr_gr_QCD1_QCD2, 1)
				ws['Y29'].value = my_round(CV_mezhdy_gr_QCD1_QCD2, 1)
				if entr_NORM_QCD1.get() != '':
					ws['Y30'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['Y30'].value = '≤15'	

			if check_var_gr_1.get() >= 5:	
				ws['Z24'].value = my_round(mean_B_V, razryad)
				ws['Z25'].value = my_round(Er_B_V, 1)
				ws['Z26'].value = my_round(sigma_B_V, 1)

				ws['Z27'].value = my_round(mean_QCE1_QCE2, razryad)
				ws['Z28'].value = my_round(CV_vnytr_gr_QCE1_QCE2, 1)
				ws['Z29'].value = my_round(CV_mezhdy_gr_QCE1_QCE2, 1)
				if entr_NORM_QCE1.get() != '':
					ws['Z30'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['Z30'].value = '≤15'	

			if check_var_gr_1.get() >= 6:	
				ws['AA24'].value = my_round(mean_B_VI, razryad)
				ws['AA25'].value = my_round(Er_B_VI, 1)
				ws['AA26'].value = my_round(sigma_B_VI, 1)

				ws['AA27'].value = my_round(mean_QCF1_QCF2, razryad)
				ws['AA28'].value = my_round(CV_vnytr_gr_QCF1_QCF2, 1)
				ws['AA29'].value = my_round(CV_mezhdy_gr_QCF1_QCF2, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AA30'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AA30'].value = '≤15'
						
		if check_var_gr_3.get() == 7:
			ws['U25'].value = 'найдено, сред.зн., \nнг/мл (n=7)'
			ws['U25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U26'].value = 'εR, % (n=7)'
			ws['U27'].value = 'σR, % (n=7)'
			
			ws['U28'].value = 'найдено, сред.зн.,\nнг/мл (n=14)'
			ws['U28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U29'].value = 'εR, % (n=14)'
			ws['U30'].value = 'σR, % (n=14)'
			ws['U31'].value = 'Норма |ε| и |σ|, %'

			ws['V25'].value = my_round(mean_B_I, razryad)
			ws['V26'].value = my_round(Er_B_I, 1)
			ws['V27'].value = my_round(sigma_B_I, 1)

			ws['V28'].value = my_round(mean_QCA1_QCA2, razryad)
			ws['V29'].value = my_round(CV_vnytr_gr_QCA1_QCA2, 1)
			ws['V30'].value = my_round(CV_mezhdy_gr_QCA1_QCA2, 1)
			if entr_NORM_QCA1.get() != '':
				ws['V31'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['V31'].value = '≤20'	

			if check_var_gr_1.get() >= 2:
				ws['W25'].value = my_round(mean_B_II, razryad)
				ws['W26'].value = my_round(Er_B_II, 1)
				ws['W27'].value = my_round(sigma_B_II, 1)

				ws['W28'].value = my_round(mean_QCB1_QCB2, razryad)
				ws['W29'].value = my_round(CV_vnytr_gr_QCB1_QCB2, 1)
				ws['W30'].value = my_round(CV_mezhdy_gr_QCB1_QCB2, 1)
				if entr_NORM_QCB1.get() != '':
					ws['W31'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['W31'].value = '≤15'	

			if check_var_gr_1.get() >= 3:	
				ws['X25'].value = my_round(mean_B_III, razryad)
				ws['X26'].value = my_round(Er_B_III, 1)
				ws['X27'].value = my_round(sigma_B_III, 1)

				ws['X28'].value = my_round(mean_QCC1_QCC2, razryad)
				ws['X29'].value = my_round(CV_vnytr_gr_QCC1_QCC2, 1)
				ws['X30'].value = my_round(CV_mezhdy_gr_QCC1_QCC2, 1)
				if entr_NORM_QCC1.get() != '':
					ws['X31'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['X31'].value = '≤15'	

			if check_var_gr_1.get() >= 4:	
				ws['Y25'].value = my_round(mean_B_IV, razryad)
				ws['Y26'].value = my_round(Er_B_IV, 1)
				ws['Y27'].value = my_round(sigma_B_IV, 1)

				ws['Y28'].value = my_round(mean_QCD1_QCD2, razryad)
				ws['Y29'].value = my_round(CV_vnytr_gr_QCD1_QCD2, 1)
				ws['Y30'].value = my_round(CV_mezhdy_gr_QCD1_QCD2, 1)
				if entr_NORM_QCD1.get() != '':
					ws['Y31'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['Y31'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['Z25'].value = my_round(mean_B_V, razryad)
				ws['Z26'].value = my_round(Er_B_V, 1)
				ws['Z27'].value = my_round(sigma_B_V, 1)

				ws['Z28'].value = my_round(mean_QCE1_QCE2, razryad)
				ws['Z29'].value = my_round(CV_vnytr_gr_QCE1_QCE2, 1)
				ws['Z30'].value = my_round(CV_mezhdy_gr_QCE1_QCE2, 1)
				if entr_NORM_QCE1.get() != '':
					ws['Z31'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['Z31'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AA25'].value = my_round(mean_B_VI, razryad)
				ws['AA26'].value = my_round(Er_B_VI, 1)
				ws['AA27'].value = my_round(sigma_B_VI, 1)

				ws['AA28'].value = my_round(mean_QCF1_QCF2, razryad)
				ws['AA29'].value = my_round(CV_vnytr_gr_QCF1_QCF2, 1)
				ws['AA30'].value = my_round(CV_mezhdy_gr_QCF1_QCF2, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AA31'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AA31'].value = '≤15'
		if check_var_gr_3.get() == 8:
			ws['U26'].value = 'найдено, сред.зн., \nнг/мл (n=8)'
			ws['U26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U27'].value = 'εR, % (n=8)'
			ws['U28'].value = 'σR, % (n=8)'
			
			ws['U29'].value = 'найдено, сред.зн.,\nнг/мл (n=16)'
			ws['U29'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U30'].value = 'εR, % (n=16)'
			ws['U30'].value = 'σR, % (n=16)'
			ws['U31'].value = 'Норма |ε| и |σ|, %'

			ws['V26'].value = my_round(mean_B_I, razryad)
			ws['V27'].value = my_round(Er_B_I, 1)
			ws['V28'].value = my_round(sigma_B_I, 1)

			ws['V29'].value = my_round(mean_QCA1_QCA2, razryad)
			ws['V30'].value = my_round(CV_vnytr_gr_QCA1_QCA2, 1)
			ws['V31'].value = my_round(CV_mezhdy_gr_QCA1_QCA2, 1)
			if entr_NORM_QCA1.get() != '':
				ws['V32'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['V32'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['W26'].value = my_round(mean_B_II, razryad)
				ws['W27'].value = my_round(Er_B_II, 1)
				ws['W28'].value = my_round(sigma_B_II, 1)

				ws['W29'].value = my_round(mean_QCB1_QCB2, razryad)
				ws['W30'].value = my_round(CV_vnytr_gr_QCB1_QCB2, 1)
				ws['W31'].value = my_round(CV_mezhdy_gr_QCB1_QCB2, 1)
				if entr_NORM_QCB1.get() != '':
					ws['W32'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['W32'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['X26'].value = my_round(mean_B_III, razryad)
				ws['X27'].value = my_round(Er_B_III, 1)
				ws['X28'].value = my_round(sigma_B_III, 1)

				ws['X29'].value = my_round(mean_QCC1_QCC2, razryad)
				ws['X30'].value = my_round(CV_vnytr_gr_QCC1_QCC2, 1)
				ws['X31'].value = my_round(CV_mezhdy_gr_QCC1_QCC2, 1)
				if entr_NORM_QCC1.get() != '':
					ws['X32'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['X32'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['Y26'].value = my_round(mean_B_IV, razryad)
				ws['Y27'].value = my_round(Er_B_IV, 1)
				ws['Y28'].value = my_round(sigma_B_IV, 1)

				ws['Y29'].value = my_round(mean_QCD1_QCD2, razryad)
				ws['Y30'].value = my_round(CV_vnytr_gr_QCD1_QCD2, 1)
				ws['Y31'].value = my_round(CV_mezhdy_gr_QCD1_QCD2, 1)
				if entr_NORM_QCD1.get() != '':
					ws['Y32'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['Y32'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['Z26'].value = my_round(mean_B_V, razryad)
				ws['Z27'].value = my_round(Er_B_V, 1)
				ws['Z28'].value = my_round(sigma_B_V, 1)

				ws['Z29'].value = my_round(mean_QCE1_QCE2, razryad)
				ws['Z30'].value = my_round(CV_vnytr_gr_QCE1_QCE2, 1)
				ws['Z31'].value = my_round(CV_mezhdy_gr_QCE1_QCE2, 1)
				if entr_NORM_QCE1.get() != '':
					ws['Z32'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['Z32'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AA26'].value = my_round(mean_B_VI, razryad)
				ws['AA27'].value = my_round(Er_B_VI, 1)
				ws['AA28'].value = my_round(sigma_B_VI, 1)

				ws['AA29'].value = my_round(mean_QCF1_QCF2, razryad)
				ws['AA30'].value = my_round(CV_vnytr_gr_QCF1_QCF2, 1)
				ws['AA31'].value = my_round(CV_mezhdy_gr_QCF1_QCF2, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AA32'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AA32'].value = '≤15'
		if check_var_gr_3.get() == 9:
			ws['U27'].value = 'найдено, сред.зн., \nнг/мл (n=9)'
			ws['U27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U28'].value = 'εR, % (n=9)'
			ws['U29'].value = 'σR, % (n=9)'
			
			ws['U30'].value = 'найдено, сред.зн.,\nнг/мл (n=18)'
			ws['U30'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U31'].value = 'εR, % (n=18)'
			ws['U32'].value = 'σR, % (n=18)'
			ws['U33'].value = 'Норма |ε| и |σ|, %'

			ws['V27'].value = my_round(mean_B_I, razryad)
			ws['V28'].value = my_round(Er_B_I, 1)
			ws['V29'].value = my_round(sigma_B_I, 1)

			ws['V30'].value = my_round(mean_QCA1_QCA2, razryad)
			ws['V31'].value = my_round(CV_vnytr_gr_QCA1_QCA2, 1)
			ws['V32'].value = my_round(CV_mezhdy_gr_QCA1_QCA2, 1)
			if entr_NORM_QCA1.get() != '':
				ws['V33'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['V33'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['W27'].value = my_round(mean_B_II, razryad)
				ws['W28'].value = my_round(Er_B_II, 1)
				ws['W29'].value = my_round(sigma_B_II, 1)

				ws['W30'].value = my_round(mean_QCB1_QCB2, razryad)
				ws['W31'].value = my_round(CV_vnytr_gr_QCB1_QCB2, 1)
				ws['W32'].value = my_round(CV_mezhdy_gr_QCB1_QCB2, 1)
				if entr_NORM_QCB1.get() != '':
					ws['W33'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['W33'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['X27'].value = my_round(mean_B_III, razryad)
				ws['X28'].value = my_round(Er_B_III, 1)
				ws['X29'].value = my_round(sigma_B_III, 1)

				ws['X30'].value = my_round(mean_QCC1_QCC2, razryad)
				ws['X31'].value = my_round(CV_vnytr_gr_QCC1_QCC2, 1)
				ws['X32'].value = my_round(CV_mezhdy_gr_QCC1_QCC2, 1)
				if entr_NORM_QCC1.get() != '':
					ws['X33'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['X33'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['Y27'].value = my_round(mean_B_IV, razryad)
				ws['Y28'].value = my_round(Er_B_IV, 1)
				ws['Y29'].value = my_round(sigma_B_IV, 1)

				ws['Y30'].value = my_round(mean_QCD1_QCD2, razryad)
				ws['Y31'].value = my_round(CV_vnytr_gr_QCD1_QCD2, 1)
				ws['Y32'].value = my_round(CV_mezhdy_gr_QCD1_QCD2, 1)
				if entr_NORM_QCD1.get() != '':
					ws['Y33'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['Y33'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['Z27'].value = my_round(mean_B_V, razryad)
				ws['Z28'].value = my_round(Er_B_V, 1)
				ws['Z29'].value = my_round(sigma_B_V, 1)

				ws['Z30'].value = my_round(mean_QCE1_QCE2, razryad)
				ws['Z31'].value = my_round(CV_vnytr_gr_QCE1_QCE2, 1)
				ws['Z32'].value = my_round(CV_mezhdy_gr_QCE1_QCE2, 1)
				if entr_NORM_QCE1.get() != '':
					ws['Z33'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['Z33'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AA27'].value = my_round(mean_B_VI, razryad)
				ws['AA28'].value = my_round(Er_B_VI, 1)
				ws['AA29'].value = my_round(sigma_B_VI, 1)

				ws['AA30'].value = my_round(mean_QCF1_QCF2, razryad)
				ws['AA31'].value = my_round(CV_vnytr_gr_QCF1_QCF2, 1)
				ws['AA32'].value = my_round(CV_mezhdy_gr_QCF1_QCF2, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AA33'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AA33'].value = '≤15'
		if check_var_gr_3.get() == 10:
			ws['U28'].value = 'найдено, сред.зн., \nнг/мл (n=10)'
			ws['U28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U29'].value = 'εR, % (n=10)'
			ws['U30'].value = 'σR, % (n=10)'
			
			ws['U31'].value = 'найдено, сред.зн.,\nнг/мл (n=20)'
			ws['U31'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['U32'].value = 'εR, % (n=20)'
			ws['U33'].value = 'σR, % (n=20)'
			ws['U34'].value = 'Норма |ε| и |σ|, %'

			ws['V28'].value = my_round(mean_B_I, razryad)
			ws['V29'].value = my_round(Er_B_I, 1)
			ws['V30'].value = my_round(sigma_B_I, 1)

			ws['V31'].value = my_round(mean_QCA1_QCA2, razryad)
			ws['V32'].value = my_round(CV_vnytr_gr_QCA1_QCA2, 1)
			ws['V33'].value = my_round(CV_mezhdy_gr_QCA1_QCA2, 1)
			if entr_NORM_QCA1.get() != '':
				ws['V34'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['V34'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['W28'].value = my_round(mean_B_II, razryad)
				ws['W29'].value = my_round(Er_B_II, 1)
				ws['W30'].value = my_round(sigma_B_II, 1)

				ws['W31'].value = my_round(mean_QCB1_QCB2, razryad)
				ws['W32'].value = my_round(CV_vnytr_gr_QCB1_QCB2, 1)
				ws['W33'].value = my_round(CV_mezhdy_gr_QCB1_QCB2, 1)
				if entr_NORM_QCB1.get() != '':
					ws['W34'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['W34'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['X28'].value = my_round(mean_B_III, razryad)
				ws['X29'].value = my_round(Er_B_III, 1)
				ws['X30'].value = my_round(sigma_B_III, 1)

				ws['X31'].value = my_round(mean_QCC1_QCC2, razryad)
				ws['X32'].value = my_round(CV_vnytr_gr_QCC1_QCC2, 1)
				ws['X33'].value = my_round(CV_mezhdy_gr_QCC1_QCC2, 1)
				if entr_NORM_QCC1.get() != '':
					ws['X34'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['X34'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['Y28'].value = my_round(mean_B_IV, razryad)
				ws['Y29'].value = my_round(Er_B_IV, 1)
				ws['Y30'].value = my_round(sigma_B_IV, 1)

				ws['Y31'].value = my_round(mean_QCD1_QCD2, razryad)
				ws['Y32'].value = my_round(CV_vnytr_gr_QCD1_QCD2, 1)
				ws['Y33'].value = my_round(CV_mezhdy_gr_QCD1_QCD2, 1)
				if entr_NORM_QCD1.get() != '':
					ws['Y34'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['Y34'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['Z28'].value = my_round(mean_B_V, razryad)
				ws['Z29'].value = my_round(Er_B_V, 1)
				ws['Z30'].value = my_round(sigma_B_V, 1)

				ws['Z31'].value = my_round(mean_QCE1_QCE2, razryad)
				ws['Z32'].value = my_round(CV_vnytr_gr_QCE1_QCE2, 1)
				ws['Z33'].value = my_round(CV_mezhdy_gr_QCE1_QCE2, 1)
				if entr_NORM_QCE1.get() != '':
					ws['Z34'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['Z34'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AA28'].value = my_round(mean_B_VI, razryad)
				ws['AA29'].value = my_round(Er_B_VI, 1)
				ws['AA30'].value = my_round(sigma_B_VI, 1)

				ws['AA31'].value = my_round(mean_QCF1_QCF2, razryad)
				ws['AA32'].value = my_round(CV_vnytr_gr_QCF1_QCF2, 1)
				ws['AA33'].value = my_round(CV_mezhdy_gr_QCF1_QCF2, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AA34'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AA34'].value = '≤15'
		if check_var_gr_1.get() >= 2:
			#QCB2
			ws['W15'].value = 'QCB2'
			ws['W16'].value = QC_II
			ws['M2'].value = 'QCB2'
			ws['W18'].value = my_round(float(entr_B1_II.get()), razryad)
			ws['M3'].value = my_round(float(entr_B1_II.get()), razryad)
			ws['W19'].value = my_round(float(entr_B2_II.get()), razryad)
			ws['M4'].value = my_round(float(entr_B2_II.get()), razryad)
			ws['W20'].value = my_round(float(entr_B3_II.get()), razryad)
			ws['M5'].value = my_round(float(entr_B3_II.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['W21'].value = my_round(float(entr_B4_II.get()), razryad)
				ws['M6'].value = my_round(float(entr_B4_II.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['W22'].value = my_round(float(entr_B5_II.get()), razryad)
				ws['M7'].value = my_round(float(entr_B5_II.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['W23'].value = my_round(float(entr_B6_II.get()), razryad)
				ws['M8'].value = my_round(float(entr_B6_II.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['W24'].value = my_round(float(entr_B7_II.get()), razryad)
				ws['M9'].value = my_round(float(entr_B7_II.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['W25'].value = my_round(float(entr_B8_II.get()), razryad)
				ws['M10'].value = my_round(float(entr_B8_II.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['W26'].value = my_round(float(entr_B9_II.get()), razryad)
				ws['M11'].value = my_round(float(entr_B9_II.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['W27'].value = my_round(float(entr_B10_II.get()), razryad)
				ws['M12'].value = my_round(float(entr_B10_II.get()), razryad)
		if check_var_gr_1.get() >= 3:
			#QCC2
			ws['X15'].value = 'QCC2'
			ws['X16'].value = QC_III
			ws['V2'].value = 'QCC2'
			ws['X18'].value = my_round(float(entr_B1_III.get()), razryad)
			ws['V3'].value = my_round(float(entr_B1_III.get()), razryad)
			ws['X19'].value = my_round(float(entr_B2_III.get()), razryad)
			ws['V4'].value = my_round(float(entr_B2_III.get()), razryad)
			ws['X20'].value = my_round(float(entr_B3_III.get()), razryad)
			ws['V5'].value = my_round(float(entr_B3_III.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['X21'].value = my_round(float(entr_B4_III.get()), razryad)
				ws['V6'].value = my_round(float(entr_B4_III.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['X22'].value = my_round(float(entr_B5_III.get()), razryad)
				ws['V7'].value = my_round(float(entr_B5_III.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['X23'].value = my_round(float(entr_B6_III.get()), razryad)
				ws['V8'].value = my_round(float(entr_B6_III.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['X24'].value = my_round(float(entr_B7_III.get()), razryad)
				ws['V9'].value = my_round(float(entr_B7_III.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['X25'].value = my_round(float(entr_B8_III.get()), razryad)
				ws['V10'].value = my_round(float(entr_B8_III.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['X26'].value = my_round(float(entr_B9_III.get()), razryad)
				ws['V11'].value = my_round(float(entr_B9_III.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['X27'].value = my_round(float(entr_B10_III.get()), razryad)
				ws['V12'].value = my_round(float(entr_B10_III.get()), razryad)
		if check_var_gr_1.get() >= 4:
			#QCD2
			ws['Y15'].value = 'QCD2'
			ws['Y16'].value = QC_IV
			ws['AE2'].value = 'QCD2'
			ws['Y18'].value = my_round(float(entr_B1_IV.get()), razryad)
			ws['AE3'].value = my_round(float(entr_B1_IV.get()), razryad)
			ws['Y19'].value = my_round(float(entr_B2_IV.get()), razryad)
			ws['AE4'].value = my_round(float(entr_B2_IV.get()), razryad)
			ws['Y20'].value = my_round(float(entr_B3_IV.get()), razryad)
			ws['AE5'].value = my_round(float(entr_B3_IV.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['Y21'].value = my_round(float(entr_B4_IV.get()), razryad)
				ws['AE6'].value = my_round(float(entr_B4_IV.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['Y22'].value = my_round(float(entr_B5_IV.get()), razryad)
				ws['AE7'].value = my_round(float(entr_B5_IV.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['Y23'].value = my_round(float(entr_B6_IV.get()), razryad)
				ws['AE8'].value = my_round(float(entr_B6_IV.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['Y24'].value = my_round(float(entr_B7_IV.get()), razryad)
				ws['AE9'].value = my_round(float(entr_B7_IV.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['Y25'].value = my_round(float(entr_B8_IV.get()), razryad)
				ws['AE10'].value = my_round(float(entr_B8_IV.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['Y26'].value = my_round(float(entr_B9_IV.get()), razryad)
				ws['AE11'].value = my_round(float(entr_B9_IV.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['Y27'].value = my_round(float(entr_B10_IV.get()), razryad)
				ws['AE12'].value = my_round(float(entr_B10_IV.get()), razryad)
		if check_var_gr_1.get() >= 5:
			#QCE2
			ws['Z15'].value = 'QCE2'
			ws['Z16'].value = QC_V
			ws['AN2'].value = 'QCE2'
			ws['Z18'].value = my_round(float(entr_B1_V.get()), razryad)
			ws['AN3'].value = my_round(float(entr_B1_V.get()), razryad)
			ws['Z19'].value = my_round(float(entr_B2_V.get()), razryad)
			ws['AN4'].value = my_round(float(entr_B2_V.get()), razryad)
			ws['Z20'].value = my_round(float(entr_B3_V.get()), razryad)
			ws['AN5'].value = my_round(float(entr_B3_V.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['Z21'].value = my_round(float(entr_B4_V.get()), razryad)
				ws['AN6'].value = my_round(float(entr_B4_V.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['Z22'].value = my_round(float(entr_B5_V.get()), razryad)
				ws['AN7'].value = my_round(float(entr_B5_V.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['Z23'].value = my_round(float(entr_B6_V.get()), razryad)
				ws['AN8'].value = my_round(float(entr_B6_V.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['Z24'].value = my_round(float(entr_B7_V.get()), razryad)
				ws['AN9'].value = my_round(float(entr_B7_V.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['Z25'].value = my_round(float(entr_B8_V.get()), razryad)
				ws['AN10'].value = my_round(float(entr_B8_V.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['Z26'].value = my_round(float(entr_B9_V.get()), razryad)
				ws['AN11'].value = my_round(float(entr_B9_V.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['Z27'].value = my_round(float(entr_B10_V.get()), razryad)
				ws['AN12'].value = my_round(float(entr_B10_V.get()), razryad)
		if check_var_gr_1.get() >= 6:
			#QCF2
			ws['AA15'].value = 'QCF2'
			ws['AA16'].value = QC_VI
			ws['AW2'].value = 'QCF2'
			ws['AA18'].value = my_round(float(entr_B1_VI.get()), razryad)
			ws['AW3'].value = my_round(float(entr_B1_VI.get()), razryad)
			ws['AA19'].value = my_round(float(entr_B2_VI.get()), razryad)
			ws['AW4'].value = my_round(float(entr_B2_VI.get()), razryad)
			ws['AA20'].value = my_round(float(entr_B3_VI.get()), razryad)
			ws['AW5'].value = my_round(float(entr_B3_VI.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AA21'].value = my_round(float(entr_B4_VI.get()), razryad)
				ws['AW6'].value = my_round(float(entr_B4_VI.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AA22'].value = my_round(float(entr_B5_VI.get()), razryad)
				ws['AW7'].value = my_round(float(entr_B5_VI.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AA23'].value = my_round(float(entr_B6_VI.get()), razryad)
				ws['AW8'].value = my_round(float(entr_B6_VI.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AA24'].value = my_round(float(entr_B7_VI.get()), razryad)
				ws['AW9'].value = my_round(float(entr_B7_VI.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AA25'].value = my_round(float(entr_B8_VI.get()), razryad)
				ws['AW10'].value = my_round(float(entr_B8_VI.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AA26'].value = my_round(float(entr_B9_VI.get()), razryad)
				ws['AW11'].value = my_round(float(entr_B9_VI.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AA27'].value = my_round(float(entr_B10_VI.get()), razryad)
				ws['AW12'].value = my_round(float(entr_B10_VI.get()), razryad)



#QC ТАБЛИЦА № 3
	if check_var_gr_2.get() >= 3:
		#выравнивание по центру]
		cols_c(ws, 'AC15:AI34')

		#границы ячеек
		if check_var_gr_1.get() == 1:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AC15:AD27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AC15:AD28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AC15:AD29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AC15:AD30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AC15:AD31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AC15:AD32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AC15:AD33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AC15:AD34')

		if check_var_gr_1.get() == 2:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AC15:AE27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AC15:AE28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AC15:AE29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AC15:AE30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AC15:AE31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AC15:AE32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AC15:AE33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AC15:AE34')

		if check_var_gr_1.get() == 3:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AC15:AF27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AC15:AF28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AC15:AF29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AC15:AF30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AC15:AF31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AC15:AF32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AC15:AF33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AC15:AF34')				

		if check_var_gr_1.get() == 4:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AC15:AG27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AC15:AG28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AC15:AG29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AC15:AG30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AC15:AG31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AC15:AG32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AC15:AG33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AC15:AG34')			

		if check_var_gr_1.get() == 5:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AC15:AH27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AC15:AH28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AC15:AH29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AC15:AH30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AC15:AH31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AC15:AH32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AC15:AH33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AC15:AH34')		

		if check_var_gr_1.get() == 6:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AC15:AI27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AC15:AI28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AC15:AI29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AC15:AI30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AC15:AI31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AC15:AI32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AC15:AI33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AC15:AI34')	

		
		#задание ширины столбца
		#ws.column_dimensions['AC'].width = 22

		#надпись и объединение ячеек под надпись
		ws['AC15'].value = 'QC'
		ws['AC16'].value = f'Введено, {entr_concent.get()}'
		ws['AC17'].value = '№ измерения'
		ws['AD17'].value = 'Найдено, последовательность 3'
		'''
		if check_var_gr_1.get() == 1:
			ws.column_dimensions['AD'].width = 31		
		if check_var_gr_1.get() == 2:
			ws.merge_cells('AD17:AE17')
		if check_var_gr_1.get() == 3:
			ws.merge_cells('AD17:AF17')
		if check_var_gr_1.get() == 4:
			ws.merge_cells('AD17:AG17')
		if check_var_gr_1.get() == 5:
			ws.merge_cells('AD17:AH17')
		if check_var_gr_1.get() == 6:
			ws.merge_cells('AD17:AI17')
		'''

	#QCA3
		ws['AD15'].value = 'QCA3'
		ws['AD16'].value = QC_I
		ws['E2'].value = 'QCA3'
		ws['AC18'].value = '1'
		ws['AD18'].value = my_round(float(entr_C1_I.get()), razryad)
		ws['E3'].value = my_round(float(entr_C1_I.get()), razryad)
		ws['AC19'].value = '2'
		ws['AD19'].value = my_round(float(entr_C2_I.get()), razryad)
		ws['E4'].value = my_round(float(entr_C2_I.get()), razryad)
		ws['AC20'].value = '3'
		ws['AD20'].value = my_round(float(entr_C3_I.get()), razryad)
		ws['E5'].value = my_round(float(entr_C3_I.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['AC21'].value = '4'
			ws['AD21'].value = my_round(float(entr_C4_I.get()), razryad)
			ws['E6'].value = my_round(float(entr_C4_I.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['AC22'].value = '5'
			ws['AD22'].value = my_round(float(entr_C5_I.get()), razryad)
			ws['E7'].value = my_round(float(entr_C5_I.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['AC23'].value = '6'
			ws['AD23'].value = my_round(float(entr_C6_I.get()), razryad)
			ws['E8'].value = my_round(float(entr_C6_I.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['AC24'].value = '7'
			ws['AD24'].value = my_round(float(entr_C7_I.get()), razryad)
			ws['E9'].value = my_round(float(entr_C7_I.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['AC25'].value = '8'
			ws['AD25'].value = my_round(float(entr_C8_I.get()), razryad)
			ws['E10'].value = my_round(float(entr_C8_I.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['AC26'].value = '9'
			ws['AD26'].value = my_round(float(entr_C9_I.get()), razryad)
			ws['E11'].value = my_round(float(entr_C9_I.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['AC27'].value = '10'
			ws['AD27'].value = my_round(float(entr_C10_I.get()), razryad)
			ws['E12'].value = my_round(float(entr_C10_I.get()), razryad)



	#############################################################################	
	#ЗАПОЛНЕНИЕ ANOVA

		#количество таблиц зависит от показателя check_var_gr_2
		if check_var_gr_2.get() >= 3:
			#QCA3
			
			mean_QCA1_QCA2_QCA3 = my_round(stat.mean([mean_A_I, mean_B_I, mean_C_I]), check_razryad.get())

			#междугруппами
			sum_kv_mezhdy_gr_QCA1_QCA2_QCA3 = len(spisok_A_I)*((mean_A_I - mean_QCA1_QCA2_QCA3)**2) + len(spisok_B_I)*((mean_B_I - mean_QCA1_QCA2_QCA3)**2) + len(spisok_C_I)*((mean_C_I - mean_QCA1_QCA2_QCA3)**2)
			CV_mezhdy_gr_QCA1_QCA2_QCA3 = (((sum_kv_mezhdy_gr_QCA1_QCA2_QCA3)/2)**0.5)/mean_QCA1_QCA2_QCA3*100
			#внутригрупп
			CV_vnytr_gr_QCA1_QCA2_QCA3 = (mean_QCA1_QCA2_QCA3 - QC_I)/QC_I * 100


			if check_var_gr_1.get() >= 2:
				#QCB3
				
				mean_QCB1_QCB2_QCB3 = my_round(stat.mean([mean_A_II, mean_B_II, mean_C_II]), check_razryad.get())
				#междугруппами
				sum_kv_mezhdy_gr_QCB1_QCB2_QCB3 = len(spisok_A_II)*((mean_A_II - mean_QCB1_QCB2_QCB3)**2) + len(spisok_B_II)*((mean_B_II - mean_QCB1_QCB2_QCB3)**2) + len(spisok_C_II)*((mean_C_II - mean_QCB1_QCB2_QCB3)**2)
				CV_mezhdy_gr_QCB1_QCB2_QCB3 = (((sum_kv_mezhdy_gr_QCB1_QCB2_QCB3)/2)**0.5)/mean_QCB1_QCB2_QCB3*100
				#внутригрупп
				CV_vnytr_gr_QCB1_QCB2_QCB3 = (mean_QCB1_QCB2_QCB3 - QC_II)/QC_II * 100


			if check_var_gr_1.get() >= 3:
				#QCC3
				
				mean_QCC1_QCC2_QCC3 = my_round(stat.mean([mean_A_III, mean_B_III, mean_C_III]), check_razryad.get())
				#междугруппами
				sum_kv_mezhdy_gr_QCC1_QCC2_QCC3 = len(spisok_A_III)*((mean_A_III - mean_QCC1_QCC2_QCC3)**2) + len(spisok_B_III)*((mean_B_III - mean_QCC1_QCC2_QCC3)**2) + len(spisok_C_III)*((mean_C_III - mean_QCC1_QCC2_QCC3)**2)
				CV_mezhdy_gr_QCC1_QCC2_QCC3 = (((sum_kv_mezhdy_gr_QCC1_QCC2_QCC3)/2)**0.5)/mean_QCC1_QCC2_QCC3*100
				#внутригрупп
				CV_vnytr_gr_QCC1_QCC2_QCC3 = (mean_QCC1_QCC2_QCC3 - QC_III)/QC_III * 100

			if check_var_gr_1.get() >= 4:
				#QCD3
				
				mean_QCD1_QCD2_QCD3 = my_round(stat.mean([mean_A_IV, mean_B_IV, mean_C_IV]), check_razryad.get())
				#междугруппами
				sum_kv_mezhdy_gr_QCD1_QCD2_QCD3 = len(spisok_A_IV)*((mean_A_IV - mean_QCD1_QCD2_QCD3)**2) + len(spisok_B_IV)*((mean_B_IV - mean_QCD1_QCD2_QCD3)**2) + len(spisok_C_IV)*((mean_C_IV - mean_QCD1_QCD2_QCD3)**2)
				CV_mezhdy_gr_QCD1_QCD2_QCD3 = (((sum_kv_mezhdy_gr_QCD1_QCD2_QCD3)/2)**0.5)/mean_QCD1_QCD2_QCD3*100
				#внутригрупп
				CV_vnytr_gr_QCD1_QCD2_QCD3 = (mean_QCD1_QCD2_QCD3 - QC_IV)/QC_IV * 100

			if check_var_gr_1.get() >= 5:
				#QCE3
				
				mean_QCE1_QCE2_QCE3 = my_round(stat.mean([mean_A_V, mean_B_V, mean_C_V]), check_razryad.get())
				#междугруппами
				sum_kv_mezhdy_gr_QCE1_QCE2_QCE3 = len(spisok_A_V)*((mean_A_V - mean_QCE1_QCE2_QCE3)**2) + len(spisok_B_V)*((mean_B_V - mean_QCE1_QCE2_QCE3)**2) + len(spisok_C_V)*((mean_C_V - mean_QCE1_QCE2_QCE3)**2)
				CV_mezhdy_gr_QCE1_QCE2_QCE3 = (((sum_kv_mezhdy_gr_QCE1_QCE2_QCE3)/2)**0.5)/mean_QCE1_QCE2_QCE3*100
				#внутригрупп
				CV_vnytr_gr_QCE1_QCE2_QCE3 = (mean_QCE1_QCE2_QCE3 - QC_V)/QC_V * 100

			if check_var_gr_1.get() >= 6:
				#QCF3
				
				mean_QCF1_QCF2_QCF3 = my_round(stat.mean([mean_A_VI, mean_B_VI, mean_C_VI]), check_razryad.get())
				#междугруппами
				sum_kv_mezhdy_gr_QCF1_QCF2_QCF3 = len(spisok_A_VI)*((mean_A_VI - mean_QCF1_QCF2_QCF3)**2) + len(spisok_B_VI)*((mean_B_VI - mean_QCF1_QCF2_QCF3)**2) + len(spisok_C_VI)*((mean_C_VI - mean_QCF1_QCF2_QCF3)**2)
				CV_mezhdy_gr_QCF1_QCF2_QCF3 = (((sum_kv_mezhdy_gr_QCF1_QCF2_QCF3)/2)**0.5)/mean_QCF1_QCF2_QCF3*100
				#внутригрупп
				CV_vnytr_gr_QCF1_QCF2_QCF3 = (mean_QCF1_QCF2_QCF3 - QC_VI)/QC_VI * 100



	#вывод среднего и остальных показателей в таблицу 
		if check_var_gr_3.get() == 3:

			ws['AC21'].value = 'найдено, сред.зн.,\nнг/мл (n=3)'
			ws['AC21'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

			ws['AC22'].value = 'εR, % (n=3)'
			ws['AC23'].value = 'σR, % (n=3)'

			ws['AC24'].value = 'найдено, сред.зн.,\nнг/мл (n=9)'
			ws['AC24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC25'].value = 'εR, % (n=9)'
			ws['AC26'].value = 'σR, % (n=9)'
			ws['AC27'].value = 'Норма |ε| и |σ|, %'


			ws['AD21'].value = my_round(mean_C_I, razryad)
			ws['AD22'].value = my_round(Er_C_I, 1)
			ws['AD23'].value = my_round(sigma_C_I, 1)

			ws['AD24'].value = my_round(mean_QCA1_QCA2_QCA3, razryad)
			ws['AD25'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3, 1)
			ws['AD26'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AD27'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AD27'].value = '≤20'

		
			if check_var_gr_1.get() >= 2:
				ws['AE21'].value = my_round(mean_C_II, razryad)
				ws['AE22'].value = my_round(Er_C_II, 1)
				ws['AE23'].value = my_round(sigma_C_II, 1)

				ws['AE24'].value = my_round(mean_QCB1_QCB2_QCB3, razryad)
				ws['AE25'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3, 1)
				ws['AE26'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AE27'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AE27'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AF21'].value = my_round(mean_C_III, razryad)
				ws['AF22'].value = my_round(Er_C_III, 1)
				ws['AF23'].value = my_round(sigma_C_III, 1)

				ws['AF24'].value = my_round(mean_QCC1_QCC2_QCC3, razryad)
				ws['AF25'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3, 1)
				ws['AF26'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AF27'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AF27'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AG21'].value = my_round(mean_C_IV, razryad)
				ws['AG22'].value = my_round(Er_C_IV, 1)
				ws['AG23'].value = my_round(sigma_C_IV, 1)

				ws['AG24'].value = my_round(mean_QCD1_QCD2_QCD3, razryad)
				ws['AG25'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3, 1)
				ws['AG26'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AG27'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AG27'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AH21'].value = my_round(mean_C_V, razryad)
				ws['AH22'].value = my_round(Er_C_V, 1)
				ws['AH23'].value = my_round(sigma_C_V, 1)

				ws['AH24'].value = my_round(mean_QCE1_QCE2_QCE3, razryad)
				ws['AH25'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3, 1)
				ws['AH26'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AH27'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AH27'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AI21'].value = my_round(mean_C_VI, razryad)
				ws['AI22'].value = my_round(Er_C_VI, 1)
				ws['AI23'].value = my_round(sigma_C_VI, 1)

				ws['AI24'].value = my_round(mean_QCF1_QCF2_QCF3, razryad)
				ws['AI25'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3, 1)
				ws['AI26'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AI27'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AI27'].value = '≤15'


		if check_var_gr_3.get() == 4:

			ws['AC22'].value = 'найдено, сред.зн.,\nнг/мл (n=4)'
			ws['AC22'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC23'].value = 'εR, % (n=4)'
			ws['AC24'].value = 'σR, % (n=4)'
			
			ws['AC25'].value = 'найдено, сред.зн.,\nнг/мл (n=12)'
			ws['AC25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC26'].value = 'εR, % (n=12)'
			ws['AC27'].value = 'σR, % (n=12)'
			ws['AC28'].value = 'Норма |ε| и |σ|, %'


			ws['AD22'].value = my_round(mean_C_I, razryad)
			ws['AD23'].value = my_round(Er_C_I, 1)
			ws['AD24'].value = my_round(sigma_C_I, 1)

			ws['AD25'].value = my_round(mean_QCA1_QCA2_QCA3, razryad)
			ws['AD26'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3, 1)
			ws['AD27'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AD28'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AD28'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AE22'].value = my_round(mean_C_II, razryad)
				ws['AE23'].value = my_round(Er_C_II, 1)
				ws['AE24'].value = my_round(sigma_C_II, 1)

				ws['AE25'].value = my_round(mean_QCB1_QCB2_QCB3, razryad)
				ws['AE26'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3, 1)
				ws['AE27'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AE28'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AE28'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AF22'].value = my_round(mean_C_III, razryad)
				ws['AF23'].value = my_round(Er_C_III, 1)
				ws['AF24'].value = my_round(sigma_C_III, 1)

				ws['AF25'].value = my_round(mean_QCC1_QCC2_QCC3, razryad)
				ws['AF26'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3, 1)
				ws['AF27'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AF28'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AF28'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AG22'].value = my_round(mean_C_IV, razryad)
				ws['AG23'].value = my_round(Er_C_IV, 1)
				ws['AG24'].value = my_round(sigma_C_IV, 1)

				ws['AG25'].value = my_round(mean_QCD1_QCD2_QCD3, razryad)
				ws['AG26'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3, 1)
				ws['AG27'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AG28'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AG28'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AH22'].value = my_round(mean_C_V, razryad)
				ws['AH23'].value = my_round(Er_C_V, 1)
				ws['AH24'].value = my_round(sigma_C_V, 1)

				ws['AH25'].value = my_round(mean_QCE1_QCE2_QCE3, razryad)
				ws['AH26'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3, 1)
				ws['AH27'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AH28'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AH28'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AI22'].value = my_round(mean_C_VI, razryad)
				ws['AI23'].value = my_round(Er_C_VI, 1)
				ws['AI24'].value = my_round(sigma_C_VI, 1)

				ws['AI25'].value = my_round(mean_QCF1_QCF2_QCF3, razryad)
				ws['AI26'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3, 1)
				ws['AI27'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AI28'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AI28'].value = '≤15'


		if check_var_gr_3.get() == 5:
			ws['AC23'].value = 'найдено, сред.зн., \nнг/мл (n=5)'
			ws['AC23'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC24'].value = 'εR, % (n=5)'
			ws['AC25'].value = 'σR, % (n=5)'
			
			ws['AC26'].value = 'найдено, сред.зн.,\nнг/мл (n=15)'
			ws['AC26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC27'].value = 'εR, % (n=15)'
			ws['AC28'].value = 'σR, % (n=15)'
			ws['AC29'].value = 'Норма |ε| и |σ|, %'

			ws['AD23'].value = my_round(mean_C_I, razryad)
			ws['AD24'].value = my_round(Er_C_I, 1)
			ws['AD25'].value = my_round(sigma_C_I, 1)

			ws['AD26'].value = my_round(mean_QCA1_QCA2_QCA3, razryad)
			ws['AD27'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3, 1)
			ws['AD28'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AD29'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AD29'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AE23'].value = my_round(mean_C_II, razryad)
				ws['AE24'].value = my_round(Er_C_II, 1)
				ws['AE25'].value = my_round(sigma_C_II, 1)

				ws['AE26'].value = my_round(mean_QCB1_QCB2_QCB3, razryad)
				ws['AE27'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3, 1)
				ws['AE28'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AE29'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AE29'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AF23'].value = my_round(mean_C_III, razryad)
				ws['AF24'].value = my_round(Er_C_III, 1)
				ws['AF25'].value = my_round(sigma_C_III, 1)

				ws['AF26'].value = my_round(mean_QCC1_QCC2_QCC3, razryad)
				ws['AF27'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3, 1)
				ws['AF28'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AF29'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AF29'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AG23'].value = my_round(mean_C_IV, razryad)
				ws['AG24'].value = my_round(Er_C_IV, 1)
				ws['AG25'].value = my_round(sigma_C_IV, 1)

				ws['AG26'].value = my_round(mean_QCD1_QCD2_QCD3, razryad)
				ws['AG27'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3, 1)
				ws['AG28'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AG29'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AG29'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AH23'].value = my_round(mean_C_V, razryad)
				ws['AH24'].value = my_round(Er_C_V, 1)
				ws['AH25'].value = my_round(sigma_C_V, 1)

				ws['AH26'].value = my_round(mean_QCE1_QCE2_QCE3, razryad)
				ws['AH27'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3, 1)
				ws['AH28'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AH29'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AH29'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AI23'].value = my_round(mean_C_VI, razryad)
				ws['AI24'].value = my_round(Er_C_VI, 1)
				ws['AI25'].value = my_round(sigma_C_VI, 1)

				ws['AI26'].value = my_round(mean_QCF1_QCF2_QCF3, razryad)
				ws['AI27'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3, 1)
				ws['AI28'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AI29'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AI29'].value = '≤15'


		if check_var_gr_3.get() == 6:
			ws['AC24'].value = 'найдено, сред.зн., \nнг/мл (n=6)'
			ws['AC24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC25'].value = 'εR, % (n=6)'
			ws['AC26'].value = 'σR, % (n=6)'
			
			ws['AC27'].value = 'найдено, сред.зн.,\nнг/мл (n=18)'
			ws['AC27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC28'].value = 'εR, % (n=18)'
			ws['AC29'].value = 'σR, % (n=18)'
			ws['AC30'].value = 'Норма |ε| и |σ|, %'

			ws['AD24'].value = my_round(mean_C_I, razryad)
			ws['AD25'].value = my_round(Er_C_I, 1)
			ws['AD26'].value = my_round(sigma_C_I, 1)

			ws['AD27'].value = my_round(mean_QCA1_QCA2_QCA3, razryad)
			ws['AD28'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3, 1)
			ws['AD29'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AD30'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AD30'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AE24'].value = my_round(mean_C_II, razryad)
				ws['AE25'].value = my_round(Er_C_II, 1)
				ws['AE26'].value = my_round(sigma_C_II, 1)

				ws['AE27'].value = my_round(mean_QCB1_QCB2_QCB3, razryad)
				ws['AE28'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3, 1)
				ws['AE29'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AE30'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AE30'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AF24'].value = my_round(mean_C_III, razryad)
				ws['AF25'].value = my_round(Er_C_III, 1)
				ws['AF26'].value = my_round(sigma_C_III, 1)

				ws['AF27'].value = my_round(mean_QCC1_QCC2_QCC3, razryad)
				ws['AF28'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3, 1)
				ws['AF29'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AF30'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AF30'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AG24'].value = my_round(mean_C_IV, razryad)
				ws['AG25'].value = my_round(Er_C_IV, 1)
				ws['AG26'].value = my_round(sigma_C_IV, 1)

				ws['AG27'].value = my_round(mean_QCD1_QCD2_QCD3, razryad)
				ws['AG28'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3, 1)
				ws['AG29'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AG30'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AG30'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AH24'].value = my_round(mean_C_V, razryad)
				ws['AH25'].value = my_round(Er_C_V, 1)
				ws['AH26'].value = my_round(sigma_C_V, 1)

				ws['AH27'].value = my_round(mean_QCE1_QCE2_QCE3, razryad)
				ws['AH28'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3, 1)
				ws['AH29'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AH30'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AH30'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AI24'].value = my_round(mean_C_VI, razryad)
				ws['AI25'].value = my_round(Er_C_VI, 1)
				ws['AI26'].value = my_round(sigma_C_VI, 1)

				ws['AI27'].value = my_round(mean_QCF1_QCF2_QCF3, razryad)
				ws['AI28'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3, 1)
				ws['AI29'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AI30'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AI30'].value = '≤15'

		if check_var_gr_3.get() == 7:
			ws['AC25'].value = 'найдено, сред.зн., \nнг/мл (n=7)'
			ws['AC25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC26'].value = 'εR, % (n=7)'
			ws['AC27'].value = 'σR, % (n=7)'
			
			ws['AC28'].value = 'найдено, сред.зн.,\nнг/мл (n=21)'
			ws['AC28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC29'].value = 'εR, % (n=21)'
			ws['AC30'].value = 'σR, % (n=21)'
			ws['AC31'].value = 'Норма |ε| и |σ|, %'

			ws['AD25'].value = my_round(mean_C_I, razryad)
			ws['AD26'].value = my_round(Er_C_I, 1)
			ws['AD27'].value = my_round(sigma_C_I, 1)

			ws['AD28'].value = my_round(mean_QCA1_QCA2_QCA3, razryad)
			ws['AD29'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3, 1)
			ws['AD30'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AD31'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AD31'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AE25'].value = my_round(mean_C_II, razryad)
				ws['AE26'].value = my_round(Er_C_II, 1)
				ws['AE27'].value = my_round(sigma_C_II, 1)

				ws['AE28'].value = my_round(mean_QCB1_QCB2_QCB3, razryad)
				ws['AE29'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3, 1)
				ws['AE30'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AE31'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AE31'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AF25'].value = my_round(mean_C_III, razryad)
				ws['AF26'].value = my_round(Er_C_III, 1)
				ws['AF27'].value = my_round(sigma_C_III, 1)

				ws['AF28'].value = my_round(mean_QCC1_QCC2_QCC3, razryad)
				ws['AF29'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3, 1)
				ws['AF30'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AF31'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AF31'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AG25'].value = my_round(mean_C_IV, razryad)
				ws['AG26'].value = my_round(Er_C_IV, 1)
				ws['AG27'].value = my_round(sigma_C_IV, 1)

				ws['AG28'].value = my_round(mean_QCD1_QCD2_QCD3, razryad)
				ws['AG29'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3, 1)
				ws['AG30'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AG31'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AG31'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AH25'].value = my_round(mean_C_V, razryad)
				ws['AH26'].value = my_round(Er_C_V, 1)
				ws['AH27'].value = my_round(sigma_C_V, 1)

				ws['AH28'].value = my_round(mean_QCE1_QCE2_QCE3, razryad)
				ws['AH29'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3, 1)
				ws['AH30'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AH31'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AH31'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AI25'].value = my_round(mean_C_VI, razryad)
				ws['AI26'].value = my_round(Er_C_VI, 1)
				ws['AI27'].value = my_round(sigma_C_VI, 1)

				ws['AI28'].value = my_round(mean_QCF1_QCF2_QCF3, razryad)
				ws['AI29'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3, 1)
				ws['AI30'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AI31'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AI31'].value = '≤15'

		if check_var_gr_3.get() == 8:
			ws['AC26'].value = 'найдено, сред.зн., \nнг/мл (n=8)'
			ws['AC26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC27'].value = 'εR, % (n=8)'
			ws['AC28'].value = 'σR, % (n=8)'
			
			ws['AC29'].value = 'найдено, сред.зн.,\nнг/мл (n=24)'
			ws['AC29'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC30'].value = 'εR, % (n=24)'
			ws['AC30'].value = 'σR, % (n=24)'
			ws['AC31'].value = 'Норма |ε| и |σ|, %'

			ws['AD26'].value = my_round(mean_C_I, razryad)
			ws['AD27'].value = my_round(Er_C_I, 1)
			ws['AD28'].value = my_round(sigma_C_I, 1)

			ws['AD29'].value = my_round(mean_QCA1_QCA2_QCA3, razryad)
			ws['AD30'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3, 1)
			ws['AD31'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3, 1)

			if entr_NORM_QCA1.get() != '':
				ws['AD32'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AD32'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AE26'].value = my_round(mean_C_II, razryad)
				ws['AE27'].value = my_round(Er_C_II, 1)
				ws['AE28'].value = my_round(sigma_C_II, 1)

				ws['AE29'].value = my_round(mean_QCB1_QCB2_QCB3, razryad)
				ws['AE30'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3, 1)
				ws['AE31'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AE32'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AE32'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AF26'].value = my_round(mean_C_III, razryad)
				ws['AF27'].value = my_round(Er_C_III, 1)
				ws['AF28'].value = my_round(sigma_C_III, 1)

				ws['AF29'].value = my_round(mean_QCC1_QCC2_QCC3, razryad)
				ws['AF30'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3, 1)
				ws['AF31'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AF32'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AF32'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AG26'].value = my_round(mean_C_IV, razryad)
				ws['AG27'].value = my_round(Er_C_IV, 1)
				ws['AG28'].value = my_round(sigma_C_IV, 1)

				ws['AG29'].value = my_round(mean_QCD1_QCD2_QCD3, razryad)
				ws['AG30'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3, 1)
				ws['AG31'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AG32'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AG32'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AH26'].value = my_round(mean_C_V, razryad)
				ws['AH27'].value = my_round(Er_C_V, 1)
				ws['AH28'].value = my_round(sigma_C_V, 1)

				ws['AH29'].value = my_round(mean_QCE1_QCE2_QCE3, razryad)
				ws['AH30'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3, 1)
				ws['AH31'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AH32'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AH32'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AI26'].value = my_round(mean_C_VI, razryad)
				ws['AI27'].value = my_round(Er_C_VI, 1)
				ws['AI28'].value = my_round(sigma_C_VI, 1)

				ws['AI29'].value = my_round(mean_QCF1_QCF2_QCF3, razryad)
				ws['AI30'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3, 1)
				ws['AI31'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AI32'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AI32'].value = '≤15'

		if check_var_gr_3.get() == 9:
			ws['AC27'].value = 'найдено, сред.зн., \nнг/мл (n=9)'
			ws['AC27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC28'].value = 'εR, % (n=9)'
			ws['AC29'].value = 'σR, % (n=9)'
			
			ws['AC30'].value = 'найдено, сред.зн.,\nнг/мл (n=27)'
			ws['AC30'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC31'].value = 'εR, % (n=27)'
			ws['AC32'].value = 'σR, % (n=27)'
			ws['AC33'].value = 'Норма |ε| и |σ|, %'

			ws['AD27'].value = my_round(mean_C_I, razryad)
			ws['AD28'].value = my_round(Er_C_I, 1)
			ws['AD29'].value = my_round(sigma_C_I, 1)

			ws['AD30'].value = my_round(mean_QCA1_QCA2_QCA3, razryad)
			ws['AD31'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3, 1)
			ws['AD32'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AD33'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AD33'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AE27'].value = my_round(mean_C_II, razryad)
				ws['AE28'].value = my_round(Er_C_II, 1)
				ws['AE29'].value = my_round(sigma_C_II, 1)

				ws['AE30'].value = my_round(mean_QCB1_QCB2_QCB3, razryad)
				ws['AE31'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3, 1)
				ws['AE32'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3, 1)

				if entr_NORM_QCB1.get() != '':
					ws['AE33'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AE33'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AF27'].value = my_round(mean_C_III, razryad)
				ws['AF28'].value = my_round(Er_C_III, 1)
				ws['AF29'].value = my_round(sigma_C_III, 1)

				ws['AF30'].value = my_round(mean_QCC1_QCC2_QCC3, razryad)
				ws['AF31'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3, 1)
				ws['AF32'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AF33'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AF33'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AG27'].value = my_round(mean_C_IV, razryad)
				ws['AG28'].value = my_round(Er_C_IV, 1)
				ws['AG29'].value = my_round(sigma_C_IV, 1)

				ws['AG30'].value = my_round(mean_QCD1_QCD2_QCD3, razryad)
				ws['AG31'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3, 1)
				ws['AG32'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AG33'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AG33'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AH27'].value = my_round(mean_C_V, razryad)
				ws['AH28'].value = my_round(Er_C_V, 1)
				ws['AH29'].value = my_round(sigma_C_V, 1)

				ws['AH30'].value = my_round(mean_QCE1_QCE2_QCE3, razryad)
				ws['AH31'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3, 1)
				ws['AH32'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AH33'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AH33'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AI27'].value = my_round(mean_C_VI, razryad)
				ws['AI28'].value = my_round(Er_C_VI, 1)
				ws['AI29'].value = my_round(sigma_C_VI, 1)

				ws['AI30'].value = my_round(mean_QCF1_QCF2_QCF3, razryad)
				ws['AI31'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3, 1)
				ws['AI32'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AI33'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AI33'].value = '≤15'


		if check_var_gr_3.get() == 10:
			ws['AC28'].value = 'найдено, сред.зн., \nнг/мл (n=10)'
			ws['AC28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC29'].value = 'εR, % (n=10)'
			ws['AC30'].value = 'σR, % (n=10)'
			
			ws['AC31'].value = 'найдено, сред.зн.,\nнг/мл (n=30)'
			ws['AC31'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AC32'].value = 'εR, % (n=30)'
			ws['AC33'].value = 'σR, % (n=30)'
			ws['AC34'].value = 'Норма |ε| и |σ|, %'

			ws['AD28'].value = my_round(mean_C_I, razryad)
			ws['AD29'].value = my_round(Er_C_I, 1)
			ws['AD30'].value = my_round(sigma_C_I, 1)

			ws['AD31'].value = my_round(mean_QCA1_QCA2_QCA3, razryad)
			ws['AD32'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3, 1)
			ws['AD33'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AD34'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AD34'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AE28'].value = my_round(mean_C_II, razryad)
				ws['AE29'].value = my_round(Er_C_II, 1)
				ws['AE30'].value = my_round(sigma_C_II, 1)

				ws['AE31'].value = my_round(mean_QCB1_QCB2_QCB3, razryad)
				ws['AE32'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3, 1)
				ws['AE33'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AE34'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AE34'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AF28'].value = my_round(mean_C_III, razryad)
				ws['AF29'].value = my_round(Er_C_III, 1)
				ws['AF30'].value = my_round(sigma_C_III, 1)

				ws['AF31'].value = my_round(mean_QCC1_QCC2_QCC3, razryad)
				ws['AF32'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3, 1)
				ws['AF33'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AF34'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AF34'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AG28'].value = my_round(mean_C_IV, razryad)
				ws['AG29'].value = my_round(Er_C_IV, 1)
				ws['AG30'].value = my_round(sigma_C_IV, 1)

				ws['AG31'].value = my_round(mean_QCD1_QCD2_QCD3, razryad)
				ws['AG32'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3, 1)
				ws['AG33'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AG34'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AG34'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AH28'].value = my_round(mean_C_V, razryad)
				ws['AH29'].value = my_round(Er_C_V, 1)
				ws['AH30'].value = my_round(sigma_C_V, 1)

				ws['AH31'].value = my_round(mean_QCE1_QCE2_QCE3, razryad)
				ws['AH32'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3, 1)
				ws['AH33'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AH34'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AH34'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AI28'].value = my_round(mean_C_VI, razryad)
				ws['AI29'].value = my_round(Er_C_VI, 1)
				ws['AI30'].value = my_round(sigma_C_VI, 1)

				ws['AI31'].value = my_round(mean_QCF1_QCF2_QCF3, razryad)
				ws['AI32'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3, 1)
				ws['AI33'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AI34'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AI34'].value = '≤15'




		if check_var_gr_1.get() >= 2:
			#QCB3
			ws['AE15'].value = 'QCB3'
			ws['AE16'].value = QC_II
			ws['N2'].value = 'QCB3'
			ws['AE18'].value = my_round(float(entr_C1_II.get()), razryad)
			ws['N3'].value = my_round(float(entr_C1_II.get()), razryad)
			ws['AE19'].value = my_round(float(entr_C2_II.get()), razryad)
			ws['N4'].value = my_round(float(entr_C2_II.get()), razryad)
			ws['AE20'].value = my_round(float(entr_C3_II.get()), razryad)
			ws['N5'].value = my_round(float(entr_C3_II.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AE21'].value = my_round(float(entr_C4_II.get()), razryad)
				ws['N6'].value = my_round(float(entr_C4_II.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AE22'].value = my_round(float(entr_C5_II.get()), razryad)
				ws['N7'].value = my_round(float(entr_C5_II.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AE23'].value = my_round(float(entr_C6_II.get()), razryad)
				ws['N8'].value = my_round(float(entr_C6_II.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AE24'].value = my_round(float(entr_C7_II.get()), razryad)
				ws['N9'].value = my_round(float(entr_C7_II.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AE25'].value = my_round(float(entr_C8_II.get()), razryad)
				ws['N10'].value = my_round(float(entr_C8_II.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AE26'].value = my_round(float(entr_C9_II.get()), razryad)
				ws['N11'].value = my_round(float(entr_C9_II.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AE27'].value = my_round(float(entr_C10_II.get()), razryad)
				ws['N12'].value = my_round(float(entr_C10_II.get()), razryad)

		if check_var_gr_1.get() >= 3:
			#QCC3
			ws['AF15'].value = 'QCC3'
			ws['AF16'].value = QC_III
			ws['W2'].value = 'QCC3'
			ws['AF18'].value = my_round(float(entr_C1_III.get()), razryad)
			ws['W3'].value = my_round(float(entr_C1_III.get()), razryad)
			ws['AF19'].value = my_round(float(entr_C2_III.get()), razryad)
			ws['W4'].value = my_round(float(entr_C2_III.get()), razryad)
			ws['AF20'].value = my_round(float(entr_C3_III.get()), razryad)
			ws['W5'].value = my_round(float(entr_C3_III.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AF21'].value = my_round(float(entr_C4_III.get()), razryad)
				ws['W6'].value = my_round(float(entr_C4_III.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AF22'].value = my_round(float(entr_C5_III.get()), razryad)
				ws['W7'].value = my_round(float(entr_C5_III.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AF23'].value = my_round(float(entr_C6_III.get()), razryad)
				ws['W8'].value = my_round(float(entr_C6_III.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AF24'].value = my_round(float(entr_C7_III.get()), razryad)
				ws['W9'].value = my_round(float(entr_C7_III.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AF25'].value = my_round(float(entr_C8_III.get()), razryad)
				ws['W10'].value = my_round(float(entr_C8_III.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AF26'].value = my_round(float(entr_C9_III.get()), razryad)
				ws['W11'].value = my_round(float(entr_C9_III.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AF27'].value = my_round(float(entr_C10_III.get()), razryad)
				ws['W12'].value = my_round(float(entr_C10_III.get()), razryad)

		if check_var_gr_1.get() >= 4:
			#QCD2
			ws['AG15'].value = 'QCD3'
			ws['AG16'].value = QC_IV
			ws['AF2'].value = 'QCD3'
			ws['AG18'].value = my_round(float(entr_C1_IV.get()), razryad)
			ws['AF3'].value = my_round(float(entr_C1_IV.get()), razryad)
			ws['AG19'].value = my_round(float(entr_C2_IV.get()), razryad)
			ws['AF4'].value = my_round(float(entr_C2_IV.get()), razryad)
			ws['AG20'].value = my_round(float(entr_C3_IV.get()), razryad)
			ws['AF5'].value = my_round(float(entr_C3_IV.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AG21'].value = my_round(float(entr_C4_IV.get()), razryad)
				ws['AF6'].value = my_round(float(entr_C4_IV.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AG22'].value = my_round(float(entr_C5_IV.get()), razryad)
				ws['AF7'].value = my_round(float(entr_C5_IV.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AG23'].value = my_round(float(entr_C6_IV.get()), razryad)
				ws['AF8'].value = my_round(float(entr_C6_IV.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AG24'].value = my_round(float(entr_C7_IV.get()), razryad)
				ws['AF9'].value = my_round(float(entr_C7_IV.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AG25'].value = my_round(float(entr_C8_IV.get()), razryad)
				ws['AF10'].value = my_round(float(entr_C8_IV.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AG26'].value = my_round(float(entr_C9_IV.get()), razryad)
				ws['AF11'].value = my_round(float(entr_C9_IV.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AG27'].value = my_round(float(entr_C10_IV.get()), razryad)
				ws['AF12'].value = my_round(float(entr_C10_IV.get()), razryad)

		if check_var_gr_1.get() >= 5:
			#QCE2
			ws['AH15'].value = 'QCE3'
			ws['AH16'].value = QC_V
			ws['AO2'].value = 'QCE3'
			ws['AH18'].value = my_round(float(entr_C1_V.get()), razryad)
			ws['AO3'].value = my_round(float(entr_C1_V.get()), razryad)
			ws['AH19'].value = my_round(float(entr_C2_V.get()), razryad)
			ws['AO4'].value = my_round(float(entr_C2_V.get()), razryad)
			ws['AH20'].value = my_round(float(entr_C3_V.get()), razryad)
			ws['AO5'].value = my_round(float(entr_C3_V.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AH21'].value = my_round(float(entr_C4_V.get()), razryad)
				ws['AO6'].value = my_round(float(entr_C4_V.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AH22'].value = my_round(float(entr_C5_V.get()), razryad)
				ws['AO7'].value = my_round(float(entr_C5_V.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AH23'].value = my_round(float(entr_C6_V.get()), razryad)
				ws['AO8'].value = my_round(float(entr_C6_V.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AH24'].value = my_round(float(entr_C7_V.get()), razryad)
				ws['AO9'].value = my_round(float(entr_C7_V.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AH25'].value = my_round(float(entr_C8_V.get()), razryad)
				ws['AO10'].value = my_round(float(entr_C8_V.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AH26'].value = my_round(float(entr_C9_V.get()), razryad)
				ws['AO11'].value = my_round(float(entr_C9_V.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AH27'].value = my_round(float(entr_C10_V.get()), razryad)
				ws['AO12'].value = my_round(float(entr_C10_V.get()), razryad)

		if check_var_gr_1.get() >= 6:
			#QCF2
			ws['AI15'].value = 'QCF3'
			ws['AI16'].value = QC_VI
			ws['AX2'].value = 'QCF3'
			ws['AI18'].value = my_round(float(entr_C1_VI.get()), razryad)
			ws['AX3'].value = my_round(float(entr_C1_VI.get()), razryad)
			ws['AI19'].value = my_round(float(entr_C2_VI.get()), razryad)
			ws['AX4'].value = my_round(float(entr_C2_VI.get()), razryad)
			ws['AI20'].value = my_round(float(entr_C3_VI.get()), razryad)
			ws['AX5'].value = my_round(float(entr_C3_VI.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AI21'].value = my_round(float(entr_C4_VI.get()), razryad)
				ws['AX6'].value = my_round(float(entr_C4_VI.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AI22'].value = my_round(float(entr_C5_VI.get()), razryad)
				ws['AX7'].value = my_round(float(entr_C5_VI.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AI23'].value = my_round(float(entr_C6_VI.get()), razryad)
				ws['AX8'].value = my_round(float(entr_C6_VI.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AI24'].value = my_round(float(entr_C7_VI.get()), razryad)
				ws['AX9'].value = my_round(float(entr_C7_VI.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AI25'].value = my_round(float(entr_C8_VI.get()), razryad)
				ws['AX10'].value = my_round(float(entr_C8_VI.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AI26'].value = my_round(float(entr_C9_VI.get()), razryad)
				ws['AX11'].value = my_round(float(entr_C9_VI.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AI27'].value = my_round(float(entr_C10_VI.get()), razryad)
				ws['AX12'].value = my_round(float(entr_C10_VI.get()), razryad)


#QC ТАБЛИЦА № 4
	if check_var_gr_2.get() >= 4:
	#выравнивание по центру]
		cols_c(ws, 'AK15:AQ34')

		#границы ячеек
		if check_var_gr_1.get() == 1:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AK15:AL27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AK15:AL28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AK15:AL29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AK15:AL30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AK15:AL31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AK15:AL32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AK15:AL33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AK15:AL34')

		if check_var_gr_1.get() == 2:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AK15:AM27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AK15:AM28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AK15:AM29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AK15:AM30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AK15:AM31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AK15:AM32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AK15:AM33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AK15:AM34')

		if check_var_gr_1.get() == 3:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AK15:AN27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AK15:AN28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AK15:AN29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AK15:AN30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AK15:AN31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AK15:AN32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AK15:AN33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AK15:AN34')


		if check_var_gr_1.get() == 4:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AK15:AO27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AK15:AO28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AK15:AO29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AK15:AO30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AK15:AO31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AK15:AO32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AK15:AO33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AK15:AO34')			

		if check_var_gr_1.get() == 5:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AK15:AP27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AK15:AP28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AK15:AP29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AK15:AP30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AK15:AP31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AK15:AP32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AK15:AP33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AK15:AP34')		

		if check_var_gr_1.get() == 6:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AK15:AQ27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AK15:AQ28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AK15:AQ29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AK15:AQ30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AK15:AQ31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AK15:AQ32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AK15:AQ33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AK15:AQ34')	

		
		#задание ширины столбца
		#ws.column_dimensions['AK'].width = 22

		#надпись и объединение ячеек под надпись
		ws['AK15'].value = 'QC'
		ws['AK16'].value = f'Введено, {entr_concent.get()}'
		ws['AK17'].value = '№ измерения'
		ws['AL17'].value = 'Найдено, последовательность 4'

		'''
		if check_var_gr_1.get() == 1:
			ws.column_dimensions['AL'].width = 31		
		if check_var_gr_1.get() == 2:
			ws.merge_cells('AL17:AM17')
		if check_var_gr_1.get() == 3:
			ws.merge_cells('AL17:AN17')
		if check_var_gr_1.get() == 4:
			ws.merge_cells('AL17:AO17')
		if check_var_gr_1.get() == 5:
			ws.merge_cells('AL17:AP17')
		if check_var_gr_1.get() == 6:
			ws.merge_cells('AL17:AQ17')
		'''





	#############################################################################	
	#ЗАПОЛНЕНИЕ ANOVA

		#количество таблиц зависит от показателя check_var_gr_2
		if check_var_gr_2.get() >= 4:
			#QCA4
			mean_QCA1_QCA2_QCA3_QCA4 = my_round(stat.mean([mean_A_I, mean_B_I, mean_C_I, mean_D_I]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCA1_QCA2_QCA3_QCA4 = len(spisok_A_I)*((mean_A_I - mean_QCA1_QCA2_QCA3_QCA4)**2) + len(spisok_B_I)*((mean_B_I - mean_QCA1_QCA2_QCA3_QCA4)**2) + len(spisok_C_I)*(mean_C_I - mean_QCA1_QCA2_QCA3_QCA4)**2 + len(spisok_D_I)*((mean_D_I - mean_QCA1_QCA2_QCA3_QCA4)**2)
			CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4 = (((sum_kv_mezhdy_gr_QCA1_QCA2_QCA3_QCA4)/3)**0.5)/mean_QCA1_QCA2_QCA3_QCA4*100
			#внутригрупп
			CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4 = (mean_QCA1_QCA2_QCA3_QCA4 - QC_I)/QC_I * 100

			if check_var_gr_1.get() >= 2:
				#QCB4
				mean_QCB1_QCB2_QCB3_QCB4 = my_round(stat.mean([mean_A_II, mean_B_II, mean_C_II, mean_D_II]), check_razryad.get())
				#междугруппами
				sum_kv_mezhdy_gr_QCB1_QCB2_QCB3_QCB4 = len(spisok_A_II)*((mean_A_II - mean_QCB1_QCB2_QCB3_QCB4)**2) + len(spisok_B_II)*((mean_B_II - mean_QCB1_QCB2_QCB3_QCB4)**2) + len(spisok_C_II)*(mean_C_II- mean_QCB1_QCB2_QCB3_QCB4)**2 + len(spisok_D_II)*((mean_D_II - mean_QCB1_QCB2_QCB3_QCB4)**2)
				CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4 = (((sum_kv_mezhdy_gr_QCB1_QCB2_QCB3_QCB4)/3)**0.5)/mean_QCB1_QCB2_QCB3_QCB4*100
				#внутригрупп
				CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4 = (mean_QCB1_QCB2_QCB3_QCB4 - QC_II)/QC_II * 100

			if check_var_gr_1.get() >= 3:
				#QCC4
				mean_QCC1_QCC2_QCC3_QCC4 = my_round(stat.mean([mean_A_III, mean_B_III, mean_C_III, mean_D_III]), check_razryad.get())
				sum_kv_mezhdy_gr_QCC1_QCC2_QCC3_QCC4 = len(spisok_A_III)*((mean_A_III - mean_QCC1_QCC2_QCC3_QCC4)**2) + len(spisok_B_III)*((mean_B_III - mean_QCC1_QCC2_QCC3_QCC4)**2) + len(spisok_C_III)*(mean_C_III - mean_QCC1_QCC2_QCC3_QCC4)**2 + len(spisok_D_III)*((mean_D_III - mean_QCC1_QCC2_QCC3_QCC4)**2)
				CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4 = (((sum_kv_mezhdy_gr_QCC1_QCC2_QCC3_QCC4)/3)**0.5)/mean_QCC1_QCC2_QCC3_QCC4*100
				CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4 = (mean_QCC1_QCC2_QCC3_QCC4 - QC_III)/QC_III * 100

			if check_var_gr_1.get() >= 4:
				mean_QCD1_QCD2_QCD3_QCD4 = my_round(stat.mean([mean_A_IV, mean_B_IV, mean_C_IV, mean_D_IV]), check_razryad.get())
				sum_kv_mezhdy_gr_QCD1_QCD2_QCD3_QCD4 = len(spisok_A_IV)*((mean_A_IV - mean_QCD1_QCD2_QCD3_QCD4)**2) + len(spisok_B_IV)*((mean_B_IV - mean_QCD1_QCD2_QCD3_QCD4)**2) + len(spisok_C_IV)*(mean_C_IV - mean_QCD1_QCD2_QCD3_QCD4)**2 + len(spisok_D_IV)*((mean_D_IV - mean_QCD1_QCD2_QCD3_QCD4)**2)
				CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4 = (((sum_kv_mezhdy_gr_QCD1_QCD2_QCD3_QCD4)/3)**0.5)/mean_QCD1_QCD2_QCD3_QCD4*100
				CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4 = (mean_QCD1_QCD2_QCD3_QCD4 - QC_IV)/QC_IV * 100

			if check_var_gr_1.get() >= 5:
				mean_QCE1_QCE2_QCE3_QCE4 = my_round(stat.mean([mean_A_V, mean_B_V, mean_C_V, mean_D_V]), check_razryad.get())
				sum_kv_mezhdy_gr_QCE1_QCE2_QCE3_QCE4 = len(spisok_A_V)*((mean_A_V - mean_QCE1_QCE2_QCE3_QCE4)**2) + len(spisok_B_V)*((mean_B_V - mean_QCE1_QCE2_QCE3_QCE4)**2) + len(spisok_C_V)*(mean_C_V - mean_QCE1_QCE2_QCE3_QCE4)**2 + len(spisok_D_V)*((mean_D_V - mean_QCE1_QCE2_QCE3_QCE4)**2)
				CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4 = (((sum_kv_mezhdy_gr_QCE1_QCE2_QCE3_QCE4)/3)**0.5)/mean_QCE1_QCE2_QCE3_QCE4*100
				CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4 = (mean_QCE1_QCE2_QCE3_QCE4 - QC_V)/QC_V * 100

			if check_var_gr_1.get() >= 6:
				mean_QCF1_QCF2_QCF3_QCF4 = my_round(stat.mean([mean_A_VI, mean_B_VI, mean_C_VI, mean_D_VI]), check_razryad.get())
				sum_kv_mezhdy_gr_QCF1_QCF2_QCF3_QCF4 = len(spisok_A_VI)*((mean_A_VI - mean_QCF1_QCF2_QCF3_QCF4)**2) + len(spisok_B_VI)*((mean_B_VI - mean_QCF1_QCF2_QCF3_QCF4)**2) + len(spisok_C_VI)*(mean_C_VI - mean_QCF1_QCF2_QCF3_QCF4)**2 + len(spisok_D_VI)*((mean_D_VI - mean_QCF1_QCF2_QCF3_QCF4)**2)
				CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4 = (((sum_kv_mezhdy_gr_QCF1_QCF2_QCF3_QCF4)/3)**0.5)/mean_QCF1_QCF2_QCF3_QCF4*100
				CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4 = (mean_QCF1_QCF2_QCF3_QCF4 - QC_VI)/QC_VI * 100


	#вывод среднего и остальных показателей в таблицу 
		if check_var_gr_3.get() == 3:
			ws['AK21'].value = 'найдено, сред.зн.,\nнг/мл (n=3)'
			ws['AK21'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK22'].value = 'εR, % (n=3)'
			ws['AK23'].value = 'σR, % (n=3)'
			ws['AK24'].value = 'найдено, сред.зн.,\nнг/мл (n=12)'
			ws['AK24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK25'].value = 'εR, % (n=12)'
			ws['AK26'].value = 'σR, % (n=12)'
			ws['AK27'].value = 'Норма |ε| и |σ|, %'

			ws['AL21'].value = my_round(mean_D_I, razryad)
			ws['AL22'].value = my_round(Er_D_I, 1)
			ws['AL23'].value = my_round(sigma_D_I, 1)

			ws['AL24'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4, razryad)
			ws['AL25'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, 1)
			ws['AL26'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AL27'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AL27'].value = '≤20'


			
			if check_var_gr_1.get() >= 2:
				ws['AM21'].value = my_round(mean_D_II, razryad)
				ws['AM22'].value = my_round(Er_D_II, 1)
				ws['AM23'].value = my_round(sigma_D_II, 1)

				ws['AM24'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4, razryad)
				ws['AM25'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, 1)
				ws['AM26'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AM27'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AM27'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AN21'].value = my_round(mean_D_III, razryad)
				ws['AN22'].value = my_round(Er_D_III, 1)
				ws['AN23'].value = my_round(sigma_D_III, 1)

				ws['AN24'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4, razryad)
				ws['AN25'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, 1)
				ws['AN26'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AN27'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AN27'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AO21'].value = my_round(mean_D_IV, razryad)
				ws['AO22'].value = my_round(Er_D_IV, 1)
				ws['AO23'].value = my_round(sigma_D_IV, 1)

				ws['AO24'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4, razryad)
				ws['AO25'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, 1)
				ws['AO26'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AO27'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AO27'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AP21'].value = my_round(mean_D_V, razryad)
				ws['AP22'].value = my_round(Er_D_V, 1)
				ws['AP23'].value = my_round(sigma_D_V, 1)

				ws['AP24'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4, razryad)
				ws['AP25'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, 1)
				ws['AP26'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AP27'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AP27'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AQ21'].value = my_round(mean_D_VI, razryad)
				ws['AQ22'].value = my_round(Er_D_VI, 1)
				ws['AQ23'].value = my_round(sigma_D_VI, 1)

				ws['AQ24'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4, razryad)
				ws['AQ25'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, 1)
				ws['AQ26'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AQ27'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AQ27'].value = '≤15'


		if check_var_gr_3.get() == 4:

			ws['AK22'].value = 'найдено, сред.зн.,\nнг/мл (n=4)'
			ws['AK22'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK23'].value = 'εR, % (n=4)'
			ws['AK24'].value = 'σR, % (n=4)'
			
			ws['AK25'].value = 'найдено, сред.зн.,\nнг/мл (n=16)'
			ws['AK25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK26'].value = 'εR, % (n=16)'
			ws['AK27'].value = 'σR, % (n=16)'
			ws['AK28'].value = 'Норма |ε| и |σ|, %'


			ws['AL22'].value = my_round(mean_D_I, razryad)
			ws['AL23'].value = my_round(Er_D_I, 1)
			ws['AL24'].value = my_round(sigma_D_I, 1)

			ws['AL25'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4, razryad)
			ws['AL26'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, 1)
			ws['AL27'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AL28'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AL28'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AM22'].value = my_round(mean_D_II, razryad)
				ws['AM23'].value = my_round(Er_D_II, 1)
				ws['AM24'].value = my_round(sigma_D_II, 1)

				ws['AM25'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4, razryad)
				ws['AM26'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, 1)
				ws['AM27'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AM28'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AM28'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AN22'].value = my_round(mean_D_III, razryad)
				ws['AN23'].value = my_round(Er_D_III, 1)
				ws['AN24'].value = my_round(sigma_D_III, 1)

				ws['AN25'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4, razryad)
				ws['AN26'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, 1)
				ws['AN27'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AN28'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AN28'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AO22'].value = my_round(mean_D_IV, razryad)
				ws['AO23'].value = my_round(Er_D_IV, 1)
				ws['AO24'].value = my_round(sigma_D_IV, 1)

				ws['AO25'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4, razryad)
				ws['AO26'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, 1)
				ws['AO27'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AO28'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AO28'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AP22'].value = my_round(mean_D_V, razryad)
				ws['AP23'].value = my_round(Er_D_V, 1)
				ws['AP24'].value = my_round(sigma_D_V, 1)

				ws['AP25'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4, razryad)
				ws['AP26'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, 1)
				ws['AP27'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AP28'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AP28'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AQ22'].value = my_round(mean_D_VI, razryad)
				ws['AQ23'].value = my_round(Er_D_VI, 1)
				ws['AQ24'].value = my_round(sigma_D_VI, 1)

				ws['AQ25'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4, razryad)
				ws['AQ26'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, 1)
				ws['AQ27'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AQ28'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AQ28'].value = '≤15'


		if check_var_gr_3.get() == 5:
			ws['AK23'].value = 'найдено, сред.зн., \nнг/мл (n=5)'
			ws['AK23'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK24'].value = 'εR, % (n=5)'
			ws['AK25'].value = 'σR, % (n=5)'
			
			ws['AK26'].value = 'найдено, сред.зн.,\nнг/мл (n=20)'
			ws['AK26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK27'].value = 'εR, % (n=20)'
			ws['AK28'].value = 'σR, % (n=20)'
			ws['AK29'].value = 'Норма |ε| и |σ|, %'

			ws['AL23'].value = my_round(mean_D_I, razryad)
			ws['AL24'].value = my_round(Er_D_I, 1)
			ws['AL25'].value = my_round(sigma_D_I, 1)

			ws['AL26'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4, razryad)
			ws['AL27'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, 1)
			ws['AL28'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AL29'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AL29'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AM23'].value = my_round(mean_D_II, razryad)
				ws['AM24'].value = my_round(Er_D_II, 1)
				ws['AM25'].value = my_round(sigma_D_II, 1)

				ws['AM26'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4, razryad)
				ws['AM27'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, 1)
				ws['AM28'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AM29'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AM29'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AN23'].value = my_round(mean_D_III, razryad)
				ws['AN24'].value = my_round(Er_D_III, 1)
				ws['AN25'].value = my_round(sigma_D_III, 1)

				ws['AN26'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4, razryad)
				ws['AN27'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, 1)
				ws['AN28'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AN29'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AN29'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AO23'].value = my_round(mean_D_IV, razryad)
				ws['AO24'].value = my_round(Er_D_IV, 1)
				ws['AO25'].value = my_round(sigma_D_IV, 1)

				ws['AO26'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4, razryad)
				ws['AO27'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, 1)
				ws['AO28'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AO29'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AO29'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AP23'].value = my_round(mean_D_V, razryad)
				ws['AP24'].value = my_round(Er_D_V, 1)
				ws['AP25'].value = my_round(sigma_D_V, 1)

				ws['AP26'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4, razryad)
				ws['AP27'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, 1)
				ws['AP28'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AP29'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AP29'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AQ23'].value = my_round(mean_D_VI, razryad)
				ws['AQ24'].value = my_round(Er_D_VI, 1)
				ws['AQ25'].value = my_round(sigma_D_VI, 1)

				ws['AQ26'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4, razryad)
				ws['AQ27'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, 1)
				ws['AQ28'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AQ29'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AQ29'].value = '≤15'


		if check_var_gr_3.get() == 6:
			ws['AK24'].value = 'найдено, сред.зн., \nнг/мл (n=6)'
			ws['AK24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK25'].value = 'εR, % (n=6)'
			ws['AK26'].value = 'σR, % (n=6)'
			
			ws['AK27'].value = 'найдено, сред.зн.,\nнг/мл (n=24)'
			ws['AK27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK28'].value = 'εR, % (n=24)'
			ws['AK28'].value = 'σR, % (n=24)'
			ws['AK29'].value = 'Норма |ε| и |σ|, %'

			ws['AL24'].value = my_round(mean_D_I, razryad)
			ws['AL25'].value = my_round(Er_D_I, 1)
			ws['AL26'].value = my_round(sigma_D_I, 1)

			ws['AL27'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4, razryad)
			ws['AL28'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, 1)
			ws['AL29'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AL30'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AL30'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AM24'].value = my_round(mean_D_II, razryad)
				ws['AM25'].value = my_round(Er_D_II, 1)
				ws['AM26'].value = my_round(sigma_D_II, 1)

				ws['AM27'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4, razryad)
				ws['AM28'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, 1)
				ws['AM29'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AM30'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AM30'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AN24'].value = my_round(mean_D_III, razryad)
				ws['AN25'].value = my_round(Er_D_III, 1)
				ws['AN26'].value = my_round(sigma_D_III, 1)

				ws['AN27'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4, razryad)
				ws['AN28'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, 1)
				ws['AN29'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AN30'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AN30'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AO24'].value = my_round(mean_D_IV, razryad)
				ws['AO25'].value = my_round(Er_D_IV, 1)
				ws['AO26'].value = my_round(sigma_D_IV, 1)

				ws['AO27'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4, razryad)
				ws['AO28'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, 1)
				ws['AO29'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AO30'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AO30'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AP24'].value = my_round(mean_D_V, razryad)
				ws['AP25'].value = my_round(Er_D_V, 1)
				ws['AP26'].value = my_round(sigma_D_V, 1)

				ws['AP27'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4, razryad)
				ws['AP28'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, 1)
				ws['AP29'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AP30'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AP30'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AQ24'].value = my_round(mean_D_VI, razryad)
				ws['AQ25'].value = my_round(Er_D_VI, 1)
				ws['AQ26'].value = my_round(sigma_D_VI, 1)

				ws['AQ27'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4, razryad)
				ws['AQ28'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, 1)
				ws['AQ29'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AQ30'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AQ30'].value = '≤15'

		if check_var_gr_3.get() == 7:
			ws['AK25'].value = 'найдено, сред.зн., \nнг/мл (n=7)'
			ws['AK25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK26'].value = 'εR, % (n=7)'
			ws['AK27'].value = 'σR, % (n=7)'
			
			ws['AK28'].value = 'найдено, сред.зн.,\nнг/мл (n=28)'
			ws['AK28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK29'].value = 'εR, % (n=28)'
			ws['AK30'].value = 'σR, % (n=28)'
			ws['AK31'].value = 'Норма |ε| и |σ|, %'

			ws['AL25'].value = my_round(mean_D_I, razryad)
			ws['AL26'].value = my_round(Er_D_I, 1)
			ws['AL27'].value = my_round(sigma_D_I, 1)

			ws['AL28'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4, razryad)
			ws['AL29'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, 1)
			ws['AL30'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AL31'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AL31'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AM25'].value = my_round(mean_D_II, razryad)
				ws['AM26'].value = my_round(Er_D_II, 1)
				ws['AM27'].value = my_round(sigma_D_II, 1)

				ws['AM28'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4, razryad)
				ws['AM29'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, 1)
				ws['AM30'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AM31'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AM31'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AN25'].value = my_round(mean_D_III, razryad)
				ws['AN26'].value = my_round(Er_D_III, 1)
				ws['AN27'].value = my_round(sigma_D_III, 1)

				ws['AN28'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4, razryad)
				ws['AN29'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, 1)
				ws['AN30'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AN31'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AN31'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AO25'].value = my_round(mean_D_IV, razryad)
				ws['AO26'].value = my_round(Er_D_IV, 1)
				ws['AO27'].value = my_round(sigma_D_IV, 1)

				ws['AO28'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4, razryad)
				ws['AO29'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, 1)
				ws['AO30'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AO31'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AO31'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AP25'].value = my_round(mean_D_V, razryad)
				ws['AP26'].value = my_round(Er_D_V, 1)
				ws['AP27'].value = my_round(sigma_D_V, 1)

				ws['AP28'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4, razryad)
				ws['AP29'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, 1)
				ws['AP30'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AP31'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AP31'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AQ25'].value = my_round(mean_D_VI, razryad)
				ws['AQ26'].value = my_round(Er_D_VI, 1)
				ws['AQ27'].value = my_round(sigma_D_VI, 1)

				ws['AQ28'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4, razryad)
				ws['AQ29'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, 1)
				ws['AQ30'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AQ31'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AQ31'].value = '≤15'

		if check_var_gr_3.get() == 8:
			ws['AK26'].value = 'найдено, сред.зн., \nнг/мл (n=8)'
			ws['AK26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK27'].value = 'εR, % (n=8)'
			ws['AK28'].value = 'σR, % (n=8)'
			
			ws['AK29'].value = 'найдено, сред.зн.,\nнг/мл (n=32)'
			ws['AK29'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK30'].value = 'εR, % (n=32)'
			ws['AK30'].value = 'σR, % (n=32)'
			ws['AK31'].value = 'Норма |ε| и |σ|, %'

			ws['AL26'].value = my_round(mean_D_I, razryad)
			ws['AL27'].value = my_round(Er_D_I, 1)
			ws['AL28'].value = my_round(sigma_D_I, 1)

			ws['AL29'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4, razryad)
			ws['AL30'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, 1)
			ws['AL31'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AL32'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AL32'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AM26'].value = my_round(mean_D_II, razryad)
				ws['AM27'].value = my_round(Er_D_II, 1)
				ws['AM28'].value = my_round(sigma_D_II, 1)

				ws['AM29'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4, razryad)
				ws['AM30'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, 1)
				ws['AM31'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AM32'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AM32'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AN26'].value = my_round(mean_D_III, razryad)
				ws['AN27'].value = my_round(Er_D_III, 1)
				ws['AN28'].value = my_round(sigma_D_III, 1)

				ws['AN29'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4, razryad)
				ws['AN30'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, 1)
				ws['AN31'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AN32'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AN32'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AO26'].value = my_round(mean_D_IV, razryad)
				ws['AO27'].value = my_round(Er_D_IV, 1)
				ws['AO28'].value = my_round(sigma_D_IV, 1)

				ws['AO29'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4, razryad)
				ws['AO30'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, 1)
				ws['AO31'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AO32'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AO32'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AP26'].value = my_round(mean_D_V, razryad)
				ws['AP27'].value = my_round(Er_D_V, 1)
				ws['AP28'].value = my_round(sigma_D_V, 1)

				ws['AP29'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4, razryad)
				ws['AP30'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, 1)
				ws['AP31'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AP32'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AP32'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AQ26'].value = my_round(mean_D_VI, razryad)
				ws['AQ27'].value = my_round(Er_D_VI, 1)
				ws['AQ28'].value = my_round(sigma_D_VI, 1)

				ws['AQ29'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4, razryad)
				ws['AQ30'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, 1)
				ws['AQ31'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AQ32'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AQ32'].value = '≤15'

		if check_var_gr_3.get() == 9:
			ws['AK27'].value = 'найдено, сред.зн., \nнг/мл (n=9)'
			ws['AK27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK28'].value = 'εR, % (n=9)'
			ws['AK29'].value = 'σR, % (n=9)'
			
			ws['AK30'].value = 'найдено, сред.зн.,\nнг/мл (n=36)'
			ws['AK30'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK31'].value = 'εR, % (n=36)'
			ws['AK32'].value = 'σR, % (n=36)'
			ws['AK33'].value = 'Норма |ε| и |σ|, %'

			ws['AL27'].value = my_round(mean_D_I, razryad)
			ws['AL28'].value = my_round(Er_D_I, 1)
			ws['AL29'].value = my_round(sigma_D_I, 1)

			ws['AL30'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4, razryad)
			ws['AL31'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, 1)
			ws['AL32'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AL33'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AL33'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AM27'].value = my_round(mean_D_II, razryad)
				ws['AM28'].value = my_round(Er_D_II, 1)
				ws['AM29'].value = my_round(sigma_D_II, 1)

				ws['AM30'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4, razryad)
				ws['AM31'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, 1)
				ws['AM32'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AM33'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AM33'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AN27'].value = my_round(mean_D_III, razryad)
				ws['AN28'].value = my_round(Er_D_III, 1)
				ws['AN29'].value = my_round(sigma_D_III, 1)

				ws['AN30'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4, razryad)
				ws['AN31'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, 1)
				ws['AN32'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AN33'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AN33'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AO27'].value = my_round(mean_D_IV, razryad)
				ws['AO28'].value = my_round(Er_D_IV, 1)
				ws['AO29'].value = my_round(sigma_D_IV, 1)

				ws['AO30'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4, razryad)
				ws['AO31'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, 1)
				ws['AO32'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AO33'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AO33'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AP27'].value = my_round(mean_D_V, razryad)
				ws['AP28'].value = my_round(Er_D_V, 1)
				ws['AP29'].value = my_round(sigma_D_V, 1)

				ws['AP30'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4, razryad)
				ws['AP31'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, 1)
				ws['AP32'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AP33'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AP33'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AQ27'].value = my_round(mean_D_VI, razryad)
				ws['AQ28'].value = my_round(Er_D_VI, 1)
				ws['AQ29'].value = my_round(sigma_D_VI, 1)

				ws['AQ30'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4, razryad)
				ws['AQ31'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, 1)
				ws['AQ32'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AQ33'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AQ33'].value = '≤15'

		if check_var_gr_3.get() == 10:
			ws['AK28'].value = 'найдено, сред.зн., \nнг/мл (n=10)'
			ws['AK28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK29'].value = 'εR, % (n=10)'
			ws['AK30'].value = 'σR, % (n=10)'
			
			ws['AK31'].value = 'найдено, сред.зн.,\nнг/мл (n=40)'
			ws['AK31'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AK32'].value = 'εR, % (n=40)'
			ws['AK33'].value = 'σR, % (n=40)'
			ws['AK34'].value = 'Норма |ε| и |σ|, %'

			ws['AL28'].value = my_round(mean_D_I, razryad)
			ws['AL29'].value = my_round(Er_D_I, 1)
			ws['AL30'].value = my_round(sigma_D_I, 1)

			ws['AL31'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4, razryad)
			ws['AL32'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, 1)
			ws['AL33'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AL34'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AL34'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AM28'].value = my_round(mean_D_II, razryad)
				ws['AM29'].value = my_round(Er_D_II, 1)
				ws['AM30'].value = my_round(sigma_D_II, 1)

				ws['AM31'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4, razryad)
				ws['AM32'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, 1)
				ws['AM33'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AM34'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AM34'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AN28'].value = my_round(mean_D_III, razryad)
				ws['AN29'].value = my_round(Er_D_III, 1)
				ws['AN30'].value = my_round(sigma_D_III, 1)

				ws['AN31'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4, razryad)
				ws['AN32'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, 1)
				ws['AN33'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AN34'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AN34'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AO28'].value = my_round(mean_D_IV, razryad)
				ws['AO29'].value = my_round(Er_D_IV, 1)
				ws['AO30'].value = my_round(sigma_D_IV, 1)

				ws['AO31'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4, razryad)
				ws['AO32'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, 1)
				ws['AO33'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AO34'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AO34'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AP28'].value = my_round(mean_D_V, razryad)
				ws['AP29'].value = my_round(Er_D_V, 1)
				ws['AP30'].value = my_round(sigma_D_V, 1)

				ws['AP31'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4, razryad)
				ws['AP32'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, 1)
				ws['AP33'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AP34'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AP34'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AQ28'].value = my_round(mean_D_VI, razryad)
				ws['AQ29'].value = my_round(Er_D_VI, 1)
				ws['AQ30'].value = my_round(sigma_D_VI, 1)

				ws['AQ31'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4, razryad)
				ws['AQ32'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, 1)
				ws['AQ33'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AQ34'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AQ34'].value = '≤15'


	#QCA4
		ws['AL15'].value = 'QCA4'
		ws['AL16'].value = QC_I
		ws['F2'].value = 'QCA4'
		ws['AK18'].value = '1'
		ws['AL18'].value = my_round(float(entr_D1_I.get()), razryad)
		ws['F3'].value = my_round(float(entr_D1_I.get()), razryad)
		ws['AK19'].value = '2'
		ws['AL19'].value = my_round(float(entr_D2_I.get()), razryad)
		ws['F4'].value = my_round(float(entr_D2_I.get()), razryad)
		ws['AK20'].value = '3'
		ws['AL20'].value = my_round(float(entr_D3_I.get()), razryad)
		ws['F5'].value = my_round(float(entr_D3_I.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['AK21'].value = '4'
			ws['AL21'].value = my_round(float(entr_D4_I.get()), razryad)
			ws['F6'].value = my_round(float(entr_D4_I.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['AK22'].value = '5'
			ws['AL22'].value = my_round(float(entr_D5_I.get()), razryad)
			ws['F7'].value = my_round(float(entr_D5_I.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['AK23'].value = '6'
			ws['AL23'].value = my_round(float(entr_D6_I.get()), razryad)
			ws['F8'].value = my_round(float(entr_D6_I.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['AK24'].value = '7'
			ws['AL24'].value = my_round(float(entr_D7_I.get()), razryad)
			ws['F9'].value = my_round(float(entr_D7_I.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['AK25'].value = '8'
			ws['AL25'].value = my_round(float(entr_D8_I.get()), razryad)
			ws['F10'].value = my_round(float(entr_D8_I.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['AK26'].value = '9'
			ws['AL26'].value = my_round(float(entr_D9_I.get()), razryad)
			ws['F11'].value = my_round(float(entr_D9_I.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['AK27'].value = '10'
			ws['AL27'].value = my_round(float(entr_D10_I.get()), razryad)
			ws['F12'].value = my_round(float(entr_D10_I.get()), razryad)



		if check_var_gr_1.get() >= 2:
			#QCB4
			ws['AM15'].value = 'QCB4'
			ws['AM16'].value = QC_II
			ws['O2'].value = 'QCB4'
			ws['AM18'].value = my_round(float(entr_D1_II.get()), razryad)
			ws['O3'].value = my_round(float(entr_D1_II.get()), razryad)
			ws['AM19'].value = my_round(float(entr_D2_II.get()), razryad)
			ws['O4'].value = my_round(float(entr_D2_II.get()), razryad)
			ws['AM20'].value = my_round(float(entr_D3_II.get()), razryad)
			ws['O5'].value = my_round(float(entr_D3_II.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AM21'].value = my_round(float(entr_D4_II.get()), razryad)
				ws['O6'].value = my_round(float(entr_D4_II.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AM22'].value = my_round(float(entr_D5_II.get()), razryad)
				ws['O7'].value = my_round(float(entr_D5_II.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AM23'].value = my_round(float(entr_D6_II.get()), razryad)
				ws['O8'].value = my_round(float(entr_D6_II.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AM24'].value = my_round(float(entr_D7_II.get()), razryad)
				ws['O9'].value = my_round(float(entr_D7_II.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AM25'].value = my_round(float(entr_D8_II.get()), razryad)
				ws['O10'].value = my_round(float(entr_D8_II.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AM26'].value = my_round(float(entr_D9_II.get()), razryad)
				ws['O11'].value = my_round(float(entr_D9_II.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AM27'].value = my_round(float(entr_D10_II.get()), razryad)
				ws['O12'].value = my_round(float(entr_D10_II.get()), razryad)

		if check_var_gr_1.get() >= 3:
			#QCC4
			ws['AN15'].value = 'QCC4'
			ws['AN16'].value = QC_III
			ws['X2'].value = 'QCC4'
			ws['AN18'].value = my_round(float(entr_D1_III.get()), razryad)
			ws['X3'].value = my_round(float(entr_D1_III.get()), razryad)
			ws['AN19'].value = my_round(float(entr_D2_III.get()), razryad)
			ws['X4'].value = my_round(float(entr_D2_III.get()), razryad)
			ws['AN20'].value = my_round(float(entr_D3_III.get()), razryad)
			ws['X5'].value = my_round(float(entr_D3_III.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AN21'].value = my_round(float(entr_D4_III.get()), razryad)
				ws['X6'].value = my_round(float(entr_D4_III.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AN22'].value = my_round(float(entr_D5_III.get()), razryad)
				ws['X7'].value = my_round(float(entr_D5_III.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AN23'].value = my_round(float(entr_D6_III.get()), razryad)
				ws['X8'].value = my_round(float(entr_D6_III.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AN24'].value = my_round(float(entr_D7_III.get()), razryad)
				ws['X9'].value = my_round(float(entr_D7_III.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AN25'].value = my_round(float(entr_D8_III.get()), razryad)
				ws['X10'].value = my_round(float(entr_D8_III.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AN26'].value = my_round(float(entr_D9_III.get()), razryad)
				ws['X11'].value = my_round(float(entr_D9_III.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AN27'].value = my_round(float(entr_D10_III.get()), razryad)
				ws['X12'].value = my_round(float(entr_D10_III.get()), razryad)

		if check_var_gr_1.get() >= 4:
			#QCD2
			ws['AO15'].value = 'QCD4'
			ws['AO16'].value = QC_IV
			ws['AG2'].value = 'QCD4'
			ws['AO18'].value = my_round(float(entr_D1_IV.get()), razryad)
			ws['AG3'].value = my_round(float(entr_D1_IV.get()), razryad)
			ws['AO19'].value = my_round(float(entr_D2_IV.get()), razryad)
			ws['AG4'].value = my_round(float(entr_D2_IV.get()), razryad)
			ws['AO20'].value = my_round(float(entr_D3_IV.get()), razryad)
			ws['AG5'].value = my_round(float(entr_D3_IV.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AO21'].value = my_round(float(entr_D4_IV.get()), razryad)
				ws['AG6'].value = my_round(float(entr_D4_IV.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AO22'].value = my_round(float(entr_D5_IV.get()), razryad)
				ws['AG7'].value = my_round(float(entr_D5_IV.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AO23'].value = my_round(float(entr_D6_IV.get()), razryad)
				ws['AG8'].value = my_round(float(entr_D6_IV.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AO24'].value = my_round(float(entr_D7_IV.get()), razryad)
				ws['AG9'].value = my_round(float(entr_D7_IV.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AO25'].value = my_round(float(entr_D8_IV.get()), razryad)
				ws['AG10'].value = my_round(float(entr_D8_IV.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AO26'].value = my_round(float(entr_D9_IV.get()), razryad)
				ws['AG11'].value = my_round(float(entr_D9_IV.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AO27'].value = my_round(float(entr_D10_IV.get()), razryad)
				ws['AG12'].value = my_round(float(entr_D10_IV.get()), razryad)

		if check_var_gr_1.get() >= 5:
			#QCE4
			ws['AP15'].value = 'QCE4'
			ws['AP16'].value = QC_V
			ws['AP2'].value = 'QCE4'
			ws['AP18'].value =  my_round(float(entr_D1_V.get()), razryad)
			ws['AP3'].value =  my_round(float(entr_D1_V.get()), razryad)
			ws['AP19'].value =  my_round(float(entr_D2_V.get()), razryad)
			ws['AP4'].value =  my_round(float(entr_D2_V.get()), razryad)
			ws['AP20'].value =  my_round(float(entr_D3_V.get()), razryad)
			ws['AP5'].value =  my_round(float(entr_D3_V.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AP21'].value = my_round(float(entr_D4_V.get()), razryad)
				ws['AP6'].value = my_round(float(entr_D4_V.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AP22'].value = my_round(float(entr_D5_V.get()), razryad)
				ws['AP7'].value = my_round(float(entr_D5_V.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AP23'].value = my_round(float(entr_D6_V.get()), razryad)
				ws['AP8'].value = my_round(float(entr_D6_V.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AP24'].value = my_round(float(entr_D7_V.get()), razryad)
				ws['AP9'].value = my_round(float(entr_D7_V.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AP25'].value = my_round(float(entr_D8_V.get()), razryad)
				ws['AP10'].value = my_round(float(entr_D8_V.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AP26'].value = my_round(float(entr_D9_V.get()), razryad)
				ws['AP11'].value = my_round(float(entr_D9_V.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AP27'].value = my_round(float(entr_D10_V.get()), razryad)
				ws['AP12'].value = my_round(float(entr_D10_V.get()), razryad)

		if check_var_gr_1.get() >= 6:
			#QCF4
			ws['AQ15'].value = 'QCF4'
			ws['AQ16'].value = QC_VI
			ws['AY2'].value = 'QCF4'
			ws['AQ18'].value = my_round(float(entr_D1_VI.get()), razryad)
			ws['AY3'].value = my_round(float(entr_D1_VI.get()), razryad)
			ws['AQ19'].value = my_round(float(entr_D2_VI.get()), razryad)
			ws['AY4'].value = my_round(float(entr_D2_VI.get()), razryad)
			ws['AQ20'].value = my_round(float(entr_D3_VI.get()), razryad)
			ws['AY5'].value = my_round(float(entr_D3_VI.get()), razryad)

			if check_var_gr_3.get() >= 4:
				ws['AQ21'].value = my_round(float(entr_D4_VI.get()), razryad)
				ws['AY6'].value = my_round(float(entr_D4_VI.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AQ22'].value = my_round(float(entr_D5_VI.get()), razryad)
				ws['AY7'].value = my_round(float(entr_D5_VI.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AQ23'].value = my_round(float(entr_D6_VI.get()), razryad)
				ws['AY8'].value = my_round(float(entr_D6_VI.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AQ24'].value = my_round(float(entr_D7_VI.get()), razryad)
				ws['AY9'].value = my_round(float(entr_D7_VI.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AQ25'].value = my_round(float(entr_D8_VI.get()), razryad)
				ws['AY10'].value = my_round(float(entr_D8_VI.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AQ26'].value = my_round(float(entr_D9_VI.get()), razryad)
				ws['AY11'].value = my_round(float(entr_D9_VI.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AQ27'].value = my_round(float(entr_D10_VI.get()), razryad)
				ws['AY12'].value = my_round(float(entr_D10_VI.get()), razryad)



#QC ТАБЛИЦА № 5
	if check_var_gr_2.get() >= 5:
	#выравнивание по центру]
		cols_c(ws, 'AS15:AY34')

		#границы ячеек
		if check_var_gr_1.get() == 1:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AS15:AT27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AS15:AT28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AS15:AT29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AS15:AT30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AS15:AT31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AS15:AT32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AS15:AT33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AS15:AT34')

		if check_var_gr_1.get() == 2:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AS15:AU27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AS15:AU28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AS15:AU29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AS15:AU30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AS15:AU31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AS15:AU32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AS15:AU33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AS15:AU34')

		if check_var_gr_1.get() == 3:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AS15:AV27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AS15:AV28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AS15:AV29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AS15:AV30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AS15:AV31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AS15:AV32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AS15:AV33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AS15:AV34')				

		if check_var_gr_1.get() == 4:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AS15:AW27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AS15:AW28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AS15:AW29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AS15:AW30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AS15:AW31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AS15:AW32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AS15:AW33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AS15:AW34')			

		if check_var_gr_1.get() == 5:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AS15:AX27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AS15:AX28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AS15:AX29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AS15:AX30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AS15:AX31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AS15:AX32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AS15:AX33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AS15:AX34')		

		if check_var_gr_1.get() == 6:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'AS15:AY27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'AS15:AY28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'AS15:AY29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'AS15:AY30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'AS15:AY31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'AS15:AY32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'AS15:AY33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'AS15:AY34')	

		
		#задание ширины столбца
		#ws.column_dimensions['AS'].width = 22

		#надпись и объединение ячеек под надпись
		ws['AS15'].value = 'QC'
		ws['AS16'].value = f'Введено, {entr_concent.get()}'
		ws['AS17'].value = '№ измерения'
		ws['AT17'].value = 'Найдено, последовательность 5'
		'''
		if check_var_gr_1.get() == 1:
			ws.column_dimensions['AS'].width = 31		
		if check_var_gr_1.get() == 2:
			ws.merge_cells('AT17:AU17')
		if check_var_gr_1.get() == 3:
			ws.merge_cells('AT17:AV17')
		if check_var_gr_1.get() == 4:
			ws.merge_cells('AT17:AW17')
		if check_var_gr_1.get() == 5:
			ws.merge_cells('AT17:AX17')
		if check_var_gr_1.get() == 6:
			ws.merge_cells('AT17:AX17')
		'''





	#############################################################################	
#ЗАПОЛНЕНИЕ ANOVA
	#количество таблиц зависит от показателя check_var_gr_2
	if check_var_gr_2.get() >= 5:
		#QCA5
		mean_QCA1_QCA2_QCA3_QCA4_QCA5 = my_round(stat.mean([mean_A_I, mean_B_I, mean_C_I, mean_D_I, mean_E_I]), check_razryad.get())
		#междугруппами
		sum_kv_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5 = len(spisok_A_I)*((mean_A_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5)**2) + len(spisok_B_I)*((mean_B_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5)**2) + len(spisok_C_I)*(mean_C_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5)**2 + len(spisok_D_I)*((mean_D_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5)**2) + len(spisok_E_I)*((mean_E_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5)**2)
		CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5 = (((sum_kv_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5)/4)**0.5)/mean_QCA1_QCA2_QCA3_QCA4_QCA5*100
		#внутригрупп
		CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5 = (mean_QCA1_QCA2_QCA3_QCA4_QCA5 - QC_I)/QC_I * 100


		if check_var_gr_1.get() >= 2:
			#QCB5
			mean_QCB1_QCB2_QCB3_QCB4_QCB5 = my_round(stat.mean([mean_A_II, mean_B_II, mean_C_II, mean_D_II, mean_E_II]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5 = len(spisok_A_II)*((mean_A_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5)**2) + len(spisok_B_II)*((mean_B_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5)**2) + len(spisok_C_II)*(mean_C_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5)**2 + len(spisok_D_II)*((mean_D_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5)**2) + len(spisok_E_II)*((mean_E_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5)**2)
			CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5 = (((sum_kv_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5)/4)**0.5)/mean_QCB1_QCB2_QCB3_QCB4_QCB5*100
			#внутригрупп
			CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5 = (mean_QCB1_QCB2_QCB3_QCB4_QCB5 - QC_II)/QC_II * 100

		if check_var_gr_1.get() >= 3:	
			#QCC5
			mean_QCC1_QCC2_QCC3_QCC4_QCC5 = my_round(stat.mean([mean_A_III, mean_B_III, mean_C_III, mean_D_III, mean_E_III]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5 = len(spisok_A_III)*((mean_A_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5)**2) + len(spisok_B_III)*((mean_B_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5)**2) + len(spisok_C_III)*(mean_C_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5)**2 + len(spisok_D_III)*((mean_D_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5)**2) + len(spisok_E_III)*((mean_E_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5)**2)
			CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5 = (((sum_kv_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5)/4)**0.5)/mean_QCC1_QCC2_QCC3_QCC4_QCC5*100
			#внутригрупп
			CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5 = (mean_QCC1_QCC2_QCC3_QCC4_QCC5 - QC_III)/QC_III * 100

		if check_var_gr_1.get() >= 4:
			#QCD5
			mean_QCD1_QCD2_QCD3_QCD4_QCD5 = my_round(stat.mean([mean_A_IV, mean_B_IV, mean_C_IV, mean_D_IV, mean_E_IV]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5 = len(spisok_A_IV)*((mean_A_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5)**2) + len(spisok_B_IV)*((mean_B_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5)**2) + len(spisok_C_IV)*(mean_C_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5)**2 + len(spisok_D_IV)*((mean_D_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5)**2) + len(spisok_E_IV)*((mean_E_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5)**2)
			CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5 = (((sum_kv_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5)/4)**0.5)/mean_QCD1_QCD2_QCD3_QCD4_QCD5*100
			#внутригрупп
			CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5 = (mean_QCD1_QCD2_QCD3_QCD4_QCD5 - QC_IV)/QC_IV * 100

		if check_var_gr_1.get() >= 5:
			#QCE5
			mean_QCE1_QCE2_QCE3_QCE4_QCE5 = my_round(stat.mean([mean_A_V, mean_B_V, mean_C_V, mean_D_V, mean_E_V]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5 = len(spisok_A_V)*((mean_A_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5)**2) + len(spisok_B_V)*((mean_B_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5)**2) + len(spisok_C_V)*(mean_C_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5)**2 + len(spisok_D_V)*((mean_D_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5)**2) + len(spisok_E_V)*((mean_E_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5)**2)
			CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5 = (((sum_kv_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5)/4)**0.5)/mean_QCE1_QCE2_QCE3_QCE4_QCE5*100
			#внутригрупп
			CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5 = (mean_QCE1_QCE2_QCE3_QCE4_QCE5 - QC_V)/QC_V * 100

		if check_var_gr_1.get() >= 6:
			#QCF5
			mean_QCF1_QCF2_QCF3_QCF4_QCF5 = my_round(stat.mean([mean_A_VI, mean_B_VI, mean_C_VI, mean_D_VI, mean_E_VI]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5 = len(spisok_A_VI)*((mean_A_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5)**2) + len(spisok_B_VI)*((mean_B_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5)**2) + len(spisok_C_VI)*(mean_C_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5)**2 + len(spisok_D_VI)*((mean_D_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5)**2) + len(spisok_E_VI)*((mean_E_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5)**2)
			CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5 = (((sum_kv_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5)/4)**0.5)/mean_QCF1_QCF2_QCF3_QCF4_QCF5*100
			#внутригрупп
			CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5 = (mean_QCF1_QCF2_QCF3_QCF4_QCF5 - QC_VI)/QC_VI * 100


#вывод среднего и остальных показателей в таблицу 
		if check_var_gr_3.get() == 3:

			ws['AS21'].value = 'найдено, сред.зн.,\nнг/мл (n=3)'
			ws['AS21'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

			ws['AS22'].value = 'εR, % (n=3)'
			ws['AS23'].value = 'σR, % (n=3)'

			ws['AS24'].value = 'найдено, сред.зн.,\nнг/мл (n=15)'
			ws['AS24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS25'].value = 'εR, % (n=15)'
			ws['AS26'].value = 'σR, % (n=15)'
			ws['AS27'].value = 'Норма |ε| и |σ|, %'

			ws['AT21'].value = my_round(mean_E_I, razryad)
			ws['AT22'].value = my_round(Er_E_I, 1)
			ws['AT23'].value = my_round(sigma_E_I, 1)

			ws['AT24'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5, razryad)
			ws['AT25'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			ws['AT26'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AT27'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AT27'].value = '≤20'

			
			if check_var_gr_1.get() >= 2:
				ws['AU21'].value = my_round(mean_E_II, razryad)
				ws['AU22'].value = my_round(Er_E_II, 1)
				ws['AU23'].value = my_round(sigma_E_II, 1)

				ws['AU24'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5, razryad)
				ws['AU25'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				ws['AU26'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AU27'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AU27'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AV21'].value = my_round(mean_E_III, razryad)
				ws['AV22'].value = my_round(Er_E_III, 1)
				ws['AV23'].value = my_round(sigma_E_III, 1)

				ws['AV24'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5, razryad)
				ws['AV25'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				ws['AV26'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AV27'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AV27'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AW21'].value = my_round(mean_E_IV, razryad)
				ws['AW22'].value = my_round(Er_E_IV, 1)
				ws['AW23'].value = my_round(sigma_E_IV, 1)

				ws['AW24'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5, razryad)
				ws['AW25'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				ws['AW26'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AW27'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AW27'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AX21'].value = my_round(mean_E_V, razryad)
				ws['AX22'].value = my_round(Er_E_V, 1)
				ws['AX23'].value = my_round(sigma_E_V, 1)

				ws['AX24'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5, razryad)
				ws['AX25'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				ws['AX26'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AX27'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AX27'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AY21'].value = my_round(mean_E_VI, razryad)
				ws['AY22'].value = my_round(Er_E_VI, 1)
				ws['AY23'].value = my_round(sigma_E_VI, 1)

				ws['AY24'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5, razryad)
				ws['AY25'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				ws['AY26'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AY27'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AY27'].value = '≤15'


		if check_var_gr_3.get() == 4:

			ws['AS22'].value = 'найдено, сред.зн.,\nнг/мл (n=4)'
			ws['AS22'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS23'].value = 'εR, % (n=4)'
			ws['AS24'].value = 'σR, % (n=4)'
			
			ws['AS25'].value = 'найдено, сред.зн.,\nнг/мл (n=20)'
			ws['AS25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS26'].value = 'εR, % (n=20)'
			ws['AS27'].value = 'σR, % (n=20)'
			ws['AS28'].value = 'Норма |ε| и |σ|, %'


			ws['AT22'].value = my_round(mean_E_I, razryad)
			ws['AT23'].value = my_round(Er_E_I, 1)
			ws['AT24'].value = my_round(sigma_E_I, 1)

			ws['AT25'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5, razryad)
			ws['AT26'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			ws['AT27'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AT28'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AT28'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AU22'].value = my_round(mean_E_II, razryad)
				ws['AU23'].value = my_round(Er_E_II, 1)
				ws['AU24'].value = my_round(sigma_E_II, 1)

				ws['AU25'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5, razryad)
				ws['AU26'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				ws['AU27'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AU28'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AU28'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AV22'].value = my_round(mean_E_III, razryad)
				ws['AV23'].value = my_round(Er_E_III, 1)
				ws['AV24'].value = my_round(sigma_E_III, 1)

				ws['AV25'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5, razryad)
				ws['AV26'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				ws['AV27'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AV28'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AV28'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AW22'].value = my_round(mean_E_IV, razryad)
				ws['AW23'].value = my_round(Er_E_IV, 1)
				ws['AW24'].value = my_round(sigma_E_IV, 1)

				ws['AW25'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5, razryad)
				ws['AW26'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				ws['AW27'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AW28'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AW28'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AX22'].value = my_round(mean_E_V, razryad)
				ws['AX23'].value = my_round(Er_E_V, 1)
				ws['AX24'].value = my_round(sigma_E_V, 1)

				ws['AX25'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5, razryad)
				ws['AX26'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				ws['AX27'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AX28'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AX28'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AY22'].value = my_round(mean_E_VI, razryad)
				ws['AY23'].value = my_round(Er_E_VI, 1)
				ws['AY24'].value = my_round(sigma_E_VI, 1)

				ws['AY25'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5, razryad)
				ws['AY26'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				ws['AY27'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AY28'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AY28'].value = '≤15'


		if check_var_gr_3.get() == 5:
			ws['AS23'].value = 'найдено, сред.зн., \nнг/мл (n=5)'
			ws['AS23'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS24'].value = 'εR, % (n=5)'
			ws['AS25'].value = 'σR, % (n=5)'
			
			ws['AS26'].value = 'найдено, сред.зн.,\nнг/мл (n=25)'
			ws['AS26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS27'].value = 'εR, % (n=25)'
			ws['AS28'].value = 'σR, % (n=25)'
			ws['AS29'].value = 'Норма |ε| и |σ|, %'

			ws['AT23'].value = my_round(mean_E_I, razryad)
			ws['AT24'].value = my_round(Er_E_I, 1)
			ws['AT25'].value = my_round(sigma_E_I, 1)

			ws['AT26'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5, razryad)
			ws['AT27'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			ws['AT28'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AT29'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AT29'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AU23'].value = my_round(mean_E_II, razryad)
				ws['AU24'].value = my_round(Er_E_II, 1)
				ws['AU25'].value = my_round(sigma_E_II, 1)

				ws['AU26'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5, razryad)
				ws['AU27'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				ws['AU28'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AU29'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AU29'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AV23'].value = my_round(mean_E_III, razryad)
				ws['AV24'].value = my_round(Er_E_III, 1)
				ws['AV25'].value = my_round(sigma_E_III, 1)

				ws['AV26'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5, razryad)
				ws['AV27'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				ws['AV28'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AV29'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AV29'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AW23'].value = my_round(mean_E_IV, razryad)
				ws['AW24'].value = my_round(Er_E_IV, 1)
				ws['AW25'].value = my_round(sigma_E_IV, 1)

				ws['AW26'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5, razryad)
				ws['AW27'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				ws['AW28'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AW29'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AW29'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AX23'].value = my_round(mean_E_V, razryad)
				ws['AX24'].value = my_round(Er_E_V, 1)
				ws['AX25'].value = my_round(sigma_E_V, 1)

				ws['AX26'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5, razryad)
				ws['AX27'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				ws['AX28'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AX29'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AX29'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AY23'].value = my_round(mean_E_VI, razryad)
				ws['AY24'].value = my_round(Er_E_VI, 1)
				ws['AY25'].value = my_round(sigma_E_VI, 1)

				ws['AY26'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5, razryad)
				ws['AY27'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				ws['AY28'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AY29'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AY29'].value = '≤15'


		if check_var_gr_3.get() == 6:
			ws['AS24'].value = 'найдено, сред.зн., \nнг/мл (n=6)'
			ws['AS24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS25'].value = 'εR, % (n=6)'
			ws['AS26'].value = 'σR, % (n=6)'
			
			ws['AS27'].value = 'найдено, сред.зн.,\nнг/мл (n=30)'
			ws['AS27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS28'].value = 'εR, % (n=30)'
			ws['AS28'].value = 'σR, % (n=30)'
			ws['AS29'].value = 'Норма |ε| и |σ|, %'

			ws['AT24'].value = my_round(mean_E_I, razryad)
			ws['AT25'].value = my_round(Er_E_I, 1)
			ws['AL26'].value = my_round(sigma_E_I, 1)

			ws['AT27'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5, razryad)
			ws['AT28'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			ws['AT29'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AT30'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AT30'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AU24'].value = my_round(mean_E_II, razryad)
				ws['AU25'].value = my_round(Er_E_II, 1)
				ws['AU26'].value = my_round(sigma_E_II, 1)

				ws['AU27'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5, razryad)
				ws['AU28'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				ws['AU29'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AU30'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AU30'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AV24'].value = my_round(mean_E_III, razryad)
				ws['AV25'].value = my_round(Er_E_III, 1)
				ws['AV26'].value = my_round(sigma_E_III, 1)

				ws['AV27'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5, razryad)
				ws['AV28'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				ws['AV29'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AV30'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AV30'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AW24'].value = my_round(mean_E_IV, razryad)
				ws['AW25'].value = my_round(Er_E_IV, 1)
				ws['AW26'].value = my_round(sigma_E_IV, 1)

				ws['AW27'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5, razryad)
				ws['AW28'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				ws['AW29'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AW30'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AW30'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AX24'].value = my_round(mean_E_V, razryad)
				ws['AX25'].value = my_round(Er_E_V, 1)
				ws['AX26'].value = my_round(sigma_E_V, 1)

				ws['AX27'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5, razryad)
				ws['AX28'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				ws['AX29'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AX30'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AX30'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AY24'].value = my_round(mean_E_VI, razryad)
				ws['AY25'].value = my_round(Er_E_VI, 1)
				ws['AY26'].value = my_round(sigma_E_VI, 1)

				ws['AY27'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5, razryad)
				ws['AY28'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				ws['AY29'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AY30'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AY30'].value = '≤15'

		if check_var_gr_3.get() == 7:
			ws['AS25'].value = 'найдено, сред.зн., \nнг/мл (n=7)'
			ws['AS25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS26'].value = 'εR, % (n=7)'
			ws['AS27'].value = 'σR, % (n=7)'
			
			ws['AS28'].value = 'найдено, сред.зн.,\nнг/мл (n=35)'
			ws['AS28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS29'].value = 'εR, % (n=35)'
			ws['AS30'].value = 'σR, % (n=35)'
			ws['AS31'].value = 'Норма |ε| и |σ|, %'

			ws['AT25'].value = my_round(mean_E_I, razryad)
			ws['AT26'].value = my_round(Er_E_I, 1)
			ws['AT27'].value = my_round(sigma_E_I, 1)

			ws['AT28'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5, razryad)
			ws['AT29'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			ws['AT30'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AT31'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AT31'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AU25'].value = my_round(mean_E_II, razryad)
				ws['AU26'].value = my_round(Er_E_II, 1)
				ws['AU27'].value = my_round(sigma_E_II, 1)

				ws['AU28'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5, razryad)
				ws['AU29'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				ws['AU30'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AU31'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AU31'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AV25'].value = my_round(mean_E_III, razryad)
				ws['AV26'].value = my_round(Er_E_III, 1)
				ws['AV27'].value = my_round(sigma_E_III, 1)

				ws['AV28'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5, razryad)
				ws['AV29'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				ws['AV30'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AV31'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AV31'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AW25'].value = my_round(mean_E_IV, razryad)
				ws['AW26'].value = my_round(Er_E_IV, 1)
				ws['AW27'].value = my_round(sigma_E_IV, 1)

				ws['AW28'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5, razryad)
				ws['AW29'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				ws['AW30'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AW31'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AW31'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AX25'].value = my_round(mean_E_V, razryad)
				ws['AX26'].value = my_round(Er_E_V, 1)
				ws['AX27'].value = my_round(sigma_E_V, 1)

				ws['AX28'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5, razryad)
				ws['AX29'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				ws['AX30'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AX31'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AX31'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AY25'].value = my_round(mean_E_VI, razryad)
				ws['AY26'].value = my_round(Er_E_VI, 1)
				ws['AY27'].value = my_round(sigma_E_VI, 1)

				ws['AY28'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5, razryad)
				ws['AY29'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				ws['AY30'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AY31'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AY31'].value = '≤15'

		if check_var_gr_3.get() == 8:
			ws['AS26'].value = 'найдено, сред.зн., \nнг/мл (n=8)'
			ws['AS26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS27'].value = 'εR, % (n=8)'
			ws['AS28'].value = 'σR, % (n=8)'
			
			ws['AS29'].value = 'найдено, сред.зн.,\nнг/мл (n=40)'
			ws['AS29'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS30'].value = 'εR, % (n=40)'
			ws['AS30'].value = 'σR, % (n=40)'
			ws['AS31'].value = 'Норма |ε| и |σ|, %'

			ws['AT26'].value = my_round(mean_E_I, razryad)
			ws['AT27'].value = my_round(Er_E_I, 1)
			ws['AT28'].value = my_round(sigma_E_I, 1)

			ws['AT29'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5, razryad)
			ws['AT30'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			ws['AT31'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AT32'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AT32'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AU26'].value = my_round(mean_E_II, razryad)
				ws['AU27'].value = my_round(Er_E_II, 1)
				ws['AU28'].value = my_round(sigma_E_II, 1)

				ws['AU29'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5, razryad)
				ws['AU30'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				ws['AU31'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AU32'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AU32'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AV26'].value = my_round(mean_E_III, razryad)
				ws['AV27'].value = my_round(Er_E_III, 1)
				ws['AV28'].value = my_round(sigma_E_III, 1)

				ws['AV29'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5, razryad)
				ws['AV30'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				ws['AV31'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AV32'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AV32'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AW26'].value = my_round(mean_E_IV, razryad)
				ws['AW27'].value = my_round(Er_E_IV, 1)
				ws['AW28'].value = my_round(sigma_E_IV, 1)

				ws['AW29'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5, razryad)
				ws['AW30'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				ws['AW31'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AW32'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AW32'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AX26'].value = my_round(mean_E_V, razryad)
				ws['AX27'].value = my_round(Er_E_V, 1)
				ws['AX28'].value = my_round(sigma_E_V, 1)

				ws['AX29'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5, razryad)
				ws['AX30'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				ws['AX31'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AX32'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AX32'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AY26'].value = my_round(mean_E_VI, razryad)
				ws['AY27'].value = my_round(Er_E_VI, 1)
				ws['AY28'].value = my_round(sigma_E_VI, 1)

				ws['AY29'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5, razryad)
				ws['AY30'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				ws['AY31'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AY32'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AY32'].value = '≤15'

		if check_var_gr_3.get() == 9:
			ws['AS27'].value = 'найдено, сред.зн., \nнг/мл (n=9)'
			ws['AS27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS28'].value = 'εR, % (n=9)'
			ws['AS29'].value = 'σR, % (n=9)'
			
			ws['AS30'].value = 'найдено, сред.зн.,\nнг/мл (n=45)'
			ws['AS30'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS31'].value = 'εR, % (n=45)'
			ws['AS32'].value = 'σR, % (n=45)'
			ws['AS33'].value = 'Норма |ε| и |σ|, %'

			ws['AT27'].value = my_round(mean_E_I, razryad)
			ws['AT28'].value = my_round(Er_E_I, 1)
			ws['AT29'].value = my_round(sigma_E_I, 1)

			ws['AT30'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5, razryad)
			ws['AT31'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			ws['AT32'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AT33'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AT33'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AU27'].value = my_round(mean_E_II, razryad)
				ws['AU28'].value = my_round(Er_E_II, 1)
				ws['AU29'].value = my_round(sigma_E_II, 1)

				ws['AU30'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5, razryad)
				ws['AU31'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				ws['AU32'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AU33'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AU33'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AV27'].value = my_round(mean_E_III, razryad)
				ws['AV28'].value = my_round(Er_E_III, 1)
				ws['AV29'].value = my_round(sigma_E_III, 1)

				ws['AV30'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5, razryad)
				ws['AV31'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				ws['AV32'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AV33'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AV33'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AW27'].value = my_round(mean_E_IV, razryad)
				ws['AW28'].value = my_round(Er_E_IV, 1)
				ws['AW29'].value = my_round(sigma_E_IV, 1)

				ws['AW30'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5, razryad)
				ws['AW31'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				ws['AW32'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AW33'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AW33'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AX27'].value = my_round(mean_E_V, razryad)
				ws['AX28'].value = my_round(Er_E_V, 1)
				ws['AX29'].value = my_round(sigma_E_V, 1)

				ws['AX30'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5, razryad)
				ws['AX31'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				ws['AX32'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AX33'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AX33'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AY27'].value = my_round(mean_E_VI, razryad)
				ws['AY28'].value = my_round(Er_E_VI, 1)
				ws['AY29'].value = my_round(sigma_E_VI, 1)

				ws['AY30'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5, razryad)
				ws['AY31'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				ws['AY32'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AY33'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AY33'].value = '≤15'

		if check_var_gr_3.get() == 10:
			ws['AS28'].value = 'найдено, сред.зн., \nнг/мл (n=10)'
			ws['AS28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS29'].value = 'εR, % (n=10)'
			ws['AS30'].value = 'σR, % (n=10)'
			
			ws['AS31'].value = 'найдено, сред.зн.,\nнг/мл (n=50)'
			ws['AS31'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['AS32'].value = 'εR, % (n=50)'
			ws['AS33'].value = 'σR, % (n=50)'
			ws['AS34'].value = 'Норма |ε| и |σ|, %'

			ws['AT28'].value = my_round(mean_E_I, razryad)
			ws['AT29'].value = my_round(Er_E_I, 1)
			ws['AT30'].value = my_round(sigma_E_I, 1)

			ws['AT31'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5, razryad)
			ws['AT32'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			ws['AT33'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, 1)
			if entr_NORM_QCA1.get() != '':
				ws['AT34'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['AT34'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['AU28'].value = my_round(mean_E_II, razryad)
				ws['AU29'].value = my_round(Er_E_II, 1)
				ws['AU30'].value = my_round(sigma_E_II, 1)

				ws['AU31'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5, razryad)
				ws['AU32'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				ws['AU33'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, 1)
				if entr_NORM_QCB1.get() != '':
					ws['AU34'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['AU34'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['AV28'].value = my_round(mean_E_III, razryad)
				ws['AV29'].value = my_round(Er_E_III, 1)
				ws['AV30'].value = my_round(sigma_E_III, 1)

				ws['AV31'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5, razryad)
				ws['AV32'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				ws['AV33'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, 1)
				if entr_NORM_QCC1.get() != '':
					ws['AV34'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['AV34'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['AW28'].value = my_round(mean_E_IV, razryad)
				ws['AW29'].value = my_round(Er_E_IV, 1)
				ws['AW30'].value = my_round(sigma_E_IV, 1)

				ws['AW31'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5, razryad)
				ws['AW32'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				ws['AW33'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, 1)
				if entr_NORM_QCD1.get() != '':
					ws['AW34'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['AW34'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['AX28'].value = my_round(mean_E_V, razryad)
				ws['AX29'].value = my_round(Er_E_V, 1)
				ws['AX30'].value = my_round(sigma_E_V, 1)

				ws['AX31'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5, razryad)
				ws['AX32'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				ws['AX33'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, 1)
				if entr_NORM_QCE1.get() != '':
					ws['AX34'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['AX34'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['AY28'].value = my_round(mean_E_VI, razryad)
				ws['AY29'].value = my_round(Er_E_VI, 1)
				ws['AY30'].value = my_round(sigma_E_VI, 1)

				ws['AY31'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5, razryad)
				ws['AY32'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				ws['AY33'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, 1)
				if entr_NORM_QCF1.get() != '':
					ws['AY34'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['AY34'].value = '≤15'


	#QCA5
		ws['AT15'].value = 'QCA5'
		ws['AT16'].value = QC_I
		ws['G2'].value = 'QCA5'
		ws['AS18'].value = '1'
		ws['AT18'].value = my_round(float(entr_E1_I.get()), razryad)
		ws['G3'].value = my_round(float(entr_E1_I.get()), razryad)
		ws['AS19'].value = '2'
		ws['AT19'].value = my_round(float(entr_E2_I.get()), razryad)
		ws['G4'].value = my_round(float(entr_E2_I.get()), razryad)
		ws['AS20'].value = '3'
		ws['AT20'].value = my_round(float(entr_E3_I.get()), razryad)
		ws['G5'].value = my_round(float(entr_E3_I.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['AS21'].value = '4'
			ws['AT21'].value = my_round(float(entr_E4_I.get()), razryad)
			ws['G6'].value = my_round(float(entr_E4_I.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['AS22'].value = '5'
			ws['AT22'].value = my_round(float(entr_E5_I.get()), razryad)
			ws['G7'].value = my_round(float(entr_E5_I.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['AS23'].value = '6'
			ws['AT23'].value = my_round(float(entr_E6_I.get()), razryad)
			ws['G8'].value = my_round(float(entr_E6_I.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['AS24'].value = '7'
			ws['AT24'].value = my_round(float(entr_E7_I.get()), razryad)
			ws['G9'].value = my_round(float(entr_E7_I.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['AS25'].value = '8'
			ws['AT25'].value = my_round(float(entr_E8_I.get()), razryad)
			ws['G10'].value = my_round(float(entr_E8_I.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['AS26'].value = '9'
			ws['AT26'].value = my_round(float(entr_E9_I.get()), razryad)
			ws['G11'].value = my_round(float(entr_E9_I.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['AS27'].value = '10'
			ws['AT27'].value = my_round(float(entr_E10_I.get()), razryad)
			ws['G12'].value = my_round(float(entr_E10_I.get()), razryad)



		if check_var_gr_1.get() >= 2:
			#QCB5
			ws['AU15'].value = 'QCB5'
			ws['AU16'].value = QC_II
			ws['P2'].value = 'QCB5'
			ws['AU18'].value = my_round(float(entr_E1_II.get()), razryad)
			ws['P3'].value = my_round(float(entr_E1_II.get()), razryad)
			ws['AU19'].value = my_round(float(entr_E2_II.get()), razryad)
			ws['P4'].value = my_round(float(entr_E2_II.get()), razryad)
			ws['AU20'].value = my_round(float(entr_E3_II.get()), razryad)
			ws['P5'].value = my_round(float(entr_E3_II.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AU21'].value = my_round(float(entr_E4_II.get()), razryad)
				ws['P6'].value = my_round(float(entr_E4_II.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AU22'].value = my_round(float(entr_E5_II.get()), razryad)
				ws['P7'].value = my_round(float(entr_E5_II.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AU23'].value = my_round(float(entr_E6_II.get()), razryad)
				ws['P8'].value = my_round(float(entr_E6_II.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AU24'].value = my_round(float(entr_E7_II.get()), razryad)
				ws['P9'].value = my_round(float(entr_E7_II.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AU25'].value = my_round(float(entr_E8_II.get()), razryad)
				ws['P10'].value = my_round(float(entr_E8_II.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AU26'].value = my_round(float(entr_E9_II.get()), razryad)
				ws['P11'].value = my_round(float(entr_E9_II.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AU27'].value = my_round(float(entr_E10_II.get()), razryad)
				ws['P12'].value = my_round(float(entr_E10_II.get()), razryad)

		if check_var_gr_1.get() >= 3:
			#QCC5
			ws['AV15'].value = 'QCC5'
			ws['AV16'].value = QC_III
			ws['Y2'].value = 'QCC5'
			ws['AV18'].value = my_round(float(entr_E1_III.get()), razryad)
			ws['Y3'].value = my_round(float(entr_E1_III.get()), razryad)
			ws['AV19'].value = my_round(float(entr_E2_III.get()), razryad)
			ws['Y4'].value = my_round(float(entr_E2_III.get()), razryad)
			ws['AV20'].value = my_round(float(entr_E3_III.get()), razryad)
			ws['Y5'].value = my_round(float(entr_E3_III.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AV21'].value = my_round(float(entr_E4_III.get()), razryad)
				ws['Y6'].value = my_round(float(entr_E4_III.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AV22'].value = my_round(float(entr_E5_III.get()), razryad)
				ws['Y7'].value = my_round(float(entr_E5_III.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AV23'].value = my_round(float(entr_E6_III.get()), razryad)
				ws['Y8'].value = my_round(float(entr_E6_III.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AV24'].value = my_round(float(entr_E7_III.get()), razryad)
				ws['Y9'].value = my_round(float(entr_E7_III.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AV25'].value = my_round(float(entr_E8_III.get()), razryad)
				ws['Y10'].value = my_round(float(entr_E8_III.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AV26'].value = my_round(float(entr_E9_III.get()), razryad)
				ws['Y11'].value = my_round(float(entr_E9_III.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AV27'].value = my_round(float(entr_E10_III.get()), razryad)
				ws['Y12'].value = my_round(float(entr_E10_III.get()), razryad)

		if check_var_gr_1.get() >= 4:
			#QCD5
			ws['AW15'].value = 'QCD5'
			ws['AW16'].value = QC_IV
			ws['AH2'].value = 'QCD5'
			ws['AW18'].value = my_round(float(entr_E1_IV.get()), razryad)
			ws['AH3'].value = my_round(float(entr_E1_IV.get()), razryad)
			ws['AW19'].value = my_round(float(entr_E2_IV.get()), razryad)
			ws['AH4'].value = my_round(float(entr_E2_IV.get()), razryad)
			ws['AW20'].value = my_round(float(entr_E3_IV.get()), razryad)
			ws['AH5'].value = my_round(float(entr_E3_IV.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AW21'].value = my_round(float(entr_E4_IV.get()), razryad)
				ws['AH6'].value = my_round(float(entr_E4_IV.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AW22'].value = my_round(float(entr_E5_IV.get()), razryad)
				ws['AH7'].value = my_round(float(entr_E5_IV.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AW23'].value = my_round(float(entr_E6_IV.get()), razryad)
				ws['AH8'].value = my_round(float(entr_E6_IV.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AW24'].value = my_round(float(entr_E7_IV.get()), razryad)
				ws['AH9'].value = my_round(float(entr_E7_IV.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AW25'].value = my_round(float(entr_E8_IV.get()), razryad)
				ws['AH10'].value = my_round(float(entr_E8_IV.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AW26'].value = my_round(float(entr_E9_IV.get()), razryad)
				ws['AH11'].value = my_round(float(entr_E9_IV.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AW27'].value = my_round(float(entr_E10_IV.get()), razryad)
				ws['AH12'].value = my_round(float(entr_E10_IV.get()), razryad)

		if check_var_gr_1.get() >= 5:
			#QCE5
			ws['AX15'].value = 'QCE5'
			ws['AX16'].value = QC_V
			ws['AQ2'].value = 'QCE5'
			ws['AX18'].value = my_round(float(entr_E1_V.get()), razryad)
			ws['AQ3'].value = my_round(float(entr_E1_V.get()), razryad)
			ws['AX19'].value = my_round(float(entr_E2_V.get()), razryad)
			ws['AQ4'].value = my_round(float(entr_E2_V.get()), razryad)
			ws['AX20'].value = my_round(float(entr_E3_V.get()), razryad)
			ws['AQ5'].value = my_round(float(entr_E3_V.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['AX21'].value = my_round(float(entr_E4_V.get()), razryad)
				ws['AQ6'].value = my_round(float(entr_E4_V.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AX22'].value = my_round(float(entr_E5_V.get()), razryad)
				ws['AQ7'].value = my_round(float(entr_E5_V.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AX23'].value = my_round(float(entr_E6_V.get()), razryad)
				ws['AQ8'].value = my_round(float(entr_E6_V.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AX24'].value = my_round(float(entr_E7_V.get()), razryad)
				ws['AQ9'].value = my_round(float(entr_E7_V.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AX25'].value = my_round(float(entr_E8_V.get()), razryad)
				ws['AQ10'].value = my_round(float(entr_E8_V.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AX26'].value = my_round(float(entr_E9_V.get()), razryad)
				ws['AQ11'].value = my_round(float(entr_E9_V.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AX27'].value = my_round(float(entr_E10_V.get()), razryad)
				ws['AQ12'].value = my_round(float(entr_E10_V.get()), razryad)

		if check_var_gr_1.get() >= 6:
			#QCF5
			ws['AY15'].value = 'QCF5'
			ws['AY16'].value = QC_VI
			ws['AZ2'].value = 'QCF5'
			ws['AY18'].value = my_round(float(entr_E1_VI.get()), razryad)
			ws['AZ3'].value = my_round(float(entr_E1_VI.get()), razryad)
			ws['AY19'].value = my_round(float(entr_E2_VI.get()), razryad)
			ws['AZ4'].value = my_round(float(entr_E2_VI.get()), razryad)
			ws['AY20'].value = my_round(float(entr_E3_VI.get()), razryad)
			ws['AZ5'].value = my_round(float(entr_E3_VI.get()), razryad)

			if check_var_gr_3.get() >= 4:
				ws['AY21'].value = my_round(float(entr_E4_VI.get()), razryad)
				ws['AZ6'].value = my_round(float(entr_E4_VI.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['AY22'].value = my_round(float(entr_E5_VI.get()), razryad)
				ws['AZ7'].value = my_round(float(entr_E5_VI.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['AY23'].value = my_round(float(entr_E6_VI.get()), razryad)
				ws['AZ8'].value = my_round(float(entr_E6_VI.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['AY24'].value = my_round(float(entr_E7_VI.get()), razryad)
				ws['AZ9'].value = my_round(float(entr_E7_VI.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['AY25'].value = my_round(float(entr_E8_VI.get()), razryad)
				ws['AZ10'].value = my_round(float(entr_E8_VI.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['AY26'].value = my_round(float(entr_E9_VI.get()), razryad)
				ws['AZ11'].value = my_round(float(entr_E9_VI.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['AY27'].value = my_round(float(entr_E10_VI.get()), razryad)
				ws['AZ12'].value = my_round(float(entr_E10_VI.get()), razryad)



#QC ТАБЛИЦА № 6
	if check_var_gr_2.get() >= 6:
	#выравнивание по центру]
		cols_c(ws, 'BA15:BG34')

		#границы ячеек
		if check_var_gr_1.get() == 1:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'BA15:BB27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'BA15:BB28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'BA15:BB29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'BA15:BB30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'BA15:BB31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'BA15:BB32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'BA15:BB33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'BA15:BB34')

		if check_var_gr_1.get() == 2:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'BA15:BC27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'BA15:BC28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'BA15:BC29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'BA15:BC30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'BA15:BC31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'BA15:BC32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'BA15:BC33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'BA15:BC34')

		if check_var_gr_1.get() == 3:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'BA15:BD27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'BA15:BD28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'BA15:BD29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'BA15:BD30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'BA15:BD31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'BA15:BD32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'BA15:BD33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'BA15:BD34')				

		if check_var_gr_1.get() == 4:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'BA15:BE27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'BA15:BE28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'BA15:BE29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'BA15:BE30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'BA15:BE31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'BA15:BE32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'BA15:BE33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'BA15:BE34')			

		if check_var_gr_1.get() == 5:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'BA15:BF27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'BA15:BF28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'BA15:BF29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'BA15:BF30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'BA15:BF31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'BA15:BF32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'BA15:BF33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'BA15:BF34')		

		if check_var_gr_1.get() == 6:
			if check_var_gr_3.get() == 3:
				thin_border(ws, 'BA15:BG27')
			if check_var_gr_3.get() == 4:
				thin_border(ws, 'BA15:BG28')
			if check_var_gr_3.get() == 5:
				thin_border(ws, 'BA15:BG29')	
			if check_var_gr_3.get() == 6:
				thin_border(ws, 'BA15:BG30')
			if check_var_gr_3.get() == 7:
				thin_border(ws, 'BA15:BG31')	
			if check_var_gr_3.get() == 8:
				thin_border(ws, 'BA15:BG32')
			if check_var_gr_3.get() == 9:
				thin_border(ws, 'BA15:BG33')
			if check_var_gr_3.get() == 10:
				thin_border(ws, 'BA15:BG34')	

		
		#задание ширины столбца
		#ws.column_dimensions['BA'].width = 22

		#надпись и объединение ячеек под надпись
		ws['BA15'].value = 'QC'
		ws['BA16'].value = f'Введено, {entr_concent.get()}'
		ws['BA17'].value = '№ измерения'
		ws['BB17'].value = 'Найдено, последовательность 6'



#ЗАПОЛНЕНИЕ ANOVA
	
	#количество таблиц зависит от показателя check_var_gr_2
	if check_var_gr_2.get() >= 6:
		#QCA6
		mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6 = my_round(stat.mean([mean_A_I, mean_B_I, mean_C_I, mean_D_I, mean_E_I, mean_F_I]), check_razryad.get())
		#междугруппами
		sum_kv_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6 = len(spisok_A_I)*((mean_A_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6)**2) + len(spisok_B_I)*((mean_B_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6)**2) + len(spisok_C_I)*(mean_C_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6)**2 + len(spisok_D_I)*((mean_D_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6)**2) + len(spisok_E_I)*((mean_E_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6)**2) + len(spisok_F_I)*((mean_F_I - mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6)**2)
		CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6 = (((sum_kv_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6)/5)**0.5)/mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6*100
		#внутригрупп
		CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6 = (mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6 - QC_I)/QC_I * 100

		if check_var_gr_1.get() >= 2:
			#QCB6
			mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6 = my_round(stat.mean([mean_A_II, mean_B_II, mean_C_II, mean_D_II, mean_E_II, mean_F_II]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6 = len(spisok_A_II)*((mean_A_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6)**2) + len(spisok_B_II)*((mean_B_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6)**2) + len(spisok_C_II)*(mean_C_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6)**2 + len(spisok_D_II)*((mean_D_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6)**2) + len(spisok_E_II)*((mean_E_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6)**2) + len(spisok_F_II)*((mean_F_II - mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6)**2)
			CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6 = (((sum_kv_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6)/5)**0.5)/mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6*100
			#внутригрупп
			CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6 = (mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6 - QC_II)/QC_II * 100

		if check_var_gr_1.get() >= 3:
			#QCC6
			mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6 = my_round(stat.mean([mean_A_III, mean_B_III, mean_C_III, mean_D_III, mean_E_III, mean_F_III]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6 = len(spisok_A_III)*((mean_A_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6)**2) + len(spisok_B_III)*((mean_B_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6)**2) + len(spisok_C_III)*(mean_C_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6)**2 + len(spisok_D_III)*((mean_D_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6)**2) + len(spisok_E_III)*((mean_E_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6)**2) + len(spisok_F_III)*((mean_F_III - mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6)**2)
			CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6 = (((sum_kv_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6)/5)**0.5)/mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6*100
			#внутригрупп
			CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6 = (mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6 - QC_III)/QC_III * 100

		if check_var_gr_1.get() >= 4:
			#QCD6
			mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6 = my_round(stat.mean([mean_A_IV, mean_B_IV, mean_C_IV, mean_D_IV, mean_E_IV, mean_F_IV]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6 = len(spisok_A_IV)*((mean_A_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6)**2) + len(spisok_B_IV)*((mean_B_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6)**2) + len(spisok_C_IV)*(mean_C_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6)**2 + len(spisok_D_IV)*((mean_D_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6)**2) + len(spisok_E_IV)*((mean_E_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6)**2) + len(spisok_F_IV)*((mean_F_IV - mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6)**2)
			CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6 = (((sum_kv_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6)/5)**0.5)/mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6*100
			#внутригрупп
			CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6 = (mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6 - QC_IV)/QC_IV * 100

		if check_var_gr_1.get() >= 5:
			#QCE6
			mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6 = my_round(stat.mean([mean_A_V, mean_B_V, mean_C_V, mean_D_V, mean_E_V, mean_F_V]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6 = len(spisok_A_V)*((mean_A_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6)**2) + len(spisok_B_V)*((mean_B_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6)**2) + len(spisok_C_V)*(mean_C_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6)**2 + len(spisok_D_V)*((mean_D_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6)**2) + len(spisok_E_V)*((mean_E_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6)**2) + len(spisok_F_V)*((mean_F_V - mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6)**2)
			CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6 = (((sum_kv_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6)/5)**0.5)/mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6*100
			#внутригрупп
			CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6 = (mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6 - QC_V)/QC_V * 100

		if check_var_gr_1.get() >= 6:
			#QCF6
			mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6 = my_round(stat.mean([mean_A_VI, mean_B_VI, mean_C_VI, mean_D_VI, mean_E_VI, mean_F_VI]), check_razryad.get())
			#междугруппами
			sum_kv_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6 = len(spisok_A_VI)*((mean_A_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6)**2) + len(spisok_B_VI)*((mean_B_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6)**2) + len(spisok_C_VI)*(mean_C_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6)**2 + len(spisok_D_VI)*((mean_D_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6)**2) + len(spisok_E_VI)*((mean_E_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6)**2) + len(spisok_F_VI)*((mean_F_VI - mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6)**2)
			CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6 = (((sum_kv_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6)/5)**0.5)/mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6*100
			#внутригрупп
			CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6 = (mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6 - QC_VI)/QC_VI * 100


#вывод среднего и остальных показателей в таблицу 
		if check_var_gr_3.get() == 3:

			ws['BA21'].value = 'найдено, сред.зн.,\nнг/мл (n=3)'
			ws['BA21'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

			ws['BA22'].value = 'εR, % (n=3)'
			ws['BA23'].value = 'σR, % (n=3)'

			ws['BA24'].value = 'найдено, сред.зн.,\nнг/мл (n=18)'
			ws['BA24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA25'].value = 'εR, % (n=18)'
			ws['BA26'].value = 'σR, % (n=18)'
			ws['BA27'].value = 'Норма |ε| и |σ|, %'

			ws['BB21'].value = my_round(mean_F_I, razryad)
			ws['BB22'].value = my_round(Er_F_I, 1)
			ws['BB23'].value = my_round(sigma_F_I, 1)

			ws['BB24'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, razryad)
			ws['BB25'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			ws['BB26'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			if entr_NORM_QCA1.get() != '':
				ws['BB27'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['BB27'].value = '≤20'
			
			if check_var_gr_1.get() >= 2:
				ws['BC21'].value = my_round(mean_F_II, razryad)
				ws['BC22'].value = my_round(Er_F_II, 1)
				ws['BC23'].value = my_round(sigma_F_II, 1)

				ws['BC24'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, razryad)
				ws['BC25'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				ws['BC26'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				if entr_NORM_QCB1.get() != '':
					ws['BC27'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['BC27'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['BD21'].value = my_round(mean_F_III, razryad)
				ws['BD22'].value = my_round(Er_F_III, 1)
				ws['BD23'].value = my_round(sigma_F_III, 1)

				ws['BD24'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, razryad)
				ws['BD25'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				ws['BD26'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				if entr_NORM_QCC1.get() != '':
					ws['BD27'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['BD27'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['BE21'].value = my_round(mean_F_IV, razryad)
				ws['BE22'].value = my_round(Er_F_IV, 1)
				ws['BE23'].value = my_round(sigma_F_IV, 1)

				ws['BE24'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, razryad)
				ws['BE25'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				ws['BE26'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				if entr_NORM_QCD1.get() != '':
					ws['BE27'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['BE27'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['BF21'].value = my_round(mean_F_V, razryad)
				ws['BF22'].value = my_round(Er_F_V, 1)
				ws['BF23'].value = my_round(sigma_F_V, 1)

				ws['BF24'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, razryad)
				ws['BF25'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				ws['BF26'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				if entr_NORM_QCE1.get() != '':
					ws['BF27'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['BF27'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['BG21'].value = my_round(mean_F_VI, razryad)
				ws['BG22'].value = my_round(Er_F_VI, 1)
				ws['BG23'].value = my_round(sigma_F_VI, 1)

				ws['BG24'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, razryad)
				ws['BG25'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				ws['BG26'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				if entr_NORM_QCF1.get() != '':
					ws['BG27'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['BG27'].value = '≤15'


		if check_var_gr_3.get() == 4:

			ws['BA22'].value = 'найдено, сред.зн.,\nнг/мл (n=4)'
			ws['BA22'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA23'].value = 'εR, % (n=4)'
			ws['BA24'].value = 'σR, % (n=4)'
			
			ws['BA25'].value = 'найдено, сред.зн.,\nнг/мл (n=24)'
			ws['BA25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA26'].value = 'εR, % (n=24)'
			ws['BA27'].value = 'σR, % (n=24)'
			ws['BA28'].value = 'Норма |ε| и |σ|, %'


			ws['BB22'].value = my_round(mean_F_I, razryad)
			ws['BB23'].value = my_round(Er_F_I, 1)
			ws['BB24'].value = my_round(sigma_F_I, 1)

			ws['BB25'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, razryad)
			ws['BB26'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			ws['BB27'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			if entr_NORM_QCA1.get() != '':
				ws['BB28'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['BB28'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['BC22'].value = my_round(mean_F_II, razryad)
				ws['BC23'].value = my_round(Er_F_II, 1)
				ws['BC24'].value = my_round(sigma_F_II, 1)

				ws['BC25'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, razryad)
				ws['BC26'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				ws['BC27'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				if entr_NORM_QCB1.get() != '':
					ws['BC28'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['BC28'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['BD22'].value = my_round(mean_F_III, razryad)
				ws['BD23'].value = my_round(Er_F_III, 1)
				ws['BD24'].value = my_round(sigma_F_III, 1)

				ws['BD25'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, razryad)
				ws['BD26'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				ws['BD27'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				if entr_NORM_QCC1.get() != '':
					ws['BD28'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['BD28'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['BE22'].value = my_round(mean_F_IV, razryad)
				ws['BE23'].value = my_round(Er_F_IV, 1)
				ws['BE24'].value = my_round(sigma_F_IV, 1)

				ws['BE25'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, razryad)
				ws['BE26'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				ws['BE27'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				if entr_NORM_QCD1.get() != '':
					ws['BE28'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['BE28'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['BF22'].value = my_round(mean_F_V, razryad)
				ws['BF23'].value = my_round(Er_F_V, 1)
				ws['BF24'].value = my_round(sigma_F_V, 1)

				ws['BF25'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, razryad)
				ws['BF26'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				ws['BF27'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				if entr_NORM_QCE1.get() != '':
					ws['BF28'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['BF28'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['BG22'].value = my_round(mean_F_VI, razryad)
				ws['BG23'].value = my_round(Er_F_VI, 1)
				ws['BG24'].value = my_round(sigma_F_VI, 1)

				ws['BG25'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, razryad)
				ws['BG26'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				ws['BG27'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				if entr_NORM_QCF1.get() != '':
					ws['BG28'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['BG28'].value = '≤15'

		if check_var_gr_3.get() == 5:
			ws['BA23'].value = 'найдено, сред.зн., \nнг/мл (n=5)'
			ws['BA23'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA24'].value = 'εR, % (n=5)'
			ws['BA25'].value = 'σR, % (n=5)'
			
			ws['BA26'].value = 'найдено, сред.зн.,\nнг/мл (n=30)'
			ws['BA26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA27'].value = 'εR, % (n=30)'
			ws['BA28'].value = 'σR, % (n=30)'
			ws['BA29'].value = 'Норма |ε| и |σ|, %'

			ws['BB23'].value = my_round(mean_F_I, razryad)
			ws['BB24'].value = my_round(Er_F_I, 1)
			ws['BB25'].value = my_round(sigma_F_I, 1)

			ws['BB26'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, razryad)
			ws['BB27'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			ws['BB28'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			if entr_NORM_QCA1.get() != '':
				ws['BB29'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['BB29'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['BC23'].value = my_round(mean_F_II, razryad)
				ws['BC24'].value = my_round(Er_F_II, 1)
				ws['BC25'].value = my_round(sigma_F_II, 1)

				ws['BC26'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, razryad)
				ws['BC27'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				ws['BC28'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				if entr_NORM_QCB1.get() != '':
					ws['BC29'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['BC29'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['BD23'].value = my_round(mean_F_III, razryad)
				ws['BD24'].value = my_round(Er_F_III, 1)
				ws['BD25'].value = my_round(sigma_F_III, 1)

				ws['BD26'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, razryad)
				ws['BD27'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				ws['BD28'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				if entr_NORM_QCC1.get() != '':
					ws['BD29'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['BD29'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['BE23'].value = my_round(mean_F_IV, razryad)
				ws['BE24'].value = my_round(Er_F_IV, 1)
				ws['BE25'].value = my_round(sigma_F_IV, 1)

				ws['BE26'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, razryad)
				ws['BE27'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				ws['BE28'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				if entr_NORM_QCD1.get() != '':
					ws['BE29'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['BE29'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['BF23'].value = my_round(mean_F_V, razryad)
				ws['BF24'].value = my_round(Er_F_V, 1)
				ws['BF25'].value = my_round(sigma_F_V, 1)

				ws['BF26'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, razryad)
				ws['BF27'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				ws['BF28'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				if entr_NORM_QCE1.get() != '':
					ws['BF29'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['BF29'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['BG23'].value = my_round(mean_F_VI, razryad)
				ws['BG24'].value = my_round(Er_F_VI, 1)
				ws['BG25'].value = my_round(sigma_F_VI, 1)

				ws['BG26'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, razryad)
				ws['BG27'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				ws['BG28'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				if entr_NORM_QCF1.get() != '':
					ws['BG29'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['BG29'].value = '≤15'

		if check_var_gr_3.get() == 6:
			ws['BA24'].value = 'найдено, сред.зн., \nнг/мл (n=6)'
			ws['BA24'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA25'].value = 'εR, % (n=6)'
			ws['BA26'].value = 'σR, % (n=6)'
			
			ws['BA27'].value = 'найдено, сред.зн.,\nнг/мл (n=36)'
			ws['BA27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA28'].value = 'εR, % (n=36)'
			ws['BA28'].value = 'σR, % (n=36)'
			ws['BA29'].value = 'Норма |ε| и |σ|, %'

			ws['BB24'].value = my_round(mean_F_I, razryad)
			ws['BB25'].value = my_round(Er_F_I, 1)
			ws['BB26'].value = my_round(sigma_F_I, 1)

			ws['BB27'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, razryad)
			ws['BB28'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			ws['BB29'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			if entr_NORM_QCA1.get() != '':
				ws['BB30'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['BB30'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['BC24'].value = my_round(mean_F_II, razryad)
				ws['BC25'].value = my_round(Er_F_II, 1)
				ws['BC26'].value = my_round(sigma_F_II, 1)

				ws['BC27'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, razryad)
				ws['BC28'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				ws['BC29'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				if entr_NORM_QCB1.get() != '':
					ws['BC30'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['BC30'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['BD24'].value = my_round(mean_F_III, razryad)
				ws['BD25'].value = my_round(Er_F_III, 1)
				ws['BD26'].value = my_round(sigma_F_III, 1)

				ws['BD27'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, razryad)
				ws['BD28'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				ws['BD29'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				if entr_NORM_QCC1.get() != '':
					ws['BD30'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['BD30'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['BE24'].value = my_round(mean_F_IV, razryad)
				ws['BE25'].value = my_round(Er_F_IV, 1)
				ws['BE26'].value = my_round(sigma_F_IV, 1)

				ws['BE27'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, razryad)
				ws['BE28'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				ws['BE29'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				if entr_NORM_QCD1.get() != '':
					ws['BE30'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['BE30'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['BF24'].value = my_round(mean_F_V, razryad)
				ws['BF25'].value = my_round(Er_F_V, 1)
				ws['BF26'].value = my_round(sigma_F_V, 1)

				ws['BF27'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, razryad)
				ws['BF28'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				ws['BF29'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				if entr_NORM_QCE1.get() != '':
					ws['BF30'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['BF30'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['BG24'].value = my_round(mean_F_VI, razryad)
				ws['BG25'].value = my_round(Er_F_VI, 1)
				ws['BG26'].value = my_round(sigma_F_VI, 1)

				ws['BG27'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, razryad)
				ws['BG28'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				ws['BG29'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				if entr_NORM_QCF1.get() != '':
					ws['BG30'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['BG30'].value = '≤15'

		if check_var_gr_3.get() == 7:
			ws['BA25'].value = 'найдено, сред.зн., \nнг/мл (n=7)'
			ws['BA25'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA26'].value = 'εR, % (n=7)'
			ws['BA27'].value = 'σR, % (n=7)'
			
			ws['BA28'].value = 'найдено, сред.зн.,\nнг/мл (n=42)'
			ws['BA28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA29'].value = 'εR, % (n=42)'
			ws['BA30'].value = 'σR, % (n=42)'
			ws['BA31'].value = 'Норма |ε| и |σ|, %'

			ws['BB25'].value = my_round(mean_F_I, razryad)
			ws['BB26'].value = my_round(Er_F_I, 1)
			ws['BB27'].value = my_round(sigma_F_I, 1)

			ws['BB28'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, razryad)
			ws['BB29'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			ws['BB30'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			if entr_NORM_QCA1.get() != '':
				ws['BB31'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['BB31'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['BC25'].value = my_round(mean_F_II, razryad)
				ws['BC26'].value = my_round(Er_F_II, 1)
				ws['BC27'].value = my_round(sigma_F_II, 1)

				ws['BC28'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, razryad)
				ws['BC29'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				ws['BC30'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				if entr_NORM_QCB1.get() != '':
					ws['BC31'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['BC31'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['BD25'].value = my_round(mean_F_III, razryad)
				ws['BD26'].value = my_round(Er_F_III, 1)
				ws['BD27'].value = my_round(sigma_F_III, 1)

				ws['BD28'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, razryad)
				ws['BD29'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				ws['BD30'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				if entr_NORM_QCC1.get() != '':
					ws['BD31'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['BD31'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['BE25'].value = my_round(mean_F_IV, razryad)
				ws['BE26'].value = my_round(Er_F_IV, 1)
				ws['BE27'].value = my_round(sigma_F_IV, 1)

				ws['BE28'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, razryad)
				ws['BE29'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				ws['BE30'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				if entr_NORM_QCD1.get() != '':
					ws['BE31'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['BE31'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['BF25'].value = my_round(mean_F_V, razryad)
				ws['BF26'].value = my_round(Er_F_V, 1)
				ws['BF27'].value = my_round(sigma_F_V, 1)

				ws['BF28'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, razryad)
				ws['BF29'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				ws['BF30'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				if entr_NORM_QCE1.get() != '':
					ws['BF31'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['BF31'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['BG25'].value = my_round(mean_F_VI, razryad)
				ws['BG26'].value = my_round(Er_F_VI, 1)
				ws['BG27'].value = my_round(sigma_F_VI, 1)

				ws['BG28'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, razryad)
				ws['BG29'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				ws['BG30'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				if entr_NORM_QCF1.get() != '':
					ws['BG31'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['BG31'].value = '≤15'

		if check_var_gr_3.get() == 8:
			ws['BA26'].value = 'найдено, сред.зн., \nнг/мл (n=8)'
			ws['BA26'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA27'].value = 'εR, % (n=8)'
			ws['BA28'].value = 'σR, % (n=8)'
			
			ws['BA29'].value = 'найдено, сред.зн.,\nнг/мл (n=48)'
			ws['BA29'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA30'].value = 'εR, % (n=48)'
			ws['BA30'].value = 'σR, % (n=48)'
			ws['BA31'].value = 'Норма |ε| и |σ|, %'

			ws['BB26'].value = my_round(mean_F_I, razryad)
			ws['BB27'].value = my_round(Er_F_I, 1)
			ws['BB28'].value = my_round(sigma_F_I, 1)

			ws['BB29'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, razryad)
			ws['BB30'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			ws['BB31'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			if entr_NORM_QCA1.get() != '':
				ws['BB32'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['BB32'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['BC26'].value = my_round(mean_F_II, razryad)
				ws['BC27'].value = my_round(Er_F_II, 1)
				ws['BC28'].value = my_round(sigma_F_II, 1)

				ws['BC29'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, razryad)
				ws['BC30'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				ws['BC31'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				if entr_NORM_QCB1.get() != '':
					ws['BC32'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['BC32'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['BD26'].value = my_round(mean_F_III, razryad)
				ws['BD27'].value = Er_F_III
				ws['BD28'].value = sigma_F_III

				ws['BD29'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, razryad)
				ws['BD30'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				ws['BD31'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				if entr_NORM_QCC1.get() != '':
					ws['BD32'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['BD32'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['BE26'].value = my_round(mean_F_IV, razryad)
				ws['BE27'].value = my_round(Er_F_IV, 1)
				ws['BE28'].value = my_round(sigma_F_IV, 1)

				ws['BE29'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, razryad)
				ws['BE30'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				ws['BE31'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				if entr_NORM_QCD1.get() != '':
					ws['BE32'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['BE32'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['BF26'].value = my_round(mean_F_V, razryad)
				ws['BF27'].value = my_round(Er_F_V, 1)
				ws['BF28'].value = my_round(sigma_F_V, 1)

				ws['BF29'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, razryad)
				ws['BF30'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				ws['BF31'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				if entr_NORM_QCE1.get() != '':
					ws['BF32'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['BF32'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['BG26'].value = my_round(mean_F_VI, razryad)
				ws['BG27'].value = my_round(Er_F_VI, 1)
				ws['BG28'].value = my_round(sigma_F_VI, 1)

				ws['BG29'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, razryad)
				ws['BG30'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				ws['BG31'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				if entr_NORM_QCF1.get() != '':
					ws['BG32'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['BG32'].value = '≤15'

		if check_var_gr_3.get() == 9:
			ws['BA27'].value = 'найдено, сред.зн., \nнг/мл (n=9)'
			ws['BA27'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA28'].value = 'εR, % (n=9)'
			ws['BA29'].value = 'σR, % (n=9)'
			
			ws['BA30'].value = 'найдено, сред.зн.,\nнг/мл (n=54)'
			ws['BA30'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA31'].value = 'εR, % (n=54)'
			ws['BA32'].value = 'σR, % (n=54)'
			ws['BA33'].value = 'Норма |ε| и |σ|, %'

			ws['BB27'].value = my_round(mean_F_I, razryad)
			ws['BB28'].value = my_round(Er_F_I, 1)
			ws['BB29'].value = my_round(sigma_F_I, 1)

			ws['BB30'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, razryad)
			ws['BB31'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			ws['BB32'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			if entr_NORM_QCA1.get() != '':
				ws['BB33'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['BB33'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['BC27'].value = my_round(mean_F_II, razryad)
				ws['BC28'].value = my_round(Er_F_II, 1)
				ws['BC29'].value = my_round(sigma_F_II, 1)

				ws['BC30'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, razryad)
				ws['BC31'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				ws['BC32'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				if entr_NORM_QCB1.get() != '':
					ws['BC33'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['BC33'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['BD27'].value = my_round(mean_F_III, razryad)
				ws['BD28'].value = my_round(Er_F_III, 1)
				ws['BD29'].value = my_round(sigma_F_III, 1)

				ws['BD30'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, razryad)
				ws['BD31'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				ws['BD32'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				if entr_NORM_QCC1.get() != '':
					ws['BD33'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['BD33'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['BE27'].value = my_round(mean_F_IV, razryad)
				ws['BE28'].value = my_round(Er_F_IV, 1)
				ws['BE29'].value = my_round(sigma_F_IV, 1)

				ws['BE30'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, razryad)
				ws['BE31'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				ws['BE32'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				if entr_NORM_QCD1.get() != '':
					ws['BE33'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['BE33'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['BF27'].value = my_round(mean_F_V, razryad)
				ws['BF28'].value = my_round(Er_F_V, 1)
				ws['BF29'].value = my_round(sigma_F_V, 1)

				ws['BF30'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, razryad)
				ws['BF31'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				ws['BF32'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				if entr_NORM_QCE1.get() != '':
					ws['BF33'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['BF33'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['BG27'].value = my_round(mean_F_VI, razryad)
				ws['BG28'].value = my_round(Er_F_VI, 1)
				ws['BG29'].value = my_round(sigma_F_VI, 1)

				ws['BG30'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, razryad)
				ws['BG31'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				ws['BG32'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				if entr_NORM_QCF1.get() != '':
					ws['BG33'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['BG33'].value = '≤15'

		if check_var_gr_3.get() == 10:
			ws['BA28'].value = 'найдено, сред.зн., \nнг/мл (n=10)'
			ws['BA28'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA29'].value = 'εR, % (n=10)'
			ws['BA30'].value = 'σR, % (n=10)'
			
			ws['BA31'].value = 'найдено, сред.зн.,\nнг/мл (n=60)'
			ws['BA31'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
			ws['BA32'].value = 'εR, % (n=60)'
			ws['BA33'].value = 'σR, % (n=60)'
			ws['BA34'].value = 'Норма |ε| и |σ|, %'

			ws['BB28'].value = my_round(mean_F_I, razryad)
			ws['BB29'].value = my_round(Er_F_I, 1)
			ws['BB30'].value = my_round(sigma_F_I, 1)

			ws['BB31'].value = my_round(mean_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, razryad)
			ws['BB32'].value = my_round(CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			ws['BB33'].value = my_round(CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6, 1)
			if entr_NORM_QCA1.get() != '':
				ws['BB34'].value = '≤{}'.format(entr_NORM_QCA1.get())
			else:
				ws['BB34'].value = '≤20'

			if check_var_gr_1.get() >= 2:
				ws['BC28'].value = my_round(mean_F_II, razryad)
				ws['BC29'].value = my_round(Er_F_II, 1)
				ws['BC30'].value = my_round(sigma_F_II, 1)

				ws['BC31'].value = my_round(mean_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, razryad)
				ws['BC32'].value = my_round(CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				ws['BC33'].value = my_round(CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6, 1)
				if entr_NORM_QCB1.get() != '':
					ws['BC34'].value = '≤{}'.format(entr_NORM_QCB1.get())
				else:
					ws['BC34'].value = '≤15'

			if check_var_gr_1.get() >= 3:	
				ws['BD28'].value = my_round(mean_F_III, razryad)
				ws['BD29'].value = my_round(Er_F_III, 1)
				ws['BD30'].value = my_round(sigma_F_III, 1)

				ws['BD31'].value = my_round(mean_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, razryad)
				ws['BD32'].value = my_round(CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				ws['BD33'].value = my_round(CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6, 1)
				if entr_NORM_QCC1.get() != '':
					ws['BD34'].value = '≤{}'.format(entr_NORM_QCC1.get())
				else:
					ws['BD34'].value = '≤15'

			if check_var_gr_1.get() >= 4:	
				ws['BE28'].value = my_round(mean_F_IV, razryad)
				ws['BE29'].value = my_round(Er_F_IV, 1)
				ws['BE30'].value = my_round(sigma_F_IV, 1)

				ws['BE31'].value = my_round(mean_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, razryad)
				ws['BE32'].value = my_round(CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				ws['BE33'].value = my_round(CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6, 1)
				if entr_NORM_QCD1.get() != '':
					ws['BE34'].value = '≤{}'.format(entr_NORM_QCD1.get())
				else:
					ws['BE34'].value = '≤15'

			if check_var_gr_1.get() >= 5:	
				ws['BF28'].value = my_round(mean_F_V, razryad)
				ws['BF29'].value = my_round(Er_F_V, 1)
				ws['BF30'].value = my_round(sigma_F_V, 1)

				ws['BF31'].value = my_round(mean_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, razryad)
				ws['BF32'].value = my_round(CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				ws['BF33'].value = my_round(CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6, 1)
				if entr_NORM_QCE1.get() != '':
					ws['BF34'].value = '≤{}'.format(entr_NORM_QCE1.get())
				else:
					ws['BF34'].value = '≤15'

			if check_var_gr_1.get() >= 6:	
				ws['BG28'].value = my_round(mean_F_VI, razryad)
				ws['BG29'].value = my_round(Er_F_VI, 1)
				ws['BG30'].value = my_round(sigma_F_VI, 1)

				ws['BG31'].value = my_round(mean_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, razryad)
				ws['BG32'].value = my_round(CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				ws['BG33'].value = my_round(CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6, 1)
				if entr_NORM_QCF1.get() != '':
					ws['BG34'].value = '≤{}'.format(entr_NORM_QCF1.get())
				else:
					ws['BG34'].value = '≤15'


	#QCA6
		ws['BB15'].value = 'QCA6'
		ws['BB16'].value = QC_I
		ws['H2'].value = 'QCA6'
		ws['BA18'].value = '1'
		ws['BB18'].value = my_round(float(entr_F1_I.get()), razryad)
		ws['H3'].value = my_round(float(entr_F1_I.get()), razryad)
		ws['BA19'].value = '2'
		ws['BB19'].value = my_round(float(entr_F2_I.get()), razryad)
		ws['H4'].value = my_round(float(entr_F2_I.get()), razryad)
		ws['BA20'].value = '3'
		ws['BB20'].value = my_round(float(entr_F3_I.get()), razryad)
		ws['H5'].value = my_round(float(entr_F3_I.get()), razryad)
		if check_var_gr_3.get() >= 4:
			ws['BA21'].value = '4'
			ws['BB21'].value = my_round(float(entr_F4_I.get()), razryad)
			ws['H6'].value = my_round(float(entr_F4_I.get()), razryad)
		if check_var_gr_3.get() >= 5:
			ws['BA22'].value = '5'
			ws['BB22'].value = my_round(float(entr_F5_I.get()), razryad)
			ws['H7'].value = my_round(float(entr_F5_I.get()), razryad)
		if check_var_gr_3.get() >= 6:	
			ws['BA23'].value = '6'
			ws['BB23'].value = my_round(float(entr_F6_I.get()), razryad)
			ws['H8'].value = my_round(float(entr_F6_I.get()), razryad)
		if check_var_gr_3.get() >= 7:	
			ws['BA24'].value = '7'
			ws['BB24'].value = my_round(float(entr_F7_I.get()), razryad)
			ws['H9'].value = my_round(float(entr_F7_I.get()), razryad)
		if check_var_gr_3.get() >= 8:	
			ws['BA25'].value = '8'
			ws['BB25'].value = my_round(float(entr_F8_I.get()), razryad)
			ws['H10'].value = my_round(float(entr_F8_I.get()), razryad)
		if check_var_gr_3.get() >= 9:	
			ws['BA26'].value = '9'
			ws['BB26'].value = my_round(float(entr_F9_I.get()), razryad)
			ws['H11'].value = my_round(float(entr_F9_I.get()), razryad)
		if check_var_gr_3.get() >= 10:	
			ws['BA27'].value = '10'
			ws['BB27'].value = my_round(float(entr_F10_I.get()), razryad)
			ws['H12'].value = my_round(float(entr_F10_I.get()), razryad)



		if check_var_gr_1.get() >= 2:
			#QCB5
			ws['BC15'].value = 'QCB6'
			ws['BC16'].value = QC_II
			ws['Q2'].value = 'QCB6'
			ws['BC18'].value = my_round(float(entr_F1_II.get()), razryad)
			ws['Q3'].value = my_round(float(entr_F1_II.get()), razryad)
			ws['BC19'].value = my_round(float(entr_F2_II.get()), razryad)
			ws['Q4'].value = my_round(float(entr_F2_II.get()), razryad)
			ws['BC20'].value = my_round(float(entr_F3_II.get()), razryad)
			ws['Q5'].value = my_round(float(entr_F3_II.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['BC21'].value = my_round(float(entr_F4_II.get()), razryad)
				ws['Q6'].value = my_round(float(entr_F4_II.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['BC22'].value = my_round(float(entr_F5_II.get()), razryad)
				ws['Q7'].value = my_round(float(entr_F5_II.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['BC23'].value = my_round(float(entr_F6_II.get()), razryad)
				ws['Q8'].value = my_round(float(entr_F6_II.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['BC24'].value = my_round(float(entr_F7_II.get()), razryad)
				ws['Q9'].value = my_round(float(entr_F7_II.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['BC25'].value = my_round(float(entr_F8_II.get()), razryad)
				ws['Q10'].value = my_round(float(entr_F8_II.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['BC26'].value = my_round(float(entr_F9_II.get()), razryad)
				ws['Q11'].value = my_round(float(entr_F9_II.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['BC27'].value = my_round(float(entr_F10_II.get()), razryad)
				ws['Q12'].value = my_round(float(entr_F10_II.get()), razryad)

		if check_var_gr_1.get() >= 3:
			#QCC5
			ws['BD15'].value = 'QCC6'
			ws['BD16'].value = QC_III
			ws['Z2'].value = 'QCC6'
			ws['BD18'].value = my_round(float(entr_F1_III.get()), razryad)
			ws['Z3'].value = my_round(float(entr_F1_III.get()), razryad)
			ws['BD19'].value = my_round(float(entr_F2_III.get()), razryad)
			ws['Z4'].value = my_round(float(entr_F2_III.get()), razryad)
			ws['BD20'].value = my_round(float(entr_F3_III.get()), razryad)
			ws['Z5'].value = my_round(float(entr_F3_III.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['BD21'].value = my_round(float(entr_F4_III.get()), razryad)
				ws['Z6'].value = my_round(float(entr_F4_III.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['BD22'].value = my_round(float(entr_F5_III.get()), razryad)
				ws['Z7'].value = my_round(float(entr_F5_III.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['BD23'].value = my_round(float(entr_F6_III.get()), razryad)
				ws['Z8'].value = my_round(float(entr_F6_III.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['BD24'].value = my_round(float(entr_F7_III.get()), razryad)
				ws['Z9'].value = my_round(float(entr_F7_III.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['BD25'].value = my_round(float(entr_F8_III.get()), razryad)
				ws['Z10'].value = my_round(float(entr_F8_III.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['BD26'].value = my_round(float(entr_F9_III.get()), razryad)
				ws['Z11'].value = my_round(float(entr_F9_III.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['BD27'].value = my_round(float(entr_F10_III.get()), razryad)
				ws['Z12'].value = my_round(float(entr_F10_III.get()), razryad)

		if check_var_gr_1.get() >= 4:
			#QCD5
			ws['BE15'].value = 'QCD6'
			ws['BE16'].value = QC_IV
			ws['AI2'].value = 'QCD6'
			ws['BE18'].value = my_round(float(entr_F1_IV.get()), razryad)
			ws['AI3'].value = my_round(float(entr_F1_IV.get()), razryad)
			ws['BE19'].value = my_round(float(entr_F2_IV.get()), razryad)
			ws['AI4'].value = my_round(float(entr_F2_IV.get()), razryad)
			ws['BE20'].value = my_round(float(entr_F3_IV.get()), razryad)
			ws['AI5'].value = my_round(float(entr_F3_IV.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['BE21'].value = my_round(float(entr_F4_IV.get()), razryad)
				ws['AI6'].value = my_round(float(entr_F4_IV.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['BE22'].value = my_round(float(entr_F5_IV.get()), razryad)
				ws['AI7'].value = my_round(float(entr_F5_IV.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['BE23'].value = my_round(float(entr_F6_IV.get()), razryad)
				ws['AI8'].value = my_round(float(entr_F6_IV.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['BE24'].value = my_round(float(entr_F7_IV.get()), razryad)
				ws['AI9'].value = my_round(float(entr_F7_IV.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['BE25'].value = my_round(float(entr_F8_IV.get()), razryad)
				ws['AI10'].value = my_round(float(entr_F8_IV.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['BE26'].value = my_round(float(entr_F9_IV.get()), razryad)
				ws['AI11'].value = my_round(float(entr_F9_IV.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['BE27'].value = my_round(float(entr_F10_IV.get()), razryad)
				ws['AI12'].value = my_round(float(entr_F10_IV.get()), razryad)

		if check_var_gr_1.get() >= 5:
			#QCE5
			ws['BF15'].value = 'QCE6'
			ws['BF16'].value = QC_V
			ws['AR2'].value = 'QCE6'
			ws['BF18'].value = my_round(float(entr_F1_V.get()), razryad)
			ws['AR3'].value = my_round(float(entr_F1_V.get()), razryad)
			ws['BF19'].value = my_round(float(entr_F2_V.get()), razryad)
			ws['AR4'].value = my_round(float(entr_F2_V.get()), razryad)
			ws['BF20'].value = my_round(float(entr_F3_V.get()), razryad)
			ws['AR5'].value = my_round(float(entr_F3_V.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['BF21'].value = my_round(float(entr_F4_V.get()), razryad)
				ws['AR6'].value = my_round(float(entr_F4_V.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['BF22'].value = my_round(float(entr_F5_V.get()), razryad)
				ws['AR7'].value = my_round(float(entr_F5_V.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['BF23'].value = my_round(float(entr_F6_V.get()), razryad)
				ws['AR8'].value = my_round(float(entr_F6_V.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['BF24'].value = my_round(float(entr_F7_V.get()), razryad)
				ws['AR9'].value = my_round(float(entr_F7_V.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['BF25'].value = my_round(float(entr_F8_V.get()), razryad)
				ws['AR10'].value = my_round(float(entr_F8_V.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['BF26'].value = my_round(float(entr_F9_V.get()), razryad)
				ws['AR11'].value = my_round(float(entr_F9_V.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['BF27'].value = my_round(float(entr_F10_V.get()), razryad)
				ws['AR12'].value = my_round(float(entr_F10_V.get()), razryad)


		if check_var_gr_1.get() >= 6:
			#QCF6
			ws['BG15'].value = 'QCF6'
			ws['BG16'].value = QC_VI
			ws['BA2'].value = 'QCF6'
			ws['BG18'].value = my_round(float(entr_F1_VI.get()), razryad)
			ws['BA3'].value = my_round(float(entr_F1_VI.get()), razryad)
			ws['BG19'].value = my_round(float(entr_F2_VI.get()), razryad)
			ws['BA4'].value = my_round(float(entr_F2_VI.get()), razryad)
			ws['BG20'].value = my_round(float(entr_F3_VI.get()), razryad)
			ws['BA5'].value = my_round(float(entr_F3_VI.get()), razryad)
			if check_var_gr_3.get() >= 4:
				ws['BG21'].value = my_round(float(entr_F4_VI.get()), razryad)
				ws['BA6'].value = my_round(float(entr_F4_VI.get()), razryad)
			if check_var_gr_3.get() >= 5:
				ws['BG22'].value = my_round(float(entr_F5_VI.get()), razryad)
				ws['BA7'].value = my_round(float(entr_F5_VI.get()), razryad)
			if check_var_gr_3.get() >= 6:	
				ws['BG23'].value = my_round(float(entr_F6_VI.get()), razryad)
				ws['BA8'].value = my_round(float(entr_F6_VI.get()), razryad)
			if check_var_gr_3.get() >= 7:	
				ws['BG24'].value = my_round(float(entr_F7_VI.get()), razryad)
				ws['BA9'].value = my_round(float(entr_F7_VI.get()), razryad)
			if check_var_gr_3.get() >= 8:	
				ws['BG25'].value = my_round(float(entr_F8_VI.get()), razryad)
				ws['BA10'].value = my_round(float(entr_F8_VI.get()), razryad)
			if check_var_gr_3.get() >= 9:	
				ws['BG26'].value = my_round(float(entr_F9_VI.get()), razryad)
				ws['BA11'].value = my_round(float(entr_F9_VI.get()), razryad)
			if check_var_gr_3.get() >= 10:	
				ws['BG27'].value = my_round(float(entr_F10_VI.get()), razryad)
				ws['BA12'].value = my_round(float(entr_F10_VI.get()), razryad)



	#ТАБЛИЦА С ПРАВИЛЬНОСТЬЮ И ПОВТОРЯЕМОСТЬЮ
	cols_c(ws, 'A1:BA12')
	cols_c(ws, 'D16:I25')
	#ws.column_dimensions['A'].width = 13.5
	ws.merge_cells('A18:C18')
	ws['A18'].value = 'Внутридневная правильность, %'
	ws.merge_cells('A20:C20')
	ws['A20'].value = 'Междневная правильность, %'
	ws.merge_cells('A22:C22')
	ws['A22'].value = 'Внутридневная повторяемость, %'
	ws.merge_cells('A24:C24')
	ws['A24'].value = 'Междневная повторяемость, %'
	if check_var_gr_1.get() >= 1:
		ws['D16'].value = 'QCA'
		
		if check_var_gr_2.get() == 2:
			ws['D18'].value = my_round(min(Er_A_I, Er_B_I), 1)
			ws['D19'].value = my_round(max(Er_A_I, Er_B_I), 1)
			ws['D22'].value = my_round(min(sigma_A_I, sigma_B_I), 1)
			ws['D23'].value = my_round(max(sigma_A_I, sigma_B_I), 1)

		if check_var_gr_2.get() == 3:
			ws['D18'].value = my_round(min(Er_A_I, Er_B_I, Er_C_I), 1)
			ws['D19'].value = my_round(max(Er_A_I, Er_B_I, Er_C_I), 1)
			ws['D20'].value = my_round(min(CV_vnytr_gr_QCA1_QCA2, CV_vnytr_gr_QCA1_QCA2_QCA3), 1)
			ws['D21'].value = my_round(max(CV_vnytr_gr_QCA1_QCA2, CV_vnytr_gr_QCA1_QCA2_QCA3), 1)
			ws['D22'].value = my_round(min(sigma_A_I, sigma_B_I, sigma_C_I), 1)
			ws['D23'].value = my_round(max(sigma_A_I, sigma_B_I, sigma_C_I), 1)
			ws['D24'].value = my_round(min(CV_mezhdy_gr_QCA1_QCA2, CV_mezhdy_gr_QCA1_QCA2_QCA3), 1)
			ws['D25'].value = my_round(max(CV_mezhdy_gr_QCA1_QCA2, CV_mezhdy_gr_QCA1_QCA2_QCA3), 1)
		if check_var_gr_2.get() == 4:
			ws['D18'].value = my_round(min(Er_A_I, Er_B_I, Er_C_I, Er_D_I), 1)
			ws['D19'].value = my_round(max(Er_A_I, Er_B_I, Er_C_I, Er_D_I), 1)
			ws['D20'].value = my_round(min(CV_vnytr_gr_QCA1_QCA2, CV_vnytr_gr_QCA1_QCA2_QCA3, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4), 1)
			ws['D21'].value = my_round(max(CV_vnytr_gr_QCA1_QCA2, CV_vnytr_gr_QCA1_QCA2_QCA3, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4), 1)
			ws['D22'].value = my_round(min(sigma_A_I, sigma_B_I, sigma_C_I, sigma_D_I), 1)
			ws['D23'].value = my_round(max(sigma_A_I, sigma_B_I, sigma_C_I, sigma_D_I), 1)
			ws['D24'].value = my_round(min(CV_mezhdy_gr_QCA1_QCA2, CV_mezhdy_gr_QCA1_QCA2_QCA3, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4), 1)
			ws['D25'].value = my_round(max(CV_mezhdy_gr_QCA1_QCA2, CV_mezhdy_gr_QCA1_QCA2_QCA3, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4), 1)
		if check_var_gr_2.get() == 5:
			ws['D18'].value = my_round(min(Er_A_I, Er_B_I, Er_C_I, Er_D_I, Er_E_I), 1)
			ws['D19'].value = my_round(max(Er_A_I, Er_B_I, Er_C_I, Er_D_I, Er_E_I), 1)
			ws['D20'].value = my_round(min(CV_vnytr_gr_QCA1_QCA2, CV_vnytr_gr_QCA1_QCA2_QCA3, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5), 1)
			ws['D21'].value = my_round(max(CV_vnytr_gr_QCA1_QCA2, CV_vnytr_gr_QCA1_QCA2_QCA3, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5), 1)
			ws['D22'].value = my_round(min(sigma_A_I, sigma_B_I, sigma_C_I, sigma_D_I, sigma_E_I), 1)
			ws['D23'].value = my_round(max(sigma_A_I, sigma_B_I, sigma_C_I, sigma_D_I, sigma_E_I), 1)
			ws['D24'].value = my_round(min(CV_mezhdy_gr_QCA1_QCA2, CV_mezhdy_gr_QCA1_QCA2_QCA3, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5), 1)
			ws['D25'].value = my_round(max(CV_mezhdy_gr_QCA1_QCA2, CV_mezhdy_gr_QCA1_QCA2_QCA3, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5), 1)

		if check_var_gr_2.get() == 6:	
			ws['D18'].value = my_round(min(Er_A_I, Er_B_I, Er_C_I, Er_D_I, Er_E_I, Er_F_I), 1)
			ws['D19'].value = my_round(max(Er_A_I, Er_B_I, Er_C_I, Er_D_I, Er_E_I, Er_F_I), 1)
			ws['D20'].value = my_round(min(CV_vnytr_gr_QCA1_QCA2, CV_vnytr_gr_QCA1_QCA2_QCA3, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6), 1)
			ws['D21'].value = my_round(max(CV_vnytr_gr_QCA1_QCA2, CV_vnytr_gr_QCA1_QCA2_QCA3, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5, CV_vnytr_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6), 1)
			ws['D22'].value = my_round(min(sigma_A_I, sigma_B_I, sigma_C_I, sigma_D_I, sigma_E_I, sigma_F_I), 1)
			ws['D23'].value = my_round(max(sigma_A_I, sigma_B_I, sigma_C_I, sigma_D_I, sigma_E_I, sigma_F_I), 1)
			ws['D24'].value = my_round(min(CV_mezhdy_gr_QCA1_QCA2, CV_mezhdy_gr_QCA1_QCA2_QCA3, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6), 1)
			ws['D25'].value = my_round(max(CV_mezhdy_gr_QCA1_QCA2, CV_mezhdy_gr_QCA1_QCA2_QCA3, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5, CV_mezhdy_gr_QCA1_QCA2_QCA3_QCA4_QCA5_QCA6), 1)



	if check_var_gr_1.get() >= 2:
		ws['E16'].value = 'QCB'

		if check_var_gr_2.get() == 2:
			ws['E18'].value = my_round(min(Er_A_II, Er_B_II), 1)
			ws['E19'].value = my_round(max(Er_A_II, Er_B_II), 1)
			ws['E22'].value = my_round(min(sigma_A_II, sigma_B_II), 1)
			ws['E23'].value = my_round(max(sigma_A_II, sigma_B_II), 1)

		if check_var_gr_2.get() == 3:
			ws['E18'].value = my_round(min(Er_A_II, Er_B_II, Er_C_II), 1)
			ws['E19'].value = my_round(max(Er_A_II, Er_B_II, Er_C_II), 1)
			ws['E20'].value = my_round(min(CV_vnytr_gr_QCB1_QCB2, CV_vnytr_gr_QCB1_QCB2_QCB3), 1)
			ws['E21'].value = my_round(max(CV_vnytr_gr_QCB1_QCB2, CV_vnytr_gr_QCB1_QCB2_QCB3), 1)
			ws['E22'].value = my_round(min(sigma_A_II, sigma_B_II, sigma_C_II), 1)
			ws['E23'].value = my_round(max(sigma_A_II, sigma_B_II, sigma_C_II), 1)
			ws['E24'].value = my_round(min(CV_mezhdy_gr_QCB1_QCB2, CV_mezhdy_gr_QCB1_QCB2_QCB3), 1)
			ws['E25'].value = my_round(max(CV_mezhdy_gr_QCB1_QCB2, CV_mezhdy_gr_QCB1_QCB2_QCB3), 1)
		if check_var_gr_2.get() == 4:
			ws['E18'].value = my_round(min(Er_A_II, Er_B_II, Er_C_II, Er_D_II), 1)
			ws['E19'].value = my_round(max(Er_A_II, Er_B_II, Er_C_II, Er_D_II), 1)
			ws['E20'].value = my_round(min(CV_vnytr_gr_QCB1_QCB2, CV_vnytr_gr_QCB1_QCB2_QCB3, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4), 1)
			ws['E21'].value = my_round(max(CV_vnytr_gr_QCB1_QCB2, CV_vnytr_gr_QCB1_QCB2_QCB3, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4), 1)
			ws['E22'].value = my_round(min(sigma_A_II, sigma_B_II, sigma_C_II, sigma_D_II), 1)
			ws['E23'].value = my_round(max(sigma_A_II, sigma_B_II, sigma_C_II, sigma_D_II), 1)
			ws['E24'].value = my_round(min(CV_mezhdy_gr_QCB1_QCB2, CV_mezhdy_gr_QCB1_QCB2_QCB3, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4), 1)
			ws['E25'].value = my_round(max(CV_mezhdy_gr_QCB1_QCB2, CV_mezhdy_gr_QCB1_QCB2_QCB3, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4), 1)
		if check_var_gr_2.get() == 5:
			ws['E18'].value = my_round(min(Er_A_II, Er_B_II, Er_C_II, Er_D_II, Er_E_II), 1)
			ws['E19'].value = my_round(max(Er_A_II, Er_B_II, Er_C_II, Er_D_II, Er_E_II), 1)
			ws['E20'].value = my_round(min(CV_vnytr_gr_QCB1_QCB2, CV_vnytr_gr_QCB1_QCB2_QCB3, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5), 1)
			ws['E21'].value = my_round(max(CV_vnytr_gr_QCB1_QCB2, CV_vnytr_gr_QCB1_QCB2_QCB3, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5), 1)
			ws['E22'].value = my_round(min(sigma_A_II, sigma_B_II, sigma_C_II, sigma_D_II, sigma_E_II), 1)
			ws['E23'].value = my_round(max(sigma_A_II, sigma_B_II, sigma_C_II, sigma_D_II, sigma_E_II), 1)
			ws['E24'].value = my_round(min(CV_mezhdy_gr_QCB1_QCB2, CV_mezhdy_gr_QCB1_QCB2_QCB3, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5), 1)
			ws['E25'].value = my_round(max(CV_mezhdy_gr_QCB1_QCB2, CV_mezhdy_gr_QCB1_QCB2_QCB3, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5), 1)

		if check_var_gr_2.get() == 6:	
			ws['E18'].value = my_round(min(Er_A_II, Er_B_II, Er_C_II, Er_D_II, Er_E_II, Er_F_II), 1)
			ws['E19'].value = my_round(max(Er_A_II, Er_B_II, Er_C_II, Er_D_II, Er_E_II, Er_F_II), 1)
			ws['E20'].value = my_round(min(CV_vnytr_gr_QCB1_QCB2, CV_vnytr_gr_QCB1_QCB2_QCB3, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6), 1)
			ws['E21'].value = my_round(max(CV_vnytr_gr_QCB1_QCB2, CV_vnytr_gr_QCB1_QCB2_QCB3, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5, CV_vnytr_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6), 1)
			ws['E22'].value = my_round(min(sigma_A_II, sigma_B_II, sigma_C_II, sigma_D_II, sigma_E_II, sigma_F_II), 1)
			ws['E23'].value = my_round(max(sigma_A_II, sigma_B_II, sigma_C_II, sigma_D_II, sigma_E_II, sigma_F_II), 1)
			ws['E24'].value = my_round(min(CV_mezhdy_gr_QCB1_QCB2, CV_mezhdy_gr_QCB1_QCB2_QCB3, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6), 1)
			ws['E25'].value = my_round(max(CV_mezhdy_gr_QCB1_QCB2, CV_mezhdy_gr_QCB1_QCB2_QCB3, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5, CV_mezhdy_gr_QCB1_QCB2_QCB3_QCB4_QCB5_QCB6), 1)

	if check_var_gr_1.get() >= 3:
		ws['F16'].value = 'QCC'

		if check_var_gr_2.get() == 2:
			ws['F18'].value = my_round(min(Er_A_III, Er_B_III), 1)
			ws['F19'].value = my_round(max(Er_A_III, Er_B_III), 1)
			ws['F22'].value = my_round(min(sigma_A_III, sigma_B_III), 1)
			ws['F23'].value = my_round(max(sigma_A_III, sigma_B_III), 1)

		if check_var_gr_2.get() == 3:
			ws['F18'].value = my_round(min(Er_A_III, Er_B_III, Er_C_III), 1)
			ws['F19'].value = my_round(max(Er_A_III, Er_B_III, Er_C_III), 1)
			ws['F20'].value = my_round(min(CV_vnytr_gr_QCC1_QCC2, CV_vnytr_gr_QCC1_QCC2_QCC3), 1)
			ws['F21'].value = my_round(max(CV_vnytr_gr_QCC1_QCC2, CV_vnytr_gr_QCC1_QCC2_QCC3), 1)
			ws['F22'].value = my_round(min(sigma_A_III, sigma_B_III, sigma_C_III), 1)
			ws['F23'].value = my_round(max(sigma_A_III, sigma_B_III, sigma_C_III), 1)
			ws['F24'].value = my_round(min(CV_mezhdy_gr_QCC1_QCC2, CV_mezhdy_gr_QCC1_QCC2_QCC3), 1)
			ws['F25'].value = my_round(max(CV_mezhdy_gr_QCC1_QCC2, CV_mezhdy_gr_QCC1_QCC2_QCC3), 1)
		if check_var_gr_2.get() == 4:
			ws['F18'].value = my_round(min(Er_A_III, Er_B_III, Er_C_III, Er_D_III), 1)
			ws['F19'].value = my_round(max(Er_A_III, Er_B_III, Er_C_III, Er_D_III), 1)
			ws['F20'].value = my_round(min(CV_vnytr_gr_QCC1_QCC2, CV_vnytr_gr_QCC1_QCC2_QCC3, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4), 1)
			ws['F21'].value = my_round(max(CV_vnytr_gr_QCC1_QCC2, CV_vnytr_gr_QCC1_QCC2_QCC3, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4), 1)
			ws['F22'].value = my_round(min(sigma_A_III, sigma_B_III, sigma_C_III, sigma_D_III), 1)
			ws['F23'].value = my_round(max(sigma_A_III, sigma_B_III, sigma_C_III, sigma_D_III), 1)
			ws['F24'].value = my_round(min(CV_mezhdy_gr_QCC1_QCC2, CV_mezhdy_gr_QCC1_QCC2_QCC3, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4), 1)
			ws['F25'].value = my_round(max(CV_mezhdy_gr_QCC1_QCC2, CV_mezhdy_gr_QCC1_QCC2_QCC3, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4), 1)
		if check_var_gr_2.get() == 5:
			ws['F18'].value = my_round(min(Er_A_III, Er_B_III, Er_C_III, Er_D_III, Er_E_III), 1)
			ws['F19'].value = my_round(max(Er_A_III, Er_B_III, Er_C_III, Er_D_III, Er_E_III), 1)
			ws['F20'].value = my_round(min(CV_vnytr_gr_QCC1_QCC2, CV_vnytr_gr_QCC1_QCC2_QCC3, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5), 1)
			ws['F21'].value = my_round(max(CV_vnytr_gr_QCC1_QCC2, CV_vnytr_gr_QCC1_QCC2_QCC3, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5), 1)
			ws['F22'].value = my_round(min(sigma_A_III, sigma_B_III, sigma_C_III, sigma_D_III, sigma_E_III), 1)
			ws['F23'].value = my_round(max(sigma_A_III, sigma_B_III, sigma_C_III, sigma_D_III, sigma_E_III), 1)
			ws['F24'].value = my_round(min(CV_mezhdy_gr_QCC1_QCC2, CV_mezhdy_gr_QCC1_QCC2_QCC3, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5), 1)
			ws['F25'].value = my_round(max(CV_mezhdy_gr_QCC1_QCC2, CV_mezhdy_gr_QCC1_QCC2_QCC3, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5), 1)

		if check_var_gr_2.get() == 6:	
			ws['F18'].value = my_round(min(Er_A_III, Er_B_III, Er_C_III, Er_D_III, Er_E_III, Er_F_III), 1)
			ws['F19'].value = my_round(max(Er_A_III, Er_B_III, Er_C_III, Er_D_III, Er_E_III, Er_F_III), 1)
			ws['F20'].value = my_round(min(CV_vnytr_gr_QCC1_QCC2, CV_vnytr_gr_QCC1_QCC2_QCC3, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6), 1)
			ws['F21'].value = my_round(max(CV_vnytr_gr_QCC1_QCC2, CV_vnytr_gr_QCC1_QCC2_QCC3, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5, CV_vnytr_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6), 1)
			ws['F22'].value = my_round(min(sigma_A_III, sigma_B_III, sigma_C_III, sigma_D_III, sigma_E_III, sigma_F_III), 1)
			ws['F23'].value = my_round(max(sigma_A_III, sigma_B_III, sigma_C_III, sigma_D_III, sigma_E_III, sigma_F_III), 1)
			ws['F24'].value = my_round(min(CV_mezhdy_gr_QCC1_QCC2, CV_mezhdy_gr_QCC1_QCC2_QCC3, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6), 1)
			ws['F25'].value = my_round(max(CV_mezhdy_gr_QCC1_QCC2, CV_mezhdy_gr_QCC1_QCC2_QCC3, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5, CV_mezhdy_gr_QCC1_QCC2_QCC3_QCC4_QCC5_QCC6), 1)

	if check_var_gr_1.get() >= 4:
		ws['G16'].value = 'QCD'

		if check_var_gr_2.get() == 2:
			ws['G18'].value = my_round(min(Er_A_IV, Er_B_IV), 1)
			ws['G19'].value = my_round(max(Er_A_IV, Er_B_IV), 1)
			ws['G22'].value = my_round(min(sigma_A_IV, sigma_B_IV), 1)
			ws['G23'].value = my_round(max(sigma_A_IV, sigma_B_IV), 1)

		if check_var_gr_2.get() == 3:
			ws['G18'].value = my_round(min(Er_A_IV, Er_B_IV, Er_C_IV), 1)
			ws['G19'].value = my_round(max(Er_A_IV, Er_B_IV, Er_C_IV), 1)
			ws['G20'].value = my_round(min(CV_vnytr_gr_QCD1_QCD2, CV_vnytr_gr_QCD1_QCD2_QCD3), 1)
			ws['G21'].value = my_round(max(CV_vnytr_gr_QCD1_QCD2, CV_vnytr_gr_QCD1_QCD2_QCD3), 1)
			ws['G22'].value = my_round(min(sigma_A_IV, sigma_B_IV, sigma_C_IV), 1)
			ws['G23'].value = my_round(max(sigma_A_IV, sigma_B_IV, sigma_C_IV), 1)
			ws['G24'].value = my_round(min(CV_mezhdy_gr_QCD1_QCD2, CV_mezhdy_gr_QCD1_QCD2_QCD3), 1)
			ws['G25'].value = my_round(max(CV_mezhdy_gr_QCD1_QCD2, CV_mezhdy_gr_QCD1_QCD2_QCD3), 1)
		if check_var_gr_2.get() == 4:
			ws['G18'].value = my_round(min(Er_A_IV, Er_B_IV, Er_C_IV, Er_D_IV), 1)
			ws['G19'].value = my_round(max(Er_A_IV, Er_B_IV, Er_C_IV, Er_D_IV), 1)
			ws['G20'].value = my_round(min(CV_vnytr_gr_QCD1_QCD2, CV_vnytr_gr_QCD1_QCD2_QCD3, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4), 1)
			ws['G21'].value = my_round(max(CV_vnytr_gr_QCD1_QCD2, CV_vnytr_gr_QCD1_QCD2_QCD3, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4), 1)
			ws['G22'].value = my_round(min(sigma_A_IV, sigma_B_IV, sigma_C_IV, sigma_D_IV), 1)
			ws['G23'].value = my_round(max(sigma_A_IV, sigma_B_IV, sigma_C_IV, sigma_D_IV), 1)
			ws['G24'].value = my_round(min(CV_mezhdy_gr_QCD1_QCD2, CV_mezhdy_gr_QCD1_QCD2_QCD3, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4), 1)
			ws['G25'].value = my_round(max(CV_mezhdy_gr_QCD1_QCD2, CV_mezhdy_gr_QCD1_QCD2_QCD3, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4), 1)
		if check_var_gr_2.get() == 5:
			ws['G18'].value = my_round(min(Er_A_IV, Er_B_IV, Er_C_IV, Er_D_IV, Er_E_IV), 1)
			ws['G19'].value = my_round(max(Er_A_IV, Er_B_IV, Er_C_IV, Er_D_IV, Er_E_IV), 1)
			ws['G20'].value = my_round(min(CV_vnytr_gr_QCD1_QCD2, CV_vnytr_gr_QCD1_QCD2_QCD3, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5), 1)
			ws['G21'].value = my_round(max(CV_vnytr_gr_QCD1_QCD2, CV_vnytr_gr_QCD1_QCD2_QCD3, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5), 1)
			ws['G22'].value = my_round(min(sigma_A_IV, sigma_B_IV, sigma_C_IV, sigma_D_IV, sigma_E_IV), 1)
			ws['G23'].value = my_round(max(sigma_A_IV, sigma_B_IV, sigma_C_IV, sigma_D_IV, sigma_E_IV), 1)
			ws['G24'].value = my_round(min(CV_mezhdy_gr_QCD1_QCD2, CV_mezhdy_gr_QCD1_QCD2_QCD3, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5), 1)
			ws['G25'].value = my_round(max(CV_mezhdy_gr_QCD1_QCD2, CV_mezhdy_gr_QCD1_QCD2_QCD3, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5), 1)

		if check_var_gr_2.get() == 6:	
			ws['G18'].value = my_round(min(Er_A_IV, Er_B_IV, Er_C_IV, Er_D_IV, Er_E_IV, Er_F_IV), 1)
			ws['G19'].value = my_round(max(Er_A_IV, Er_B_IV, Er_C_IV, Er_D_IV, Er_E_IV, Er_F_IV), 1)
			ws['G20'].value = my_round(min(CV_vnytr_gr_QCD1_QCD2, CV_vnytr_gr_QCD1_QCD2_QCD3, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6), 1)
			ws['G21'].value = my_round(max(CV_vnytr_gr_QCD1_QCD2, CV_vnytr_gr_QCD1_QCD2_QCD3, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5, CV_vnytr_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6), 1)
			ws['G22'].value = my_round(min(sigma_A_IV, sigma_B_IV, sigma_C_IV, sigma_D_IV, sigma_E_IV, sigma_F_IV), 1)
			ws['G23'].value = my_round(max(sigma_A_IV, sigma_B_IV, sigma_C_IV, sigma_D_IV, sigma_E_IV, sigma_F_IV), 1)
			ws['G24'].value = my_round(min(CV_mezhdy_gr_QCD1_QCD2, CV_mezhdy_gr_QCD1_QCD2_QCD3, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6), 1)
			ws['G25'].value = my_round(max(CV_mezhdy_gr_QCD1_QCD2, CV_mezhdy_gr_QCD1_QCD2_QCD3, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5, CV_mezhdy_gr_QCD1_QCD2_QCD3_QCD4_QCD5_QCD6), 1)
			
	if check_var_gr_1.get() >= 5:
		ws['H16'].value = 'QCE'

		if check_var_gr_2.get() == 2:
			ws['H18'].value = my_round(min(Er_A_V, Er_B_V), 1)
			ws['H19'].value = my_round(max(Er_A_V, Er_B_V), 1)
			ws['H22'].value = my_round(min(sigma_A_V, sigma_B_V), 1)
			ws['H23'].value = my_round(max(sigma_A_V, sigma_B_V), 1)

		if check_var_gr_2.get() == 3:
			ws['H18'].value = my_round(min(Er_A_V, Er_B_V, Er_C_V), 1)
			ws['H19'].value = my_round(max(Er_A_V, Er_B_V, Er_C_V), 1)
			ws['H20'].value = my_round(min(CV_vnytr_gr_QCE1_QCE2, CV_vnytr_gr_QCE1_QCE2_QCE3), 1)
			ws['H21'].value = my_round(max(CV_vnytr_gr_QCE1_QCE2, CV_vnytr_gr_QCE1_QCE2_QCE3), 1)
			ws['H22'].value = my_round(min(sigma_A_V, sigma_B_V, sigma_C_V), 1)
			ws['H23'].value = my_round(max(sigma_A_V, sigma_B_V, sigma_C_V), 1)
			ws['H24'].value = my_round(min(CV_mezhdy_gr_QCE1_QCE2, CV_mezhdy_gr_QCE1_QCE2_QCE3), 1)
			ws['H25'].value = my_round(max(CV_mezhdy_gr_QCE1_QCE2, CV_mezhdy_gr_QCE1_QCE2_QCE3), 1)
		if check_var_gr_2.get() == 4:
			ws['H18'].value = my_round(min(Er_A_V, Er_B_V, Er_C_V, Er_D_V), 1)
			ws['H19'].value = my_round(max(Er_A_V, Er_B_V, Er_C_V, Er_D_V), 1)
			ws['H20'].value = my_round(min(CV_vnytr_gr_QCE1_QCE2, CV_vnytr_gr_QCE1_QCE2_QCE3, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4), 1)
			ws['H21'].value = my_round(max(CV_vnytr_gr_QCE1_QCE2, CV_vnytr_gr_QCE1_QCE2_QCE3, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4), 1)
			ws['H22'].value = my_round(min(sigma_A_V, sigma_B_V, sigma_C_V, sigma_D_V), 1)
			ws['H23'].value = my_round(max(sigma_A_V, sigma_B_V, sigma_C_V, sigma_D_V), 1)
			ws['H24'].value = my_round(min(CV_mezhdy_gr_QCE1_QCE2, CV_mezhdy_gr_QCE1_QCE2_QCE3, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4), 1)
			ws['H25'].value = my_round(max(CV_mezhdy_gr_QCE1_QCE2, CV_mezhdy_gr_QCE1_QCE2_QCE3, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4), 1)
		if check_var_gr_2.get() == 5:
			ws['H18'].value = my_round(min(Er_A_V, Er_B_V, Er_C_V, Er_D_V, Er_E_V), 1)
			ws['H19'].value = my_round(max(Er_A_V, Er_B_V, Er_C_V, Er_D_V, Er_E_V), 1)
			ws['H20'].value = my_round(min(CV_vnytr_gr_QCE1_QCE2, CV_vnytr_gr_QCE1_QCE2_QCE3, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5), 1)
			ws['H21'].value = my_round(max(CV_vnytr_gr_QCE1_QCE2, CV_vnytr_gr_QCE1_QCE2_QCE3, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5), 1)
			ws['H22'].value = my_round(min(sigma_A_V, sigma_B_V, sigma_C_V, sigma_D_V, sigma_E_V), 1)
			ws['H23'].value = my_round(max(sigma_A_V, sigma_B_V, sigma_C_V, sigma_D_V, sigma_E_V), 1)
			ws['H24'].value = my_round(min(CV_mezhdy_gr_QCE1_QCE2, CV_mezhdy_gr_QCE1_QCE2_QCE3, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5), 1)
			ws['H25'].value = my_round(max(CV_mezhdy_gr_QCE1_QCE2, CV_mezhdy_gr_QCE1_QCE2_QCE3, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5), 1)

		if check_var_gr_2.get() == 6:	
			ws['H18'].value = my_round(min(Er_A_V, Er_B_V, Er_C_V, Er_D_V, Er_E_V, Er_F_V), 1)
			ws['H19'].value = my_round(max(Er_A_V, Er_B_V, Er_C_V, Er_D_V, Er_E_V, Er_F_V), 1)
			ws['H20'].value = my_round(min(CV_vnytr_gr_QCE1_QCE2, CV_vnytr_gr_QCE1_QCE2_QCE3, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6), 1)
			ws['H21'].value = my_round(max(CV_vnytr_gr_QCE1_QCE2, CV_vnytr_gr_QCE1_QCE2_QCE3, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5, CV_vnytr_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6), 1)
			ws['H22'].value = my_round(min(sigma_A_V, sigma_B_V, sigma_C_V, sigma_D_V, sigma_E_V, sigma_F_V), 1)
			ws['H23'].value = my_round(max(sigma_A_V, sigma_B_V, sigma_C_V, sigma_D_V, sigma_E_V, sigma_F_V), 1)
			ws['H24'].value = my_round(min(CV_mezhdy_gr_QCE1_QCE2, CV_mezhdy_gr_QCE1_QCE2_QCE3, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6), 1)
			ws['H25'].value = my_round(max(CV_mezhdy_gr_QCE1_QCE2, CV_mezhdy_gr_QCE1_QCE2_QCE3, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5, CV_mezhdy_gr_QCE1_QCE2_QCE3_QCE4_QCE5_QCE6), 1)
			
	if check_var_gr_1.get() >= 6:
		ws['I16'].value = 'QCF'

		if check_var_gr_2.get() == 2:
			ws['I18'].value = my_round(min(Er_A_VI, Er_B_VI), 1)
			ws['I19'].value = my_round(max(Er_A_VI, Er_B_VI), 1)
			ws['I22'].value = my_round(min(sigma_A_VI, sigma_B_VI), 1)
			ws['I23'].value = my_round(max(sigma_A_VI, sigma_B_VI), 1)

		if check_var_gr_2.get() == 3:
			ws['I18'].value = my_round(min(Er_A_VI, Er_B_VI, Er_C_VI), 1)
			ws['I19'].value = my_round(max(Er_A_VI, Er_B_VI, Er_C_VI), 1)
			ws['I20'].value = my_round(min(CV_vnytr_gr_QCF1_QCF2, CV_vnytr_gr_QCF1_QCF2_QCF3), 1)
			ws['I21'].value = my_round(max(CV_vnytr_gr_QCF1_QCF2, CV_vnytr_gr_QCF1_QCF2_QCF3), 1)
			ws['I22'].value = my_round(min(sigma_A_VI, sigma_B_VI, sigma_C_VI), 1)
			ws['I23'].value = my_round(max(sigma_A_VI, sigma_B_VI, sigma_C_VI), 1)
			ws['I24'].value = my_round(min(CV_mezhdy_gr_QCF1_QCF2, CV_mezhdy_gr_QCF1_QCF2_QCF3), 1)
			ws['I25'].value = my_round(max(CV_mezhdy_gr_QCF1_QCF2, CV_mezhdy_gr_QCF1_QCF2_QCF3), 1)
		if check_var_gr_2.get() == 4:
			ws['I18'].value = my_round(min(Er_A_VI, Er_B_VI, Er_C_VI, Er_D_VI), 1)
			ws['I19'].value = my_round(max(Er_A_VI, Er_B_VI, Er_C_VI, Er_D_VI), 1)
			ws['I20'].value = my_round(min(CV_vnytr_gr_QCF1_QCF2, CV_vnytr_gr_QCF1_QCF2_QCF3, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4), 1)
			ws['I21'].value = my_round(max(CV_vnytr_gr_QCF1_QCF2, CV_vnytr_gr_QCF1_QCF2_QCF3, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4), 1)
			ws['I22'].value = my_round(min(sigma_A_VI, sigma_B_VI, sigma_C_VI, sigma_D_VI), 1)
			ws['I23'].value = my_round(max(sigma_A_VI, sigma_B_VI, sigma_C_VI, sigma_D_VI), 1)
			ws['I24'].value = my_round(min(CV_mezhdy_gr_QCF1_QCF2, CV_mezhdy_gr_QCF1_QCF2_QCF3, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4), 1)
			ws['I25'].value = my_round(max(CV_mezhdy_gr_QCF1_QCF2, CV_mezhdy_gr_QCF1_QCF2_QCF3, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4), 1)
		if check_var_gr_2.get() == 5:
			ws['I18'].value = my_round(min(Er_A_VI, Er_B_VI, Er_C_VI, Er_D_VI, Er_E_VI), 1)
			ws['I19'].value = my_round(max(Er_A_VI, Er_B_VI, Er_C_VI, Er_D_VI, Er_E_VI), 1)
			ws['I20'].value = my_round(min(CV_vnytr_gr_QCF1_QCF2, CV_vnytr_gr_QCF1_QCF2_QCF3, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5), 1)
			ws['I21'].value = my_round(max(CV_vnytr_gr_QCF1_QCF2, CV_vnytr_gr_QCF1_QCF2_QCF3, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5), 1)
			ws['I22'].value = my_round(min(sigma_A_VI, sigma_B_VI, sigma_C_VI, sigma_D_VI, sigma_E_VI), 1)
			ws['I23'].value = my_round(max(sigma_A_VI, sigma_B_VI, sigma_C_VI, sigma_D_VI, sigma_E_VI), 1)
			ws['I24'].value = my_round(min(CV_mezhdy_gr_QCF1_QCF2, CV_mezhdy_gr_QCF1_QCF2_QCF3, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5), 1)
			ws['I25'].value = my_round(max(CV_mezhdy_gr_QCF1_QCF2, CV_mezhdy_gr_QCF1_QCF2_QCF3, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5), 1)
		if check_var_gr_2.get() == 6:	
			ws['I18'].value = my_round(min(Er_A_VI, Er_B_VI, Er_C_VI, Er_D_VI, Er_E_VI, Er_F_VI), 1)
			ws['I19'].value = my_round(max(Er_A_VI, Er_B_VI, Er_C_VI, Er_D_VI, Er_E_VI, Er_F_VI), 1)
			ws['I20'].value = my_round(min(CV_vnytr_gr_QCF1_QCF2, CV_vnytr_gr_QCF1_QCF2_QCF3, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6), 1)
			ws['I21'].value = my_round(max(CV_vnytr_gr_QCF1_QCF2, CV_vnytr_gr_QCF1_QCF2_QCF3, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5, CV_vnytr_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6), 1)
			ws['I22'].value = my_round(min(sigma_A_VI, sigma_B_VI, sigma_C_VI, sigma_D_VI, sigma_E_VI, sigma_F_VI), 1)
			ws['I23'].value = my_round(max(sigma_A_VI, sigma_B_VI, sigma_C_VI, sigma_D_VI, sigma_E_VI, sigma_F_VI), 1)
			ws['I24'].value = my_round(min(CV_mezhdy_gr_QCF1_QCF2, CV_mezhdy_gr_QCF1_QCF2_QCF3, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6), 1)
			ws['I25'].value = my_round(max(CV_mezhdy_gr_QCF1_QCF2, CV_mezhdy_gr_QCF1_QCF2_QCF3, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5, CV_mezhdy_gr_QCF1_QCF2_QCF3_QCF4_QCF5_QCF6), 1)
			 




	#СДВИГ ТАБЛИЦ С ОФОРМЛЕНИЕМ

	if check_var_gr_1.get() == 1:
		#сдвиг первой таблицы		
		ws.move_range("M15:N34", cols=-7)
		ws.column_dimensions['G'].width = 32
		ws.column_dimensions['F'].width = 18	

		if check_var_gr_2.get() >= 2:
			#сдвиг второй таблицы
			ws.move_range("U15:Y34", cols=-12)
			ws.column_dimensions['J'].width = 32
			ws.column_dimensions['I'].width = 18

		if check_var_gr_2.get() >= 3:
			#сдвиг третьей таблицы
			ws.move_range("AC15:AG34", cols=-17)
			ws.column_dimensions['M'].width = 32
			ws.column_dimensions['L'].width = 18

		if check_var_gr_2.get() >= 4:
			#сдвиг четвертой таблицы
			ws.move_range("AK15:AO34", cols=-22)
			ws.column_dimensions['P'].width = 32
			ws.column_dimensions['O'].width = 18

		if check_var_gr_2.get() >= 5:
			#сдвиг пятой таблицы
			ws.move_range("AS15:AW34", cols=-27)
			ws.column_dimensions['S'].width = 32
			ws.column_dimensions['R'].width = 18

		if check_var_gr_2.get() >= 6:
			#сдвиг шестой таблицы
			ws.move_range("BA15:BB34", cols=-32)
			ws.column_dimensions['V'].width = 32
			ws.column_dimensions['U'].width = 18


	if check_var_gr_1.get() == 2:
		#сдвиг первой таблицы		
		ws.move_range("M15:Q34", cols=-6)
		ws.merge_cells('H17:I17')
		ws.column_dimensions['G'].width = 18	

		ws.column_dimensions['H'].width = 15
		ws.column_dimensions['I'].width = 15

		if check_var_gr_2.get() >= 2:
			#сдвиг второй таблицы
			ws.move_range("U15:Y34", cols=-10)
			ws.merge_cells('L17:M17')
			ws.column_dimensions['K'].width = 18

			ws.column_dimensions['L'].width = 15
			ws.column_dimensions['M'].width = 15

		if check_var_gr_2.get() >= 3:
			#сдвиг третьей таблицы
			ws.move_range("AC15:AG34", cols=-14)
			ws.merge_cells('P17:Q17')
			ws.column_dimensions['O'].width = 18

			ws.column_dimensions['P'].width = 15
			ws.column_dimensions['Q'].width = 15

		if check_var_gr_2.get() >= 4:
			#сдвиг четвертой таблицы
			ws.move_range("AK15:AO34", cols=-18)
			ws.merge_cells('T17:U17')
			ws.column_dimensions['S'].width = 18

			ws.column_dimensions['T'].width = 15
			ws.column_dimensions['U'].width = 15

		if check_var_gr_2.get() >= 5:
			#сдвиг пятой таблицы
			ws.move_range("AS15:AW34", cols=-22)
			ws.merge_cells('W17:Y17')
			ws.column_dimensions['W'].width = 18

			ws.column_dimensions['X'].width = 15
			ws.column_dimensions['Y'].width = 15

		if check_var_gr_2.get() >= 6:
			#сдвиг шестой таблицы
			ws.move_range("BA15:BC34", cols=-26)
			ws.merge_cells('AB17:AC17')
			ws.column_dimensions['AA'].width = 18

			ws.column_dimensions['AB'].width = 15
			ws.column_dimensions['AC'].width = 15


	if check_var_gr_1.get() == 3:
		#сдвиг первой таблицы		
		ws.move_range("M15:P34", cols=-5)
		ws.merge_cells('I17:K17')
		ws.column_dimensions['H'].width = 18

		ws.column_dimensions['I'].width = 11	
		ws.column_dimensions['J'].width = 11	
		ws.column_dimensions['K'].width = 11		

		if check_var_gr_2.get() >= 2:
			#сдвиг второй таблицы
			ws.move_range("U15:X34", cols=-8)
			ws.merge_cells('N17:P17')
			ws.column_dimensions['M'].width = 18

			ws.column_dimensions['N'].width = 11	
			ws.column_dimensions['O'].width = 11	
			ws.column_dimensions['P'].width = 11

		if check_var_gr_2.get() >= 3:
			#сдвиг третьей таблицы
			ws.move_range("AC15:AG34", cols=-11)
			ws.merge_cells('S17:U17')
			ws.column_dimensions['R'].width = 18

			ws.column_dimensions['S'].width = 11	
			ws.column_dimensions['T'].width = 11	
			ws.column_dimensions['U'].width = 11

		if check_var_gr_2.get() >= 4:
			#сдвиг четвертой таблицы
			ws.move_range("AK15:AO34", cols=-14)
			ws.merge_cells('X17:Z17')
			ws.column_dimensions['W'].width = 18

			ws.column_dimensions['X'].width = 11	
			ws.column_dimensions['Y'].width = 11	
			ws.column_dimensions['Z'].width = 11

		if check_var_gr_2.get() >= 5:
			#сдвиг пятой таблицы
			ws.move_range("AS15:AW34", cols=-17)
			ws.merge_cells('AC17:AE17')
			ws.column_dimensions['AB'].width = 18

			ws.column_dimensions['AC'].width = 11	
			ws.column_dimensions['AD'].width = 11	
			ws.column_dimensions['AE'].width = 11

		if check_var_gr_2.get() >= 6:
			#сдвиг шестой таблицы
			ws.move_range("BA15:BD34", cols=-20)
			ws.merge_cells('AH17:AJ17')
			ws.column_dimensions['AG'].width = 18

			ws.column_dimensions['AH'].width = 11	
			ws.column_dimensions['AI'].width = 11	
			ws.column_dimensions['AJ'].width = 11

	
	if check_var_gr_1.get() == 4:
		#сдвиг первой таблицы		
		ws.move_range("M15:Q34", cols=-4)
		ws.merge_cells('J17:M17')
		ws.column_dimensions['I'].width = 18	

		if check_var_gr_2.get() >= 2:
			#сдвиг второй таблицы
			ws.move_range("U15:Y34", cols=-6)
			ws.merge_cells('P17:S17')
			ws.column_dimensions['O'].width = 18

		if check_var_gr_2.get() >= 3:
			#сдвиг третьей таблицы
			ws.move_range("AC15:AG34", cols=-8)
			ws.merge_cells('V17:Y17')
			ws.column_dimensions['U'].width = 18

		if check_var_gr_2.get() >= 4:
			#сдвиг четвертой таблицы
			ws.move_range("AK15:AO34", cols=-10)
			ws.merge_cells('AB17:AE17')
			ws.column_dimensions['AA'].width = 18

		if check_var_gr_2.get() >= 5:
			#сдвиг пятой таблицы
			ws.move_range("AS15:AW34", cols=-12)
			ws.merge_cells('AH17:AK17')
			ws.column_dimensions['AG'].width = 18

		if check_var_gr_2.get() >= 6:
			#сдвиг шестой таблицы
			ws.move_range("BK15:BO34", cols=-14)
			ws.merge_cells('AN17:AQ17')
			ws.column_dimensions['AG'].width = 18


	if check_var_gr_1.get() == 5:
		#сдвиг первой таблицы		
		ws.move_range("M15:R34", cols=-3)
		ws.merge_cells('K17:O17')
		ws.column_dimensions['J'].width = 18	

		if check_var_gr_2.get() >= 2:
			#сдвиг второй таблицы
			ws.move_range("U15:Z34", cols=-4)
			ws.merge_cells('R17:V17')
			ws.column_dimensions['Q'].width = 18

		if check_var_gr_2.get() >= 3:
			#сдвиг третьей таблицы
			ws.move_range("AC15:AH34", cols=-5)
			ws.merge_cells('Y17:AC17')
			ws.column_dimensions['X'].width = 18

		if check_var_gr_2.get() >= 4:
			#сдвиг четвертой таблицы
			ws.move_range("AK15:AP34", cols=-6)
			ws.merge_cells('AF17:AJ17')
			ws.column_dimensions['AE'].width = 18

		if check_var_gr_2.get() >= 5:
			#сдвиг пятой таблицы
			ws.move_range("AS15:AX34", cols=-7)
			ws.merge_cells('AM17:AQ17')
			ws.column_dimensions['AL'].width = 18

		if check_var_gr_2.get() >= 6:
			#сдвиг шестой таблицы
			ws.move_range("BA15:BF34", cols=-8)
			ws.merge_cells('AT17:AX17')
			ws.column_dimensions['AS'].width = 18



	if check_var_gr_1.get() == 6:
		#сдвиг первой таблицы		
		ws.move_range("M15:S34", cols=-2)
		ws.merge_cells('L17:Q17')
		ws.column_dimensions['K'].width = 18	

		if check_var_gr_2.get() >= 2:
			#сдвиг второй таблицы
			ws.move_range("U15:AA34", cols=-2)
			ws.merge_cells('T17:Y17')
			ws.column_dimensions['S'].width = 18

		if check_var_gr_2.get() >= 3:
			#сдвиг третьей таблицы
			ws.move_range("AC15:AI34", cols=-2)
			ws.merge_cells('AB17:AG17')
			ws.column_dimensions['AA'].width = 18

		if check_var_gr_2.get() >= 4:
			#сдвиг четвертой таблицы
			ws.move_range("AK15:AQ34", cols=-2)
			ws.merge_cells('AJ17:AO17')
			ws.column_dimensions['AI'].width = 18

		if check_var_gr_2.get() >= 5:
			#сдвиг пятой таблицы
			ws.move_range("AS15:AY34", cols=-2)
			ws.merge_cells('AR17:AW17')
			ws.column_dimensions['AQ'].width = 18

		if check_var_gr_2.get() >= 6:
			#сдвиг шестой таблицы
			ws.move_range("BA15:BG34", cols=-2)
			ws.merge_cells('AZ17:BE17')
			ws.column_dimensions['AY'].width = 18



	#СДВИГ ВВОДА

	if check_var_gr_2.get() == 1:
		A = 5
		ws.move_range("J1:Q12", cols=-A) #СДВИГ QCB
		ws.move_range("S1:Z12", cols=-A-5) #СДВИГ QCC
		ws.move_range("AB1:AI12", cols=-A-10) #СДВИГ QCD
		ws.move_range("AK1:AR12", cols=-A-15) #СДВИГ QCE
		ws.move_range("AT1:BA12", cols=-A-20) #СДВИГ QCF
	elif check_var_gr_2.get() == 2:
		A = 4
		ws.move_range("J1:Q12", cols=-A) #СДВИГ QCB
		ws.move_range("S1:Z12", cols=-A-4) #СДВИГ QCC
		ws.move_range("AB1:AI12", cols=-A-8) #СДВИГ QCD
		ws.move_range("AK1:AR12", cols=-A-12) #СДВИГ QCE
		ws.move_range("AT1:BA12", cols=-A-16) #СДВИГ QCF
	elif check_var_gr_2.get() == 3:
		A = 3
		ws.move_range("J1:Q12", cols=-A) #СДВИГ QCB
		ws.move_range("S1:Z12", cols=-A-3) #СДВИГ QCC
		ws.move_range("AB1:AI12", cols=-A-6) #СДВИГ QCD
		ws.move_range("AK1:AR12", cols=-A-9) #СДВИГ QCE
		ws.move_range("AT1:BA12", cols=-A-12) #СДВИГ QCF
	elif check_var_gr_2.get() == 4:
		A = 2
		ws.move_range("J1:Q12", cols=-A) #СДВИГ QCB
		ws.move_range("S1:Z12", cols=-A-2) #СДВИГ QCC
		ws.move_range("AB1:AI12", cols=-A-4) #СДВИГ QCD
		ws.move_range("AK1:AR12", cols=-A-6) #СДВИГ QCE
		ws.move_range("AT1:BA12", cols=-A-8) #СДВИГ QCF
	elif check_var_gr_2.get() == 5:
		A = 1
		ws.move_range("J1:Q12", cols=-A) #СДВИГ QCB
		ws.move_range("S1:Z12", cols=-A-1) #СДВИГ QCC
		ws.move_range("AB1:AI12", cols=-A-2) #СДВИГ QCD
		ws.move_range("AK1:AR12", cols=-A-3) #СДВИГ QCE
		ws.move_range("AT1:BA12", cols=-A-4) #СДВИГ QCF

	

	#расширение столбцов для правильностей и повторяемостей	
	ws.column_dimensions['A'].width = 16


	wb.save(filename = txt)




def spravka_open():
	toplevel_spravka.deiconify()
	toplevel_spravka.geometry(f"401x300+{(root_open.winfo_x())+7}+{(root_open.winfo_y()+283)}")

	




def spravka_close():
	toplevel_spravka.withdraw()
	

def back_to_first_widow():

	#обнуление переменных отвечающих за сохранение (в этом нет смысла т.к. при вызове все функции переназначаются)

	#сохранение значений со ввода
	#QCA
	#1 колонка
	global save_MAIN_I
	save_MAIN_I = entr_MAIN_I.get()
	global save_A1_I
	save_A1_I = entr_A1_I.get()
	global save_A2_I
	save_A2_I = entr_A2_I.get()
	global save_A3_I
	save_A3_I = entr_A3_I.get()
	if check_var_gr_3.get() >= 4:
		global save_A4_I
		save_A4_I = entr_A4_I.get()
	if check_var_gr_3.get() >= 5:
		global save_A5_I
		save_A5_I = entr_A5_I.get()
	if check_var_gr_3.get() >= 6:
		global save_A6_I
		save_A6_I = entr_A6_I.get()
	if check_var_gr_3.get() >= 7:
		global save_A7_I
		save_A7_I = entr_A7_I.get()
	if check_var_gr_3.get() >= 8:
		global save_A8_I
		save_A8_I = entr_A8_I.get()
	if check_var_gr_3.get() >= 9:
		global save_A9_I
		save_A9_I = entr_A9_I.get()
	if check_var_gr_3.get() >= 10:
		global save_A10_I
		save_A10_I = entr_A10_I.get()

	#2 колонка

	if check_var_gr_2.get() >= 2:
		global save_B1_I
		save_B1_I = entr_B1_I.get()
		global save_B2_I
		save_B2_I = entr_B2_I.get()
		global save_B3_I
		save_B3_I = entr_B3_I.get()
		if check_var_gr_3.get() >= 4:
			global save_B4_I
			save_B4_I = entr_B4_I.get()
		if check_var_gr_3.get() >= 5:
			global save_B5_I
			save_B5_I = entr_B5_I.get()
		if check_var_gr_3.get() >= 6:
			global save_B6_I
			save_B6_I = entr_B6_I.get()
		if check_var_gr_3.get() >= 7:
			global save_B7_I
			save_B7_I = entr_B7_I.get()
		if check_var_gr_3.get() >= 8:
			global save_B8_I
			save_B8_I = entr_B8_I.get()
		if check_var_gr_3.get() >= 9:
			global save_B9_I
			save_B9_I = entr_B9_I.get()
		if check_var_gr_3.get() >= 10:
			global save_B10_I
			save_B10_I = entr_B10_I.get()


	#3 колонка

	if check_var_gr_2.get() >= 3:
		global save_C1_I
		save_C1_I = entr_C1_I.get()
		global save_C2_I
		save_C2_I = entr_C2_I.get()
		global save_C3_I
		save_C3_I = entr_C3_I.get()
		if check_var_gr_3.get() >= 4:
			global save_C4_I
			save_C4_I = entr_C4_I.get()
		if check_var_gr_3.get() >= 5:
			global save_C5_I
			save_C5_I = entr_C5_I.get()
		if check_var_gr_3.get() >= 6:
			global save_C6_I
			save_C6_I = entr_C6_I.get()
		if check_var_gr_3.get() >= 7:
			global save_C7_I
			save_C7_I = entr_C7_I.get()
		if check_var_gr_3.get() >= 8:
			global save_C8_I
			save_C8_I = entr_C8_I.get()
		if check_var_gr_3.get() >= 9:
			global save_C9_I
			save_C9_I = entr_C9_I.get()
		if check_var_gr_3.get() >= 10:
			global save_C10_I
			save_C10_I = entr_C10_I.get()


	#4 колонка

	if check_var_gr_2.get() >= 4:
		global save_D1_I
		save_D1_I = entr_D1_I.get()
		global save_D2_I
		save_D2_I = entr_D2_I.get()
		global save_D3_I
		save_D3_I = entr_D3_I.get()
		if check_var_gr_3.get() >= 4:
			global save_D4_I
			save_D4_I = entr_D4_I.get()
		if check_var_gr_3.get() >= 5:
			global save_D5_I
			save_D5_I = entr_D5_I.get()
		if check_var_gr_3.get() >= 6:
			global save_D6_I
			save_D6_I = entr_D6_I.get()
		if check_var_gr_3.get() >= 7:
			global save_D7_I
			save_D7_I = entr_D7_I.get()
		if check_var_gr_3.get() >= 8:
			global save_D8_I
			save_D8_I = entr_D8_I.get()
		if check_var_gr_3.get() >= 9:
			global save_D9_I
			save_D9_I = entr_D9_I.get()
		if check_var_gr_3.get() >= 10:
			global save_D10_I
			save_D10_I = entr_D10_I.get()


	#5 колонка

	if check_var_gr_2.get() >= 5:
		global save_E1_I
		save_E1_I = entr_E1_I.get()
		global save_E2_I
		save_E2_I = entr_E2_I.get()
		global save_E3_I
		save_E3_I = entr_E3_I.get()
		if check_var_gr_3.get() >= 4:
			global save_E4_I
			save_E4_I = entr_E4_I.get()
		if check_var_gr_3.get() >= 5:
			global save_E5_I
			save_E5_I = entr_E5_I.get()
		if check_var_gr_3.get() >= 6:
			global save_E6_I
			save_E6_I = entr_E6_I.get()
		if check_var_gr_3.get() >= 7:
			global save_E7_I
			save_E7_I = entr_E7_I.get()
		if check_var_gr_3.get() >= 8:
			global save_E8_I
			save_E8_I = entr_E8_I.get()
		if check_var_gr_3.get() >= 9:
			global save_E9_I
			save_E9_I = entr_E9_I.get()
		if check_var_gr_3.get() >= 10:
			global save_E10_I
			save_E10_I = entr_E10_I.get()

	#6 колонка

	if check_var_gr_2.get() >= 6:
		global save_F1_I
		save_F1_I = entr_F1_I.get()
		global save_F2_I
		save_F2_I = entr_F2_I.get()
		global save_F3_I
		save_F3_I = entr_F3_I.get()
		if check_var_gr_3.get() >= 4:
			global save_F4_I
			save_F4_I = entr_F4_I.get()
		if check_var_gr_3.get() >= 5:
			global save_F5_I
			save_F5_I = entr_F5_I.get()
		if check_var_gr_3.get() >= 6:
			global save_F6_I
			save_F6_I = entr_F6_I.get()
		if check_var_gr_3.get() >= 7:
			global save_F7_I
			save_F7_I = entr_F7_I.get()
		if check_var_gr_3.get() >= 8:
			global save_F8_I
			save_F8_I = entr_F8_I.get()
		if check_var_gr_3.get() >= 9:
			global save_F9_I
			save_F9_I = entr_F9_I.get()
		if check_var_gr_3.get() >= 10:
			global save_F10_I
			save_F10_I = entr_F10_I.get()


	#QCB
		#1 колонка
	if check_var_gr_2.get() >= 2:
		global save_MAIN_II
		save_MAIN_II = entr_MAIN_II.get()
		global save_A1_II
		save_A1_II = entr_A1_II.get()
		global save_A2_II
		save_A2_II = entr_A2_II.get()
		global save_A3_II
		save_A3_II = entr_A3_II.get()
		if check_var_gr_3.get() >= 4:
			global save_A4_II
			save_A4_II = entr_A4_II.get()
		if check_var_gr_3.get() >= 5:
			global save_A5_II
			save_A5_II = entr_A5_II.get()
		if check_var_gr_3.get() >= 6:
			global save_A6_II
			save_A6_II = entr_A6_II.get()
		if check_var_gr_3.get() >= 7:
			global save_A7_II
			save_A7_II = entr_A7_II.get()
		if check_var_gr_3.get() >= 8:
			global save_A8_II
			save_A8_II = entr_A8_II.get()
		if check_var_gr_3.get() >= 9:
			global save_A9_II
			save_A9_II = entr_A9_II.get()
		if check_var_gr_3.get() >= 10:
			global save_A10_II
			save_A10_II = entr_A10_II.get()

		#2 колонка

		if check_var_gr_2.get() >= 2:
			global save_B1_II
			save_B1_II = entr_B1_II.get()
			global save_B2_II
			save_B2_II = entr_B2_II.get()
			global save_B3_II
			save_B3_II = entr_B3_II.get()
			if check_var_gr_3.get() >= 4:
				global save_B4_II
				save_B4_II = entr_B4_II.get()
			if check_var_gr_3.get() >= 5:
				global save_B5_II
				save_B5_II = entr_B5_II.get()
			if check_var_gr_3.get() >= 6:
				global save_B6_II
				save_B6_II = entr_B6_II.get()
			if check_var_gr_3.get() >= 7:
				global save_B7_II
				save_B7_II = entr_B7_II.get()
			if check_var_gr_3.get() >= 8:
				global save_B8_II
				save_B8_II = entr_B8_II.get()
			if check_var_gr_3.get() >= 9:
				global save_B9_II
				save_B9_II = entr_B9_II.get()
			if check_var_gr_3.get() >= 10:
				global save_B10_II
				save_B10_II = entr_B10_II.get()


		#3 колонка

		if check_var_gr_2.get() >= 3:
			global save_C1_II
			save_C1_II = entr_C1_II.get()
			global save_C2_II
			save_C2_II = entr_C2_II.get()
			global save_C3_II
			save_C3_II = entr_C3_II.get()
			if check_var_gr_3.get() >= 4:
				global save_C4_II
				save_C4_II = entr_C4_II.get()
			if check_var_gr_3.get() >= 5:
				global save_C5_II
				save_C5_II = entr_C5_II.get()
			if check_var_gr_3.get() >= 6:
				global save_C6_II
				save_C6_II = entr_C6_II.get()
			if check_var_gr_3.get() >= 7:
				global save_C7_II
				save_C7_II = entr_C7_II.get()
			if check_var_gr_3.get() >= 8:
				global save_C8_II
				save_C8_II = entr_C8_II.get()
			if check_var_gr_3.get() >= 9:
				global save_C9_II
				save_C9_II = entr_C9_II.get()
			if check_var_gr_3.get() >= 10:
				global save_C10_II
				save_C10_II = entr_C10_II.get()


		#4 колонка

		if check_var_gr_2.get() >= 4:
			global save_D1_II
			save_D1_II = entr_D1_II.get()
			global save_D2_II
			save_D2_II = entr_D2_II.get()
			global save_D3_II
			save_D3_II = entr_D3_II.get()
			if check_var_gr_3.get() >= 4:
				global save_D4_II
				save_D4_II = entr_D4_II.get()
			if check_var_gr_3.get() >= 5:
				global save_D5_II
				save_D5_II = entr_D5_II.get()
			if check_var_gr_3.get() >= 6:
				global save_D6_II
				save_D6_II = entr_D6_II.get()
			if check_var_gr_3.get() >= 7:
				global save_D7_II
				save_D7_II = entr_D7_II.get()
			if check_var_gr_3.get() >= 8:
				global save_D8_II
				save_D8_II = entr_D8_II.get()
			if check_var_gr_3.get() >= 9:
				global save_D9_II
				save_D9_II = entr_D9_II.get()
			if check_var_gr_3.get() >= 10:
				global save_D10_II
				save_D10_II = entr_D10_II.get()


		#5 колонка

		if check_var_gr_2.get() >= 5:
			global save_E1_II
			save_E1_II = entr_E1_II.get()
			global save_E2_II
			save_E2_II = entr_E2_II.get()
			global save_E3_II
			save_E3_II = entr_E3_II.get()
			if check_var_gr_3.get() >= 4:
				global save_E4_II
				save_E4_II = entr_E4_II.get()
			if check_var_gr_3.get() >= 5:
				global save_E5_II
				save_E5_II = entr_E5_II.get()
			if check_var_gr_3.get() >= 6:
				global save_E6_II
				save_E6_II = entr_E6_II.get()
			if check_var_gr_3.get() >= 7:
				global save_E7_II
				save_E7_II = entr_E7_II.get()
			if check_var_gr_3.get() >= 8:
				global save_E8_II
				save_E8_II = entr_E8_II.get()
			if check_var_gr_3.get() >= 9:
				global save_E9_II
				save_E9_II = entr_E9_II.get()
			if check_var_gr_3.get() >= 10:
				global save_E10_II
				save_E10_II = entr_E10_II.get()

		#6 колонка

		if check_var_gr_2.get() >= 6:
			global save_F1_II
			save_F1_II = entr_F1_II.get()
			global save_F2_II
			save_F2_II = entr_F2_II.get()
			global save_F3_II
			save_F3_II = entr_F3_II.get()
			if check_var_gr_3.get() >= 4:
				global save_F4_II
				save_F4_II = entr_F4_II.get()
			if check_var_gr_3.get() >= 5:
				global save_F5_II
				save_F5_II = entr_F5_II.get()
			if check_var_gr_3.get() >= 6:
				global save_F6_II
				save_F6_II = entr_F6_II.get()
			if check_var_gr_3.get() >= 7:
				global save_F7_II
				save_F7_II = entr_F7_II.get()
			if check_var_gr_3.get() >= 8:
				global save_F8_II
				save_F8_II = entr_F8_II.get()
			if check_var_gr_3.get() >= 9:
				global save_F9_II
				save_F9_II = entr_F9_II.get()
			if check_var_gr_3.get() >= 10:
				global save_F10_II
				save_F10_II = entr_F10_II.get()


	#QCC
		#1 колонка
	if check_var_gr_2.get() >= 3:
		global save_MAIN_III
		save_MAIN_III = entr_MAIN_III.get()
		global save_A1_III
		save_A1_III = entr_A1_III.get()
		global save_A2_III
		save_A2_III = entr_A2_III.get()
		global save_A3_III
		save_A3_III = entr_A3_III.get()
		if check_var_gr_3.get() >= 4:
			global save_A4_III
			save_A4_III = entr_A4_III.get()
		if check_var_gr_3.get() >= 5:
			global save_A5_III
			save_A5_III = entr_A5_III.get()
		if check_var_gr_3.get() >= 6:
			global save_A6_III
			save_A6_III = entr_A6_III.get()
		if check_var_gr_3.get() >= 7:
			global save_A7_III
			save_A7_III = entr_A7_III.get()
		if check_var_gr_3.get() >= 8:
			global save_A8_III
			save_A8_III = entr_A8_III.get()
		if check_var_gr_3.get() >= 9:
			global save_A9_III
			save_A9_III = entr_A9_III.get()
		if check_var_gr_3.get() >= 10:
			global save_A10_III
			save_A10_III = entr_A10_III.get()

		#2 колонка

		if check_var_gr_2.get() >= 2:
			global save_B1_III
			save_B1_III = entr_B1_III.get()
			global save_B2_III
			save_B2_III = entr_B2_III.get()
			global save_B3_III
			save_B3_III = entr_B3_III.get()
			if check_var_gr_3.get() >= 4:
				global save_B4_III
				save_B4_III = entr_B4_III.get()
			if check_var_gr_3.get() >= 5:
				global save_B5_III
				save_B5_III = entr_B5_III.get()
			if check_var_gr_3.get() >= 6:
				global save_B6_III
				save_B6_III = entr_B6_III.get()
			if check_var_gr_3.get() >= 7:
				global save_B7_III
				save_B7_III = entr_B7_III.get()
			if check_var_gr_3.get() >= 8:
				global save_B8_III
				save_B8_III = entr_B8_III.get()
			if check_var_gr_3.get() >= 9:
				global save_B9_III
				save_B9_III = entr_B9_III.get()
			if check_var_gr_3.get() >= 10:
				global save_B10_III
				save_B10_III = entr_B10_III.get()


		#3 колонка

		if check_var_gr_2.get() >= 3:
			global save_C1_III
			save_C1_III = entr_C1_III.get()
			global save_C2_III
			save_C2_III = entr_C2_III.get()
			global save_C3_III
			save_C3_III = entr_C3_III.get()
			if check_var_gr_3.get() >= 4:
				global save_C4_III
				save_C4_III = entr_C4_III.get()
			if check_var_gr_3.get() >= 5:
				global save_C5_III
				save_C5_III = entr_C5_III.get()
			if check_var_gr_3.get() >= 6:
				global save_C6_III
				save_C6_III = entr_C6_III.get()
			if check_var_gr_3.get() >= 7:
				global save_C7_III
				save_C7_III = entr_C7_III.get()
			if check_var_gr_3.get() >= 8:
				global save_C8_III
				save_C8_III = entr_C8_III.get()
			if check_var_gr_3.get() >= 9:
				global save_C9_III
				save_C9_III = entr_C9_III.get()
			if check_var_gr_3.get() >= 10:
				global save_C10_III
				save_C10_III = entr_C10_III.get()


		#4 колонка

		if check_var_gr_2.get() >= 4:
			global save_D1_III
			save_D1_III = entr_D1_III.get()
			global save_D2_III
			save_D2_III = entr_D2_III.get()
			global save_D3_III
			save_D3_III = entr_D3_III.get()
			if check_var_gr_3.get() >= 4:
				global save_D4_III
				save_D4_III = entr_D4_III.get()
			if check_var_gr_3.get() >= 5:
				global save_D5_III
				save_D5_III = entr_D5_III.get()
			if check_var_gr_3.get() >= 6:
				global save_D6_III
				save_D6_III = entr_D6_III.get()
			if check_var_gr_3.get() >= 7:
				global save_D7_III
				save_D7_III = entr_D7_III.get()
			if check_var_gr_3.get() >= 8:
				global save_D8_III
				save_D8_III = entr_D8_III.get()
			if check_var_gr_3.get() >= 9:
				global save_D9_III
				save_D9_III = entr_D9_III.get()
			if check_var_gr_3.get() >= 10:
				global save_D10_III
				save_D10_III = entr_D10_III.get()


		#5 колонка

		if check_var_gr_2.get() >= 5:
			global save_E1_III
			save_E1_III = entr_E1_III.get()
			global save_E2_III
			save_E2_III = entr_E2_III.get()
			global save_E3_III
			save_E3_III = entr_E3_III.get()
			if check_var_gr_3.get() >= 4:
				global save_E4_III
				save_E4_III = entr_E4_III.get()
			if check_var_gr_3.get() >= 5:
				global save_E5_III
				save_E5_III = entr_E5_III.get()
			if check_var_gr_3.get() >= 6:
				global save_E6_III
				save_E6_III = entr_E6_III.get()
			if check_var_gr_3.get() >= 7:
				global save_E7_III
				save_E7_III = entr_E7_III.get()
			if check_var_gr_3.get() >= 8:
				global save_E8_III
				save_E8_III = entr_E8_III.get()
			if check_var_gr_3.get() >= 9:
				global save_E9_III
				save_E9_III = entr_E9_III.get()
			if check_var_gr_3.get() >= 10:
				global save_E10_III
				save_E10_III = entr_E10_III.get()

		#6 колонка

		if check_var_gr_2.get() >= 6:
			global save_F1_III
			save_F1_III = entr_F1_III.get()
			global save_F2_III
			save_F2_III = entr_F2_III.get()
			global save_F3_III
			save_F3_III = entr_F3_III.get()
			if check_var_gr_3.get() >= 4:
				global save_F4_III
				save_F4_III = entr_F4_III.get()
			if check_var_gr_3.get() >= 5:
				global save_F5_III
				save_F5_III = entr_F5_III.get()
			if check_var_gr_3.get() >= 6:
				global save_F6_III
				save_F6_III = entr_F6_III.get()
			if check_var_gr_3.get() >= 7:
				global save_F7_III
				save_F7_III = entr_F7_III.get()
			if check_var_gr_3.get() >= 8:
				global save_F8_III
				save_F8_III = entr_F8_III.get()
			if check_var_gr_3.get() >= 9:
				global save_F9_III
				save_F9_III = entr_F9_III.get()
			if check_var_gr_3.get() >= 10:
				global save_F10_III
				save_F10_III = entr_F10_III.get()


	#QCD
		#1 колонка
	if check_var_gr_2.get() >= 4:
		global save_MAIN_IV
		save_MAIN_IV = entr_MAIN_IV.get()
		global save_A1_IV
		save_A1_IV = entr_A1_IV.get()
		global save_A2_IV
		save_A2_IV = entr_A2_IV.get()
		global save_A3_IV
		save_A3_IV = entr_A3_IV.get()
		if check_var_gr_3.get() >= 4:
			global save_A4_IV
			save_A4_IV = entr_A4_IV.get()
		if check_var_gr_3.get() >= 5:
			global save_A5_IV
			save_A5_IV = entr_A5_IV.get()
		if check_var_gr_3.get() >= 6:
			global save_A6_IV
			save_A6_IV = entr_A6_IV.get()
		if check_var_gr_3.get() >= 7:
			global save_A7_IV
			save_A7_IV = entr_A7_IV.get()
		if check_var_gr_3.get() >= 8:
			global save_A8_IV
			save_A8_IV = entr_A8_IV.get()
		if check_var_gr_3.get() >= 9:
			global save_A9_IV
			save_A9_IV = entr_A9_IV.get()
		if check_var_gr_3.get() >= 10:
			global save_A10_IV
			save_A10_IV = entr_A10_IV.get()

		#2 колонка

		if check_var_gr_2.get() >= 2:
			global save_B1_IV
			save_B1_IV = entr_B1_IV.get()
			global save_B2_IV
			save_B2_IV = entr_B2_IV.get()
			global save_B3_IV
			save_B3_IV = entr_B3_IV.get()
			if check_var_gr_3.get() >= 4:
				global save_B4_IV
				save_B4_IV = entr_B4_IV.get()
			if check_var_gr_3.get() >= 5:
				global save_B5_IV
				save_B5_IV = entr_B5_IV.get()
			if check_var_gr_3.get() >= 6:
				global save_B6_IV
				save_B6_IV = entr_B6_IV.get()
			if check_var_gr_3.get() >= 7:
				global save_B7_IV
				save_B7_IV = entr_B7_IV.get()
			if check_var_gr_3.get() >= 8:
				global save_B8_IV
				save_B8_IV = entr_B8_IV.get()
			if check_var_gr_3.get() >= 9:
				global save_B9_IV
				save_B9_IV = entr_B9_IV.get()
			if check_var_gr_3.get() >= 10:
				global save_B10_IV
				save_B10_IV = entr_B10_IV.get()


		#3 колонка

		if check_var_gr_2.get() >= 3:
			global save_C1_IV
			save_C1_IV = entr_C1_IV.get()
			global save_C2_IV
			save_C2_IV = entr_C2_IV.get()
			global save_C3_IV
			save_C3_IV = entr_C3_IV.get()
			if check_var_gr_3.get() >= 4:
				global save_C4_IV
				save_C4_IV = entr_C4_IV.get()
			if check_var_gr_3.get() >= 5:
				global save_C5_IV
				save_C5_IV = entr_C5_IV.get()
			if check_var_gr_3.get() >= 6:
				global save_C6_IV
				save_C6_IV = entr_C6_IV.get()
			if check_var_gr_3.get() >= 7:
				global save_C7_IV
				save_C7_IV = entr_C7_IV.get()
			if check_var_gr_3.get() >= 8:
				global save_C8_IV
				save_C8_IV = entr_C8_IV.get()
			if check_var_gr_3.get() >= 9:
				global save_C9_IV
				save_C9_IV = entr_C9_IV.get()
			if check_var_gr_3.get() >= 10:
				global save_C10_IV
				save_C10_IV = entr_C10_IV.get()


		#4 колонка

		if check_var_gr_2.get() >= 4:
			global save_D1_IV
			save_D1_IV = entr_D1_IV.get()
			global save_D2_IV
			save_D2_IV = entr_D2_IV.get()
			global save_D3_IV
			save_D3_IV = entr_D3_IV.get()
			if check_var_gr_3.get() >= 4:
				global save_D4_IV
				save_D4_IV = entr_D4_IV.get()
			if check_var_gr_3.get() >= 5:
				global save_D5_IV
				save_D5_IV = entr_D5_IV.get()
			if check_var_gr_3.get() >= 6:
				global save_D6_IV
				save_D6_IV = entr_D6_IV.get()
			if check_var_gr_3.get() >= 7:
				global save_D7_IV
				save_D7_IV = entr_D7_IV.get()
			if check_var_gr_3.get() >= 8:
				global save_D8_IV
				save_D8_IV = entr_D8_IV.get()
			if check_var_gr_3.get() >= 9:
				global save_D9_IV
				save_D9_IV = entr_D9_IV.get()
			if check_var_gr_3.get() >= 10:
				global save_D10_IV
				save_D10_IV = entr_D10_IV.get()


		#5 колонка

		if check_var_gr_2.get() >= 5:
			global save_E1_IV
			save_E1_IV = entr_E1_IV.get()
			global save_E2_IV
			save_E2_IV = entr_E2_IV.get()
			global save_E3_IV
			save_E3_IV = entr_E3_IV.get()
			if check_var_gr_3.get() >= 4:
				global save_E4_IV
				save_E4_IV = entr_E4_IV.get()
			if check_var_gr_3.get() >= 5:
				global save_E5_IV
				save_E5_IV = entr_E5_IV.get()
			if check_var_gr_3.get() >= 6:
				global save_E6_IV
				save_E6_IV = entr_E6_IV.get()
			if check_var_gr_3.get() >= 7:
				global save_E7_IV
				save_E7_IV = entr_E7_IV.get()
			if check_var_gr_3.get() >= 8:
				global save_E8_IV
				save_E8_IV = entr_E8_IV.get()
			if check_var_gr_3.get() >= 9:
				global save_E9_IV
				save_E9_IV = entr_E9_IV.get()
			if check_var_gr_3.get() >= 10:
				global save_E10_IV
				save_E10_IV = entr_E10_IV.get()

		#6 колонка

		if check_var_gr_2.get() >= 6:
			global save_F1_IV
			save_F1_IV = entr_F1_IV.get()
			global save_F2_IV
			save_F2_IV = entr_F2_IV.get()
			global save_F3_IV
			save_F3_IV = entr_F3_IV.get()
			if check_var_gr_3.get() >= 4:
				global save_F4_IV
				save_F4_IV = entr_F4_IV.get()
			if check_var_gr_3.get() >= 5:
				global save_F5_IV
				save_F5_IV = entr_F5_IV.get()
			if check_var_gr_3.get() >= 6:
				global save_F6_IV
				save_F6_IV = entr_F6_IV.get()
			if check_var_gr_3.get() >= 7:
				global save_F7_IV
				save_F7_IV = entr_F7_IV.get()
			if check_var_gr_3.get() >= 8:
				global save_F8_IV
				save_F8_IV = entr_F8_IV.get()
			if check_var_gr_3.get() >= 9:
				global save_F9_IV
				save_F9_IV = entr_F9_IV.get()
			if check_var_gr_3.get() >= 10:
				global save_F10_IV
				save_F10_IV = entr_F10_IV.get()


	#QCE
		#1 колонка
	if check_var_gr_2.get() >= 5:
		global save_MAIN_V
		save_MAIN_V = entr_MAIN_V.get()
		global save_A1_V
		save_A1_V = entr_A1_V.get()
		global save_A2_V
		save_A2_V = entr_A2_V.get()
		global save_A3_V
		save_A3_V = entr_A3_V.get()
		if check_var_gr_3.get() >= 4:
			global save_A4_V
			save_A4_V = entr_A4_V.get()
		if check_var_gr_3.get() >= 5:
			global save_A5_V
			save_A5_V = entr_A5_V.get()
		if check_var_gr_3.get() >= 6:
			global save_A6_V
			save_A6_V = entr_A6_V.get()
		if check_var_gr_3.get() >= 7:
			global save_A7_V
			save_A7_V = entr_A7_V.get()
		if check_var_gr_3.get() >= 8:
			global save_A8_V
			save_A8_V = entr_A8_V.get()
		if check_var_gr_3.get() >= 9:
			global save_A9_V
			save_A9_V = entr_A9_V.get()
		if check_var_gr_3.get() >= 10:
			global save_A10_V
			save_A10_V = entr_A10_V.get()

		#2 колонка

		if check_var_gr_2.get() >= 2:
			global save_B1_V
			save_B1_V = entr_B1_V.get()
			global save_B2_V
			save_B2_V = entr_B2_V.get()
			global save_B3_V
			save_B3_V = entr_B3_V.get()
			if check_var_gr_3.get() >= 4:
				global save_B4_V
				save_B4_V = entr_B4_V.get()
			if check_var_gr_3.get() >= 5:
				global save_B5_V
				save_B5_V = entr_B5_V.get()
			if check_var_gr_3.get() >= 6:
				global save_B6_V
				save_B6_V = entr_B6_V.get()
			if check_var_gr_3.get() >= 7:
				global save_B7_V
				save_B7_V = entr_B7_V.get()
			if check_var_gr_3.get() >= 8:
				global save_B8_V
				save_B8_V = entr_B8_V.get()
			if check_var_gr_3.get() >= 9:
				global save_B9_V
				save_B9_V = entr_B9_V.get()
			if check_var_gr_3.get() >= 10:
				global save_B10_V
				save_B10_V = entr_B10_V.get()


		#3 колонка
		if check_var_gr_2.get() >= 3:
			global save_C1_V
			save_C1_V = entr_C1_V.get()
			global save_C2_V
			save_C2_V = entr_C2_V.get()
			global save_C3_V
			save_C3_V = entr_C3_V.get()
			if check_var_gr_3.get() >= 4:
				global save_C4_V
				save_C4_V = entr_C4_V.get()
			if check_var_gr_3.get() >= 5:
				global save_C5_V
				save_C5_V = entr_C5_V.get()
			if check_var_gr_3.get() >= 6:
				global save_C6_V
				save_C6_V = entr_C6_V.get()
			if check_var_gr_3.get() >= 7:
				global save_C7_V
				save_C7_V = entr_C7_V.get()
			if check_var_gr_3.get() >= 8:
				global save_C8_V
				save_C8_V = entr_C8_V.get()
			if check_var_gr_3.get() >= 9:
				global save_C9_V
				save_C9_V = entr_C9_V.get()
			if check_var_gr_3.get() >= 10:
				global save_C10_V
				save_C10_V = entr_C10_V.get()

		#4 колонка
		if check_var_gr_2.get() >= 4:
			global save_D1_V
			save_D1_V = entr_D1_V.get()
			global save_D2_V
			save_D2_V = entr_D2_V.get()
			global save_D3_V
			save_D3_V = entr_D3_V.get()
			if check_var_gr_3.get() >= 4:
				global save_D4_V
				save_D4_V = entr_D4_V.get()
			if check_var_gr_3.get() >= 5:
				global save_D5_V
				save_D5_V = entr_D5_V.get()
			if check_var_gr_3.get() >= 6:
				global save_D6_V
				save_D6_V = entr_D6_V.get()
			if check_var_gr_3.get() >= 7:
				global save_D7_V
				save_D7_V = entr_D7_V.get()
			if check_var_gr_3.get() >= 8:
				global save_D8_V
				save_D8_V = entr_D8_V.get()
			if check_var_gr_3.get() >= 9:
				global save_D9_V
				save_D9_V = entr_D9_V.get()
			if check_var_gr_3.get() >= 10:
				global save_D10_V
				save_D10_V = entr_D10_V.get()


		#5 колонка

		if check_var_gr_2.get() >= 5:
			global save_E1_V
			save_E1_V = entr_E1_V.get()
			global save_E2_V
			save_E2_V = entr_E2_V.get()
			global save_E3_V
			save_E3_V = entr_E3_V.get()
			if check_var_gr_3.get() >= 4:
				global save_E4_V
				save_E4_V = entr_E4_V.get()
			if check_var_gr_3.get() >= 5:
				global save_E5_V
				save_E5_V = entr_E5_V.get()
			if check_var_gr_3.get() >= 6:
				global save_E6_V
				save_E6_V = entr_E6_V.get()
			if check_var_gr_3.get() >= 7:
				global save_E7_V
				save_E7_V = entr_E7_V.get()
			if check_var_gr_3.get() >= 8:
				global save_E8_V
				save_E8_V = entr_E8_V.get()
			if check_var_gr_3.get() >= 9:
				global save_E9_V
				save_E9_V = entr_E9_V.get()
			if check_var_gr_3.get() >= 10:
				global save_E10_V
				save_E10_V = entr_E10_V.get()

		#6 колонка

		if check_var_gr_2.get() >= 6:
			global save_F1_V
			save_F1_V = entr_F1_V.get()
			global save_F2_V
			save_F2_V = entr_F2_V.get()
			global save_F3_V
			save_F3_V = entr_F3_V.get()
			if check_var_gr_3.get() >= 4:
				global save_F4_V
				save_F4_V = entr_F4_V.get()
			if check_var_gr_3.get() >= 5:
				global save_F5_V
				save_F5_V = entr_F5_V.get()
			if check_var_gr_3.get() >= 6:
				global save_F6_V
				save_F6_V = entr_F6_V.get()
			if check_var_gr_3.get() >= 7:
				global save_F7_V
				save_F7_V = entr_F7_V.get()
			if check_var_gr_3.get() >= 8:
				global save_F8_V
				save_F8_V = entr_F8_V.get()
			if check_var_gr_3.get() >= 9:
				global save_F9_V
				save_F9_V = entr_F9_V.get()
			if check_var_gr_3.get() >= 10:
				global save_F10_V
				save_F10_V = entr_F10_V.get()


	#QCF
		#1 колонка
	if check_var_gr_2.get() >= 6:
		global save_MAIN_VI
		save_MAIN_VI = entr_MAIN_VI.get()
		global save_A1_VI
		save_A1_VI = entr_A1_VI.get()
		global save_A2_VI
		save_A2_VI = entr_A2_VI.get()
		global save_A3_VI
		save_A3_VI = entr_A3_VI.get()
		if check_var_gr_3.get() >= 4:
			global save_A4_VI
			save_A4_VI = entr_A4_VI.get()
		if check_var_gr_3.get() >= 5:
			global save_A5_VI
			save_A5_VI = entr_A5_VI.get()
		if check_var_gr_3.get() >= 6:
			global save_A6_VI
			save_A6_VI = entr_A6_VI.get()
		if check_var_gr_3.get() >= 7:
			global save_A7_VI
			save_A7_VI = entr_A7_VI.get()
		if check_var_gr_3.get() >= 8:
			global save_A8_VI
			save_A8_VI = entr_A8_VI.get()
		if check_var_gr_3.get() >= 9:
			global save_A9_VI
			save_A9_VI = entr_A9_VI.get()
		if check_var_gr_3.get() >= 10:
			global save_A10_VI
			save_A10_VI = entr_A10_VI.get()

		#2 колонка

		if check_var_gr_2.get() >= 2:
			global save_B1_VI
			save_B1_VI = entr_B1_VI.get()
			global save_B2_VI
			save_B2_VI = entr_B2_VI.get()
			global save_B3_VI
			save_B3_VI = entr_B3_VI.get()
			if check_var_gr_3.get() >= 4:
				global save_B4_VI
				save_B4_VI = entr_B4_VI.get()
			if check_var_gr_3.get() >= 5:
				global save_B5_VI
				save_B5_VI = entr_B5_VI.get()
			if check_var_gr_3.get() >= 6:
				global save_B6_VI
				save_B6_VI = entr_B6_VI.get()
			if check_var_gr_3.get() >= 7:
				global save_B7_VI
				save_B7_VI = entr_B7_VI.get()
			if check_var_gr_3.get() >= 8:
				global save_B8_VI
				save_B8_VI = entr_B8_VI.get()
			if check_var_gr_3.get() >= 9:
				global save_B9_VI
				save_B9_VI = entr_B9_VI.get()
			if check_var_gr_3.get() >= 10:
				global save_B10_VI
				save_B10_VI = entr_B10_VI.get()


		#3 колонка

		if check_var_gr_2.get() >= 3:
			global save_C1_VI
			save_C1_VI = entr_C1_VI.get()
			global save_C2_VI
			save_C2_VI = entr_C2_VI.get()
			global save_C3_VI
			save_C3_VI = entr_C3_VI.get()
			if check_var_gr_3.get() >= 4:
				global save_C4_VI
				save_C4_VI = entr_C4_VI.get()
			if check_var_gr_3.get() >= 5:
				global save_C5_VI
				save_C5_VI = entr_C5_VI.get()
			if check_var_gr_3.get() >= 6:
				global save_C6_VI
				save_C6_VI = entr_C6_VI.get()
			if check_var_gr_3.get() >= 7:
				global save_C7_VI
				save_C7_VI = entr_C7_VI.get()
			if check_var_gr_3.get() >= 8:
				global save_C8_VI
				save_C8_VI = entr_C8_VI.get()
			if check_var_gr_3.get() >= 9:
				global save_C9_VI
				save_C9_VI = entr_C9_VI.get()
			if check_var_gr_3.get() >= 10:
				global save_C10_VI
				save_C10_VI = entr_C10_VI.get()


		#4 колонка

		if check_var_gr_2.get() >= 4:
			global save_D1_VI
			save_D1_VI = entr_D1_VI.get()
			global save_D2_VI
			save_D2_VI = entr_D2_VI.get()
			global save_D3_VI
			save_D3_VI = entr_D3_VI.get()
			if check_var_gr_3.get() >= 4:
				global save_D4_VI
				save_D4_VI = entr_D4_VI.get()
			if check_var_gr_3.get() >= 5:
				global save_D5_VI
				save_D5_VI = entr_D5_VI.get()
			if check_var_gr_3.get() >= 6:
				global save_D6_VI
				save_D6_VI = entr_D6_VI.get()
			if check_var_gr_3.get() >= 7:
				global save_D7_VI
				save_D7_VI = entr_D7_VI.get()
			if check_var_gr_3.get() >= 8:
				global save_D8_VI
				save_D8_VI = entr_D8_VI.get()
			if check_var_gr_3.get() >= 9:
				global save_D9_VI
				save_D9_VI = entr_D9_VI.get()
			if check_var_gr_3.get() >= 10:
				global save_D10_VI
				save_D10_VI = entr_D10_VI.get()


		#5 колонка

		if check_var_gr_2.get() >= 5:
			global save_E1_VI
			save_E1_VI = entr_E1_VI.get()
			global save_E2_VI
			save_E2_VI = entr_E2_VI.get()
			global save_E3_VI
			save_E3_VI = entr_E3_VI.get()
			if check_var_gr_3.get() >= 4:
				global save_E4_VI
				save_E4_VI = entr_E4_VI.get()
			if check_var_gr_3.get() >= 5:
				global save_E5_VI
				save_E5_VI = entr_E5_VI.get()
			if check_var_gr_3.get() >= 6:
				global save_E6_VI
				save_E6_VI = entr_E6_VI.get()
			if check_var_gr_3.get() >= 7:
				global save_E7_VI
				save_E7_VI = entr_E7_VI.get()
			if check_var_gr_3.get() >= 8:
				global save_E8_VI
				save_E8_VI = entr_E8_VI.get()
			if check_var_gr_3.get() >= 9:
				global save_E9_VI
				save_E9_VI = entr_E9_VI.get()
			if check_var_gr_3.get() >= 10:
				global save_E10_VI
				save_E10_VI = entr_E10_VI.get()

		#6 колонка

		if check_var_gr_2.get() >= 6:
			global save_F1_VI
			save_F1_VI = entr_F1_VI.get()
			global save_F2_VI
			save_F2_VI = entr_F2_VI.get()
			global save_F3_VI
			save_F3_VI = entr_F3_VI.get()
			if check_var_gr_3.get() >= 4:
				global save_F4_VI
				save_F4_VI = entr_F4_VI.get()
			if check_var_gr_3.get() >= 5:
				global save_F5_VI
				save_F5_VI = entr_F5_VI.get()
			if check_var_gr_3.get() >= 6:
				global save_F6_VI
				save_F6_VI = entr_F6_VI.get()
			if check_var_gr_3.get() >= 7:
				global save_F7_VI
				save_F7_VI = entr_F7_VI.get()
			if check_var_gr_3.get() >= 8:
				global save_F8_VI
				save_F8_VI = entr_F8_VI.get()
			if check_var_gr_3.get() >= 9:
				global save_F9_VI
				save_F9_VI = entr_F9_VI.get()
			if check_var_gr_3.get() >= 10:
				global save_F10_VI
				save_F10_VI = entr_F10_VI.get()




	check_var_gr_1.set(1)
	check_var_gr_2.set(1)
	check_var_gr_3.set(3)
	root.destroy()
	first_widow()



def first_widow():
	#head prog
	global root_open
	root_open = tk.Tk()
	root_open.title('AnovaD')
	root_open.geometry('400x250+500+200')
	#root_open.configure(bg = 'white')
	root_open.resizable(False,False)

	btn_spravka = tk.Button(root_open, text = '?', command = spravka_open, width = 2, bg = 'pink')
	btn_spravka.place(x = 370, y = 10)



	global check_var_gr_1
	check_var_gr_1 = tk.IntVar()
	check_var_gr_1.set(1)
	check1_1 = tk.Radiobutton(root_open,text = '1', variable = check_var_gr_1, value = 1)
	check1_1.place(x = 25, y = 30)
	check2_1 = tk.Radiobutton(root_open,text = '2', variable = check_var_gr_1, value = 2)
	check2_1.place(x = 25, y = 50)
	check3_1 = tk.Radiobutton(root_open,text = '3', variable = check_var_gr_1, value = 3)
	check3_1.place(x = 25, y = 70)
	check4_1 = tk.Radiobutton(root_open,text = '4', variable = check_var_gr_1, value = 4)
	check4_1.place(x = 25, y = 90)
	check5_1 = tk.Radiobutton(root_open,text = '5', variable = check_var_gr_1, value = 5)
	check5_1.place(x = 25, y = 110)
	check6_1 = tk.Radiobutton(root_open,text = '6', variable = check_var_gr_1, value = 6)
	check6_1.place(x = 25, y = 130)

	labl_start_1 = tk.Label(text = 'QC')
	labl_start_1.place(x = 30, y = 10)


	global check_var_gr_2
	check_var_gr_2 = tk.IntVar()
	check_var_gr_2.set(1)
	check1_2 = tk.Radiobutton(root_open,text = '1', variable = check_var_gr_2, value = 1)
	check1_2.place(x = 80, y = 30)
	check2_2 = tk.Radiobutton(root_open,text = '2', variable = check_var_gr_2, value = 2)
	check2_2.place(x = 80, y = 50)
	check3_2 = tk.Radiobutton(root_open,text = '3', variable = check_var_gr_2, value = 3)
	check3_2.place(x = 80, y = 70)
	check4_2 = tk.Radiobutton(root_open,text = '4', variable = check_var_gr_2, value = 4)
	check4_2.place(x = 80, y = 90)
	check5_2 = tk.Radiobutton(root_open,text = '5', variable = check_var_gr_2, value = 5)
	check5_2.place(x = 80, y = 110)
	check6_2 = tk.Radiobutton(root_open,text = '6', variable = check_var_gr_2, value = 6)
	check6_2.place(x = 80, y = 130)

	labl_start_2 = tk.Label(text = 'n групп')
	labl_start_2.place(x = 73, y = 10)


	global check_var_gr_3
	check_var_gr_3 = tk.IntVar()
	check_var_gr_3.set(3)
	check1_3 = tk.Radiobutton(root_open,text = '3', variable = check_var_gr_3, value = 3)
	check1_3.place(x = 135, y = 30)
	check2_3 = tk.Radiobutton(root_open,text = '4', variable = check_var_gr_3, value = 4)
	check2_3.place(x = 135, y = 50)
	check3_3 = tk.Radiobutton(root_open,text = '5', variable = check_var_gr_3, value = 5)
	check3_3.place(x = 135, y = 70)
	check4_3 = tk.Radiobutton(root_open,text = '6', variable = check_var_gr_3, value = 6)
	check4_3.place(x = 135, y = 90)
	check5_3 = tk.Radiobutton(root_open,text = '7', variable = check_var_gr_3, value = 7)
	check5_3.place(x = 135, y = 110)
	check6_3 = tk.Radiobutton(root_open,text = '8', variable = check_var_gr_3, value = 8)
	check6_3.place(x = 135, y = 130)
	check7_3 = tk.Radiobutton(root_open,text = '9', variable = check_var_gr_3, value = 9)
	check7_3.place(x = 135, y = 150)
	check8_3 = tk.Radiobutton(root_open,text = '10', variable = check_var_gr_3, value = 10)
	check8_3.place(x = 135, y = 170)


	labl_start_3 = tk.Label(text = 'n повт.')
	labl_start_3.place(x = 131, y = 10)




	labl_info_1 = tk.Label(text = 'Для начала работы\nвыберите настраиваемые\nпараметры\nлибо нажмите кнопку\n "стандарт".\n\nЕсли остались вопросы\nнажмите кнопку\n"?".')
	labl_info_1.place(x = 200, y = 30)



	butt_check = tk.Button(root_open, text = 'Начать работу\nпо выбранному', command = start, bg = '#c9bee8')
	butt_check.place(x = 40, y = 200)

	butt_check_1 = tk.Button(root_open, text = 'Начать работу\n"стандарт" 4x3x5', command = start_std, bg = '#c9bee8')
	butt_check_1.place(x = 225, y = 200)





	def peremeshenie_toplevel(event):
		x = root_open.winfo_x() + 7
		y = root_open.winfo_y() + 283
		toplevel_spravka.geometry("+%d+%d" % (x,y))

		#toplevel_spravka.geometry(f"401x300+{(root_open.winfo_x())+7}+{(root_open.winfo_y()+283)}")

	root_open.bind("<Configure>", peremeshenie_toplevel)

	global toplevel_spravka
	toplevel_spravka = tk.Toplevel()
	toplevel_spravka.title('Справка')
	toplevel_spravka.resizable(False, False)
	toplevel_spravka.withdraw()
	toplevel_spravka.overrideredirect(True)
	spravka_text = tk.Text(toplevel_spravka, width = 43, height = 19)
	spravka_text.place(x = 1, y = 1)
	btn_spravka_close = tk.Button(toplevel_spravka, text = 'close', command = spravka_close, bg = 'pink')
	btn_spravka_close.place(x = 355, y = 15)




	##########################################
	#текстовая часть справки 
	listbox_insert_text = '''
1. Зачем нужна программа?
  Программа "AnovaD" автоматизирует  
получение данных о внутригрупповой и
межгрупповойпрецизионности методом
однофакторного дисперсионного анализа,
а также сопутствующих показателей. 
  Программа работает с Excel-файлами
расширения ".xlsx". 

2. Как работать с программой?
Для начала работы выберите параметры на
начальном окне. 
В зависимости от вашего выбора программ
создаст окна ввода для ваших  данных для
расчета. Можно начать работу как по 
выбранным параметрам, так и по "стандарту"
принятом в лаборатории. 
Будьте внимательны, если после выбора
параметов, при вводе данных вы обнаружили,
что неправильно сделали выбор, возврат
на предыдущий экран возможен и вам
не придется делать перезапуск ПО и 
заполнять все заного. 
После выбора параметров программа
предложит вам в зависимости от выбора
форму для заполнения. Заполнять нужно
СТРОГО все ячейки  которые будут
участвовать в расчетах: ячейки QCA, QCB и
тд, ячейки под наименованиями Обр.1, Обр.2
и тд.
Также укажите вашу концентрацию, например
"ng/mL", если не указать программа все
равно расчитает, но поле "Concentation"
будет  заполнено без концентрации,
придется   дописывать вручную. пропуски
недопустимы,  программа не сохранит
результат. 
Ячейки норм QCA, QCB, QCC, QCD и тд,
заполняются по необходимости, если оставить
эти ячейки пустыми, то они примут значения
по-умолчанию, QCA - 20 и все остальные по
15 для всех таблиц.

После ввода данных результат расчета 
необходимо сохранить в файл. Для этого
нажмите кнопку, которая находится в 
правом верхнем углу программы
"open file to save".
После нажатия клавиши - будет открыто
меню проводника. 
С помощью данного меню выберите уже
существующий файл, в который будет
произведено сохранение, либо создайте
новый, путем нажатия ПКМ(правой клавиши
мыши), на свободном месте внутри
директории. После нажатия на ПКМ, 
появится окно выбора.
Выберите пункт --> создать -->
--> выбираем "Лист Microsoft Excel"
(Убедитесь, что создается .xlsx файл),
после выбора данного пункта, проводник
предложит вам ввести имя файла -->
--> вводим имя файла --> после ввода
имени файла --> нажмите на
клавишу клавиатуры "Enter" --> затем 
выберите ЛКМ(левой кнопкой мыши)
только что созданный файл и нажмите
на кнопку окна проводника "открыть".
Теперь файл открыт внутри программы
(т.е. визуально он не отобразится
для пользователя, но программа будет
понимать, в какой файл идет сохранение)
Теперь нажмите кнопку "save", для
сохранения расчетов в файл.

Также программа поддерживает автозаполнение
ячеек Кнопка "input" - свойственна только 
для окна "стандарт", при нажатии на нее
откроется окно проводника, в котором нужно
выбрать excel файл в котором будут
содержаться входные данные для
автозаполнения.
ПРОГРАММА ПОДДЕРЖИВАЕТ ТОЛЬКО xlsx файлы на
вход и выход, будьте внимательны! 
Кнопка "+" рядом с кнопкой "input" также
открывает меню проводника в котором нужно
выбрать пустой ".xlsx" файл, данная кнопка в
этом файле выделит те ячейки, которые
необходимо заполнить для "стандарта", чтобы
произвести автозаполнение из Excel-файла.
По точно такому же алгоритму заполняется
Excel файл на вход для любого другого
размера программы, соответственно:
QC - заполняется в левом верхнем углу в
столбце A, далее идет отступ на 1 строку
вниз и на один столбец вправо. 
Соответственно заполнение значений QCA1,
QCA2 и тд начинается с ячейки "C2" и 
смещается влево(в зависимости от количества
QC) и вниз(в зависисмости от n), после
заполнения необходимого количества
показателей QCA, идет смещение на одну
строку вниз и начинается заполнения
показателя QCB. 
Программа для заполнения ориентируется по
концентрации, которая указана в столбце "A",
поэтому ее правильное заполнение
обязательно.
Кнопка "open file to save" позволяет
выбрать необходимый файл .xlsx для 
проведения в него расчетов. 
Кнопка "save" - проводит расчеты и
сохраняет данные в выбранном файле.

Связь с разработчиком
email - daniil.popkov@gmail.com
WhatsApp - +7 985 187 81-24
	'''

	for i in listbox_insert_text:
		spravka_text.insert(tk.END, str(i))

	root_open.mainloop()




#запуск программы
first_widow()


















